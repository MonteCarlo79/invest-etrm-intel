"""
Monthly installed capacity scanner.

Scans data/market-fundamentals/各省份装机数据/ for province-level Excel files,
parses wind / solar / thermal / hydro / nuclear / BESS capacity (MW) per month,
and upserts into marketdata.province_installed_monthly.

Supported Excel layouts:
  - Wide format  : rows = months, columns = energy types
  - Long format  : rows = (month, type), separate column for type name

Run manually:
    python scripts/scan_installed_capacity.py

Run with dry-run (no DB writes):
    python scripts/scan_installed_capacity.py --dry-run

Scheduled monthly via setup_monthly_capacity_scan.ps1.
"""
from __future__ import annotations

import argparse
import datetime as dt
import io
import logging
import re
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Optional

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

_REPO = Path(__file__).resolve().parents[1]
_DATA_DIR = _REPO / "data" / "market-fundamentals" / "各省份装机数据"

# ── Province name extraction ─────────────────────────────────────────────────

# Ordered longest-first so "河北南网" matches before "河北"
_PROVINCE_PATTERNS: list[tuple[str, str]] = [
    ("河北南网", "冀南"),
    ("冀北",   "冀北"),
    ("黑龙江",  "黑龙江"),
    ("内蒙古东", "蒙东"),
    ("内蒙古西", "蒙西"),
    ("蒙东",   "蒙东"),
    ("蒙西",   "蒙西"),
    ("北京",   "北京"),
    ("天津",   "天津"),
    ("山西",   "山西"),
    ("山东",   "山东"),
    ("辽宁",   "辽宁"),
    ("吉林",   "吉林"),
    ("上海",   "上海"),
    ("江苏",   "江苏"),
    ("浙江",   "浙江"),
    ("安徽",   "安徽"),
    ("福建",   "福建"),
    ("江西",   "江西"),
    ("河南",   "河南"),
    ("湖北",   "湖北"),
    ("湖南",   "湖南"),
    ("广东",   "广东"),
    ("广西",   "广西"),
    ("海南",   "海南"),
    ("重庆",   "重庆"),
    ("四川",   "四川"),
    ("贵州",   "贵州"),
    ("云南",   "云南"),
    ("陕西",   "陕西"),
    ("甘肃",   "甘肃"),
    ("青海",   "青海"),
    ("宁夏",   "宁夏"),
    ("新疆",   "新疆"),
]


def _province_from_filename(name: str) -> Optional[str]:
    for pattern, canonical in _PROVINCE_PATTERNS:
        if pattern in name:
            return canonical
    return None


# ── Unit detection ────────────────────────────────────────────────────────────

def _detect_unit_factor(header_text: str) -> float:
    """
    Return the multiplier to convert source values → MW.

    万kW  / 万千瓦 : 1万kW = 10,000 kW = 10 MW  →  ×10
    千kW  / 千千瓦 : 1千kW = 1,000  kW =  1 MW  →  ×1
    MW             :                               →  ×1
    """
    t = header_text.lower()
    if "万kw" in t or "万千瓦" in t or "万kw" in t:
        return 10.0
    if "千kw" in t or "千千瓦" in t:
        return 1.0
    if "mw" in t:
        return 1.0
    # Default: in Chinese grid data 万kW is by far the most common unit.
    return 10.0


# ── Column keyword → fuel type mapping ───────────────────────────────────────

_FUEL_KEYWORDS: list[tuple[list[str], str]] = [
    (["储能", "电化学"],      "bess"),
    (["核电", "核"],          "nuclear"),
    (["水电", "水力", "水"],   "hydro"),
    (["太阳能", "光伏", "光"], "solar"),
    (["风电", "风力", "风"],  "wind"),
    # thermal last — broad match; must come after the others
    (["火电", "燃煤", "煤电", "气电", "油电", "燃气", "火力"], "thermal"),
]

_ROW_TYPE_KEYWORDS: list[tuple[list[str], str]] = _FUEL_KEYWORDS + [
    (["总计", "合计", "全网", "全省"],  "total"),
]


def _classify_col(text: str) -> Optional[str]:
    # "储能容量" means energy in MWh (kWh), NOT power in MW — skip it.
    # We only want "储能装机" (installed power capacity).
    if "储能" in text and "容量" in text and "装机" not in text:
        return None
    for kwds, fuel in _FUEL_KEYWORDS:
        for kw in kwds:
            if kw in text:
                return fuel
    return None


def _classify_row_type(text: str) -> Optional[str]:
    for kwds, fuel in _ROW_TYPE_KEYWORDS:
        for kw in kwds:
            if kw in text:
                return fuel
    return None


# ── Date parsing ──────────────────────────────────────────────────────────────

_RE_YYYY_MM = re.compile(r"(\d{4})[-年/](\d{1,2})")
_RE_N_MONTH = re.compile(r"^(\d{1,2})[月]")   # "1月" "12月"


def _parse_year_month(cell_val, context_year: Optional[int] = None) -> Optional[dt.date]:
    """Parse a cell value to the first day of its month."""
    if isinstance(cell_val, dt.datetime):
        return cell_val.date().replace(day=1)
    if isinstance(cell_val, dt.date):
        return cell_val.replace(day=1)

    s = str(cell_val).strip()
    m = _RE_YYYY_MM.search(s)
    if m:
        return dt.date(int(m.group(1)), int(m.group(2)), 1)

    # "1月"..."12月" — needs context year
    m2 = _RE_N_MONTH.match(s)
    if m2 and context_year:
        return dt.date(context_year, int(m2.group(1)), 1)

    return None


# ── Sheet selection ───────────────────────────────────────────────────────────

def _best_sheet(wb):
    """Return the worksheet most likely to contain monthly capacity data."""
    # Ordered from most specific to least specific
    cap_keywords = [
        "装机发电", "装机与发电", "装机结构",
        "装机数据", "装机明细", "装机_",
        "装机",
    ]
    for kw in cap_keywords:
        for name in wb.sheetnames:
            if kw in name:
                return wb[name]
    return wb.active


# ── Wide-format parser ────────────────────────────────────────────────────────

def _parse_wide(ws, province: str, source_file: str) -> list[dict]:
    """
    Wide format: rows = months, columns = energy-type capacity values.
    Scans rows 0-8 for the header row (contains fuel-type keywords).
    """
    rows_cache: list[tuple] = []
    for row in ws.iter_rows(values_only=True):
        rows_cache.append(row)

    if not rows_cache:
        return []

    n_cols = max(len(r) for r in rows_cache)

    # ── Find header row ───────────────────────────────────────────────────────
    header_row_idx: Optional[int] = None
    col_fuel: dict[int, str] = {}   # col_index → fuel type
    unit_factor = 10.0              # default: 万kW

    for ri, row in enumerate(rows_cache[:10]):
        found: dict[int, str] = {}
        unit_text = ""
        for ci, val in enumerate(row):
            if val is None:
                continue
            s = str(val)
            unit_text += s  # accumulate for unit detection
            fuel = _classify_col(s)
            if fuel:
                found[ci] = fuel
        if len(found) >= 2:   # need at least 2 energy types to call it a header
            header_row_idx = ri
            col_fuel = found
            unit_factor = _detect_unit_factor(unit_text)
            break

    # Also check the row above header for unit info
    if header_row_idx is not None and header_row_idx > 0:
        prev = rows_cache[header_row_idx - 1]
        unit_text2 = " ".join(str(v) for v in prev if v is not None)
        if unit_text2.strip():
            unit_factor = _detect_unit_factor(unit_text2)

    if header_row_idx is None or not col_fuel:
        return []

    # ── Check for year+month split date columns (e.g. Jilin, Hunan, Gansu) ──
    year_col:  Optional[int] = None
    month_col: Optional[int] = None
    for ci, val in enumerate(rows_cache[header_row_idx]):
        if val is None:
            continue
        s = str(val)
        if s in ("年份", "年") or (s == "年" and ci not in col_fuel):
            year_col = ci
        elif s in ("月份", "月") or (s == "月" and ci not in col_fuel):
            month_col = ci

    # ── Find date column ──────────────────────────────────────────────────────
    date_col: Optional[int] = None
    if year_col is None or month_col is None:
        # Standard: single column with YYYY-MM formatted dates
        for ri in range(header_row_idx + 1, min(header_row_idx + 5, len(rows_cache))):
            for ci, val in enumerate(rows_cache[ri]):
                if ci in col_fuel:
                    continue
                d = _parse_year_month(val)
                if d is not None:
                    date_col = ci
                    break
            if date_col is not None:
                break

        if date_col is None:
            date_col = 0

    # ── Determine context year for N月 format ────────────────────────────────
    context_year: Optional[int] = None
    if year_col is not None:
        # Extract from year column directly
        for row in rows_cache[header_row_idx + 1:]:
            yv = row[year_col] if year_col < len(row) else None
            if yv is not None:
                try:
                    yr = int(str(yv).strip())
                    if 2000 <= yr <= 2099:
                        context_year = yr
                        break
                except (ValueError, TypeError):
                    pass
    elif date_col is not None:
        for row in rows_cache[header_row_idx + 1:]:
            val = row[date_col] if date_col < len(row) else None
            d = _parse_year_month(val)
            if d:
                context_year = d.year
                break

    # ── Extract data rows ─────────────────────────────────────────────────────
    records: list[dict] = []
    current_year = context_year

    for row in rows_cache[header_row_idx + 1:]:
        if not any(row):
            continue

        # Build date from either split year/month cols or a single date col
        d: Optional[dt.date] = None
        if year_col is not None and month_col is not None:
            yv = row[year_col] if year_col < len(row) else None
            mv = row[month_col] if month_col < len(row) else None
            if yv is not None and mv is not None:
                try:
                    yr = int(str(yv).strip())
                    mo_s = str(mv).strip()
                    mo_m = re.match(r"(\d{1,2})", mo_s)
                    if mo_m and 2000 <= yr <= 2099:
                        d = dt.date(yr, int(mo_m.group(1)), 1)
                        current_year = yr
                except (ValueError, TypeError):
                    pass
        else:
            date_val = row[date_col] if date_col < len(row) else None
            d = _parse_year_month(date_val, current_year)
            if d is None:
                continue
            if isinstance(date_val, (dt.datetime, dt.date)):
                current_year = d.year
            elif isinstance(date_val, str) and _RE_YYYY_MM.search(date_val):
                current_year = d.year

        if d is None:
            continue

        rec: dict = {
            "province":    province,
            "year_month":  d,
            "source_file": source_file,
        }
        for ci, fuel in col_fuel.items():
            if ci < len(row) and row[ci] is not None:
                try:
                    v = float(str(row[ci]).replace(",", "")) * unit_factor
                    # Don't overwrite if multiple columns map to same fuel; keep larger
                    if fuel not in rec or v > rec[fuel]:
                        rec[fuel] = round(v, 2)
                except (TypeError, ValueError):
                    pass

        if any(f in rec for f in ("wind", "solar", "thermal", "hydro", "nuclear", "bess")):
            records.append(rec)

    # ── Auto-detect kW units (no unit in header, values in raw kW) ──────────
    # China's largest provinces can reach ~200 GW = 20,000 万kW = 200,000 MW.
    # Tianjin's raw-kW values multiply to ~190,000,000 MW — clearly implausible.
    # Threshold: if any value > 1,000,000 MW (1 TW) after unit_factor, values are kW.
    if records and unit_factor == 10.0:
        max_val = max(
            v for rec in records
            for k, v in rec.items()
            if k in ("wind", "solar", "thermal", "hydro", "nuclear", "bess") and isinstance(v, (int, float))
        ) if records else 0
        if max_val > 1_000_000:
            # Re-scale: was ×10 (万kW→MW), should be ×0.001 (kW→MW)
            rescale = 0.001 / 10.0  # = 0.0001
            for rec in records:
                for k in ("wind", "solar", "thermal", "hydro", "nuclear", "bess"):
                    if k in rec:
                        rec[k] = round(rec[k] * rescale, 2)

    return records


# ── Long-format parser ────────────────────────────────────────────────────────

def _parse_long(ws, province: str, source_file: str) -> list[dict]:
    """
    Long format: rows = (date, ..., type_label, capacity_value, ...).
    Groups rows by date and accumulates capacity by fuel type.
    """
    rows_cache: list[tuple] = []
    for row in ws.iter_rows(values_only=True):
        rows_cache.append(row)

    if not rows_cache:
        return []

    # ── Detect header + capacity column + type column ─────────────────────────
    header_row_idx: Optional[int] = None
    type_col:     Optional[int] = None
    cap_col:      Optional[int] = None
    date_col:     Optional[int] = None
    unit_factor   = 10.0

    unit_col: Optional[int] = None    # Guizhou-style: per-row unit column
    metric_col: Optional[int] = None  # Guizhou-style: filter column (指标组)

    # ── Pre-scan: Guizhou-style (header has 指标组 + 数值 + 单位 columns) ─────
    for ri, row in enumerate(rows_cache[:3]):
        tm = {ci: str(val) for ci, val in enumerate(row) if val is not None}
        vals = list(tm.values())
        if "指标组" in vals and "数值" in vals:
            header_row_idx = ri
            for ci, txt in tm.items():
                if type_col is None and txt in ("类别", "类"):
                    type_col = ci
                elif type_col is None and "类" in txt and "月" not in txt:
                    type_col = ci
                if cap_col is None and txt == "数值":
                    cap_col = ci
                if unit_col is None and txt == "单位":
                    unit_col = ci
                if metric_col is None and txt == "指标组":
                    metric_col = ci
                if date_col is None and "月份" in txt:
                    date_col = ci
            break

    for ri, row in enumerate(rows_cache[:10]):
        if header_row_idx is not None:
            break  # already found via pre-scan
        text_map: dict[int, str] = {}
        for ci, val in enumerate(row):
            if val is not None:
                text_map[ci] = str(val)
        # Skip rows that look like merged title cells (≤2 non-None values)
        if len(text_map) <= 2:
            # Still extract unit info from title rows (e.g. "单位：万千瓦")
            full_title = " ".join(text_map.values())
            if "万千瓦" in full_title or "万kW" in full_title.lower():
                unit_factor = _detect_unit_factor(full_title)
            continue
        full_text = " ".join(text_map.values())
        # Look for a header row that references capacity keywords
        # Includes "容量" for Anhui-style "期末发电设备容量" headers
        if "装机" in full_text or "容量" in full_text or "installed" in full_text.lower():
            header_row_idx = ri
            unit_factor = _detect_unit_factor(full_text)
            # Identify columns — don't overwrite once found; first match wins
            for ci, txt in text_map.items():
                if type_col is None and any(kw in txt for kw in ["指标", "类型", "能源类型", "发电类型", "类别"]):
                    type_col = ci
                # cap_col: match capacity column; exclude "供热" (heat supply) columns
                if cap_col is None and "供热" not in txt:
                    # Prefer "装机" + unit indicator; also accept "设备容量" + unit
                    is_cap = ("装机" in txt and (
                        "容量" in txt or "kW" in txt.lower() or "千瓦" in txt
                        or "末月" in txt or "月末" in txt
                    ))
                    is_equip_cap = ("设备容量" in txt and "千瓦" in txt)
                    if is_cap or is_equip_cap:
                        cap_col = ci
                # date_col: "年月" preferred over plain "月份"
                if txt == "年月":
                    date_col = ci   # highest-priority date column
                elif date_col is None and any(kw in txt for kw in ["月份", "日期", "时间"]):
                    date_col = ci
                # Guizhou-style: per-row unit and metric filter columns
                if unit_col is None and txt == "单位":
                    unit_col = ci
                if metric_col is None and txt == "指标组":
                    metric_col = ci
            break

    if header_row_idx is None:
        return []

    # Fallback column positions if header parsing was incomplete
    if date_col is None:
        date_col = 0
    if type_col is None:
        # Scan early data rows to find a column containing energy-type keywords
        for row in rows_cache[max(0, header_row_idx or 0): (header_row_idx or 0) + 8]:
            for ci, val in enumerate(row):
                if val and ci != date_col and _classify_row_type(str(val)):
                    type_col = ci
                    break
            if type_col is not None:
                break
        if type_col is None:
            type_col = 1   # last resort
    if cap_col is None:
        # Find first numeric-looking column after type_col
        for row in rows_cache[(header_row_idx or 0) + 1: (header_row_idx or 0) + 5]:
            for ci in range(type_col + 1, len(row)):
                if row[ci] is not None:
                    try:
                        float(str(row[ci]).replace(",", ""))
                        cap_col = ci
                        break
                    except (TypeError, ValueError):
                        pass
            if cap_col is not None:
                break
        if cap_col is None:
            cap_col = type_col + 1

    # ── Extract data rows ─────────────────────────────────────────────────────
    by_date: dict[dt.date, dict] = {}

    # Bootstrap context_year by scanning all columns for any YYYY-MM date
    context_year: Optional[int] = None
    for row in rows_cache[(header_row_idx or 0) + 1:]:
        for ci in range(min(5, len(row))):
            d = _parse_year_month(row[ci])  # no context_year
            if d and d.year >= 2020:
                context_year = d.year
                break
        if context_year:
            break

    # Fallback: extract year from filename (e.g. "江苏-装机数据_2026年5月.xlsx")
    if context_year is None:
        m_fn = _RE_YYYY_MM.search(source_file)
        if m_fn:
            context_year = int(m_fn.group(1))

    last_date: Optional[dt.date] = None  # cascade for Yunnan-style None date rows

    for row in rows_cache[(header_row_idx or 0) + 1:]:
        if not any(row):
            continue

        # Guizhou-style: if metric_col present, skip rows that aren't capacity rows
        if metric_col is not None and metric_col < len(row):
            metric_val = str(row[metric_col] or "")
            if "装机" not in metric_val and metric_val.strip():
                continue

        # Date: first try date_col, then scan row for date-like values
        date_val = row[date_col] if date_col < len(row) else None
        d = _parse_year_month(date_val, context_year)
        if d is None:
            # Scan other early cols
            for ci in range(min(4, len(row))):
                d = _parse_year_month(row[ci], context_year)
                if d:
                    break
        if d is None:
            # Yunnan-style: date only on aggregate row; fuel rows have None date
            d = last_date
        if d is None:
            continue
        if d.year >= 2000:
            context_year = d.year
            last_date = d

        type_val = row[type_col] if type_col < len(row) else None
        if type_val is None:
            continue
        fuel = _classify_row_type(str(type_val))
        if fuel is None:
            continue

        cap_val = row[cap_col] if cap_col < len(row) else None
        if cap_val is None:
            continue
        try:
            # Guizhou-style: unit may be per-row (万千瓦, 万kWh, etc.)
            if unit_col is not None and unit_col < len(row) and row[unit_col]:
                row_unit = _detect_unit_factor(str(row[unit_col]))
            else:
                row_unit = unit_factor
            cap_mw = float(str(cap_val).replace(",", "")) * row_unit
        except (TypeError, ValueError):
            continue

        if d not in by_date:
            by_date[d] = {"province": province, "year_month": d, "source_file": source_file}
        # Don't overwrite with lower value (e.g. sub-category vs aggregate)
        if fuel not in by_date[d] or cap_mw > by_date[d][fuel]:
            by_date[d][fuel] = round(cap_mw, 2)

    # Remove "total" key before returning
    records = []
    for rec in by_date.values():
        rec.pop("total", None)
        if any(f in rec for f in ("wind", "solar", "thermal", "hydro", "nuclear", "bess")):
            records.append(rec)
    return records


# ── Format detection ──────────────────────────────────────────────────────────

def _detect_format(ws) -> str:
    """Return 'wide' or 'long'."""
    rows: list[tuple] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(row)
        if i >= 12:
            break

    # If any of the first 3 rows (header rows) has 2+ fuel keywords in distinct columns → wide
    # We restrict to rows[:3] to avoid counting fuel-type values in data rows (long format).
    for row in rows[:3]:
        fuel_count = sum(
            1 for val in row
            if val and any(kw in str(val) for kwds, _ in _FUEL_KEYWORDS for kw in kwds)
        )
        if fuel_count >= 2:
            return "wide"

    # Look for data rows where a single column (not col 0) contains an energy keyword
    # AND the next non-None column is numeric — that's long format.
    for row in rows[1:]:
        for ci, val in enumerate(row):
            if ci == 0 or val is None:
                continue
            if _classify_row_type(str(val)) in ("wind", "solar", "thermal", "hydro"):
                # Check that there's a numeric-looking cell in the same row
                for nv in row[ci + 1: ci + 5]:
                    try:
                        float(str(nv).replace(",", "")) if nv else None
                        return "long"
                    except (ValueError, TypeError):
                        pass

    return "wide"   # default


# ── File parser entry point ───────────────────────────────────────────────────

def parse_file(path: Path, province: str) -> list[dict]:
    """Parse one Excel file and return a list of record dicts."""
    try:
        raw = path.read_bytes()
    except (PermissionError, OSError):
        tmp = tempfile.mktemp(suffix=".xlsx")
        shutil.copy2(str(path), tmp)
        with open(tmp, "rb") as fh:
            raw = fh.read()
        try:
            Path(tmp).unlink()
        except OSError:
            pass

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Cannot open %s: %s", path.name, e)
        return []

    ws = _best_sheet(wb)
    fmt = _detect_format(ws)
    source = path.name

    logger.info("  %-40s  format=%-4s  province=%s", path.name[:40], fmt, province)

    if fmt == "wide":
        records = _parse_wide(ws, province, source)
        if not records:
            # Fallback: some files are misdetected; try long format
            ws2 = _best_sheet(wb)
            records = _parse_long(ws2, province, source)
    else:
        records = _parse_long(ws, province, source)
        if not records:
            ws2 = _best_sheet(wb)
            records = _parse_wide(ws2, province, source)

    wb.close()

    if not records:
        logger.warning("  No records extracted from %s", path.name)

    return records


# ── DB upsert ─────────────────────────────────────────────────────────────────

_UPSERT_SQL = """
INSERT INTO marketdata.province_installed_monthly
    (province, year_month, wind_mw, solar_mw, thermal_mw, hydro_mw,
     nuclear_mw, bess_mw, total_mw, source_file, ingested_at)
VALUES
    (%(province)s, %(year_month)s, %(wind)s, %(solar)s, %(thermal)s, %(hydro)s,
     %(nuclear)s, %(bess)s, %(total)s, %(source_file)s, NOW())
ON CONFLICT (province, year_month)
DO UPDATE SET
    wind_mw     = EXCLUDED.wind_mw,
    solar_mw    = EXCLUDED.solar_mw,
    thermal_mw  = EXCLUDED.thermal_mw,
    hydro_mw    = EXCLUDED.hydro_mw,
    nuclear_mw  = EXCLUDED.nuclear_mw,
    bess_mw     = EXCLUDED.bess_mw,
    total_mw    = EXCLUDED.total_mw,
    source_file = EXCLUDED.source_file,
    ingested_at = NOW()
"""


def _to_db_row(rec: dict) -> dict:
    """Convert a parsed record dict to the DB param dict."""
    wind    = rec.get("wind")
    solar   = rec.get("solar")
    total   = rec.get("total")
    if total is None and wind is not None and solar is not None:
        pass  # don't synthesise total — leave as NULL if not in source
    return {
        "province":    rec["province"],
        "year_month":  rec["year_month"],
        "wind":        rec.get("wind"),
        "solar":       rec.get("solar"),
        "thermal":     rec.get("thermal"),
        "hydro":       rec.get("hydro"),
        "nuclear":     rec.get("nuclear"),
        "bess":        rec.get("bess"),
        "total":       rec.get("total"),
        "source_file": rec.get("source_file"),
    }


def upsert_records(records: list[dict], dsn: str) -> int:
    import psycopg2
    conn = psycopg2.connect(dsn)
    try:
        with conn.cursor() as cur:
            rows = [_to_db_row(r) for r in records]
            cur.executemany(_UPSERT_SQL, rows)
        conn.commit()
        return len(records)
    finally:
        conn.close()


# ── Table bootstrap ───────────────────────────────────────────────────────────

def _ensure_table(dsn: str) -> None:
    """Create province_installed_monthly if it doesn't exist."""
    import psycopg2
    ddl_path = (
        Path(__file__).resolve().parents[1]
        / "db" / "ddl" / "marketdata" / "province_installed_monthly.sql"
    )
    ddl = ddl_path.read_text(encoding="utf-8") if ddl_path.exists() else None
    if not ddl:
        ddl = """
        CREATE TABLE IF NOT EXISTS marketdata.province_installed_monthly (
            id SERIAL PRIMARY KEY, province TEXT NOT NULL, year_month DATE NOT NULL,
            wind_mw NUMERIC, solar_mw NUMERIC, thermal_mw NUMERIC,
            hydro_mw NUMERIC, nuclear_mw NUMERIC, bess_mw NUMERIC,
            total_mw NUMERIC, source_file TEXT, ingested_at TIMESTAMPTZ DEFAULT NOW(),
            UNIQUE (province, year_month)
        );
        CREATE INDEX IF NOT EXISTS idx_pim_province_ym
            ON marketdata.province_installed_monthly (province, year_month DESC);
        """
    conn = psycopg2.connect(dsn)
    try:
        with conn.cursor() as cur:
            cur.execute(ddl)
        conn.commit()
    finally:
        conn.close()


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Scan province installed-capacity Excel files and upsert to DB."
    )
    parser.add_argument(
        "--data-dir", type=Path, default=_DATA_DIR,
        help="Folder containing province Excel files (default: %(default)s)",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Parse files but do not write to DB.",
    )
    parser.add_argument(
        "--province", type=str, default=None,
        help="Only process this one province (Chinese name, e.g. 山东).",
    )
    parser.add_argument(
        "--dsn", type=str, default=None,
        help="DB connection string. Falls back to PGURL env var.",
    )
    args = parser.parse_args()

    # Resolve DSN
    if not args.dry_run:
        import os
        dsn = args.dsn or os.environ.get("PGURL") or os.environ.get("DATABASE_URL")
        if not dsn:
            # Try loading config/.env
            try:
                from dotenv import load_dotenv
                load_dotenv(_REPO / "config" / ".env")
                dsn = os.environ.get("PGURL")
            except ImportError:
                pass
        if not dsn:
            logger.error("No DB DSN found. Set PGURL env var or pass --dsn.")
            sys.exit(1)
        _ensure_table(dsn)
    else:
        dsn = None

    data_dir: Path = args.data_dir
    if not data_dir.exists():
        logger.error("Data dir not found: %s", data_dir)
        sys.exit(1)

    xlsx_files = sorted(data_dir.glob("*.xlsx"))
    if not xlsx_files:
        logger.warning("No .xlsx files found in %s", data_dir)
        return

    total_upserted = 0
    total_skipped  = 0

    for path in xlsx_files:
        province = _province_from_filename(path.name)
        if province is None:
            logger.warning("Cannot determine province from filename: %s — skipping", path.name)
            total_skipped += 1
            continue

        if args.province and province != args.province:
            continue

        records = parse_file(path, province)
        if not records:
            total_skipped += 1
            continue

        if args.dry_run:
            logger.info(
                "  DRY-RUN  %-8s  %d records  first=%s  last=%s",
                province,
                len(records),
                min(r["year_month"] for r in records),
                max(r["year_month"] for r in records),
            )
            # Print a sample
            sample = records[-1]
            logger.info(
                "    sample: wind=%.0f solar=%.0f thermal=%.0f hydro=%.0f bess=%.0f",
                sample.get("wind") or 0,
                sample.get("solar") or 0,
                sample.get("thermal") or 0,
                sample.get("hydro") or 0,
                sample.get("bess") or 0,
            )
        else:
            n = upsert_records(records, dsn)
            logger.info("  %-8s  upserted %d rows", province, n)
            total_upserted += n

    if args.dry_run:
        logger.info("Dry-run complete — no DB changes made.")
    else:
        logger.info("Done. Total upserted: %d  |  files skipped: %d", total_upserted, total_skipped)


def parse_and_upsert_bytes(file_bytes: bytes, filename: str, dsn: str) -> dict:
    """Parse a province capacity Excel from raw bytes and upsert to DB.

    Returns {"upserted": int, "provinces": list[str], "year_month": str, "errors": list[str]}
    Compatible with capacity_etl.upsert_capacity() return format.

    Province is extracted from the filename (e.g. "江苏-装机数据_2026年5月.xlsx" → 江苏).
    Year is extracted from filename too (e.g. "_2026年5月" → context_year=2026).
    """
    import tempfile, os as _os

    province = _province_from_filename(filename)
    if province is None:
        return {"upserted": 0, "provinces": [], "year_month": "", "errors": [f"Cannot determine province from filename: {filename}"]}

    # Write bytes to a temp file so parse_file() can open it via openpyxl
    suffix = ".xlsx" if filename.lower().endswith(".xlsx") else ".xls"
    tmp = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
    try:
        tmp.write(file_bytes)
        tmp.close()
        records = parse_file(Path(tmp.name), province)
    finally:
        try:
            _os.unlink(tmp.name)
        except OSError:
            pass

    if not records:
        return {"upserted": 0, "provinces": [province], "year_month": "", "errors": ["No records extracted from file"]}

    # Patch source_file to the original filename (not the temp path)
    for r in records:
        r["source_file"] = filename

    errors: list[str] = []
    try:
        _ensure_table(dsn)
        n = upsert_records(records, dsn)
        year_months = sorted({str(r["year_month"]) for r in records})
        return {"upserted": n, "provinces": [province], "year_month": year_months[-1] if year_months else "", "errors": errors}
    except Exception as exc:
        return {"upserted": 0, "provinces": [province], "year_month": "", "errors": [str(exc)]}


if __name__ == "__main__":
    # Load .env so PGURL is available when run directly
    try:
        from dotenv import load_dotenv as _lde
        _lde(Path(__file__).resolve().parents[1] / "config" / ".env")
    except ImportError:
        pass
    main()
