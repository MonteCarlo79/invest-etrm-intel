"""
China Spot Market Price Cockpit
Visualises daily DA / RT clearing prices from spot_daily.

Run:
    py -m streamlit run apps/spot-market/app.py
"""
from __future__ import annotations

import os
import sys
import time
from datetime import date, timedelta
from pathlib import Path
import re as _re

import json
import requests

import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import matplotlib.font_manager as _mfm
from matplotlib.patches import Polygon as MplPolygon

# Pick the first CJK-capable font available on this system.
# Linux (Docker) typically has Noto CJK; Windows has YaHei/SimHei.
_CJK_FONTS = ["Noto Sans CJK SC", "Microsoft YaHei", "SimHei", "SimSun",
               "STHeiti", "WenQuanYi Micro Hei", "Arial Unicode MS"]
_CJK_FONT: str | None = None
for _f in _CJK_FONTS:
    try:
        if _mfm.findfont(_mfm.FontProperties(family=_f), fallback_to_default=False):
            _CJK_FONT = _f
            break
    except (ValueError, OSError):
        pass

# File-path fallback: scan all system font files for Noto/WQY CJK fonts.
# Handles cases where the font is installed but not yet in the family-name index.
if _CJK_FONT is None:
    for _fp in _mfm.findSystemFonts():
        _bn = os.path.basename(_fp).lower()
        if any(k in _bn for k in ("notocjk", "notosanscjk", "noto_cjk",
                                   "wqymicro", "wenquanyi")):
            try:
                _mfm.fontManager.addfont(_fp)
                _CJK_FONT = _mfm.FontProperties(fname=_fp).get_name()
                break
            except Exception:
                pass
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import psycopg2
import streamlit as st
from dotenv import load_dotenv

# ── path / env setup ─────────────────────────────────────────────────────────
_REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(_REPO))

for _env in [_REPO / "config" / ".env", _REPO / ".env"]:
    if _env.exists():
        load_dotenv(_env)
_spot_env = _REPO / "apps" / "spot-agent" / ".env"
if _spot_env.exists():
    load_dotenv(_spot_env)

# ── translations ──────────────────────────────────────────────────────────────
_T: dict[str, dict[str, str]] = {
    "en": {
        # app
        "app_title":            "⚡ China Spot Market Price Cockpit",
        # sidebar
        "lang_label":           "🌐 Language",
        "filters":              "Filters",
        "date_range":           "Date range",
        "provinces":            "Provinces (multi-select)",
        "show_band":            "Show min/max band",
        "filter_bad_data":      "Filter bad data",
        "filter_bad_data_help": "Exclude rows where avg is outside [min, max] bounds — caused by early-Jan PDF format differences",
        "data_caption":         "Data: spot_daily · units: ¥/kWh",
        "select_prov_info":     "Select at least one province in the sidebar.",
        # KPIs
        "latest_date":          "Latest Date",
        "dates_in_db":          "Dates in DB",
        "provinces_kpi":        "Provinces",
        "complete_rows":        "Complete Rows",
        "coverage":             "Coverage",
        # tabs
        "tab_overview":         "Overview",
        "tab_spread":           "DA−RT Spread",
        "tab_heatmap":          "Heatmap",
        "tab_province":         "Province Deep-Dive",
        "tab_dist":             "Distributions",
        "tab_geo":              "Geo Map",
        "tab_interprov":        "Inter-Provincial Flow",
        "tab_fundamentals":     "Market Fundamentals",
        "tab_agent":            "Strategist",
        "tab_news":             "📡 News Sources",
        "tab_mgmt":             "Data Management",
        "tab_intraday":         "Intraday Analysis",
        # intraday
        "intraday_shape_title": "Average Hourly Price Shape",
        "intraday_spread_title":"Intraday Spread Ranking (¥/kWh)",
        "intraday_spread_help": "Max − Min of hourly avg prices per province. Higher = more BESS arbitrage potential.",
        "intraday_pdc_title":   "Price Duration Curve",
        "intraday_heat_title":  "Hour × Province Heatmap",
        "intraday_price_type":  "Price type",
        "intraday_select_prov": "Select province for duration curve",
        # intraday fundamentals
        "intraday_fund_title":  "Fundamentals — Market Drivers",
        "intraday_bid_title":   "Bidding Space vs RT Price",
        "intraday_bid_caption": "Each point = one hour. Higher bidding space → lower RT price (supply glut suppresses clearing price).",
        "intraday_wap_title":   "Wind & Solar WAP vs Avg RT Price",
        "intraday_wap_caption": "WAP = Σ(generation_MW × price) / Σ(generation_MW). WAP below avg RT = renewable generation concentrated in low-price hours.",
        "intraday_no_fund":     "No fundamentals data found in DB for the selected period. Use bess-map Data Management → Run Fundamentals Ingest to populate.",
        # market fundamentals
        "fund_provinces":       "Provinces",
        "fund_year":            "Year",
        "fund_capacity_title":  "Installed Capacity by Fuel Type (万kW)",
        "fund_generation_title":"Generation by Fuel Type (亿kWh)",
        "fund_peak_title":      "Peak Load (MW)",
        "fund_summer_peak":     "Summer Peak",
        "fund_winter_peak":     "Winter Peak",
        "fund_other_peak":      "Off-Peak",
        "fund_renewables_share":"Renewables share",
        "fund_lf_title":        "Load Factor by Fuel Type (%)",
        "fund_lf_caption":      "LF = Generation (亿kWh) ÷ Capacity (万kW) ÷ 8760. All provinces ranked.",
        "fund_lf_sort":         "Sort ranking by",
        "fund_tightness_title": "System Tightness",
        "fund_tightness_caption": "Effective capacity = Σ(Installed capacity × standard EOH ÷ 8760). Green = surplus, red = deficit.",
        "fund_eff_cap":         "Eff. Cap (MW)",
        "fund_avg_demand":      "Avg Demand (MW)",
        "fund_tight_avg":       "vs Avg Demand (MW)",
        "fund_tight_summer":    "vs Summer Peak (MW)",
        "fund_tight_winter":    "vs Winter Peak (MW)",
        "fund_no_data":         "No market fundamentals data available. Ensure the Excel file is in data/market-fundamentals/.",
        "fund_total":           "Total",
        "fund_share_label":     "Share of capacity (%)",
        "fund_capacity_unit":   "万kW",
        "fund_generation_unit": "亿kWh",
        "fund_select_prompt":   "Select at least one province.",
        # agent
        "agent_title":          "Spot Market AI Agent",
        "agent_caption":        "Ask questions about prices, trends, inter-provincial flows, or trigger ingestion.",
        "agent_welcome":        "Hi! I can query spot market prices, inter-provincial flows, and daily summaries from the database, and run the ingestion pipeline for new PDFs. What would you like to know?",
        "agent_placeholder":    "e.g. What were Shandong DA prices in April 2026?",
        "agent_thinking":       "Thinking...",
        "agent_tool_call":      "Tool call: {tool}",
        "agent_tool_result":    "Result ({n} rows)",
        "agent_no_key":         "ANTHROPIC_API_KEY is not set. Please add it to your .env file.",
        "agent_clear":          "Clear chat",
        "agent_error":          "Agent error: {err}",
        # knowledge base
        "kb_title":             "Knowledge Base",
        "kb_caption":           "Upload market rules, annual reports, and policy documents for the agent to reference.",
        "kb_upload_label":      "Upload documents (PDF / PPTX / Word / Excel / TXT / Image)",
        "kb_category_label":    "Category override",
        "kb_category_auto":     "Auto-detect",
        "kb_upload_btn":        "Process & Add",
        "kb_success":           "{n} document(s) added to knowledge base.",
        "kb_duplicate":         "{fname} is already in the knowledge base.",
        "kb_failed":            "Failed to process {fname}: {err}",
        "kb_doc_list_title":    "Registered Reference Documents",
        "kb_doc_list_empty":    "No reference documents uploaded yet.",
        "kb_delete":            "Remove",
        "kb_pages":             "{n} pages",
        "kb_chunks":            "chunks indexed",
        # memory
        "mem_saved_ok":         "Saved {n} memory item(s).",
        "mem_confirm_title":    "💡 Suggested memories from this conversation",
        "mem_save_selected":    "Save selected",
        "mem_dismiss":          "Dismiss",
        "mem_manage":           "Memory Management",
        "mem_caption":          "Active memories are injected into every conversation as domain context.",
        "mem_empty":            "No memories stored yet.",
        "mem_delete":           "Delete",
        # overview
        "latest_prices":        "Latest prices",
        "col_province":         "Province",
        "col_province_cn":      "CN Name",
        "col_date":             "Date",
        # spread
        "spread_stats":         "Spread statistics (¥/kWh)",
        "col_mean":             "Mean",
        "col_std":              "Std",
        "col_min":              "Min",
        "col_max":              "Max",
        "col_da_gt_rt":         "DA > RT (%)",
        "col_days":             "Days",
        # heatmap
        "metric_label":         "Metric",
        # province
        "select_province":      "Select province",
        "raw_data":             "raw data",
        # distributions
        "market_label":         "Market",
        "hist_bins":            "Histogram bins",
        "kde_label":            "Overlay KDE curve",
        "both_label":           "Both",
        "desc_stats":           "Descriptive statistics (¥/kWh)",
        "col_n":                "N",
        "col_median":           "Median",
        "col_p10":              "P10",
        "col_p25":              "P25",
        "col_p75":              "P75",
        "col_p90":              "P90",
        # geo
        "avg_by_province":      "Average Prices by Province",
        "geo_color_caption":    "Green = Low (<0.20 ¥/kWh) · Yellow = Medium (0.20–0.30) · Red = High (>0.30)",
        "geo_maps_title":       "Geographic Price Maps",
        "geo_color_scale":      "Color scale: 🟢 **< 0.20 ¥/kWh** (low) · 🟡 **0.20–0.30** (medium) · 🔴 **> 0.30 ¥/kWh** (high)",
        "geo_unavailable":      "Province boundaries unavailable — showing bubble fallback.",
        "da_caption":           "Day-Ahead (DA)",
        "rt_caption":           "Real-Time (RT)",
        "col_avg_da":           "Avg DA (¥/kWh)",
        "col_da_level":         "DA Level",
        "col_avg_rt":           "Avg RT (¥/kWh)",
        "col_rt_level":         "RT Level",
        "col_days_da":          "Days (DA)",
        "col_days_rt":          "Days (RT)",
        "level_low":            "Low",
        "level_medium":         "Medium",
        "level_high":           "High",
        # inter-provincial flow
        "interprov_title":      "Inter-Provincial Spot Trading (省间现货交易)",
        "interprov_no_data":    "No inter-provincial data for the selected period.",
        "interprov_price_trend":"Inter-Provincial Clearing Price Trend (¥/kWh)",
        "interprov_vol_trend":  "Total Inter-Provincial Volume (亿kWh)",
        "direction_export":     "Exporting (送端)",
        "direction_import":     "Importing (受端)",
        "col_direction":        "Direction",
        "col_metric_type":      "Metric",
        "col_share":            "Share (%)",
        "col_price_kwh":        "Price (¥/kWh)",
        "col_price_chg":        "Day-on-day (%)",
        "col_time_period":      "Active period",
        "col_volume_gwh":       "Volume (亿kWh)",
        "col_source":           "Source PDF",
        "interprov_price_hi":   "Peak Avg (最高均价)",
        "interprov_price_lo":   "Floor Avg (最低均价)",
        "interprov_prov_leaders":"Province Leaders — Peak Avg Price",
        "hover_province":       "Province",
        "hover_chg":            "Day-on-day",
        "hover_period":         "Active period",
        "hover_share":          "Market share",
        "hover_volume":         "Volume",
        # province summaries
        "summaries_title":      "Market Summaries",
        "summaries_no_data":    "No summaries available for this period.",
        "summary_label":        "{date}",
        # data management
        "data_mgmt_title":      "Data Management",
        "report_year":          "Report year",
        "mode_label":           "Mode",
        "mode_fill_gaps":       "Fill gaps (ingest missing dates only)",
        "mode_backfill":        "Backfill date range (ingest all PDFs covering the range)",
        "additional_steps":     "Additional steps",
        "chk_interprov":        "Parse 省间现货交易 data",
        "chk_interprov_help":   "Extract inter-provincial trading data and save to staging.spot_interprov_flow",
        "chk_ai":               "Generate AI summaries",
        "chk_ai_help":          "Generate Claude daily market summaries (requires ANTHROPIC_API_KEY)",
        "start_date":           "Start date",
        "end_date":             "End date",
        "col_pdf":              "PDF",
        "col_covers":           "Covers",
        "col_dates_range":      "Dates in range",
        "col_missing":          "Missing from DB",
        "col_partial":          "Partial (DA or RT=0)",
        "col_status":           "Status",
        "status_missing":       "Missing",
        "status_partial":       "Partial",
        "status_ok":            "OK",
        "btn_fill_gaps":        "Backfill {n} PDF(s) with missing dates",
        "btn_reingest":         "Re-ingest all {n} PDF(s) in range",
        "warn_partial":         "{n} PDF(s) have partial data (DA or RT missing). Switch to 'Backfill date range' mode to re-ingest them.",
        "all_present":          "All dates in range are present in DB.",
        "no_pdfs":              "No PDFs found in the selected date range.",
        "upload_pdf":           "Upload PDF report(s)",
        "upload_help":          "Upload PDFs here when running on AWS (no local data folder). Files are stored in S3 and immediately available for ingestion.",
        "upload_btn":           "Upload {n} file(s) to S3",
        "upload_success":       "Uploaded {n} file(s) to S3.",
        "prog_starting":        "Starting…",
        "prog_parsing":         "Parsing {fname}…",
        "prog_interprov":       "省间 data: {fname}…",
        "prog_ai":              "AI summary {rdate}…",
        "prog_done":            "Done.",
        "backfill_complete":    "Backfill complete — processed {n} PDF(s).",
        "col_dates":            "Dates",
        "col_rows":             "Rows upserted",
        "col_interprov":        "Interprov rows",
        "col_ai":               "AI summaries",
        "col_error":            "Error",
        # knowledge base sync
        "kb_sync_title":        "Sync Knowledge Base from Market Fundamentals",
        "kb_sync_caption":      "Scans data/market-fundamentals/ and ingests any new documents into the knowledge pool. Already-ingested files are skipped automatically.",
        "kb_sync_unavailable":  "Knowledge base sync is only available in local mode (data/market-fundamentals/ not found).",
        "kb_sync_no_files":     "No new files found — knowledge base is up to date.",
        "kb_sync_btn":          "Sync {n} new file(s)",
        "kb_sync_progress":     "Ingesting {i}/{n}: {fname}…",
        "kb_sync_done":         "Sync complete — added: {added}  skipped: {skipped}  errors: {errors}",
        # data export
        "export_title":         "Download Price Data",
        "export_caption":       "Export DA / RT average prices from the database to Excel (.xlsx). "
                                "Rows = dates, columns = provinces (Chinese names), units = ¥/kWh.",
        "export_btn":           "Download Excel",
        "export_filename":      "电力现货市场日均价格.xlsx",
        # wechat batch ingest
        "wechat_title":         "WeChat Article Batch Import",
        "wechat_caption":       "Paste one WeChat article URL per line. The app fetches each article server-side and ingests the text into the Strategist knowledge pool.",
        "wechat_url_label":     "Article URLs (one per line)",
        "wechat_url_placeholder": "https://mp.weixin.qq.com/s/...\nhttps://mp.weixin.qq.com/s/...",
        "wechat_run_btn":       "Fetch & Ingest {n} article(s)",
        "wechat_no_urls":       "Paste at least one URL above.",
        "wechat_fetching":      "Fetching {i}/{n}: {url}…",
        "wechat_done":          "Done — added: {added}  skipped: {skipped}  errors: {errors}",
        "wechat_digest_btn":    "Digest articles → Insights",
        # daily report
        "report_section_title": "Daily Market Report",
        "report_section_caption": "Scheduled daily PDF report of DA/RT provincial prices with AI commentary. "
                                  "Sent via email at 06:00 SGT and optionally to WeCom groups.",
        "report_schedule_status": "Scheduler status",
        "report_next_run":      "Next run",
        "report_send_now":      "Send Report Now",
        "report_send_success":  "Report sent — {size:,} bytes, date: {rdate}",
        "report_send_wecom":    "WeCom: {result}",
        "report_send_error":    "Report failed: {err}",
        "report_email_label":   "Recipient email(s)",
        "report_email_help":    "Comma-separated. Defaults to REPORT_TO_EMAIL env var.",
        "report_webhook_title": "WeCom Webhook Groups",
        "report_webhook_caption": "Add multiple WeCom bot webhook URLs. Saved to DB and survive restarts.",
        "report_webhook_add_label": "Webhook URL",
        "report_webhook_add_label_label": "Label (optional)",
        "report_webhook_add_btn": "Add Webhook",
        "report_webhook_added": "Webhook added.",
        "report_webhook_empty": "No webhooks configured — WeCom delivery disabled.",
        "report_webhook_delete": "Remove",
        "report_webhook_enabled": "Enabled",
        # chart labels
        "da_label":             "Day-Ahead (DA)",
        "rt_label":             "Real-Time (RT)",
        "da_avg_label":         "DA avg",
        "rt_avg_label":         "RT avg",
        "price_unit":           "¥/kWh",
        "prob_density":         "Probability density",
        "price_axis":           "Price (¥/kWh)",
        "spread_title":         "DA − RT Spread  (¥/kWh)  |  +ve = DA premium, −ve = RT spike",
        "da_clearing":          "Day-Ahead (DA) Clearing Price  (¥/kWh)",
        "rt_clearing":          "Real-Time (RT) Clearing Price  (¥/kWh)",
        "da_dist_title":        "Day-Ahead (DA) Price Distribution  (¥/kWh)",
        "rt_dist_title":        "Real-Time (RT) Price Distribution  (¥/kWh)",
        "da_violin_title":      "Day-Ahead (DA) — Violin / Box Plot  (¥/kWh)",
        "rt_violin_title":      "Real-Time (RT) — Violin / Box Plot  (¥/kWh)",
        "da_heatmap_title":     "Day-Ahead Average Clearing Price — Province × Date Heatmap",
        "rt_heatmap_title":     "Real-Time Average Clearing Price — Province × Date Heatmap",
        "geo_title_da":         "Day-Ahead (DA) — Average Price by Province (¥/kWh)",
        "geo_title_rt":         "Real-Time (RT) — Average Price by Province (¥/kWh)",
        # geo animation
        "anim_title":           "Monthly RT Price Animation",
        "anim_range":           "Animation period",
        "anim_start_year":      "Start year",
        "anim_start_month":     "Start month",
        "anim_end_year":        "End year",
        "anim_end_month":       "End month",
        "anim_play":            "▶ Play",
        "anim_pause":           "⏸ Pause",
        "anim_speed":           "Seconds per frame",
        "anim_slider":          "Select month",
        "anim_no_data":         "No data for this month.",
        "anim_map_title":       "RT Avg Price — {month}",
        # geo comparison
        "cmp_title":            "Period Comparison",
        "cmp_metric":           "Metric",
        "cmp_period_a":         "Period A",
        "cmp_period_b":         "Period B",
        "cmp_start":            "Start",
        "cmp_end":              "End",
        "cmp_no_data":          "No data for this period.",
        "cmp_map_title":        "{metric} Avg — {start} → {end}",
    },
    "zh": {
        # app
        "app_title":            "⚡ 中国电力现货市场价格驾驶舱",
        # sidebar
        "lang_label":           "🌐 语言",
        "filters":              "筛选条件",
        "date_range":           "日期范围",
        "provinces":            "省份（多选）",
        "show_band":            "显示最大/最小区间",
        "filter_bad_data":      "过滤异常数据",
        "filter_bad_data_help": "排除均值不在最大/最小区间内的行——由1月初PDF格式差异引起",
        "data_caption":         "数据来源：spot_daily · 单位：元/千瓦时",
        "select_prov_info":     "请在侧边栏选择至少一个省份。",
        # KPIs
        "latest_date":          "最新日期",
        "dates_in_db":          "数据库日期数",
        "provinces_kpi":        "省份数",
        "complete_rows":        "完整行数",
        "coverage":             "覆盖率",
        # tabs
        "tab_overview":         "总览",
        "tab_spread":           "日前-实时价差",
        "tab_heatmap":          "热力图",
        "tab_province":         "省份深度分析",
        "tab_dist":             "价格分布",
        "tab_geo":              "地理分布图",
        "tab_interprov":        "省间现货交易",
        "tab_fundamentals":     "市场基础数据",
        "tab_agent":            "策略分析师",
        "tab_news":             "📡 新闻来源",
        "tab_mgmt":             "数据管理",
        "tab_intraday":         "日内价格分析",
        # intraday
        "intraday_shape_title": "平均小时价格曲线",
        "intraday_spread_title":"日内价差排名 (元/千瓦时)",
        "intraday_spread_help": "各省小时均价最大值减最小值。越高代表储能套利潜力越大。",
        "intraday_pdc_title":   "价格持续时间曲线",
        "intraday_heat_title":  "小时×省份热力图",
        "intraday_price_type":  "价格类型",
        "intraday_select_prov": "选择省份（持续时间曲线）",
        # intraday fundamentals
        "intraday_fund_title":  "基本面——市场驱动因素",
        "intraday_bid_title":   "竞价空间 vs 实时价格",
        "intraday_bid_caption": "每点代表一小时。竞价空间越大，实时价格越低（供给宽松压低出清价格）。",
        "intraday_wap_title":   "风光加权平均价 vs 均价",
        "intraday_wap_caption": "加权均价 = Σ(发电量×价格) / Σ(发电量)。加权均价低于均价说明新能源发电集中于低价时段。",
        "intraday_no_fund":     "所选时段数据库中暂无基本面数据，请在储能地图→数据管理→运行基本面导入后再查看。",
        # market fundamentals
        "fund_provinces":       "省份",
        "fund_year":            "年份",
        "fund_capacity_title":  "各燃料类型装机容量（万千瓦）",
        "fund_generation_title":"各燃料类型发电量（亿千瓦时）",
        "fund_peak_title":      "最大负荷（兆瓦）",
        "fund_summer_peak":     "度夏峰值",
        "fund_winter_peak":     "度冬峰值",
        "fund_other_peak":      "其余月份",
        "fund_renewables_share":"可再生能源占比",
        "fund_lf_title":        "各类型负荷率（%）",
        "fund_lf_caption":      "负荷率 = 年发电量（亿千瓦时）÷ 装机容量（万千瓦）÷ 8760。显示全部省份排名。",
        "fund_lf_sort":         "排序依据",
        "fund_tightness_title": "电力系统紧张程度",
        "fund_tightness_caption": "有效容量 = Σ（装机容量 × 标准利用小时 ÷ 8760）。绿色 = 盈余，红色 = 缺口。",
        "fund_eff_cap":         "有效容量（兆瓦）",
        "fund_avg_demand":      "平均负荷（兆瓦）",
        "fund_tight_avg":       "对均值盈余（兆瓦）",
        "fund_tight_summer":    "对度夏盈余（兆瓦）",
        "fund_tight_winter":    "对度冬盈余（兆瓦）",
        "fund_no_data":         "暂无市场基础数据。请确保Excel文件位于 data/market-fundamentals/ 目录下。",
        "fund_total":           "合计",
        "fund_share_label":     "装机占比（%）",
        "fund_capacity_unit":   "万千瓦",
        "fund_generation_unit": "亿千瓦时",
        "fund_select_prompt":   "请至少选择一个省份。",
        # agent
        "agent_title":          "现货市场策略分析师",
        "agent_caption":        "查询价格、走势、省间交易数据，或触发PDF导入流程。",
        "agent_welcome":        "您好！我可以查询现货市场价格、省间交易数据及每日市场摘要，也可以为新PDF运行导入流程。请问您想了解什么？",
        "agent_placeholder":    "例如：2026年4月山东日前价格是多少？",
        "agent_thinking":       "思考中…",
        "agent_tool_call":      "工具调用：{tool}",
        "agent_tool_result":    "结果（{n} 行）",
        "agent_no_key":         "未设置 ANTHROPIC_API_KEY，请在 .env 文件中添加。",
        "agent_clear":          "清除对话",
        "agent_error":          "助手出错：{err}",
        # knowledge base
        "kb_title":             "知识库",
        "kb_caption":           "上传交易规则、年度报告、政策文件等，供智能助手参考使用。",
        "kb_upload_label":      "上传文档（PDF / PPTX / Word / Excel / TXT / 图片）",
        "kb_category_label":    "手动指定类别",
        "kb_category_auto":     "自动识别",
        "kb_upload_btn":        "处理并添加",
        "kb_success":           "已添加 {n} 个文档至知识库。",
        "kb_duplicate":         "{fname} 已存在于知识库中。",
        "kb_failed":            "处理 {fname} 失败：{err}",
        "kb_doc_list_title":    "已注册的参考文档",
        "kb_doc_list_empty":    "暂无参考文档，请上传。",
        "kb_delete":            "删除",
        "kb_pages":             "{n} 页",
        "kb_chunks":            "个片段已索引",
        # memory
        "mem_saved_ok":         "已保存 {n} 条记忆。",
        "mem_confirm_title":    "💡 本次对话中发现的记忆建议",
        "mem_save_selected":    "保存所选",
        "mem_dismiss":          "忽略",
        "mem_manage":           "记忆管理",
        "mem_caption":          "已激活的记忆将在每次对话开始时注入系统提示。",
        "mem_empty":            "暂无已保存的记忆。",
        "mem_delete":           "删除",
        # overview
        "latest_prices":        "最新价格",
        "col_province":         "省份",
        "col_province_cn":      "中文名",
        "col_date":             "日期",
        # spread
        "spread_stats":         "价差统计（元/千瓦时）",
        "col_mean":             "均值",
        "col_std":              "标准差",
        "col_min":              "最小值",
        "col_max":              "最大值",
        "col_da_gt_rt":         "日前>实时（%）",
        "col_days":             "天数",
        # heatmap
        "metric_label":         "指标",
        # province
        "select_province":      "选择省份",
        "raw_data":             "原始数据",
        # distributions
        "market_label":         "市场",
        "hist_bins":            "直方图组数",
        "kde_label":            "叠加KDE曲线",
        "both_label":           "两者",
        "desc_stats":           "描述性统计（元/千瓦时）",
        "col_n":                "N",
        "col_median":           "中位数",
        "col_p10":              "P10",
        "col_p25":              "P25",
        "col_p75":              "P75",
        "col_p90":              "P90",
        # geo
        "avg_by_province":      "各省平均价格",
        "geo_color_caption":    "绿色 = 低价（<0.20元/千瓦时）· 黄色 = 中等（0.20–0.30）· 红色 = 高价（>0.30）",
        "geo_maps_title":       "地理价格分布图",
        "geo_color_scale":      "色阶：🟢 **< 0.20 元/千瓦时**（低）· 🟡 **0.20–0.30**（中）· 🔴 **> 0.30 元/千瓦时**（高）",
        "geo_unavailable":      "省级边界数据不可用——显示气泡图替代。",
        "da_caption":           "日前（DA）",
        "rt_caption":           "实时（RT）",
        "col_avg_da":           "日前均价（元/千瓦时）",
        "col_da_level":         "日前价格水平",
        "col_avg_rt":           "实时均价（元/千瓦时）",
        "col_rt_level":         "实时价格水平",
        "col_days_da":          "日前天数",
        "col_days_rt":          "实时天数",
        "level_low":            "低",
        "level_medium":         "中",
        "level_high":           "高",
        # inter-provincial flow
        "interprov_title":      "省间现货交易情况",
        "interprov_no_data":    "所选时段内无省间交易数据。",
        "interprov_price_trend":"省间出清价格走势（元/千瓦时）",
        "interprov_vol_trend":  "省间总交易量（亿kWh）",
        "direction_export":     "送端（出力）",
        "direction_import":     "受端（受入）",
        "col_direction":        "方向",
        "col_metric_type":      "指标类型",
        "col_share":            "占比（%）",
        "col_price_kwh":        "价格（元/千瓦时）",
        "col_price_chg":        "日环比（%）",
        "col_time_period":      "活跃时段",
        "col_volume_gwh":       "电量（亿kWh）",
        "col_source":           "数据来源",
        "interprov_price_hi":   "最高均价",
        "interprov_price_lo":   "最低均价",
        "interprov_prov_leaders":"每日最高均价省份",
        "hover_province":       "省份",
        "hover_chg":            "日环比",
        "hover_period":         "活跃时段",
        "hover_share":          "市场占比",
        "hover_volume":         "电量",
        # province summaries
        "summaries_title":      "市场日报摘要",
        "summaries_no_data":    "所选时段内暂无市场摘要。",
        "summary_label":        "{date}",
        # data management
        "data_mgmt_title":      "数据管理",
        "report_year":          "报告年份",
        "mode_label":           "模式",
        "mode_fill_gaps":       "补全缺口（仅录入缺失日期）",
        "mode_backfill":        "回填日期范围（录入覆盖该范围的所有PDF）",
        "additional_steps":     "附加步骤",
        "chk_interprov":        "解析省间现货交易数据",
        "chk_interprov_help":   "提取省间交易数据并保存至 staging.spot_interprov_flow",
        "chk_ai":               "生成AI摘要",
        "chk_ai_help":          "生成Claude每日市场摘要（需设置 ANTHROPIC_API_KEY）",
        "start_date":           "开始日期",
        "end_date":             "结束日期",
        "col_pdf":              "PDF文件",
        "col_covers":           "覆盖日期",
        "col_dates_range":      "范围内日期数",
        "col_missing":          "数据库缺失",
        "col_partial":          "部分缺失（日前或实时=0）",
        "col_status":           "状态",
        "status_missing":       "缺失",
        "status_partial":       "部分",
        "status_ok":            "正常",
        "btn_fill_gaps":        "回填 {n} 个PDF（含缺失日期）",
        "btn_reingest":         "重新录入范围内全部 {n} 个PDF",
        "warn_partial":         "{n} 个PDF存在部分数据（日前或实时缺失）。切换至「回填日期范围」模式可重新录入。",
        "all_present":          "所选范围内所有日期均已存在于数据库中。",
        "no_pdfs":              "所选日期范围内未找到PDF文件。",
        "upload_pdf":           "上传PDF报告",
        "upload_help":          "在AWS环境下（无本地数据文件夹时）上传PDF。文件保存至S3后即可导入。",
        "upload_btn":           "上传 {n} 个文件至S3",
        "upload_success":       "已上传 {n} 个文件至S3。",
        "prog_starting":        "启动中…",
        "prog_parsing":         "解析 {fname}…",
        "prog_interprov":       "省间数据：{fname}…",
        "prog_ai":              "AI摘要 {rdate}…",
        "prog_done":            "完成。",
        "backfill_complete":    "回填完成——已处理 {n} 个PDF。",
        "col_dates":            "日期",
        "col_rows":             "已写入行数",
        "col_interprov":        "省间行数",
        "col_ai":               "AI摘要数",
        "col_error":            "错误",
        # knowledge base sync
        "kb_sync_title":        "从市场基础数据同步知识库",
        "kb_sync_caption":      "扫描 data/market-fundamentals/ 文件夹，将新文档录入知识库。已录入的文件自动跳过。",
        "kb_sync_unavailable":  "知识库同步仅在本地模式下可用（未找到 data/market-fundamentals/ 文件夹）。",
        "kb_sync_no_files":     "未发现新文件——知识库已是最新。",
        "kb_sync_btn":          "同步 {n} 个新文件",
        "kb_sync_progress":     "正在录入 {i}/{n}：{fname}…",
        "kb_sync_done":         "同步完成——新增：{added}  跳过：{skipped}  错误：{errors}",
        # data export
        "export_title":         "下载价格数据",
        "export_caption":       "将数据库中的日前/实时均价导出为 Excel (.xlsx)。行=日期，列=省份（中文），单位=元/千瓦时。",
        "export_btn":           "下载 Excel",
        "export_filename":      "电力现货市场日均价格.xlsx",
        # wechat batch ingest
        "wechat_title":         "微信文章批量导入",
        "wechat_caption":       "每行粘贴一个微信公众号文章链接，系统将自动抓取正文并录入策略师知识库。",
        "wechat_url_label":     "文章链接（每行一个）",
        "wechat_url_placeholder": "https://mp.weixin.qq.com/s/...\nhttps://mp.weixin.qq.com/s/...",
        "wechat_run_btn":       "抓取并录入 {n} 篇文章",
        "wechat_no_urls":       "请在上方粘贴至少一个链接。",
        "wechat_fetching":      "正在处理 {i}/{n}：{url}…",
        "wechat_done":          "完成——新增：{added}  跳过：{skipped}  错误：{errors}",
        "wechat_digest_btn":    "摘要文章 → 洞察",
        # daily report
        "report_section_title": "每日市场报告",
        "report_section_caption": "每日自动生成含AI分析的日前/实时省份价格PDF报告，新加坡时间06:00发送邮件，并可推送至企业微信群。",
        "report_schedule_status": "调度状态",
        "report_next_run":      "下次运行",
        "report_send_now":      "立即发送报告",
        "report_send_success":  "报告已发送 — {size:,} 字节，日期：{rdate}",
        "report_send_wecom":    "企业微信：{result}",
        "report_send_error":    "报告发送失败：{err}",
        "report_email_label":   "收件人邮箱",
        "report_email_help":    "多个地址用逗号分隔。默认使用 REPORT_TO_EMAIL 环境变量。",
        "report_webhook_title": "企业微信群机器人",
        "report_webhook_caption": "配置多个企业微信群机器人 Webhook，保存至数据库（重启后不丢失）。",
        "report_webhook_add_label": "Webhook 链接",
        "report_webhook_add_label_label": "标签（可选）",
        "report_webhook_add_btn": "添加",
        "report_webhook_added": "Webhook 已添加。",
        "report_webhook_empty": "暂无 Webhook，企业微信推送已禁用。",
        "report_webhook_delete": "删除",
        "report_webhook_enabled": "启用",
        # chart labels
        "da_label":             "日前（DA）",
        "rt_label":             "实时（RT）",
        "da_avg_label":         "日前均价",
        "rt_avg_label":         "实时均价",
        "price_unit":           "元/千瓦时",
        "prob_density":         "概率密度",
        "price_axis":           "价格（元/千瓦时）",
        "spread_title":         "日前−实时价差（元/千瓦时）| 正值=日前溢价，负值=实时峰值",
        "da_clearing":          "日前（DA）出清价格（元/千瓦时）",
        "rt_clearing":          "实时（RT）出清价格（元/千瓦时）",
        "da_dist_title":        "日前（DA）价格分布（元/千瓦时）",
        "rt_dist_title":        "实时（RT）价格分布（元/千瓦时）",
        "da_violin_title":      "日前（DA）— 小提琴/箱线图（元/千瓦时）",
        "rt_violin_title":      "实时（RT）— 小提琴/箱线图（元/千瓦时）",
        "da_heatmap_title":     "日前平均出清价格 — 省份 × 日期热力图",
        "rt_heatmap_title":     "实时平均出清价格 — 省份 × 日期热力图",
        "geo_title_da":         "日前（DA）— 各省平均价格（元/千瓦时）",
        "geo_title_rt":         "实时（RT）— 各省平均价格（元/千瓦时）",
        # geo animation
        "anim_title":           "各月实时价格动画",
        "anim_range":           "动画时段",
        "anim_start_year":      "起始年",
        "anim_start_month":     "起始月",
        "anim_end_year":        "结束年",
        "anim_end_month":       "结束月",
        "anim_play":            "▶ 播放",
        "anim_pause":           "⏸ 暂停",
        "anim_speed":           "每帧秒数",
        "anim_slider":          "选择月份",
        "anim_no_data":         "该月无数据。",
        "anim_map_title":       "实时均价 — {month}",
        # geo comparison
        "cmp_title":            "时段对比",
        "cmp_metric":           "指标",
        "cmp_period_a":         "时段 A",
        "cmp_period_b":         "时段 B",
        "cmp_start":            "开始日期",
        "cmp_end":              "结束日期",
        "cmp_no_data":          "该时段无数据。",
        "cmp_map_title":        "{metric} 均价 — {start} → {end}",
    },
}


def _t(key: str, **kwargs) -> str:
    """Return translated string for the current language selection."""
    lang = "zh" if st.session_state.get("lang_radio") == "中文" else "en"
    s = _T[lang].get(key, _T["en"].get(key, key))
    return s.format(**kwargs) if kwargs else s


# ── page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Spot Market Cockpit",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── DB connection ─────────────────────────────────────────────────────────────
@st.cache_resource
def __conn():
    url = (
        os.environ.get("PGURL")
        or os.environ.get("DATABASE_URL")
        or os.environ.get("DB_URL")
        or "postgresql://postgres:root@127.0.0.1:5433/marketdata"
    )
    return psycopg2.connect(url, keepalives=1, keepalives_idle=60,
                            keepalives_interval=10, keepalives_count=5)

def _conn():
    conn = __conn()
    try:
        conn.cursor().execute("SELECT 1")
    except Exception:
        __conn.clear()
        conn = __conn()
    return conn


# ── Process-level caches for knowledge-pool calls that open fresh connections ──
# knowledge_pool/db.py's _conn() opens a NEW TCP connection on every call.
# These @st.cache_data wrappers limit that to once per 5 minutes per process,
# regardless of how many Streamlit sessions are active.  Without this, each new
# browser session triggered 2-3 fresh DB connections before tab_mgmt could render.
@st.cache_data(ttl=300, show_spinner=False)
def _cached_memory_stats() -> dict:
    try:
        from services.knowledge_pool.expert_memory import get_memory_stats as _gms
        return _gms()
    except Exception:
        return {}


@st.cache_data(ttl=300, show_spinner=False)
def _cached_kb_docs() -> list:
    try:
        from services.knowledge_pool.knowledge_docs import list_knowledge_docs as _lkd
        return _lkd()
    except Exception:
        return []


@st.cache_data(ttl=300, show_spinner=False)
def _spot_kb_stats() -> dict:
    """Return spot market KB + insight statistics for the metrics dashboard."""
    docs_row = _query(
        "SELECT COUNT(*)::int AS total, "
        "COUNT(*) FILTER (WHERE parse_error IS NULL AND active = TRUE)::int AS parsed, "
        "COUNT(*) FILTER (WHERE active = TRUE)::int AS active_docs "
        "FROM staging.spot_knowledge_docs"
    )
    chunks_row = _query("SELECT COUNT(*)::int AS n FROM staging.spot_knowledge_chunks")
    insights_row = _query(
        "SELECT COUNT(*)::int AS n FROM staging.kp_expert_insights WHERE active = TRUE"
    )
    total   = int(docs_row.iloc[0]["total"])       if not docs_row.empty     else 0
    parsed  = int(docs_row.iloc[0]["parsed"])      if not docs_row.empty     else 0
    active  = int(docs_row.iloc[0]["active_docs"]) if not docs_row.empty     else 0
    n_chunks    = int(chunks_row.iloc[0]["n"])     if not chunks_row.empty   else 0
    n_insights  = int(insights_row.iloc[0]["n"])   if not insights_row.empty else 0
    return {
        "total":       total,
        "parsed":      parsed,
        "active":      active,
        "parse_pct":   parsed / total if total > 0 else 0.0,
        "n_chunks":    n_chunks,
        "n_insights":  n_insights,
    }


# ── APScheduler — daily spot market report ────────────────────────────────────
import os as _os
_DEFAULT_RECIPIENT = "chen_dpeng@hotmail.com"

@st.cache_resource
def _start_spot_scheduler():
    """Start background scheduler for daily spot market report (runs once per process)."""
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
    except ImportError:
        return None

    def _daily_spot_report_job():
        """06:00 SGT — generate PDF, email, WeCom."""
        try:
            import importlib.util, pathlib as _pl
            _spec = importlib.util.spec_from_file_location(
                "spot_report",
                _pl.Path(__file__).parent / "spot_report.py",
            )
            _mod = importlib.util.module_from_spec(_spec)
            _spec.loader.exec_module(_mod)
            result = _mod.run_daily_report()
            import logging as _logging
            _logging.getLogger(__name__).info(
                "Scheduled spot report: %s", result
            )
        except Exception as _exc:
            import logging as _logging
            _logging.getLogger(__name__).error(
                "Scheduled spot report failed: %s", _exc, exc_info=True
            )

    scheduler = BackgroundScheduler(timezone="Asia/Singapore")
    scheduler.add_job(
        _daily_spot_report_job,
        "cron", hour=6, minute=0,
        id="spot_daily_report",
        misfire_grace_time=3600,
    )
    scheduler.start()
    return scheduler


_start_spot_scheduler()  # no-op after first call (cache_resource)


@st.cache_resource
def _get_spot_report_mod():
    """Load spot_report module once per process (exec_module is expensive — loads ReportLab etc.)."""
    import importlib.util as _ilu, pathlib as _pl2
    _spec = _ilu.spec_from_file_location(
        "spot_report",
        _pl2.Path(__file__).parent / "spot_report.py",
    )
    _mod = _ilu.module_from_spec(_spec)
    _spec.loader.exec_module(_mod)
    try:
        _mod.ensure_webhook_table()   # DDL once per process
    except Exception:
        pass
    return _mod


@st.cache_data(ttl=300, show_spinner=False)
def _cached_webhooks() -> list:
    """Cached webhook list — avoids a fresh DB connection on every rerun."""
    try:
        return _get_spot_report_mod().list_webhooks()
    except Exception:
        return []


# ── intraday hourly price loaders ─────────────────────────────────────────────
@st.cache_data(ttl=3600)
def _load_intraday_shape(_conn_fn, provinces: tuple, start: str, end: str, price_col: str):
    sql = f"""
        SELECT province,
               EXTRACT(hour FROM datetime)::int AS hour,
               AVG({price_col}) AS avg_price
        FROM marketdata.spot_prices_hourly
        WHERE province = ANY(%s) AND datetime BETWEEN %s AND %s
        GROUP BY province, hour ORDER BY province, hour
    """
    return pd.read_sql(sql, _conn_fn(), params=[list(provinces), start, end])

@st.cache_data(ttl=3600)
def _load_intraday_spread(_conn_fn, start: str, end: str, price_col: str):
    sql = f"""
        SELECT province, MAX(avg_price) - MIN(avg_price) AS spread
        FROM (
            SELECT province, EXTRACT(hour FROM datetime)::int AS hour,
                   AVG({price_col}) AS avg_price
            FROM marketdata.spot_prices_hourly
            WHERE datetime BETWEEN %s AND %s
            GROUP BY province, hour
        ) t GROUP BY province ORDER BY spread DESC
    """
    return pd.read_sql(sql, _conn_fn(), params=[start, end])

@st.cache_data(ttl=3600)
def _load_hourly_series(_conn_fn, province: str, start: str, end: str, price_col: str):
    sql = f"""
        SELECT {price_col} AS price FROM marketdata.spot_prices_hourly
        WHERE province = %s AND datetime BETWEEN %s AND %s ORDER BY datetime
    """
    return pd.read_sql(sql, _conn_fn(), params=[province, start, end])["price"].dropna()


@st.cache_data(ttl=3600)
def _load_bidding_vs_price(_conn_fn, provinces: tuple, start: str, end: str):
    """Hourly bidding space (RT outturn) vs RT price — for scatter/OLS."""
    sql = """
        SELECT f.province,
               f.bidding_space_mw,
               p.rt_price
        FROM marketdata.spot_fundamentals_hourly f
        JOIN marketdata.spot_prices_hourly p
          ON p.province = f.province AND p.datetime = f.datetime
        WHERE f.province = ANY(%s)
          AND f.datetime BETWEEN %s AND %s
          AND f.bidding_space_mw IS NOT NULL
          AND p.rt_price IS NOT NULL
        ORDER BY f.province, f.datetime
    """
    return pd.read_sql(sql, _conn_fn(), params=[list(provinces), start, end])


@st.cache_data(ttl=3600)
def _load_wap(_conn_fn, provinces: tuple, start: str, end: str):
    """Wind and solar weighted average price vs avg RT price, by province."""
    sql = """
        SELECT f.province,
               SUM(f.wind_mw  * p.rt_price) / NULLIF(SUM(f.wind_mw),  0) AS wind_wap,
               SUM(f.solar_mw * p.rt_price) / NULLIF(SUM(f.solar_mw), 0) AS solar_wap,
               AVG(p.rt_price)                                             AS avg_rt_price
        FROM marketdata.spot_fundamentals_hourly f
        JOIN marketdata.spot_prices_hourly p
          ON p.province = f.province AND p.datetime = f.datetime
        WHERE f.province = ANY(%s)
          AND f.datetime BETWEEN %s AND %s
          AND (f.wind_mw IS NOT NULL OR f.solar_mw IS NOT NULL)
        GROUP BY f.province
        ORDER BY f.province
    """
    return pd.read_sql(sql, _conn_fn(), params=[list(provinces), start, end])


# ── price forecast data loaders ───────────────────────────────────────────────
@st.cache_data(ttl=3600)
def _load_price_matrix(_conn_fn, province: str, start: str, end: str, price_col: str):
    """Load hourly prices for a province, pivot to (date × 24h) matrix for PCA."""
    sql = f"""
        SELECT DATE(datetime) AS trade_date,
               EXTRACT(hour FROM datetime)::int AS hour,
               AVG({price_col}) AS avg_price
        FROM marketdata.spot_prices_hourly
        WHERE province = %s AND datetime BETWEEN %s AND %s
        GROUP BY trade_date, hour
        ORDER BY trade_date, hour
    """
    df = pd.read_sql(sql, _conn_fn(), params=[province, start, end])
    if df.empty:
        return pd.DataFrame()
    pivot = df.pivot(index='trade_date', columns='hour', values='avg_price')
    pivot.columns = [int(c) for c in pivot.columns]
    # Keep only days with all 24 hours present
    pivot = pivot.dropna(thresh=20)
    return pivot.ffill(axis=1).bfill(axis=1)


@st.cache_data(ttl=3600)
def _load_forecast_fundamentals(_conn_fn, province: str, start: str, end: str):
    """Load hourly load/wind/solar for stack model input."""
    sql = """
        SELECT EXTRACT(hour FROM datetime)::int AS hour,
               AVG(load_mw) AS avg_load,
               AVG(COALESCE(wind_mw, 0)) AS avg_wind,
               AVG(COALESCE(solar_mw, 0)) AS avg_solar
        FROM marketdata.spot_fundamentals_hourly
        WHERE province = %s AND datetime BETWEEN %s AND %s
          AND load_mw > 0
        GROUP BY hour ORDER BY hour
    """
    return pd.read_sql(sql, _conn_fn(), params=[province, start, end])


@st.cache_data(ttl=3600)
def _load_price_holdout(_conn_fn, province: str, start: str, end: str, price_col: str):
    """Load actual hourly prices for backtest evaluation window."""
    sql = f"""
        SELECT DATE(datetime) AS trade_date,
               EXTRACT(hour FROM datetime)::int AS hour,
               AVG({price_col}) AS actual_price
        FROM marketdata.spot_prices_hourly
        WHERE province = %s AND datetime BETWEEN %s AND %s
          AND {price_col} IS NOT NULL AND {price_col} > 0
        GROUP BY trade_date, hour
        ORDER BY trade_date, hour
    """
    return pd.read_sql(sql, _conn_fn(), params=[province, start, end])


@st.cache_data(ttl=3600)
def _load_hourly_price_provinces(_conn_fn):
    """Provinces that have at least 30 days of hourly price data (da_price or rt_price)."""
    sql = """
        SELECT DISTINCT province
        FROM marketdata.spot_prices_hourly
        WHERE da_price IS NOT NULL OR rt_price IS NOT NULL
        GROUP BY province
        HAVING COUNT(DISTINCT DATE(datetime)) >= 30
        ORDER BY province
    """
    df = pd.read_sql(sql, _conn_fn())
    return sorted(df['province'].tolist()) if not df.empty else []


# ── data quality filter ───────────────────────────────────────────────────────
def _apply_quality_filter(df: pd.DataFrame) -> pd.DataFrame:
    # Only remove rows where avg itself is impossible (bad_range).
    # When avg > max or avg < min, the max/min is wrong (parser column mismatch)
    # but the avg is still valid — nullify the bad bound instead of dropping the row.
    df = df.copy()
    for m in ("da", "rt"):
        avg, mx, mn = f"{m}_avg", f"{m}_max", f"{m}_min"
        bad_range = df[avg].notna() & ((df[avg] < -0.5) | (df[avg] > 2.0))
        df = df[~bad_range]
        bad_hi = df[avg].notna() & df[mx].notna() & (df[avg] > df[mx] + 0.001)
        bad_lo = df[avg].notna() & df[mn].notna() & (df[avg] < df[mn] - 0.001)
        df.loc[bad_hi, mx] = None
        df.loc[bad_lo, mn] = None
        for col in (mx, mn):
            df.loc[df[col].notna() & ((df[col] > 2.0) | (df[col] < -1.0)), col] = None
    return df

# ── data loaders ─────────────────────────────────────────────────────────────
@st.cache_data(ttl=120, show_spinner=False)
def load_all(start: date, end: date, quality_filter: bool) -> pd.DataFrame:
    q = """
        SELECT report_date::date AS report_date,
               province_en, province_cn,
               da_avg, da_max, da_min,
               rt_avg, rt_max, rt_min
        FROM spot_daily
        WHERE report_date BETWEEN %s AND %s
          AND (da_avg IS NOT NULL OR rt_avg IS NOT NULL)
        ORDER BY report_date, province_en
    """
    df = pd.read_sql(q, _conn(), params=(start, end), parse_dates=["report_date"])
    for c in ["da_avg","da_max","da_min","rt_avg","rt_max","rt_min"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    if quality_filter:
        df = _apply_quality_filter(df)
    return df


@st.cache_data(ttl=120, show_spinner=False)
def load_provinces() -> list[str]:
    cur = _conn().cursor()
    cur.execute(
        "SELECT DISTINCT province_en FROM spot_daily "
        "WHERE report_date >= '2026-01-01' ORDER BY 1"
    )
    return [r[0] for r in cur.fetchall()]


@st.cache_data(ttl=60, show_spinner=False)
def load_kpis(quality_filter: bool) -> dict:
    cur = _conn().cursor()
    cur.execute("""
        SELECT
            MAX(report_date)                              AS latest_date,
            COUNT(DISTINCT report_date)                   AS total_dates,
            COUNT(DISTINCT province_en)                   AS total_provinces,
            SUM(CASE WHEN da_avg IS NOT NULL AND rt_avg IS NOT NULL THEN 1 ELSE 0 END) AS complete_rows,
            COUNT(*)                                      AS total_rows
        FROM spot_daily
        WHERE report_date >= '2026-01-01'
    """)
    r = cur.fetchone()
    return {
        "latest_date":     r[0],
        "total_dates":     r[1],
        "total_provinces": r[2],
        "complete_rows":   r[3],
        "total_rows":      r[4],
    }


@st.cache_data(ttl=60, show_spinner=False)
def load_interprov(start: date, end: date) -> tuple[pd.DataFrame, str]:
    try:
        cur = _conn().cursor()
        cur.execute("""
            SELECT report_date::date, direction, metric_type,
                   province_cn, province_share,
                   price_yuan_kwh, price_chg_pct,
                   time_period, total_vol_100gwh, source_pdf
            FROM staging.spot_interprov_flow
            WHERE report_date BETWEEN %s AND %s
            ORDER BY report_date, direction, metric_type
        """, (start, end))
        cols = [d[0] for d in cur.description]
        df = pd.DataFrame(cur.fetchall(), columns=cols)
        for c in ["price_yuan_kwh", "price_chg_pct", "total_vol_100gwh", "province_share"]:
            df[c] = pd.to_numeric(df[c], errors="coerce")
        return df, ""
    except Exception as e:
        return pd.DataFrame(), str(e)


@st.cache_data(ttl=60, show_spinner=False)
def load_summaries(start: date, end: date) -> tuple[pd.DataFrame, str]:
    try:
        cur = _conn().cursor()
        cur.execute("""
            SELECT report_date::date, summary_text, model, source_pdf
            FROM staging.spot_report_summaries
            WHERE report_date BETWEEN %s AND %s
            ORDER BY report_date DESC
        """, (start, end))
        cols = [d[0] for d in cur.description]
        return pd.DataFrame(cur.fetchall(), columns=cols), ""
    except Exception as e:
        return pd.DataFrame(), str(e)


@st.cache_data(ttl=86400, show_spinner=False)
def _translate_to_zh(text: str) -> str:
    """Translate an English market summary to Chinese. Cached per text (one API call per unique summary)."""
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return text
    try:
        import anthropic as _ant_tr
        msg = _ant_tr.Anthropic(api_key=api_key).messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=800,
            messages=[{
                "role": "user",
                "content": (
                    "Translate the following China electricity market daily summary to "
                    "Chinese (简体中文). Keep all numbers, units (¥/kWh, MW, GWh, 亿kWh), "
                    "and province names unchanged. Output only the translated text.\n\n"
                    + text
                ),
            }],
        )
        return msg.content[0].text
    except Exception:
        return text


# ── colour helpers ────────────────────────────────────────────────────────────
_PALETTE = px.colors.qualitative.Plotly + px.colors.qualitative.Dark24

def _prov_colour(provinces: list[str]) -> dict[str, str]:
    return {p: _PALETTE[i % len(_PALETTE)] for i, p in enumerate(sorted(provinces))}


# ── chart builders ────────────────────────────────────────────────────────────
def chart_timeseries(df: pd.DataFrame, provinces: list[str],
                     metric: str, show_band: bool) -> go.Figure:
    fig = go.Figure()
    colours = _prov_colour(provinces)
    avg_col, max_col, min_col = f"{metric}_avg", f"{metric}_max", f"{metric}_min"

    for prov in sorted(provinces):
        sub = df[df["province_en"] == prov].sort_values("report_date")
        if sub.empty or sub[avg_col].isna().all():
            continue
        col = colours[prov]

        sub_band = sub[sub[avg_col].notna()]
        if show_band and sub_band[max_col].notna().any():
            x_band = pd.concat([sub_band["report_date"], sub_band["report_date"].iloc[::-1]])
            y_band = pd.concat([sub_band[max_col], sub_band[min_col].iloc[::-1]])
            fig.add_trace(go.Scatter(
                x=x_band, y=y_band,
                fill="toself", fillcolor=col, opacity=0.10,
                line=dict(width=0), showlegend=False, hoverinfo="skip",
            ))

        fig.add_trace(go.Scatter(
            x=sub["report_date"], y=sub[avg_col],
            name=prov, mode="lines+markers",
            line=dict(color=col, width=1.8), marker=dict(size=4),
            hovertemplate="%{x|%Y-%m-%d}<br>%{y:.4f} ¥/kWh<extra>" + prov + "</extra>",
        ))

    title_key = "da_clearing" if metric == "da" else "rt_clearing"
    fig.update_layout(
        height=430,
        title=dict(text=_t(title_key), font=dict(size=14)),
        margin=dict(l=10, r=10, t=45, b=90),
        legend=dict(orientation="h", yanchor="top", y=-0.18,
                    xanchor="center", x=0.5, font=dict(size=11)),
        xaxis=dict(showgrid=True, gridcolor="#f0f0f0"),
        yaxis=dict(showgrid=True, gridcolor="#f0f0f0", tickformat=".3f"),
        plot_bgcolor="white", paper_bgcolor="white",
        hovermode="x unified",
    )
    return fig


def chart_da_rt_overlay(df: pd.DataFrame, province: str) -> go.Figure:
    sub = df[df["province_en"] == province].sort_values("report_date")
    fig = go.Figure()

    for metric, label_key, colour in [
        ("da", "da_avg_label", "#1f77b4"),
        ("rt", "rt_avg_label", "#ff7f0e"),
    ]:
        avg_col, max_col, min_col = f"{metric}_avg", f"{metric}_max", f"{metric}_min"
        if sub[avg_col].isna().all():
            continue
        label = _t(label_key)
        if sub[max_col].notna().any():
            fig.add_trace(go.Scatter(
                x=pd.concat([sub["report_date"], sub["report_date"].iloc[::-1]]),
                y=pd.concat([sub[max_col], sub[min_col].iloc[::-1]]),
                fill="toself", fillcolor=colour, opacity=0.12,
                line=dict(width=0), showlegend=False, hoverinfo="skip",
            ))
        fig.add_trace(go.Scatter(
            x=sub["report_date"], y=sub[avg_col],
            name=label, mode="lines+markers",
            line=dict(color=colour, width=2), marker=dict(size=4),
            hovertemplate="%{x|%Y-%m-%d}<br>%{y:.4f} ¥/kWh<extra>" + label + "</extra>",
        ))

    fig.update_layout(
        height=390,
        title=dict(text=f"{province} — {_t('da_avg_label')} vs {_t('rt_avg_label')}  ({_t('price_unit')})",
                   font=dict(size=13)),
        margin=dict(l=10, r=10, t=45, b=60),
        legend=dict(orientation="h", yanchor="top", y=-0.15,
                    xanchor="center", x=0.5, font=dict(size=11)),
        xaxis=dict(showgrid=True, gridcolor="#f0f0f0"),
        yaxis=dict(showgrid=True, gridcolor="#f0f0f0", tickformat=".3f"),
        plot_bgcolor="white", paper_bgcolor="white",
        hovermode="x unified",
    )
    return fig


def chart_spread(df: pd.DataFrame, provinces: list[str]) -> go.Figure:
    fig = go.Figure()
    colours = _prov_colour(provinces)

    for prov in sorted(provinces):
        sub = df[df["province_en"] == prov].dropna(subset=["da_avg", "rt_avg"]).copy()
        if sub.empty:
            continue
        sub["spread"] = sub["da_avg"] - sub["rt_avg"]
        fig.add_trace(go.Bar(
            x=sub["report_date"], y=sub["spread"],
            name=prov, marker_color=colours[prov], opacity=0.8,
            hovertemplate="%{x|%Y-%m-%d}<br>Spread: %{y:.4f} ¥/kWh<extra>" + prov + "</extra>",
        ))

    fig.add_hline(y=0, line_width=1, line_color="black", opacity=0.5)
    fig.update_layout(
        height=360, barmode="group",
        title=dict(text=_t("spread_title"), font=dict(size=13)),
        margin=dict(l=10, r=10, t=45, b=90),
        legend=dict(orientation="h", yanchor="top", y=-0.22,
                    xanchor="center", x=0.5, font=dict(size=11)),
        xaxis=dict(showgrid=False),
        yaxis=dict(showgrid=True, gridcolor="#f0f0f0", tickformat=".3f"),
        plot_bgcolor="white", paper_bgcolor="white",
        hovermode="x unified",
    )
    return fig


def chart_heatmap(df_sel: pd.DataFrame, metric: str) -> go.Figure:
    """Province × Date daily average price heatmap from spot_daily.

    df_sel: DataFrame with columns province_en, report_date, da_avg, rt_avg.
    """
    if df_sel.empty:
        return go.Figure()
    val_col = "da_avg" if metric == "da" else "rt_avg"
    pivot = df_sel.pivot_table(index="province_en", columns="report_date", values=val_col)
    if pivot.empty:
        return go.Figure()

    title_key = "da_heatmap_title" if metric == "da" else "rt_heatmap_title"
    _vals = pivot.values.flatten()
    _vals = _vals[~np.isnan(_vals)]
    zmin = float(np.percentile(_vals, 2)) if len(_vals) else None
    zmax = float(np.percentile(_vals, 98)) if len(_vals) else None

    x_labels = [d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d) for d in pivot.columns]
    fig = go.Figure(go.Heatmap(
        z=pivot.values,
        x=x_labels,
        y=pivot.index.tolist(),
        colorscale="RdYlGn_r",
        zmin=zmin,
        zmax=zmax,
        colorbar=dict(title=_t("price_unit"), thickness=12),
        hoverongaps=False,
        hovertemplate="Date: %{x}<br>Province: %{y}<br>Avg Price: %{z:.4f} ¥/kWh<extra></extra>",
    ))
    fig.update_layout(
        height=max(350, len(pivot) * 28),
        title=dict(text=_t(title_key), font=dict(size=13)),
        margin=dict(l=120, r=20, t=45, b=60),
        xaxis=dict(tickfont=dict(size=9), tickangle=-45),
        yaxis=dict(tickfont=dict(size=11)),
    )
    return fig


def chart_distributions(df: pd.DataFrame, provinces: list[str],
                         metric: str, nbins: int, show_kde: bool) -> go.Figure:
    avg_col = f"{metric}_avg"
    colours = _prov_colour(provinces)
    fig = go.Figure()

    for prov in sorted(provinces):
        vals = df[df["province_en"] == prov][avg_col].dropna().values
        if len(vals) < 2:
            continue
        col = colours[prov]

        fig.add_trace(go.Histogram(
            x=vals,
            name=prov,
            nbinsx=nbins,
            marker_color=col,
            opacity=0.45,
            histnorm="probability density",
            hovertemplate="Price: %{x:.4f} ¥/kWh<br>Density: %{y:.3f}<extra>" + prov + "</extra>",
        ))

        if show_kde and len(vals) >= 5:
            std = vals.std()
            if std > 0:
                bw = 1.06 * std * len(vals) ** (-0.2)
                x_grid = np.linspace(vals.min() - 2 * bw, vals.max() + 2 * bw, 300)
                kde = np.zeros_like(x_grid)
                for v in vals:
                    kde += np.exp(-0.5 * ((x_grid - v) / bw) ** 2)
                kde /= len(vals) * bw * np.sqrt(2 * np.pi)
                fig.add_trace(go.Scatter(
                    x=x_grid, y=kde,
                    name=f"{prov} KDE",
                    mode="lines",
                    line=dict(color=col, width=2, dash="solid"),
                    showlegend=False,
                    hovertemplate="%{x:.4f} ¥/kWh<br>KDE: %{y:.3f}<extra>" + prov + "</extra>",
                ))

    title_key = "da_dist_title" if metric == "da" else "rt_dist_title"
    fig.update_layout(
        height=430,
        barmode="overlay",
        title=dict(text=_t(title_key), font=dict(size=14)),
        margin=dict(l=10, r=10, t=45, b=90),
        legend=dict(orientation="h", yanchor="top", y=-0.18,
                    xanchor="center", x=0.5, font=dict(size=11)),
        xaxis=dict(title=_t("price_axis"), showgrid=True, gridcolor="#f0f0f0"),
        yaxis=dict(title=_t("prob_density"), showgrid=True, gridcolor="#f0f0f0"),
        plot_bgcolor="white", paper_bgcolor="white",
    )
    return fig


def chart_violin(df: pd.DataFrame, provinces: list[str], metric: str) -> go.Figure:
    avg_col = f"{metric}_avg"
    colours = _prov_colour(provinces)
    fig = go.Figure()

    for prov in sorted(provinces):
        vals = df[df["province_en"] == prov][avg_col].dropna().values
        if len(vals) < 3:
            continue
        fig.add_trace(go.Violin(
            y=vals, name=prov,
            box_visible=True,
            meanline_visible=True,
            fillcolor=colours[prov],
            opacity=0.65,
            line_color=colours[prov],
            hoverinfo="y+name",
        ))

    title_key = "da_violin_title" if metric == "da" else "rt_violin_title"
    fig.update_layout(
        height=430,
        title=dict(text=_t(title_key), font=dict(size=14)),
        margin=dict(l=10, r=10, t=45, b=90),
        legend=dict(orientation="h", yanchor="top", y=-0.18,
                    xanchor="center", x=0.5, font=dict(size=11)),
        yaxis=dict(title=_t("price_axis"), showgrid=True, gridcolor="#f0f0f0", tickformat=".3f"),
        plot_bgcolor="white", paper_bgcolor="white",
        violinmode="group",
    )
    return fig


def _dist_stats(df: pd.DataFrame, provinces: list[str], metric: str) -> pd.DataFrame:
    avg_col = f"{metric}_avg"
    rows = []
    for prov in sorted(provinces):
        vals = df[df["province_en"] == prov][avg_col].dropna()
        if vals.empty:
            continue
        rows.append({
            _t("col_province"): prov,
            _t("col_n"):        len(vals),
            _t("col_mean"):     f"{vals.mean():.4f}",
            _t("col_median"):   f"{vals.median():.4f}",
            _t("col_std"):      f"{vals.std():.4f}",
            _t("col_p10"):      f"{vals.quantile(0.10):.4f}",
            _t("col_p25"):      f"{vals.quantile(0.25):.4f}",
            _t("col_p75"):      f"{vals.quantile(0.75):.4f}",
            _t("col_p90"):      f"{vals.quantile(0.90):.4f}",
            _t("col_min"):      f"{vals.min():.4f}",
            _t("col_max"):      f"{vals.max():.4f}",
        })
    return pd.DataFrame(rows)


# ── Geo map helpers ───────────────────────────────────────────────────────────

_PROV_ADCODE: dict[str, str] = {
    "Beijing":      "110000", "Tianjin":     "120000",
    "Hebei":        "130000", "Hebei-North": "130000", "Hebei-South": "130000",
    "Shanxi":       "140000",
    "Mengxi":       "150000", "Mengdong":    "150000",
    "Liaoning":     "210000", "Jilin":       "220000", "Heilongjiang": "230000",
    "Shanghai":     "310000", "Jiangsu":     "320000", "Zhejiang":     "330000",
    "Anhui":        "340000", "Fujian":      "350000", "Jiangxi":      "360000",
    "Shandong":     "370000", "Henan":       "410000", "Hubei":        "420000",
    "Hunan":        "430000", "Guangdong":   "440000", "Guangxi":      "450000",
    "Hainan":       "460000", "Chongqing":   "500000", "Sichuan":      "510000",
    "Guizhou":      "520000", "Yunnan":      "530000",
    "Shaanxi":      "610000", "Gansu":       "620000", "Qinghai":      "630000",
    "Ningxia":      "640000", "Xinjiang":    "650000",
}

_PROV_CENTROIDS: dict[str, tuple[float, float]] = {
    "110000": (39.90, 116.40), "120000": (39.13, 117.20),
    "130000": (38.04, 114.47), "140000": (37.87, 112.56),
    "150000": (44.09, 113.09), "210000": (41.80, 123.43),
    "220000": (43.89, 125.32), "230000": (47.85, 127.57),
    "310000": (31.23, 121.47), "320000": (32.06, 119.59),
    "330000": (30.27, 120.15), "340000": (31.86, 117.29),
    "350000": (26.10, 118.31), "360000": (27.62, 115.70),
    "370000": (36.67, 117.02), "410000": (34.76, 113.75),
    "420000": (30.60, 114.30), "430000": (28.23, 112.94),
    "440000": (23.37, 113.50), "450000": (23.73, 108.38),
    "460000": (20.02, 110.35), "500000": (29.56, 106.54),
    "510000": (30.57, 103.99), "520000": (26.82, 106.83),
    "530000": (25.05, 101.71), "610000": (34.27, 108.95),
    "620000": (36.06, 103.83), "630000": (36.62, 101.74),
    "640000": (38.47, 106.26), "650000": (41.17,  85.29),
}

_ADCODE_LABEL: dict[str, str] = {
    "110000": "Beijing",        "120000": "Tianjin",
    "130000": "Hebei",          "140000": "Shanxi",
    "150000": "Inner Mongolia", "210000": "Liaoning",
    "220000": "Jilin",          "230000": "Heilongjiang",
    "310000": "Shanghai",       "320000": "Jiangsu",
    "330000": "Zhejiang",       "340000": "Anhui",
    "350000": "Fujian",         "360000": "Jiangxi",
    "370000": "Shandong",       "410000": "Henan",
    "420000": "Hubei",          "430000": "Hunan",
    "440000": "Guangdong",      "450000": "Guangxi",
    "460000": "Hainan",         "500000": "Chongqing",
    "510000": "Sichuan",        "520000": "Guizhou",
    "530000": "Yunnan",         "610000": "Shaanxi",
    "620000": "Gansu",          "630000": "Qinghai",
    "640000": "Ningxia",        "650000": "Xinjiang",
}

_LOW_PRICE  = 0.20
_HIGH_PRICE = 0.30

_GEO_FILE = Path(__file__).parent / "data" / "china_provinces.geojson"

_GEO_COLORSCALE = [
    [0.00, "#00aa44"],
    [0.40, "#ffe000"],
    [0.60, "#ff6600"],
    [1.00, "#cc0000"],
]
_GEO_ZMIN, _GEO_ZMAX = 0.0, 0.5


def _price_level(v: float | None) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "—"
    if v < _LOW_PRICE:
        return _t("level_low")
    if v <= _HIGH_PRICE:
        return _t("level_medium")
    return _t("level_high")


def _level_bg(level: str) -> str:
    return {
        "Low": "#d4edda", "低": "#d4edda",
        "Medium": "#fff3cd", "中": "#fff3cd",
        "High": "#ffe0e0", "高": "#ffe0e0",
    }.get(level, "")


@st.cache_data(ttl=None, show_spinner=False)
def _load_china_geojson() -> tuple[dict | None, str | None]:
    if _GEO_FILE.exists():
        try:
            return json.loads(_GEO_FILE.read_text(encoding="utf-8")), None
        except Exception:
            pass

    try:
        resp = requests.get(
            "https://geo.datav.aliyun.com/areas_v3/bound/100000_full.json",
            timeout=20,
        )
        resp.raise_for_status()
        gj = resp.json()
        if len(gj.get("features", [])) < 10:
            return None, "GeoJSON has too few features — unexpected format"
        _GEO_FILE.parent.mkdir(parents=True, exist_ok=True)
        _GEO_FILE.write_text(json.dumps(gj), encoding="utf-8")
        return gj, None
    except Exception as exc:
        return None, str(exc)


def _geo_agg(df: pd.DataFrame, metric: str) -> pd.DataFrame:
    avg_col = f"{metric}_avg"
    df2 = df.copy()
    df2["adcode"] = df2["province_en"].map(_PROV_ADCODE)
    df2 = df2.dropna(subset=["adcode", avg_col])
    if df2.empty:
        return pd.DataFrame(columns=["adcode", "avg", "label", "price_str"])
    agg = df2.groupby("adcode", as_index=False)[avg_col].mean()
    agg.columns = ["adcode", "avg"]
    agg["label"]     = agg["adcode"].map(_ADCODE_LABEL)
    agg["price_str"] = agg["avg"].map(lambda v: f"{v:.2f}")
    return agg


def _make_china_cmap() -> mcolors.LinearSegmentedColormap:
    stops = [(pos, mcolors.to_rgb(hex_col)) for pos, hex_col in _GEO_COLORSCALE]
    return mcolors.LinearSegmentedColormap.from_list("china_price", stops)


def chart_geo_map(df: pd.DataFrame, metric: str, geojson: dict | None,
                  title: str | None = None) -> plt.Figure:
    agg = _geo_agg(df, metric)
    title_key = "geo_title_da" if metric == "da" else "geo_title_rt"
    display_title = title if title is not None else _t(title_key)

    # Use a CJK-capable font for Chinese labels when one is available
    _lang = st.session_state.get("lang_radio", "English")
    _rc_font = ({"font.family": _CJK_FONT} if _lang == "中文" and _CJK_FONT else {})

    cmap = _make_china_cmap()
    norm = mcolors.Normalize(vmin=_GEO_ZMIN, vmax=_GEO_ZMAX)

    price_map: dict[int, float] = {}
    if not agg.empty:
        for _, row in agg.iterrows():
            try:
                price_map[int(row["adcode"])] = float(row["avg"])
            except (ValueError, TypeError):
                pass

    with plt.rc_context(_rc_font):
        fig, ax = plt.subplots(figsize=(9, 6), facecolor="white")
        ax.set_facecolor("#b8d4f0")

        if geojson:
            for feat in geojson.get("features", []):
                adcode_int = feat.get("properties", {}).get("adcode")
                price = price_map.get(adcode_int)
                fc = cmap(norm(price)) if price is not None else "#d0d0d0"

                geom = feat.get("geometry", {})
                rings: list = []
                if geom.get("type") == "Polygon":
                    rings = [geom["coordinates"][0]]
                elif geom.get("type") == "MultiPolygon":
                    rings = [p[0] for p in geom["coordinates"]]

                for ring in rings:
                    coords = np.array(ring)
                    ax.add_patch(MplPolygon(
                        coords, closed=True,
                        facecolor=fc, edgecolor="white", linewidth=0.8,
                    ))

        if not agg.empty:
            for _, row in agg.iterrows():
                coord = _PROV_CENTROIDS.get(row["adcode"])
                if coord:
                    lat, lon = coord
                    ax.text(lon, lat, row["price_str"], ha="center", va="center",
                            fontsize=7, fontweight="bold", color="black")

        ax.set_xlim(72, 137)
        ax.set_ylim(16, 54)
        ax.set_aspect("equal")
        ax.axis("off")

        sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
        sm.set_array([])
        cbar = fig.colorbar(sm, ax=ax, orientation="vertical", fraction=0.025, pad=0.01, aspect=25)
        cbar.set_label(_t("price_unit"), fontsize=9)
        cbar.set_ticks([0.0, 0.1, 0.2, 0.3, 0.4, 0.5])
        cbar.set_ticklabels(["0.0", "0.1", "0.2", "0.3", "0.4", "0.5+"])
        cbar.ax.tick_params(labelsize=8)

        ax.set_title(display_title, fontsize=11, pad=10)
        plt.tight_layout(pad=0.5)
    return fig


# ─────────────────────────────────────────────────────────────────────────────
# LAYOUT
# ─────────────────────────────────────────────────────────────────────────────

st.title(_t("app_title"))

# ── KPI strip ────────────────────────────────────────────────────────────────
with st.spinner("Loading…"):
    try:
        provinces_all = load_provinces()
    except Exception as e:
        st.error(f"DB connection failed: {e}")
        st.stop()

# ── Sidebar controls ──────────────────────────────────────────────────────────
with st.sidebar:
    # Language toggle — must come first so _t() works for everything below
    st.radio("🌐", ["English", "中文"], horizontal=True,
             key="lang_radio", label_visibility="collapsed")

    st.header(_t("filters"))

    _today = date.today()
    date_range = st.date_input(
        _t("date_range"),
        value=(date(2026, 1, 1), _today),
        min_value=date(2024, 1, 1),
        max_value=_today,
    )
    if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
        d_start, d_end = date_range
    else:
        d_start, d_end = date(2026, 1, 1), _today

    prov_options = sorted(provinces_all)
    _fallback_provs = [p for p in ["Shandong", "Shanxi", "Mengxi", "Guangdong", "Sichuan"]
                       if p in prov_options] or prov_options[:5]
    # Restore saved selection from URL query params (survives browser refresh)
    _qp_raw = st.query_params.get("provs", "")
    _qp_saved = [p for p in _qp_raw.split(",") if p in prov_options] if _qp_raw else []
    default_provs = _qp_saved if _qp_saved else _fallback_provs
    selected_provs = st.multiselect(
        _t("provinces"),
        prov_options,
        default=default_provs,
        help="Select one or more provinces to compare",
    )
    # Persist selection in URL so it survives page reload
    _provs_qp = ",".join(selected_provs)
    if st.query_params.get("provs", "") != _provs_qp:
        st.query_params["provs"] = _provs_qp

    show_band = st.checkbox(_t("show_band"), value=True)
    quality_filter = st.checkbox(
        _t("filter_bad_data"),
        value=True,
        help=_t("filter_bad_data_help"),
    )

    st.divider()
    st.caption(_t("data_caption"))

if not selected_provs:
    st.info(_t("select_prov_info"))
    st.stop()

# ── Load data ─────────────────────────────────────────────────────────────────
kpis = load_kpis(quality_filter)
df = load_all(d_start, d_end, quality_filter)
df_sel = df[df["province_en"].isin(selected_provs)]

# ── KPI strip ─────────────────────────────────────────────────────────────────
k1, k2, k3, k4, k5 = st.columns(5)
k1.metric(_t("latest_date"),   str(kpis["latest_date"]) if kpis["latest_date"] else "—")
k2.metric(_t("dates_in_db"),   kpis["total_dates"])
k3.metric(_t("provinces_kpi"), kpis["total_provinces"])
k4.metric(_t("complete_rows"), kpis["complete_rows"],
          delta=f"/ {kpis['total_rows']} total", delta_color="off")
k5.metric(_t("coverage"),
          f"{100*kpis['complete_rows']/kpis['total_rows']:.0f}%" if kpis["total_rows"] else "—")

if quality_filter:
    n_bad = load_kpis(False)["total_rows"] - kpis["total_rows"]
    if n_bad > 0:
        st.caption(f"ℹ️ {n_bad} rows with invalid avg/min/max values hidden (toggle '{_t('filter_bad_data')}' in sidebar to include)")

st.divider()

# ─────────────────────────────────────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────────────────────────────────────
tab_overview, tab_spread, tab_heatmap, tab_intraday, tab_province, tab_dist, tab_geo, \
tab_interprov, tab_fundamentals, tab_agent, tab_news, tab_library, tab_jizhi, tab_supply, \
tab_forecast, tab_mgmt = st.tabs([
    _t("tab_overview"), _t("tab_spread"), _t("tab_heatmap"), _t("tab_intraday"),
    _t("tab_province"), _t("tab_dist"), _t("tab_geo"),
    _t("tab_interprov"), _t("tab_fundamentals"), _t("tab_agent"), _t("tab_news"),
    "Library", "机制竞价", "供需结构", "价格预测", _t("tab_mgmt"),
])

# ── Tab 1: Overview ───────────────────────────────────────────────────────────
with tab_overview:
    col_da, col_rt = st.columns(2)
    with col_da:
        st.plotly_chart(chart_timeseries(df_sel, selected_provs, "da", show_band),
                        use_container_width=True)
    with col_rt:
        st.plotly_chart(chart_timeseries(df_sel, selected_provs, "rt", show_band),
                        use_container_width=True)

    st.subheader(_t("latest_prices"))
    latest = (
        df[df["province_en"].isin(selected_provs)]
        .sort_values("report_date", ascending=False)
        .groupby("province_en")
        .first()
        .reset_index()
        [["province_en", "province_cn", "report_date",
          "da_avg", "da_max", "da_min",
          "rt_avg", "rt_max", "rt_min"]]
        .rename(columns={
            "province_en": _t("col_province"),
            "province_cn": _t("col_province_cn"),
            "report_date": _t("col_date"),
        })
        .sort_values(_t("col_province"))
    )
    latest[_t("col_date")] = pd.to_datetime(latest[_t("col_date")]).dt.date
    for c in ["da_avg","da_max","da_min","rt_avg","rt_max","rt_min"]:
        latest[c] = latest[c].map(lambda v: f"{v:.4f}" if pd.notna(v) else "—")
    st.dataframe(latest, use_container_width=True, hide_index=True)

# ── Tab 2: Spread ─────────────────────────────────────────────────────────────
with tab_spread:
    st.plotly_chart(chart_spread(df_sel, selected_provs), use_container_width=True)

    st.subheader(_t("spread_stats"))
    spread_rows = []
    for prov in sorted(selected_provs):
        sub = df_sel[df_sel["province_en"] == prov].dropna(subset=["da_avg", "rt_avg"])
        if sub.empty:
            continue
        s = sub["da_avg"] - sub["rt_avg"]
        spread_rows.append({
            _t("col_province"): prov,
            _t("col_mean"):     f"{s.mean():.4f}",
            _t("col_std"):      f"{s.std():.4f}",
            _t("col_min"):      f"{s.min():.4f}",
            _t("col_max"):      f"{s.max():.4f}",
            _t("col_da_gt_rt"): f"{(s > 0).mean()*100:.0f}%",
            _t("col_days"):     len(s),
        })
    if spread_rows:
        st.dataframe(pd.DataFrame(spread_rows), use_container_width=True, hide_index=True)

# ── Tab 3: Heatmap ────────────────────────────────────────────────────────────
with tab_heatmap:
    hm_metric = st.radio(_t("metric_label"), ["DA", "RT"], horizontal=True)
    fig_hm = chart_heatmap(df_sel, hm_metric.lower())
    if fig_hm.data:
        st.plotly_chart(fig_hm, use_container_width=True)
    else:
        st.info("No data for selected range / provinces.")

# ── Tab 4: Intraday Analysis ──────────────────────────────────────────────────
with tab_intraday:
    pt_sel = st.radio(_t("intraday_price_type"), ["RT", "DA"], horizontal=True, key="intraday_pt")
    price_col = "rt_price" if pt_sel == "RT" else "da_price"

    h_start = str(df_sel["report_date"].min()) if not df_sel.empty else "2025-01-01"
    h_end   = str(df_sel["report_date"].max()) if not df_sel.empty else "2026-01-31"

    # spot_prices_hourly stores Chinese province names; selected_provs are English.
    # Build a mapping from df (which has both columns) and translate before querying.
    _en_to_zh = (
        df[["province_en", "province_cn"]]
        .drop_duplicates()
        .set_index("province_en")["province_cn"]
        .to_dict()
    ) if not df.empty else {}
    h_provs_zh = tuple(sorted(
        _en_to_zh[p] for p in selected_provs if p in _en_to_zh
    )) if selected_provs else ()

    if not h_provs_zh:
        st.info(_t("select_prov_info"))
    else:
        # 1) Intraday shape
        st.subheader(_t("intraday_shape_title"))
        shape_df = _load_intraday_shape(_conn, h_provs_zh, h_start, h_end, price_col)
        if not shape_df.empty:
            fig_intra = px.line(
                shape_df, x="hour", y="avg_price", color="province",
                labels={"hour": "Hour of day", "avg_price": "Avg price (¥/kWh)",
                        "province": _t("col_province")},
            )
            fig_intra.update_layout(height=340, margin=dict(t=20, b=20))
            st.plotly_chart(fig_intra, use_container_width=True)

        # 2) Intraday spread ranking (all provinces)
        st.subheader(_t("intraday_spread_title"))
        st.caption(_t("intraday_spread_help"))
        spread_df = _load_intraday_spread(_conn, h_start, h_end, price_col)
        if not spread_df.empty:
            fig_sp = px.bar(
                spread_df, x="spread", y="province", orientation="h",
                color="spread", color_continuous_scale="RdYlGn_r",
                labels={"spread": "Intraday spread (¥/kWh)", "province": ""},
            )
            fig_sp.update_layout(
                height=max(300, len(spread_df) * 22),
                margin=dict(t=10, b=10),
                showlegend=False, coloraxis_showscale=False,
                yaxis=dict(autorange="reversed"),
            )
            st.plotly_chart(fig_sp, use_container_width=True)

        # 3) Price Duration Curve + Hour×Province heatmap
        col_pdc, col_heat = st.columns(2)
        with col_pdc:
            st.subheader(_t("intraday_pdc_title"))
            pdc_prov = st.selectbox(_t("intraday_select_prov"), sorted(h_provs_zh), key="intraday_pdc")
            pdc_s = _load_hourly_series(_conn, pdc_prov, h_start, h_end, price_col)
            if not pdc_s.empty:
                sorted_p = pdc_s.sort_values(ascending=False).reset_index(drop=True)
                pct = pd.Series(range(len(sorted_p))) / len(sorted_p) * 100
                fig_pdc = px.line(
                    x=pct, y=sorted_p.values,
                    labels={"x": "Percentile (%)", "y": "Price (¥/kWh)"},
                )
                fig_pdc.update_layout(height=300, margin=dict(t=10, b=10))
                st.plotly_chart(fig_pdc, use_container_width=True)
        with col_heat:
            st.subheader(_t("intraday_heat_title"))
            if not shape_df.empty:
                pivot = shape_df.pivot_table(
                    index="hour", columns="province", values="avg_price", aggfunc="mean"
                )
                fig_heat = px.imshow(
                    pivot.T, aspect="auto", color_continuous_scale="RdYlGn_r",
                    labels=dict(x="Hour", y="", color="¥/kWh"),
                )
                fig_heat.update_layout(height=300, margin=dict(t=10, b=10))
                st.plotly_chart(fig_heat, use_container_width=True)

        # 4) Fundamentals analysis (bidding space + WAP)
        st.divider()
        st.subheader(_t("intraday_fund_title"))
        bid_df = _load_bidding_vs_price(_conn, h_provs_zh, h_start, h_end)

        if bid_df.empty:
            st.info(_t("intraday_no_fund"))
        else:
            # Bidding space vs RT price scatter + OLS trendline
            st.subheader(_t("intraday_bid_title"))
            st.caption(_t("intraday_bid_caption"))
            try:
                fig_bid = px.scatter(
                    bid_df, x="bidding_space_mw", y="rt_price", color="province",
                    opacity=0.35, trendline="ols",
                    labels={"bidding_space_mw": "Bidding Space (MW)",
                            "rt_price": "RT Price (¥/kWh)",
                            "province": _t("col_province")},
                )
            except Exception:
                # statsmodels not installed — fall back to scatter without trendline
                fig_bid = px.scatter(
                    bid_df, x="bidding_space_mw", y="rt_price", color="province",
                    opacity=0.35,
                    labels={"bidding_space_mw": "Bidding Space (MW)",
                            "rt_price": "RT Price (¥/kWh)",
                            "province": _t("col_province")},
                )
            fig_bid.update_layout(height=400, margin=dict(t=20, b=20))
            st.plotly_chart(fig_bid, use_container_width=True)

            # Wind & Solar WAP vs avg RT price
            wap_df = _load_wap(_conn, h_provs_zh, h_start, h_end)
            if not wap_df.empty:
                st.subheader(_t("intraday_wap_title"))
                st.caption(_t("intraday_wap_caption"))
                wap_long = wap_df.melt(
                    id_vars=["province", "avg_rt_price"],
                    value_vars=["wind_wap", "solar_wap"],
                    var_name="type", value_name="wap",
                )
                wap_long["type"] = wap_long["type"].map(
                    {"wind_wap": "Wind WAP", "solar_wap": "Solar WAP"}
                )
                fig_wap = px.bar(
                    wap_long, x="province", y="wap", color="type", barmode="group",
                    labels={"wap": "WAP (¥/kWh)", "province": "", "type": ""},
                    color_discrete_map={"Wind WAP": "#4C9BE8", "Solar WAP": "#F5A623"},
                )
                fig_wap.add_scatter(
                    x=wap_df["province"], y=wap_df["avg_rt_price"],
                    mode="markers",
                    marker=dict(symbol="diamond", size=10, color="black"),
                    name="Avg RT price",
                )
                fig_wap.update_layout(height=360, margin=dict(t=20, b=20))
                st.plotly_chart(fig_wap, use_container_width=True)

# ── Tab 5: Province Deep-Dive ────────────────────────────────────────────────
with tab_province:
    dive_prov = st.selectbox(_t("select_province"), sorted(selected_provs))
    if dive_prov:
        st.plotly_chart(chart_da_rt_overlay(df_sel, dive_prov), use_container_width=True)

        sub = df_sel[df_sel["province_en"] == dive_prov].sort_values("report_date").copy()
        sub["report_date"] = pd.to_datetime(sub["report_date"]).dt.date
        st.subheader(f"{dive_prov} — {_t('raw_data')}")
        st.dataframe(
            sub[["report_date","da_avg","da_max","da_min","rt_avg","rt_max","rt_min"]]
            .rename(columns={"report_date": _t("col_date")})
            .style.format(
                {c: "{:.4f}" for c in ["da_avg","da_max","da_min","rt_avg","rt_max","rt_min"]},
                na_rep="—",
            ),
            use_container_width=True, hide_index=True,
        )

        # ── Market summaries for the selected period ──────────────────────────
        st.divider()
        st.subheader(_t("summaries_title"))
        df_summ, _summ_err = load_summaries(d_start, d_end)
        if _summ_err:
            st.warning(f"summaries query failed: {_summ_err}")
        # Filter to dates where this province has data
        prov_dates = set(sub["report_date"].astype(str))
        if df_summ.empty or "report_date" not in df_summ.columns:
            df_summ_prov = pd.DataFrame()
        else:
            df_summ_prov = df_summ[df_summ["report_date"].astype(str).isin(prov_dates)]
        if df_summ_prov.empty:
            st.info(_t("summaries_no_data"))
        else:
            _is_zh = st.session_state.get("lang_radio") == "中文"
            # Translations stored in session state so they survive rerenders
            # without triggering new API calls on every interaction.
            if "translated_summaries" not in st.session_state:
                st.session_state["translated_summaries"] = {}

            for _, row in df_summ_prov.iterrows():
                _rkey = str(row["report_date"])
                with st.expander(_t("summary_label", date=_rkey)):
                    if _is_zh and _rkey in st.session_state["translated_summaries"]:
                        st.markdown(st.session_state["translated_summaries"][_rkey])
                    else:
                        st.markdown(row["summary_text"] or "")
                        if _is_zh:
                            if st.button("🌐 翻译", key=f"tr_{_rkey}"):
                                with st.spinner("翻译中…"):
                                    _raw = row["summary_text"] or ""
                                    st.session_state["translated_summaries"][_rkey] = (
                                        _translate_to_zh(_raw) if _raw else ""
                                    )
                                st.rerun()
                    st.caption(f"{row['model']} · {row['source_pdf']}")

# ── Tab 5: Distributions ─────────────────────────────────────────────────────
with tab_dist:
    dc1, dc2, dc3 = st.columns([2, 1, 1])
    with dc1:
        both_opt = _t("both_label")
        dist_metric = st.radio(_t("market_label"), ["DA", "RT", both_opt],
                               horizontal=True, key="dist_metric")
    with dc2:
        nbins = st.slider(_t("hist_bins"), 10, 80, 30, key="dist_bins")
    with dc3:
        show_kde = st.checkbox(_t("kde_label"), value=True, key="dist_kde")

    metrics_to_show = ["da", "rt"] if dist_metric == both_opt else [dist_metric.lower()]

    for m in metrics_to_show:
        st.plotly_chart(
            chart_distributions(df_sel, selected_provs, m, nbins, show_kde),
            use_container_width=True,
        )
        st.plotly_chart(
            chart_violin(df_sel, selected_provs, m),
            use_container_width=True,
        )
        st.subheader(f"{'DA' if m == 'da' else 'RT'} — {_t('desc_stats')}")
        stats_df = _dist_stats(df_sel, selected_provs, m)
        if not stats_df.empty:
            st.dataframe(stats_df, use_container_width=True, hide_index=True)
        if dist_metric == both_opt and m == "da":
            st.divider()

# ── Tab 6: Geo Map ────────────────────────────────────────────────────────────
with tab_geo:
    st.caption(f"{_t('avg_by_province')} · **{d_start}** → **{d_end}**")

    df_geo = load_all(d_start, d_end, quality_filter)

    if df_geo.empty:
        st.info("No data for selected period.")
    else:
        st.subheader(_t("avg_by_province"))
        st.caption(_t("geo_color_caption"))

        tbl_rows = []
        for prov_en in sorted(df_geo["province_en"].unique()):
            sub = df_geo[df_geo["province_en"] == prov_en]
            da_vals = sub["da_avg"].dropna()
            rt_vals = sub["rt_avg"].dropna()
            da_avg  = da_vals.mean() if not da_vals.empty else None
            rt_avg  = rt_vals.mean() if not rt_vals.empty else None
            tbl_rows.append({
                _t("col_province"):  prov_en,
                _t("col_avg_da"):    round(da_avg, 4) if da_avg is not None else None,
                _t("col_da_level"):  _price_level(da_avg),
                _t("col_avg_rt"):    round(rt_avg, 4) if rt_avg is not None else None,
                _t("col_rt_level"):  _price_level(rt_avg),
                _t("col_days_da"):   len(da_vals),
                _t("col_days_rt"):   len(rt_vals),
            })

        tbl_df = pd.DataFrame(tbl_rows)
        da_level_col = _t("col_da_level")
        rt_level_col = _t("col_rt_level")
        avg_da_col   = _t("col_avg_da")
        avg_rt_col   = _t("col_avg_rt")

        def _style_level(col: pd.Series) -> list[str]:
            return [f"background-color: {_level_bg(v)}" for v in col]

        styled = (
            tbl_df.style
            .apply(_style_level, subset=[da_level_col])
            .apply(_style_level, subset=[rt_level_col])
            .format({
                avg_da_col: lambda v: f"{v:.4f}" if pd.notna(v) else "—",
                avg_rt_col: lambda v: f"{v:.4f}" if pd.notna(v) else "—",
            })
        )
        st.dataframe(styled, use_container_width=True, hide_index=True)

        st.divider()

        st.subheader(_t("geo_maps_title"))
        st.caption(_t("geo_color_scale"))

        _geojson, _geo_err = _load_china_geojson()
        if _geo_err:
            st.warning(f"{_t('geo_unavailable')} ({_geo_err})")

        col_map_da, col_map_rt = st.columns(2)
        with col_map_da:
            st.caption(f"**{_t('da_caption')}** · {d_start} → {d_end}")
            fig_da = chart_geo_map(df_geo, "da", _geojson)
            st.pyplot(fig_da, use_container_width=True)
            plt.close(fig_da)
        with col_map_rt:
            st.caption(f"**{_t('rt_caption')}** · {d_start} → {d_end}")
            fig_rt = chart_geo_map(df_geo, "rt", _geojson)
            st.pyplot(fig_rt, use_container_width=True)
            plt.close(fig_rt)

        # ── Section: Monthly RT Animation ────────────────────────────────────
        st.divider()
        st.subheader(_t("anim_title"))

        _all_years = list(range(2020, date.today().year + 2))
        anim_c1, anim_c2, anim_c3, anim_c4 = st.columns(4)
        with anim_c1:
            _anim_sy = st.selectbox(
                _t("anim_start_year"), _all_years,
                index=_all_years.index(min(d_start.year, _all_years[-1])),
                key="anim_sy",
            )
        with anim_c2:
            _anim_sm = st.selectbox(
                _t("anim_start_month"), list(range(1, 13)),
                index=d_start.month - 1, key="anim_sm",
                format_func=lambda m: f"{m:02d}",
            )
        with anim_c3:
            _anim_ey = st.selectbox(
                _t("anim_end_year"), _all_years,
                index=_all_years.index(min(d_end.year, _all_years[-1])),
                key="anim_ey",
            )
        with anim_c4:
            _anim_em = st.selectbox(
                _t("anim_end_month"), list(range(1, 13)),
                index=d_end.month - 1, key="anim_em",
                format_func=lambda m: f"{m:02d}",
            )

        _anim_period_start = date(_anim_sy, _anim_sm, 1)
        _anim_period_end   = date(_anim_ey, _anim_em, 1)

        # Build ordered list of month start dates
        _anim_months: list[date] = []
        _m = _anim_period_start
        while _m <= _anim_period_end:
            _anim_months.append(_m)
            _m = (_m.replace(day=28) + timedelta(days=4)).replace(day=1)

        if _anim_months:
            _month_labels = [m.strftime("%Y-%m") for m in _anim_months]
            _n_frames = len(_anim_months)

            # Initialise session state for animation
            if "anim_playing" not in st.session_state:
                st.session_state["anim_playing"] = False
            if "anim_frame_idx" not in st.session_state:
                st.session_state["anim_frame_idx"] = 0
            # If this rerun was NOT triggered by the animation loop itself,
            # stop the animation. This prevents the loop from running
            # indefinitely in the background when the user navigates away
            # or interacts with any other widget.
            if not st.session_state.pop("_anim_loop_rerun", False):
                st.session_state["anim_playing"] = False
            # Clamp index in case period changed
            st.session_state["anim_frame_idx"] = (
                st.session_state["anim_frame_idx"] % _n_frames
            )

            # Controls row
            ctrl_c1, ctrl_c2, ctrl_c3 = st.columns([1, 1, 3])
            with ctrl_c1:
                if st.button(_t("anim_play"), key="anim_play_btn"):
                    st.session_state["anim_playing"] = True
                    st.session_state["_anim_loop_rerun"] = True
            with ctrl_c2:
                if st.button(_t("anim_pause"), key="anim_pause_btn"):
                    st.session_state["anim_playing"] = False
            with ctrl_c3:
                anim_speed = st.slider(
                    _t("anim_speed"), min_value=1, max_value=10,
                    value=5, step=1, key="anim_speed",
                )

            # Manual scrub slider — value= drives position; no key so Streamlit
            # always uses the value we pass (avoids stale session-state reads)
            _cur_idx = st.session_state["anim_frame_idx"]
            _sel_label = st.select_slider(
                _t("anim_slider"), options=_month_labels,
                value=_month_labels[_cur_idx],
            )
            # Detect manual scrub: slider returned something different from
            # what we passed in (user dragged it) → jump and stop auto-play
            _slider_idx = _month_labels.index(_sel_label)
            if _slider_idx != _cur_idx:
                st.session_state["anim_frame_idx"] = _slider_idx
                st.session_state["anim_playing"] = False
                _cur_idx = _slider_idx

            _sel_m = _anim_months[_cur_idx]
            _sel_m_end = (_sel_m.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

            df_anim = load_all(_sel_m, _sel_m_end, quality_filter)
            if df_anim.empty:
                st.info(_t("anim_no_data"))
            else:
                _anim_title = _t("anim_map_title", month=_month_labels[_cur_idx])
                fig_anim = chart_geo_map(df_anim, "rt", _geojson, title=_anim_title)
                st.pyplot(fig_anim, use_container_width=True)
                plt.close(fig_anim)

            # Auto-advance: sleep then advance index and rerun.
            # Set _anim_loop_rerun so the next rerun keeps the animation alive.
            if st.session_state["anim_playing"]:
                time.sleep(anim_speed)
                _next_idx = (_cur_idx + 1) % _n_frames
                st.session_state["anim_frame_idx"] = _next_idx
                st.session_state["_anim_loop_rerun"] = True
                st.rerun()

        # ── Section: Period Comparison ────────────────────────────────────────
        st.divider()
        st.subheader(_t("cmp_title"))

        _cmp_metric_opt = st.radio(
            _t("cmp_metric"), ["DA", "RT"], horizontal=True, key="cmp_metric"
        )
        _cmp_m = _cmp_metric_opt.lower()

        cmp_a_col, cmp_b_col = st.columns(2)
        with cmp_a_col:
            st.markdown(f"**{_t('cmp_period_a')}**")
            cmp_a_start = st.date_input(_t("cmp_start"), value=d_start, key="cmp_a_start")
            cmp_a_end   = st.date_input(_t("cmp_end"),   value=d_end,   key="cmp_a_end")
        with cmp_b_col:
            st.markdown(f"**{_t('cmp_period_b')}**")
            _b_default_end   = d_end - timedelta(days=365)
            _b_default_start = d_start - timedelta(days=365)
            cmp_b_start = st.date_input(_t("cmp_start"), value=_b_default_start, key="cmp_b_start")
            cmp_b_end   = st.date_input(_t("cmp_end"),   value=_b_default_end,   key="cmp_b_end")

        df_cmp_a = load_all(cmp_a_start, cmp_a_end, quality_filter)
        df_cmp_b = load_all(cmp_b_start, cmp_b_end, quality_filter)

        cmp_map_a, cmp_map_b = st.columns(2)
        with cmp_map_a:
            _title_a = _t("cmp_map_title",
                          metric=_cmp_metric_opt, start=cmp_a_start, end=cmp_a_end)
            if df_cmp_a.empty:
                st.info(_t("cmp_no_data"))
            else:
                fig_cmp_a = chart_geo_map(df_cmp_a, _cmp_m, _geojson, title=_title_a)
                st.pyplot(fig_cmp_a, use_container_width=True)
                plt.close(fig_cmp_a)
        with cmp_map_b:
            _title_b = _t("cmp_map_title",
                          metric=_cmp_metric_opt, start=cmp_b_start, end=cmp_b_end)
            if df_cmp_b.empty:
                st.info(_t("cmp_no_data"))
            else:
                fig_cmp_b = chart_geo_map(df_cmp_b, _cmp_m, _geojson, title=_title_b)
                st.pyplot(fig_cmp_b, use_container_width=True)
                plt.close(fig_cmp_b)

# ── Tab 7: Inter-Provincial Flow ─────────────────────────────────────────────
with tab_interprov:
    st.subheader(_t("interprov_title"))
    st.caption(f"**{d_start}** → **{d_end}**")

    df_ip, _ip_err = load_interprov(d_start, d_end)

    if _ip_err:
        st.error(f"Query failed: {_ip_err}")
    elif df_ip.empty:
        st.info(_t("interprov_no_data"))
    else:
        _dir_export = "送端"
        _dir_import = "受端"

        # ── Clearing price chart: 最高均价 (solid) + 最低均价 (dashed) ──────────
        _hi_rows  = df_ip[df_ip["metric_type"] == "最高均价"].copy()
        _lo_rows  = df_ip[df_ip["metric_type"] == "最低均价"].copy()
        _has_price = not _hi_rows.empty or not _lo_rows.empty

        if _has_price:
            fig_ip_price = go.Figure()
            _dir_colors = {_dir_export: "#1f77b4", _dir_import: "#ff7f0e"}

            for _dir, _dlabel, _dshort in [
                (_dir_export, _t("direction_export"), "送端"),
                (_dir_import, _t("direction_import"), "受端"),
            ]:
                _clr = _dir_colors[_dir]
                # ── Peak avg (solid) ────────────────────────────────────────
                _sh = _hi_rows[_hi_rows["direction"] == _dir].sort_values("report_date")
                if not _sh.empty:
                    _cd = _sh[["province_cn", "price_chg_pct", "time_period"]].values
                    fig_ip_price.add_trace(go.Scatter(
                        x=_sh["report_date"], y=_sh["price_yuan_kwh"],
                        mode="lines+markers", name=f"{_dlabel} — {_t('interprov_price_hi')}",
                        line=dict(color=_clr, width=2),
                        marker=dict(size=6),
                        customdata=_cd,
                        hovertemplate=(
                            "<b>%{x}</b><br>"
                            f"{_t('hover_province')}: %{{customdata[0]}}<br>"
                            f"{_t('interprov_price_hi')}: %{{y:.4f}} ¥/kWh<br>"
                            f"{_t('hover_chg')}: %{{customdata[1]:.2f}}%<br>"
                            f"{_t('hover_period')}: %{{customdata[2]}}"
                            "<extra></extra>"
                        ),
                    ))
                # ── Floor avg (dashed) ──────────────────────────────────────
                _sl = _lo_rows[_lo_rows["direction"] == _dir].sort_values("report_date")
                if not _sl.empty:
                    _cd2 = _sl[["province_cn", "price_chg_pct", "time_period"]].values
                    fig_ip_price.add_trace(go.Scatter(
                        x=_sl["report_date"], y=_sl["price_yuan_kwh"],
                        mode="lines+markers", name=f"{_dlabel} — {_t('interprov_price_lo')}",
                        line=dict(color=_clr, width=2, dash="dash"),
                        marker=dict(size=5, symbol="diamond"),
                        customdata=_cd2,
                        hovertemplate=(
                            "<b>%{x}</b><br>"
                            f"{_t('hover_province')}: %{{customdata[0]}}<br>"
                            f"{_t('interprov_price_lo')}: %{{y:.4f}} ¥/kWh<br>"
                            f"{_t('hover_chg')}: %{{customdata[1]:.2f}}%<br>"
                            f"{_t('hover_period')}: %{{customdata[2]}}"
                            "<extra></extra>"
                        ),
                    ))

            fig_ip_price.update_layout(
                title=dict(text=_t("interprov_price_trend"), x=0, xanchor="left"),
                xaxis_title="", yaxis_title="¥/kWh",
                height=380,
                margin=dict(l=50, r=20, t=55, b=110),
                legend=dict(
                    orientation="h", xanchor="center", x=0.5,
                    yanchor="top", y=-0.18,
                    font=dict(size=11),
                ),
            )
            st.plotly_chart(fig_ip_price, use_container_width=True)

        # ── Province leaders: which province had the peak avg price each day ─
        if not _hi_rows.empty:
            st.subheader(_t("interprov_prov_leaders"))
            _c1, _c2 = st.columns(2)
            for _col_widget, _dir, _dlabel in [
                (_c1, _dir_export, _t("direction_export")),
                (_c2, _dir_import, _t("direction_import")),
            ]:
                _spl = _hi_rows[_hi_rows["direction"] == _dir].sort_values("report_date")
                if _spl.empty:
                    _col_widget.info(_dlabel + ": —")
                    continue
                _cd_pl = _spl[["price_yuan_kwh", "province_share", "time_period"]].values
                fig_pl = go.Figure(go.Scatter(
                    x=_spl["report_date"],
                    y=_spl["province_cn"],
                    mode="markers",
                    marker=dict(
                        size=10,
                        color=_spl["price_yuan_kwh"],
                        colorscale="RdYlGn_r",
                        showscale=True,
                        colorbar=dict(title="¥/kWh", thickness=10, len=0.8),
                    ),
                    customdata=_cd_pl,
                    hovertemplate=(
                        "<b>%{x}</b> · %{y}<br>"
                        f"{_t('interprov_price_hi')}: %{{customdata[0]:.4f}} ¥/kWh<br>"
                        f"{_t('hover_share')}: %{{customdata[1]:.1f}}%<br>"
                        f"{_t('hover_period')}: %{{customdata[2]}}"
                        "<extra></extra>"
                    ),
                    showlegend=False,
                ))
                fig_pl.update_layout(
                    title=dict(text=_dlabel, x=0, xanchor="left"),
                    xaxis_title="", yaxis_title="",
                    height=300,
                    margin=dict(l=90, r=60, t=45, b=30),
                    yaxis=dict(categoryorder="total ascending"),
                )
                _col_widget.plotly_chart(fig_pl, use_container_width=True)

        # ── Volume trend chart (total_vol_100gwh, one bar per direction/date) ─
        _vol_rows = df_ip[df_ip["total_vol_100gwh"].notna()].copy()
        if not _vol_rows.empty:
            fig_ip_vol = go.Figure()
            for _dir, _label in [(_dir_export, _t("direction_export")),
                                  (_dir_import, _t("direction_import"))]:
                _sv = _vol_rows[_vol_rows["direction"] == _dir].sort_values("report_date")
                if not _sv.empty:
                    # attach province name from 最高均价 row for that date (if available)
                    _sv_cd = _sv[["province_cn"]].values
                    fig_ip_vol.add_trace(go.Bar(
                        x=_sv["report_date"], y=_sv["total_vol_100gwh"],
                        name=_label,
                        customdata=_sv_cd,
                        hovertemplate=(
                            "<b>%{x}</b><br>"
                            f"{_t('hover_volume')}: %{{y:.2f}} 亿kWh<br>"
                            f"{_t('hover_province')}: %{{customdata[0]}}"
                            "<extra></extra>"
                        ),
                    ))
            fig_ip_vol.update_layout(
                title=dict(text=_t("interprov_vol_trend"), x=0, xanchor="left"),
                xaxis_title="", yaxis_title="亿kWh",
                barmode="group", height=300,
                margin=dict(l=50, r=20, t=55, b=90),
                legend=dict(
                    orientation="h", xanchor="center", x=0.5,
                    yanchor="top", y=-0.2,
                    font=dict(size=11),
                ),
            )
            st.plotly_chart(fig_ip_vol, use_container_width=True)

        # ── Detail tables per direction ───────────────────────────────────────
        st.divider()
        _col_map = {
            "report_date":      _t("col_date"),
            "metric_type":      _t("col_metric_type"),
            "province_cn":      _t("col_province_cn"),
            "province_share":   _t("col_share"),
            "price_yuan_kwh":   _t("col_price_kwh"),
            "price_chg_pct":    _t("col_price_chg"),
            "time_period":      _t("col_time_period"),
            "total_vol_100gwh": _t("col_volume_gwh"),
            "source_pdf":       _t("col_source"),
        }
        _display_cols = list(_col_map.keys())

        for _dir, _label in [(_dir_export, _t("direction_export")),
                              (_dir_import, _t("direction_import"))]:
            _sub = (df_ip[df_ip["direction"] == _dir][_display_cols]
                    .rename(columns=_col_map)
                    .sort_values(_t("col_date"), ascending=False))
            if not _sub.empty:
                st.subheader(_label)
                _fmt = {
                    _t("col_price_kwh"):   "{:.4f}",
                    _t("col_price_chg"):   "{:.2f}",
                    _t("col_volume_gwh"):  "{:.4f}",
                    _t("col_share"):       "{:.2f}",
                }
                st.dataframe(
                    _sub.style.format(_fmt, na_rep="—"),
                    use_container_width=True, hide_index=True,
                )


# ── Tab 8: Market Fundamentals ───────────────────────────────────────────────
with tab_fundamentals:
    from services.market_fundamentals.loader import (
        load_province_data as _load_fund,
        FUEL_EN as _FUEL_EN,
        FUEL_COLORS as _FUEL_COLORS,
        PROVINCE_EN as _PROVINCE_EN,
    )

    _fund_data = _load_fund()

    if not _fund_data:
        st.warning(_t("fund_no_data"))
    else:
        # ── Controls ──────────────────────────────────────────────────────────
        _all_provs_cn = sorted(_fund_data.keys(), key=lambda p: _PROVINCE_EN.get(p, p))
        _all_provs_en = [_PROVINCE_EN.get(p, p) for p in _all_provs_cn]
        _prov_cn_to_en = {p: _PROVINCE_EN.get(p, p) for p in _all_provs_cn}
        _prov_en_to_cn = {v: k for k, v in _prov_cn_to_en.items()}

        _ctrl1, _ctrl2 = st.columns([5, 1])
        with _ctrl1:
            _default_provs_en = [
                _PROVINCE_EN.get(p, p) for p in ["山东", "广东", "江苏", "蒙西", "四川"]
                if p in _fund_data
            ]
            _sel_provs_en = st.multiselect(
                _t("fund_provinces"), _all_provs_en, default=_default_provs_en,
                key="fund_provs",
            )
        with _ctrl2:
            _fund_year = st.radio(_t("fund_year"), [2025, 2024], key="fund_year")

        if not _sel_provs_en:
            st.info(_t("fund_select_prompt"))
        else:
            _sel_provs_cn = [_prov_en_to_cn[e] for e in _sel_provs_en if e in _prov_en_to_cn]

            # ── Helper: build DataFrame for one metric ─────────────────────
            def _fund_df(metric: str, year: int) -> pd.DataFrame:
                """metric = 'capacity' or 'generation'"""
                rows = []
                for pcn in _sel_provs_cn:
                    info  = _fund_data.get(pcn, {})
                    raw   = info.get(metric, {}).get(year, {})
                    pen   = _prov_cn_to_en[pcn]
                    for fuel_cn in ["风电", "光伏", "水电", "核电", "储能", "火电"]:
                        fuel_en = _FUEL_EN[fuel_cn]
                        val = raw.get(fuel_cn, {}).get("value")
                        if val is not None and val > 0:
                            rows.append({"Province": pen, "Fuel": fuel_en, "Value": val})
                return pd.DataFrame(rows)

            # ── Helper: stacked-bar chart (multi-province) ─────────────────
            def _stacked_bar(df: pd.DataFrame, title: str, unit: str) -> go.Figure:
                fuel_order = ["Wind", "Solar", "Hydro", "Nuclear", "Storage", "Thermal"]
                fig = go.Figure()
                for fuel in fuel_order:
                    sub = df[df["Fuel"] == fuel]
                    if sub.empty:
                        continue
                    fig.add_trace(go.Bar(
                        name=fuel,
                        x=sub["Province"],
                        y=sub["Value"],
                        marker_color=_FUEL_COLORS.get(fuel, "#aaa"),
                        hovertemplate=f"<b>%{{x}}</b><br>{fuel}: %{{y:,.1f}} {unit}<extra></extra>",
                    ))
                fig.update_layout(
                    barmode="stack",
                    title=title,
                    xaxis_title="",
                    yaxis_title=unit,
                    legend_title="Fuel Type",
                    height=420,
                    margin=dict(t=50, b=40),
                )
                return fig

            # ── Helper: donut chart (single province) ─────────────────────
            def _donut(df: pd.DataFrame, title: str, unit: str) -> go.Figure:
                fuel_order = ["Wind", "Solar", "Hydro", "Nuclear", "Storage", "Thermal"]
                df_sorted = df.set_index("Fuel").reindex(fuel_order).dropna()
                fig = go.Figure(go.Pie(
                    labels=df_sorted.index,
                    values=df_sorted["Value"],
                    hole=0.45,
                    marker_colors=[_FUEL_COLORS.get(f, "#aaa") for f in df_sorted.index],
                    textinfo="label+percent",
                    hovertemplate="<b>%{label}</b><br>%{value:,.1f} " + unit + "<br>%{percent}<extra></extra>",
                ))
                fig.update_layout(title=title, height=380, margin=dict(t=50, b=10))
                return fig

            # ── Installed Capacity & Generation (side by side) ─────────────
            _df_cap = _fund_df("capacity", _fund_year)
            _df_gen = _fund_df("generation", _fund_year)

            _cap_title = f"{_t('fund_capacity_title')} — {_fund_year}"
            _gen_title = f"{_t('fund_generation_title')} — {_fund_year}"

            _chart_col1, _chart_col2 = st.columns(2)

            if len(_sel_provs_en) == 1 and not _df_cap.empty:
                with _chart_col1:
                    st.plotly_chart(_donut(_df_cap, _cap_title, "万kW"), use_container_width=True)
            elif not _df_cap.empty:
                with _chart_col1:
                    st.plotly_chart(_stacked_bar(_df_cap, _cap_title, "万kW"), use_container_width=True)

            if len(_sel_provs_en) == 1 and not _df_gen.empty:
                with _chart_col2:
                    st.plotly_chart(_donut(_df_gen, _gen_title, "亿kWh"), use_container_width=True)
            elif not _df_gen.empty:
                with _chart_col2:
                    st.plotly_chart(_stacked_bar(_df_gen, _gen_title, "亿kWh"), use_container_width=True)

            # ── Renewables share bar ───────────────────────────────────────
            _renew_fuels = {"Wind", "Solar", "Hydro", "Nuclear"}
            _renew_rows = []
            for pcn in _sel_provs_cn:
                pen = _prov_cn_to_en[pcn]
                raw = _fund_data.get(pcn, {}).get("capacity", {}).get(_fund_year, {})
                total = sum(
                    v["value"] for v in raw.values()
                    if v.get("value") is not None and v["value"] > 0
                )
                renew = sum(
                    raw.get(fc, {}).get("value") or 0
                    for fc, fe in _FUEL_EN.items()
                    if fe in _renew_fuels
                )
                if total > 0:
                    _renew_rows.append({"Province": pen, "Share (%)": round(renew / total * 100, 1)})

            if _renew_rows:
                _df_renew = pd.DataFrame(sorted(_renew_rows, key=lambda r: -r["Share (%)"]))
                _fig_renew = px.bar(
                    _df_renew, x="Province", y="Share (%)",
                    title=f"{_t('fund_renewables_share')} — {_fund_year}",
                    color_discrete_sequence=["#4C9BE8"],
                    text="Share (%)",
                )
                _fig_renew.update_traces(texttemplate="%{text:.1f}%", textposition="outside")
                _fig_renew.update_layout(height=320, margin=dict(t=50, b=40), yaxis_range=[0, 105])
                st.plotly_chart(_fig_renew, use_container_width=True)

            # ── Peak Load Ranking Table (both years, all selected provinces) ──
            st.subheader(_t("fund_peak_title"))
            _peak_rows = []
            for _pcn in _sel_provs_cn:
                _pen = _prov_cn_to_en.get(_pcn, _pcn)
                _row = {
                    "_prov_en": _pen,
                    "_prov_cn": _pcn,
                }
                for _yr in [2024, 2025]:
                    _pl = _fund_data.get(_pcn, {}).get("peak_load", {}).get(_yr, {})
                    _row[f"{_t('fund_summer_peak')} {_yr}"] = _pl.get("summer")
                    _row[f"{_t('fund_winter_peak')} {_yr}"] = _pl.get("winter")
                _peak_rows.append(_row)

            if _peak_rows:
                _peak_df = pd.DataFrame(_peak_rows)
                # Sort by 2025 summer peak descending
                _sort_col = f"{_t('fund_summer_peak')} 2025"
                if _sort_col in _peak_df.columns and _peak_df[_sort_col].notna().any():
                    _peak_df = _peak_df.sort_values(_sort_col, ascending=False, na_position="last")
                _peak_df = _peak_df.reset_index(drop=True)
                _peak_df.insert(0, "#", range(1, len(_peak_df) + 1))
                _prov_col = "省份" if _is_zh else "Province"
                _peak_df.insert(1, _prov_col, _peak_df["_prov_cn"] if _is_zh else _peak_df["_prov_en"])
                _peak_df = _peak_df.drop(columns=["_prov_en", "_prov_cn"])
                # Format numeric columns
                _num_cols = [c for c in _peak_df.columns if c not in ("#", _prov_col)]
                for _nc in _num_cols:
                    _peak_df[_nc] = _peak_df[_nc].apply(
                        lambda v: f"{int(v):,}" if pd.notna(v) and v else "—"
                    )
                st.dataframe(_peak_df, use_container_width=True, hide_index=True)

            # ── Standard EOH (expected operating hours/year) ──────────────────
            # Used for both LF and system tightness calculations.
            # Thermal blended: coal-dominated Chinese fleet (~5 500 h).
            # Storage excluded — net-zero generation asset.
            _STD_EOH = {
                "Wind":    2000,
                "Solar":   1100,
                "Thermal": 5500,
                "Hydro":   3500,
                "Nuclear": 7500,
            }
            _EOH_NOTE = (
                f"Wind {_STD_EOH['Wind']} h · Solar {_STD_EOH['Solar']} h · "
                f"Thermal {_STD_EOH['Thermal']} h · Hydro {_STD_EOH['Hydro']} h · "
                f"Nuclear {_STD_EOH['Nuclear']} h"
            )

            # ── Load Factor ranking ───────────────────────────────────────────
            st.divider()
            st.subheader(f"{_t('fund_lf_title')} — {_fund_year}")
            st.caption(_t("fund_lf_caption"))

            _lf_rows = []
            for _pcn, _pdata in _fund_data.items():
                _pen = _prov_cn_to_en.get(_pcn, _pcn)
                _cap_yr = _pdata.get("capacity", {}).get(_fund_year, {})
                _gen_yr = _pdata.get("generation", {}).get(_fund_year, {})
                _row = {"_prov_en": _pen, "_prov_cn": _pcn}
                for _fcn, _fen in _FUEL_EN.items():
                    if _fen not in _STD_EOH:
                        continue
                    _cap_v = (_cap_yr.get(_fcn) or {}).get("value")
                    _gen_v = (_gen_yr.get(_fcn) or {}).get("value")
                    if _cap_v and _cap_v > 0 and _gen_v is not None and _gen_v >= 0:
                        # LF = gen(亿kWh)×10^8 / (cap(万kW)×10^4 × 8760)
                        #     = gen × 10^4 / (cap × 8760)
                        _row[_fen] = round(_gen_v * 1e4 / (_cap_v * 8760) * 100, 1)
                    else:
                        _row[_fen] = None
                _lf_rows.append(_row)

            if _lf_rows:
                _lf_df = pd.DataFrame(_lf_rows)
                _fuel_cols_lf = [f for f in _STD_EOH if f in _lf_df.columns]
                _lf_sort_col = st.selectbox(
                    _t("fund_lf_sort"), _fuel_cols_lf,
                    index=0, key="lf_sort_fuel",
                )
                if _lf_sort_col and _lf_df[_lf_sort_col].notna().any():
                    _lf_df = _lf_df.sort_values(_lf_sort_col, ascending=False, na_position="last")
                _lf_df = _lf_df.reset_index(drop=True)
                _lf_df.insert(0, "#", range(1, len(_lf_df) + 1))
                _prov_col_lf = "省份" if _is_zh else "Province"
                _lf_df.insert(1, _prov_col_lf,
                              _lf_df["_prov_cn"] if _is_zh else _lf_df["_prov_en"])
                _lf_df = _lf_df.drop(columns=["_prov_en", "_prov_cn"])
                # Interleave LF% and equivalent hours (LF×8760) for each fuel
                _lf_out = _lf_df[["#", _prov_col_lf]].copy()
                for _fc in _fuel_cols_lf:
                    _lf_out[f"{_fc} (%)"] = _lf_df[_fc].apply(
                        lambda v: f"{v:.1f}%" if pd.notna(v) and v is not None else "—"
                    )
                    _lf_out[f"{_fc} (h)"] = _lf_df[_fc].apply(
                        lambda v: f"{int(round(v * 8760 / 100)):,}" if pd.notna(v) and v is not None else "—"
                    )
                st.dataframe(_lf_out, use_container_width=True, hide_index=True)

            # ── System Tightness ranking ──────────────────────────────────────
            st.divider()
            st.subheader(f"{_t('fund_tightness_title')} — {_fund_year}")
            st.caption(f"{_t('fund_tightness_caption')}  |  {_EOH_NOTE}")

            _tight_rows = []
            for _pcn, _pdata in _fund_data.items():
                _pen = _prov_cn_to_en.get(_pcn, _pcn)
                _cap_yr = _pdata.get("capacity", {}).get(_fund_year, {})
                _gen_yr = _pdata.get("generation", {}).get(_fund_year, {})
                _pl_yr  = _pdata.get("peak_load", {}).get(_fund_year, {})

                # Effective generation capacity (MW)
                # = Σ_type [cap(万kW) × 10 MW/万kW × EOH / 8760]
                # Use `or 0` after .get("value") to guard against {"value": None}.
                _eff_mw = sum(
                    ((_cap_yr.get(_fcn) or {}).get("value") or 0) * 10 * _STD_EOH[_fen] / 8760
                    for _fcn, _fen in _FUEL_EN.items()
                    if _fen in _STD_EOH
                    and ((_cap_yr.get(_fcn) or {}).get("value") or 0) > 0
                )

                # Average (baseload) demand (MW)
                # = total annual generation(亿kWh) × 10^8 kWh/亿kWh / 8760 h / 1000 kW/MW
                _total_gen = sum(
                    (_gen_yr.get(_fcn) or {}).get("value") or 0
                    for _fcn in _FUEL_EN
                )
                _avg_mw = _total_gen * 1e8 / 8760 / 1000 if _total_gen > 0 else None

                _sum_pk = _pl_yr.get("summer")
                _win_pk = _pl_yr.get("winter")

                _tight_rows.append({
                    "_prov_en":            _pen,
                    "_prov_cn":            _pcn,
                    _t("fund_eff_cap"):    round(_eff_mw)   if _eff_mw  else None,
                    _t("fund_avg_demand"): round(_avg_mw)   if _avg_mw  else None,
                    "Summer Peak (MW)":    round(_sum_pk)   if _sum_pk  else None,
                    "Winter Peak (MW)":    round(_win_pk)   if _win_pk  else None,
                    _t("fund_tight_avg"):  round(_eff_mw - _avg_mw) if (_eff_mw and _avg_mw)  else None,
                    _t("fund_tight_summer"): round(_eff_mw - _sum_pk) if (_eff_mw and _sum_pk) else None,
                    _t("fund_tight_winter"): round(_eff_mw - _win_pk) if (_eff_mw and _win_pk) else None,
                })

            if _tight_rows:
                _tdf = pd.DataFrame(_tight_rows)
                # Rename raw peak columns to translated labels
                _tdf = _tdf.rename(columns={
                    "Summer Peak (MW)": f"{_t('fund_summer_peak')} (MW)",
                    "Winter Peak (MW)": f"{_t('fund_winter_peak')} (MW)",
                })
                # Sort by tightness vs summer peak ascending (tightest = most stressed first)
                _tight_sort = _t("fund_tight_summer")
                if _tight_sort in _tdf.columns and _tdf[_tight_sort].notna().any():
                    _tdf = _tdf.sort_values(_tight_sort, ascending=True, na_position="last")
                _tdf = _tdf.reset_index(drop=True)
                _tdf.insert(0, "#", range(1, len(_tdf) + 1))
                _prov_col_t = "省份" if _is_zh else "Province"
                _tdf.insert(1, _prov_col_t,
                            _tdf["_prov_cn"] if _is_zh else _tdf["_prov_en"])
                _tdf = _tdf.drop(columns=["_prov_en", "_prov_cn"])

                # Format all numeric columns; prefix surplus cols with + / − for clarity
                _surplus_col_keys = {
                    _t("fund_tight_avg"),
                    _t("fund_tight_summer"),
                    _t("fund_tight_winter"),
                }
                _num_cols_t = [c for c in _tdf.columns if c not in ("#", _prov_col_t)]
                for _nc in _num_cols_t:
                    if _nc in _surplus_col_keys:
                        _tdf[_nc] = _tdf[_nc].apply(
                            lambda v: (f"+{int(v):,}" if v > 0 else f"{int(v):,}")
                            if pd.notna(v) and v is not None else "—"
                        )
                    else:
                        _tdf[_nc] = _tdf[_nc].apply(
                            lambda v: f"{int(v):,}" if pd.notna(v) and v is not None else "—"
                        )

                st.dataframe(_tdf, use_container_width=True, hide_index=True)


# ── Tab 9: Agent ──────────────────────────────────────────────────────────────
with tab_agent:
    import os as _os
    import json as _json
    import uuid as _uuid
    import anthropic as _anthropic

    # ── Memory infrastructure ──────────────────────────────────────────────────
    _SPOT_MEM_KEY = "spot_v1"
    _SPOT_APP_NAME = "spot_market"

    @st.cache_resource
    def _ensure_spot_memory_table():
        conn = _conn()
        with conn.cursor() as _cur:
            _cur.execute("""
                CREATE TABLE IF NOT EXISTS marketdata.agent_memory (
                    id SERIAL PRIMARY KEY,
                    app TEXT NOT NULL DEFAULT 'spot_market',
                    category TEXT NOT NULL,
                    subject TEXT NOT NULL,
                    content TEXT NOT NULL,
                    source TEXT DEFAULT 'manual',
                    created_at TIMESTAMPTZ DEFAULT NOW(),
                    active BOOLEAN DEFAULT TRUE
                )
            """)
            _cur.execute("""
                ALTER TABLE marketdata.agent_memory
                ADD COLUMN IF NOT EXISTS app TEXT NOT NULL DEFAULT 'spot_market'
            """)
        conn.commit()
        return True

    @st.cache_data(ttl=60)
    def _load_spot_memories(_key) -> pd.DataFrame:
        try:
            return pd.read_sql(
                "SELECT id, category, subject, content, source "
                "FROM marketdata.agent_memory WHERE active AND app=%s ORDER BY id",
                _conn(),
                params=(_SPOT_APP_NAME,),
            )
        except Exception:
            return pd.DataFrame(columns=["id", "category", "subject", "content", "source"])

    def _save_spot_memory(category: str, subject: str, content: str, source: str = "manual"):
        conn = _conn()
        with conn.cursor() as _cur:
            _cur.execute(
                "INSERT INTO marketdata.agent_memory (app,category,subject,content,source) "
                "VALUES (%s,%s,%s,%s,%s)",
                (_SPOT_APP_NAME, category, subject, content, source),
            )
        conn.commit()
        _load_spot_memories.clear()

    def _delete_spot_memory(memory_id: int):
        conn = _conn()
        with conn.cursor() as _cur:
            _cur.execute(
                "UPDATE marketdata.agent_memory SET active=FALSE WHERE id=%s AND app=%s",
                (memory_id, _SPOT_APP_NAME),
            )
        conn.commit()
        _load_spot_memories.clear()

    try:
        _ensure_spot_memory_table()
    except Exception:
        pass  # non-fatal: agent works without this legacy table

    # ── Session persistence ────────────────────────────────────────────────────

    def _ensure_spot_sessions_table():
        if st.session_state.get("_spot_sessions_table_ok"):
            return
        conn = _conn()
        with conn.cursor() as _cur:
            _cur.execute("""
                CREATE TABLE IF NOT EXISTS staging.spot_analyst_sessions (
                    session_id TEXT PRIMARY KEY,
                    messages JSONB NOT NULL DEFAULT '[]',
                    created_at TIMESTAMPTZ DEFAULT NOW(),
                    updated_at TIMESTAMPTZ DEFAULT NOW()
                )
            """)
        conn.commit()
        st.session_state["_spot_sessions_table_ok"] = True

    def _save_spot_session(session_id: str, display_messages: list):
        try:
            _ensure_spot_sessions_table()
            conn = _conn()
            with conn.cursor() as _cur:
                _payload = _json.dumps(display_messages)
                _cur.execute(
                    "INSERT INTO staging.spot_analyst_sessions "
                    "(session_id, messages, updated_at) VALUES (%s, %s, NOW()) "
                    "ON CONFLICT (session_id) DO UPDATE "
                    "SET messages=%s, updated_at=NOW()",
                    (session_id, _payload, _payload),
                )
            conn.commit()
        except Exception:
            pass

    def _load_spot_session(session_id: str) -> list:
        try:
            _ensure_spot_sessions_table()
            conn = _conn()
            with conn.cursor() as _cur:
                _cur.execute(
                    "SELECT messages FROM staging.spot_analyst_sessions "
                    "WHERE session_id = %s",
                    (session_id,),
                )
                row = _cur.fetchone()
                return _json.loads(row[0]) if row else []
        except Exception:
            return []

    def _list_recent_spot_sessions(limit: int = 3) -> pd.DataFrame:
        try:
            _ensure_spot_sessions_table()
            return pd.read_sql(
                "SELECT session_id, jsonb_array_length(messages) AS msg_count, "
                "updated_at AT TIME ZONE 'Asia/Singapore' AS updated_at "
                "FROM staging.spot_analyst_sessions "
                "WHERE jsonb_array_length(messages) > 0 "
                "ORDER BY updated_at DESC LIMIT %s",
                _conn(), params=(limit,),
            )
        except Exception:
            return pd.DataFrame()

    def _display_from_session(session_messages: list) -> list:
        """Reconstruct agent_display list from saved display messages."""
        return session_messages if session_messages else [
            {"role": "assistant", "content": _t("agent_welcome"), "tool": None}
        ]

    def _api_messages_from_display(display: list) -> list:
        """Reconstruct lean API message list (text only) from display messages for context."""
        return [
            {"role": m["role"], "content": m["content"]}
            for m in display
            if m.get("role") in ("user", "assistant") and m.get("content")
        ]

    # ── Knowledge gap interview helpers ────────────────────────────────────────

    def _generate_spot_interview_questions() -> list[dict]:
        """Audit kp_expert_insights pool, identify gaps, return up to 5 questions."""
        try:
            _summary = pd.read_sql(
                "SELECT insight_type, confidence, COUNT(*) AS n "
                "FROM staging.kp_expert_insights WHERE active = TRUE "
                "GROUP BY insight_type, confidence ORDER BY n DESC",
                _conn(),
            )
        except Exception:
            _summary = pd.DataFrame()
        try:
            _sample = pd.read_sql(
                "SELECT insight_text, insight_type "
                "FROM staging.kp_expert_insights WHERE active = TRUE "
                "ORDER BY id DESC LIMIT 15",
                _conn(),
            )
        except Exception:
            _sample = pd.DataFrame()

        _ctx = ["Current expert insight pool:"]
        if not _summary.empty:
            for _, _r in _summary.iterrows():
                _ctx.append(f"  {_r['insight_type']} ({_r['confidence']}): {int(_r['n'])} insights")
        else:
            _ctx.append("  (empty — knowledge gaps in all areas)")
        _ctx.append("\nSample of already-known insights (do NOT duplicate):")
        if not _sample.empty:
            for _, _r in _sample.iterrows():
                _ctx.append(f"  [{_r['insight_type']}] {str(_r['insight_text'])[:120]}")

        _system = """\
You are the China electricity market strategist agent auditing your own knowledge base.
Identify the 5 most valuable areas where knowledge is THIN, UNCERTAIN, or MISSING.
Generate one precise expert interview question per gap — something only a practitioner
with hands-on China electricity market experience can answer from observation.

Prioritise gaps in these areas (in order):
1. Province-specific dispatch mechanics (curtailment triggers, peak-shaving rules, settlement quirks)
2. FM/ancillary market nuances in Inner Mongolia, Shanxi, Shandong, or Guangdong
3. BESS capacity payment rules or recent policy changes with concrete operational implications
4. Counterintuitive DA/RT spread patterns the expert has personally observed
5. Revenue stacking strategies that differentiate top-performing BESS assets from median

Do NOT generate questions already answered in the sample insights above.
Do NOT generate generic textbook questions about Chinese power markets.

Respond ONLY with valid JSON:
{"questions": [{"question": "...", "topic": "market_structure|regulation|operations|dispatch_economics|investment", "why_asking": "one sentence on what knowledge gap this fills"}]}
"""
        try:
            _api_key = _os.environ.get("ANTHROPIC_API_KEY", "")
            if not _api_key:
                return []
            _haiku = _anthropic.Anthropic(api_key=_api_key)
            _resp = _haiku.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=900,
                system=_system,
                messages=[{"role": "user", "content": "\n".join(_ctx)}],
            )
            _raw = _resp.content[0].text.strip()
            if _raw.startswith("```"):
                _raw = _raw.split("```", 2)[1]
                if _raw.startswith("json"):
                    _raw = _raw[4:]
            return _json.loads(_raw).get("questions", [])[:5]
        except Exception:
            return []

    def _answer_from_kb(question: str) -> str | None:
        """Try to answer a gap question from the knowledge base. Returns text answer or None."""
        try:
            from services.knowledge_pool.knowledge_docs import search_reference_docs as _srd
            _chunks = _srd(query=question, app="strategist", limit=5)
            if not _chunks or len(_chunks) < 2:
                return None
            _combined = "\n\n".join(c.get("chunk_text", "") for c in _chunks[:5])
            if len(_combined) < 150:
                return None
            _api_key = _os.environ.get("ANTHROPIC_API_KEY", "")
            if not _api_key:
                return None
            _haiku = _anthropic.Anthropic(api_key=_api_key)
            _resp = _haiku.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=400,
                system=(
                    "Answer the question using only the provided knowledge base excerpts. "
                    "Be concise and specific (2-4 sentences). If the excerpts do not contain "
                    "enough information to answer confidently, respond with exactly: INSUFFICIENT"
                ),
                messages=[{"role": "user", "content": (
                    f"Question: {question}\n\nKB Excerpts:\n{_combined[:3000]}"
                )}],
            )
            _answer = _resp.content[0].text.strip()
            if "INSUFFICIENT" in _answer.upper() or len(_answer) < 40:
                return None
            return _answer
        except Exception:
            return None

    def _store_spot_interview_answer(question: str, answer: str, topic: str,
                                      confidence: str = "high") -> None:
        """Store a user or KB-sourced expert interview answer as an insight."""
        _text = f"[Expert interview] Q: {question[:150]} | A: {answer}"
        try:
            conn = _conn()
            with conn.cursor() as _cur:
                _cur.execute(
                    "INSERT INTO staging.kp_expert_insights "
                    "(insight_text, insight_type, confidence, source_session, validated_at) "
                    "VALUES (%s, %s, %s, %s, NOW())",
                    (_text[:1000], topic, confidence, date.today().isoformat()),
                )
            conn.commit()
        except Exception as _e:
            raise _e

    # ── Base system prompt ─────────────────────────────────────────────────────
    _SPOT_AGENT_BASE_SYSTEM = """\
You are a specialist analyst for China's spot electricity market. \
Your knowledge comes **exclusively** from the data tools below — never from general training data or external information. \
Do not state any price level, trend, or market event unless it was returned by a tool call in this conversation.

## Domain definitions
- **DA price**: Day-Ahead electricity clearing price (¥/kWh). Set one day ahead via auction.
- **RT price**: Real-Time electricity clearing price (¥/kWh). Reflects actual intraday supply/demand balance.
- **Spread**: DA − RT; positive = DA premium (normal); negative = RT spike (intraday supply stress).
- **送端**: Exporting province in inter-provincial spot market.
- **受端**: Importing province in inter-provincial spot market.
- Province names in DB: Shandong, Guangdong, Mengxi, Shanxi, Gansu, Sichuan, Yunnan, Guizhou, \
Guangxi, Hunan, Hubei, Anhui, Zhejiang, Jiangsu, Fujian, Henan, Shaanxi, Ningxia, Xinjiang, \
Liaoning, Jilin, Heilongjiang, Mengdong, Hebei, Hebei-North, Hebei-South, Qinghai, Jiangxi, \
Hainan, Chongqing, Shanghai, Beijing, Tianjin.

## Analytical framework
1. Always call a tool before stating any price, spread, volume, or trend.
2. When comparing provinces or time periods, call get_spot_prices with the full range then compute statistics from the returned rows.
3. For structural questions (fuel mix, capacity, renewables share, peak demand), call get_market_fundamentals.
4. For inter-provincial flow questions (volumes, sending/receiving provinces), call get_interprov_flow.
5. For qualitative market colour or key drivers, call get_market_summaries.
6. For questions about market rules, trading procedures, settlement mechanisms, annual exchange reports, \
regulatory policy, or any uploaded reference data (Excel spreadsheets with trading volumes, network losses, \
contract data, etc.), call search_reference_docs. \
The knowledge base indexes ALL file types: PDF, Excel (.xlsx/.xls), PowerPoint, Word, and text files.
7. Use markdown tables for multi-province or multi-period comparisons.
8. To ingest a new reference document (PDF, Excel, PPTX, DOCX — any format), call ingest_kb_document \
with s3_key (file in the S3 uploads bucket) or file_path (local/repo-relative path). \
For spot market daily price PDFs specifically, use run_pipeline instead (it also extracts structured DA/RT price data).
9. If the user says they already uploaded a file via the UI, call search_reference_docs directly — \
the file is already ingested and searchable without calling ingest_kb_document again.
10. For questions about Inner Mongolia BESS asset performance, P&L, dispatch cycles, or strategy \
comparison across suyou/hangjinqi/siziwangqi/gushanliang, call get_bess_pnl. \
It returns daily P&L and dispatch metrics across all 5 strategy scenarios.
"""

    def _build_spot_system(query: str = "") -> str:
        base = _SPOT_AGENT_BASE_SYSTEM
        _api_key = _os.environ.get("ANTHROPIC_API_KEY", "")

        if query:
            # Inject structured expert insights from kp_expert_insights (FTS-retrieved)
            try:
                from services.knowledge_pool.expert_memory import (
                    get_relevant_insights as _get_insights,
                    inject_expert_memory as _inject_memory,
                )
                _insights = _get_insights(query=query, limit=5)
                _mem_block = _inject_memory(_insights)
                if _mem_block:
                    base += f"\n\n{_mem_block}"
            except Exception:
                pass

            # Inject advanced retrieval context (HyDE + reranking) from knowledge pool
            if _api_key:
                try:
                    from services.knowledge_pool.advanced_retrieval import retrieve_for_agent as _retrieve
                    _kb_ctx = _retrieve(
                        query=query, api_key=_api_key, app="shared",
                        use_hyde=True, use_rerank=True, top_k=5,
                    )
                    if _kb_ctx:
                        base += f"\n\n{_kb_ctx}"
                except Exception:
                    pass

        # Inject flat analyst preferences from agent_memory
        mems = _load_spot_memories(_SPOT_MEM_KEY)
        if not mems.empty:
            mem_lines = "\n".join(
                f"- [{r.category}] {r.subject}: {r.content}"
                for r in mems.itertuples()
            )
            base += f"\n\n## Analyst preferences & domain knowledge\n{mem_lines}"

        lang_suffix = "\n\n请用中文（简体）回复所有问题。" if st.session_state.get("lang_radio") == "中文" else "\n\nRespond in English."
        return base + lang_suffix

    def _extract_spot_memories(user_msg: str, agent_reply: str) -> list[dict]:
        """Use Haiku to extract memorable facts/preferences from a conversation turn."""
        api_key = _os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            return []
        try:
            haiku = _anthropic.Anthropic(api_key=api_key)
            resp = haiku.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=512,
                system=(
                    "Extract memorable analyst preferences or domain facts from the conversation. "
                    "Return a JSON array of objects with keys: category (string, e.g. 'preference', 'market_view', 'methodology'), "
                    "subject (short title ≤8 words), content (one sentence). "
                    "Only extract genuinely reusable insights — not one-off data points. "
                    "Return [] if nothing is worth remembering."
                ),
                messages=[{
                    "role": "user",
                    "content": f"User: {user_msg}\n\nAgent: {agent_reply}",
                }],
            )
            raw = next((b.text for b in resp.content if hasattr(b, "text")), "[]")
            start, end = raw.find("["), raw.rfind("]")
            if start == -1:
                return []
            return _json.loads(raw[start:end + 1])
        except Exception:
            return []

    # ── Tool definitions ───────────────────────────────────────────────────────
    _AGENT_TOOLS = [
        {
            "name": "get_spot_prices",
            "description": (
                "Fetch day-ahead (DA) and real-time (RT) spot electricity clearing "
                "prices from public.spot_daily. Prices in ¥/kWh."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "start_date": {"type": "string", "description": "ISO date, e.g. '2026-01-01'"},
                    "end_date":   {"type": "string", "description": "ISO date, e.g. '2026-04-30'"},
                    "provinces":  {
                        "type": "array", "items": {"type": "string"},
                        "description": "Optional list of province_en names, e.g. ['Shandong','Guangdong']",
                    },
                },
                "required": ["start_date", "end_date"],
            },
        },
        {
            "name": "get_interprov_flow",
            "description": (
                "Fetch inter-provincial spot trading data (省间现货交易情况). "
                "Returns daily peak/floor average prices and volumes for exporting "
                "(送端) and importing (受端) provinces."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "start_date": {"type": "string", "description": "ISO date"},
                    "end_date":   {"type": "string", "description": "ISO date"},
                },
                "required": ["start_date", "end_date"],
            },
        },
        {
            "name": "get_market_summaries",
            "description": (
                "Fetch AI-generated daily market narrative summaries. Each summary "
                "covers price levels, key drivers, inter-provincial flows, and "
                "notable events for that trading day."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "start_date": {"type": "string", "description": "ISO date"},
                    "end_date":   {"type": "string", "description": "ISO date"},
                },
                "required": ["start_date", "end_date"],
            },
        },
        {
            "name": "run_pipeline",
            "description": (
                "Run the full spot-market ingestion pipeline for one PDF report. "
                "Parses DA/RT prices, inter-provincial data, generates AI summary, "
                "and populates the knowledge pool."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "pdf_path": {"type": "string", "description": "Absolute or repo-relative path to the PDF"},
                    "dry_run":  {"type": "boolean", "description": "If true, parse only — no DB writes"},
                },
                "required": ["pdf_path"],
            },
        },
        {
            "name": "ingest_kb_document",
            "description": (
                "Add any reference document — Excel (.xlsx/.xls), PDF, PowerPoint (.pptx), "
                "Word (.docx), plain text, or image — to the knowledge base so it can be "
                "searched with search_reference_docs. "
                "Use s3_key to fetch a file from the uploads S3 bucket, or file_path for a "
                "local/repo-relative path. "
                "Supports all file types including Excel spreadsheets with trading data, "
                "network loss tables, contract volumes, or any structured market data."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "s3_key": {
                        "type": "string",
                        "description": "S3 object key in the uploads bucket, e.g. 'uploads/跨区交易.xlsx'",
                    },
                    "file_path": {
                        "type": "string",
                        "description": "Repo-relative or absolute path, e.g. 'data/market-fundamentals/report.xlsx'",
                    },
                    "category": {
                        "type": "string",
                        "description": "Optional: market_rules | annual_report | policy_doc | technical_spec | research_report | other",
                    },
                    "app": {
                        "type": "string",
                        "description": "'shared' (visible to all agents, default) or 'strategist'",
                    },
                },
                "required": [],
            },
        },
        {
            "name": "get_market_fundamentals",
            "description": (
                "Fetch market fundamentals for Chinese electricity provinces: "
                "installed capacity by fuel type (万kW), generation mix (亿kWh), "
                "and seasonal peak loads (MW). Data covers 2024 and 2025. "
                "Use this to answer questions about fuel mix, renewables penetration, "
                "storage capacity, peak demand, or structural comparisons between provinces."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "provinces": {
                        "type": "array", "items": {"type": "string"},
                        "description": "Optional list of province_en names, e.g. ['Shandong','Guangdong']. Omit for all.",
                    },
                    "year": {
                        "type": "integer",
                        "description": "2024 or 2025 (default 2025)",
                    },
                },
                "required": [],
            },
        },
        {
            "name": "search_reference_docs",
            "description": (
                "Search the uploaded reference document knowledge base. Covers market rules, "
                "annual exchange reports, policy documents, technical specs, research reports, "
                "AND Excel spreadsheets with trading data (cross-regional volumes, network losses, "
                "contract quantities, etc.). Supports all file types: PDF, Excel (.xlsx/.xls), "
                "PPTX, DOCX, TXT. Use this whenever the user asks about trading rules, "
                "settlement procedures, market mechanisms, regulatory requirements, "
                "or any data from documents they have uploaded."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Search query in Chinese or English, e.g. '结算周期' or 'DA price cap'",
                    },
                    "category": {
                        "type": "string",
                        "description": (
                            "Optional category filter: market_rules | annual_report | "
                            "policy_doc | technical_spec | research_report | other. "
                            "Omit to search all categories."
                        ),
                    },
                    "limit": {
                        "type": "integer",
                        "description": "Max number of text chunks to return (default 5, max 10).",
                    },
                },
                "required": ["query"],
            },
        },
        {
            "name": "get_bess_pnl",
            "description": (
                "Fetch daily P&L and dispatch metrics for Inner Mongolia BESS assets "
                "across all strategy scenarios. Returns total_pnl, market_revenue, "
                "compensation_revenue, discharge_mwh, charge_mwh, avg_daily_cycles. "
                "The 4 assets: suyou, hangjinqi, siziwangqi, gushanliang. "
                "Scenarios: perfect_foresight_hourly (LP upper bound), "
                "forecast_ols_rt_time_v1 (LP forecast), nominated_dispatch (ops), "
                "cleared_actual (ops actual), trading_cleared (id market). "
                "Use this for BESS performance analysis, strategy comparison, or P&L attribution."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "start_date": {"type": "string", "description": "ISO date, e.g. '2026-01-01'"},
                    "end_date":   {"type": "string", "description": "ISO date, e.g. '2026-04-30'"},
                    "asset_codes": {
                        "type": "array", "items": {"type": "string"},
                        "description": "Optional list of asset codes. Omit for all 4 IM assets.",
                    },
                },
                "required": ["start_date", "end_date"],
            },
        },
    ]

    # ── Tool dispatcher ────────────────────────────────────────────────────────
    def _dispatch_tool(name: str, inputs: dict) -> str:
        from services.spot_mcp.tools import (
            get_spot_prices as _gsp,
            get_interprov_flow as _gif,
            get_market_summaries as _gms,
            run_pipeline as _rp,
            get_market_fundamentals as _gmf,
            ingest_kb_document as _ikd,
        )
        try:
            if name == "get_spot_prices":
                result = _gsp(**inputs)
            elif name == "get_interprov_flow":
                result = _gif(**inputs)
            elif name == "get_market_summaries":
                result = _gms(**inputs)
            elif name == "run_pipeline":
                result = _rp(**inputs)
            elif name == "get_market_fundamentals":
                result = _gmf(**inputs)
            elif name == "search_reference_docs":
                from services.knowledge_pool.knowledge_docs import search_reference_docs as _srd
                _limit = min(int(inputs.get("limit", 5)), 10)
                rows = _srd(
                    query=inputs["query"],
                    category=inputs.get("category"),
                    app="strategist",
                    limit=_limit,
                )
                result = {"count": len(rows), "chunks": rows}
            elif name == "ingest_kb_document":
                result = _ikd(**inputs)
            elif name == "get_bess_pnl":
                from services.bess_mcp.tools import bess_get_portfolio_pnl as _bgms
                result = _bgms(
                    asset_codes=inputs.get("asset_codes"),
                    start_date=inputs["start_date"],
                    end_date=inputs["end_date"],
                )
            else:
                result = {"error": f"Unknown tool: {name}"}
        except Exception as _e:
            result = {"error": str(_e)}
        return _json.dumps(result, default=str)

    # ── Agent turn (handles multi-step tool-use loop) ──────────────────────────
    def _run_agent_turn(
        messages: list, system: str, text_placeholder=None
    ) -> tuple[str, list, list]:
        _api_key = _os.environ.get("ANTHROPIC_API_KEY", "")
        if not _api_key:
            return _t("agent_no_key"), messages, []

        client = _anthropic.Anthropic(api_key=_api_key)
        tool_events: list[dict] = []
        # Status line shown during tool calls (lives inside the same chat message)
        _status_ph = st.empty() if text_placeholder is not None else None

        while True:
            streamed_text = ""

            if text_placeholder is not None:
                if _status_ph:
                    _status_ph.caption("⏳ Thinking…")
                with client.messages.stream(
                    model="claude-sonnet-4-6",
                    max_tokens=4096,
                    system=system,
                    tools=_AGENT_TOOLS,
                    messages=messages,
                ) as _stream:
                    for _chunk in _stream.text_stream:
                        streamed_text += _chunk
                        if _status_ph:
                            _status_ph.empty()
                        text_placeholder.markdown(streamed_text + "▌")
                    _final = _stream.get_final_message()
            else:
                _final = client.messages.create(
                    model="claude-sonnet-4-6",
                    max_tokens=4096,
                    system=system,
                    tools=_AGENT_TOOLS,
                    messages=messages,
                )
                streamed_text = next(
                    (b.text for b in _final.content if hasattr(b, "text")), ""
                )

            messages = messages + [{"role": "assistant", "content": _final.content}]

            if _final.stop_reason == "end_turn":
                if _status_ph:
                    _status_ph.empty()
                if text_placeholder is not None:
                    text_placeholder.markdown(streamed_text)
                return streamed_text, messages, tool_events

            if _final.stop_reason != "tool_use":
                if _status_ph:
                    _status_ph.empty()
                return f"Unexpected stop_reason: {_final.stop_reason}", messages, tool_events

            tool_results = []
            for block in _final.content:
                if block.type == "tool_use":
                    if _status_ph:
                        _icon = _TOOL_ICONS.get(block.name, "⚙️")
                        _status_ph.caption(f"{_icon} Calling `{block.name}`…")
                    result_str = _dispatch_tool(block.name, block.input)
                    tool_events.append({"tool": block.name, "result": result_str})
                    tool_results.append({
                        "type": "tool_result",
                        "tool_use_id": block.id,
                        "content": result_str,
                    })

            if _status_ph:
                _status_ph.empty()

            messages = messages + [{"role": "user", "content": tool_results}]

    # ── Session state ──────────────────────────────────────────────────────────
    if "spot_strat_session_id" not in st.session_state:
        st.session_state["spot_strat_session_id"] = str(_uuid.uuid4())
    if "agent_messages" not in st.session_state:
        st.session_state["agent_messages"] = []
    if "agent_display" not in st.session_state:
        st.session_state["agent_display"] = [
            {"role": "assistant", "content": _t("agent_welcome"), "tool": None}
        ]
    if "spot_mem_suggestions" not in st.session_state:
        st.session_state["spot_mem_suggestions"] = []  # kept for back-compat, unused

    st.subheader(_t("agent_title"))
    _n_ins = (_cached_memory_stats() or {}).get("total", 0) or 0
    if _n_ins:
        st.caption(
            f"{_t('agent_caption')} · "
            f"{_n_ins} expert insight{'s' if _n_ins != 1 else ''} accumulated"
        )
    else:
        st.caption(_t("agent_caption"))

    # ── Clear button (generates new session UUID, old session preserved in DB) ─
    if st.button(_t("agent_clear"), key="agent_clear_btn"):
        st.session_state["spot_strat_session_id"] = str(_uuid.uuid4())
        st.session_state["agent_messages"] = []
        st.session_state["agent_display"] = [
            {"role": "assistant", "content": _t("agent_welcome"), "tool": None}
        ]
        st.rerun()

    # ── Resume previous session (only when current chat is empty) ─────────────
    if not st.session_state["agent_messages"]:
        _recent_sessions = _list_recent_spot_sessions()
        if not _recent_sessions.empty:
            with st.expander("Resume a previous conversation?", expanded=False):
                for _, _srow in _recent_sessions.iterrows():
                    _sess_label = (
                        f"{_srow['updated_at'].strftime('%Y-%m-%d %H:%M')} — "
                        f"{int(_srow['msg_count'])} messages"
                    )
                    if st.button(_sess_label, key=f"resume_spot_{_srow['session_id']}"):
                        _loaded_display = _load_spot_session(_srow["session_id"])
                        st.session_state["spot_strat_session_id"] = _srow["session_id"]
                        st.session_state["agent_display"] = (
                            _display_from_session(_loaded_display)
                        )
                        st.session_state["agent_messages"] = (
                            _api_messages_from_display(_loaded_display)
                        )
                        st.rerun()

    # ── Tool result display helper ─────────────────────────────────────────────
    _TOOL_ICONS = {
        "get_spot_prices":        "📊",
        "get_interprov_flow":     "🔀",
        "get_market_summaries":   "📝",
        "run_pipeline":           "⚙️",
        "get_market_fundamentals": "🏭",
        "search_reference_docs":  "🔍",
        "ingest_kb_document":     "📥",
        "get_bess_pnl":           "⚡",
    }

    def _render_tool_result(tool_name: str, content_str: str):
        icon = _TOOL_ICONS.get(tool_name, "🔧")
        try:
            parsed = _json.loads(content_str)
        except Exception:
            st.code(content_str, language="json")
            return

        if "error" in parsed:
            st.error(parsed["error"])
            return

        if tool_name == "get_spot_prices":
            rows = parsed.get("rows", [])
            n = len(rows)
            provs = sorted({r.get("province_en", "") for r in rows})
            dates = sorted({r.get("report_date", "") for r in rows})
            date_range = f"{dates[0]} → {dates[-1]}" if dates else "—"
            c1, c2, c3 = st.columns(3)
            c1.metric("Rows", f"{n:,}")
            c2.metric("Provinces", f"{len(provs)}")
            c3.metric("Date range", date_range)
            if provs:
                st.caption("Provinces: " + ", ".join(provs[:8]) + ("…" if len(provs) > 8 else ""))
            with st.expander("Raw data", expanded=False):
                st.json(parsed, expanded=1)

        elif tool_name == "get_interprov_flow":
            rows = parsed.get("rows", [])
            n = len(rows)
            dates = sorted({r.get("report_date", "") for r in rows})
            date_range = f"{dates[0]} → {dates[-1]}" if dates else "—"
            c1, c2 = st.columns(2)
            c1.metric("Rows", f"{n:,}")
            c2.metric("Date range", date_range)
            with st.expander("Raw data", expanded=False):
                st.json(parsed, expanded=1)

        elif tool_name == "get_market_summaries":
            items = parsed.get("summaries", [])
            n = len(items)
            dates = sorted({r.get("report_date", "") for r in items})
            date_range = f"{dates[0]} → {dates[-1]}" if dates else "—"
            c1, c2 = st.columns(2)
            c1.metric("Summaries", f"{n:,}")
            c2.metric("Date range", date_range)
            if items:
                with st.expander("Latest summary", expanded=False):
                    st.markdown(items[0].get("summary_text", "")[:600])
            with st.expander("All summaries (raw)", expanded=False):
                st.json(parsed, expanded=1)

        elif tool_name == "get_market_fundamentals":
            provs = parsed.get("provinces", [])
            n = len(provs)
            year = parsed.get("year", "")
            c1, c2 = st.columns(2)
            c1.metric("Provinces", f"{n:,}")
            c2.metric("Year", str(year))
            with st.expander("Raw data", expanded=False):
                st.json(parsed, expanded=1)

        elif tool_name == "search_reference_docs":
            chunks = parsed.get("chunks", parsed.get("results", []))
            n = len(chunks)
            docs = sorted({c.get("file_name", "") for c in chunks})
            st.metric("Chunks found", f"{n}")
            if docs:
                st.caption("From: " + ", ".join(docs[:4]) + ("…" if len(docs) > 4 else ""))
            if chunks:
                with st.expander("Top result", expanded=False):
                    top = chunks[0]
                    st.caption(f"{top.get('file_name', '')} · p{top.get('page_no', '?')} · {top.get('category', '')}")
                    st.markdown(top.get("chunk_text", "")[:500])
            with st.expander("All chunks (raw)", expanded=False):
                st.json(parsed, expanded=1)

        elif tool_name == "ingest_kb_document":
            status = parsed.get("status", "")
            fname = parsed.get("filename", "")
            cat = parsed.get("category", "")
            if status == "ingested":
                st.success(f"Ingested **{fname}** (category: {cat})")
            elif status == "duplicate":
                st.info(f"**{fname}** already in KB (doc_id={parsed.get('doc_id')})")
            else:
                st.warning(parsed.get("message", str(parsed)))

        elif tool_name == "run_pipeline":
            upserted = parsed.get("upserted", 0)
            dates = parsed.get("dates", [])
            errs = parsed.get("errors", [])
            c1, c2, c3 = st.columns(3)
            c1.metric("Rows upserted", upserted)
            c2.metric("Dates", len(dates))
            c3.metric("Errors", len(errs))
            if errs:
                st.warning("\n".join(errs[:3]))
            with st.expander("Raw data", expanded=False):
                st.json(parsed, expanded=1)

        elif tool_name == "get_bess_pnl":
            rows = parsed.get("rows", [])
            n = len(rows)
            assets = sorted({r.get("asset_code", "") for r in rows})
            dates = sorted({r.get("trade_date", "") for r in rows})
            date_range = f"{dates[0]} → {dates[-1]}" if dates else "—"
            scenarios = sorted({r.get("scenario_name", "") for r in rows})
            c1, c2, c3 = st.columns(3)
            c1.metric("Rows", f"{n:,}")
            c2.metric("Assets", len(assets))
            c3.metric("Date range", date_range)
            if assets:
                st.caption("Assets: " + ", ".join(assets))
            if scenarios:
                st.caption("Scenarios: " + ", ".join(scenarios))
            with st.expander("Raw data", expanded=False):
                st.json(parsed, expanded=1)

        else:
            _n = parsed.get("count", parsed.get("rows", ""))
            _n = len(_n) if isinstance(_n, list) else _n
            st.caption(_t("agent_tool_result", n=_n) if isinstance(_n, int) else "")
            st.json(parsed, expanded=1)

    # ── Chat-area file upload ──────────────────────────────────────────────────
    with st.expander("📎 Upload file to knowledge base", expanded=False):
        _chat_uploads = st.file_uploader(
            "Drop a file here to ingest it immediately — then ask the agent about it",
            type=["pdf", "pptx", "ppt", "txt", "docx", "doc",
                  "xlsx", "xls", "png", "jpg", "jpeg", "webp"],
            accept_multiple_files=True,
            key="agent_chat_uploader",
            label_visibility="collapsed",
        )
        if _chat_uploads:
            from services.knowledge_pool.knowledge_docs import register_and_ingest as _rai_chat
            _api_key_up = _os.environ.get("ANTHROPIC_API_KEY")
            _processed = st.session_state.setdefault("_agent_processed_uploads", set())
            _new_files = [
                _f for _f in _chat_uploads
                if f"{_f.name}_{_f.size}" not in _processed
            ]
            if _new_files and st.button(
                f"Ingest {len(_new_files)} file(s)", key="agent_ingest_btn", type="primary"
            ):
                _ingest_results = []
                _prog_up = st.progress(0, text="Ingesting…")
                for _fi, _f in enumerate(_new_files):
                    _prog_up.progress((_fi + 1) / len(_new_files), text=f"{_f.name}…")
                    try:
                        _doc_id, _is_new, _cat = _rai_chat(
                            _f.read(), _f.name, api_key=_api_key_up, app="strategist"
                        )
                        _ingest_results.append((_f.name, _cat, _is_new))
                        _processed.add(f"{_f.name}_{_f.size}")
                    except Exception as _upe:
                        st.error(f"{_f.name}: {_upe}")
                _prog_up.empty()
                if _ingest_results:
                    _summary = "; ".join(
                        f"**{n}** ({c}, {'new' if isnew else 'already indexed'})"
                        for n, c, isnew in _ingest_results
                    )
                    _up_user_msg = f"📎 I uploaded: {_summary}"
                    _up_asst_msg = (
                        f"File(s) ingested into the knowledge base: {_summary}. "
                        "You can now ask me to search or analyse their contents."
                    )
                    st.session_state["agent_display"].append(
                        {"role": "user", "content": _up_user_msg, "tool": None}
                    )
                    st.session_state["agent_messages"].append(
                        {"role": "user", "content": _up_user_msg}
                    )
                    st.session_state["agent_display"].append(
                        {"role": "assistant", "content": _up_asst_msg, "tool": None}
                    )
                    st.session_state["agent_messages"].append(
                        {"role": "assistant", "content": _up_asst_msg}
                    )
                    st.rerun()

    # ── Render existing chat history ───────────────────────────────────────────
    for _msg in st.session_state["agent_display"]:
        if _msg["role"] == "tool":
            _icon = _TOOL_ICONS.get(_msg["tool"], "🔧")
            with st.expander(f"{_icon} {_t('agent_tool_call', tool=_msg['tool'])}", expanded=False):
                _render_tool_result(_msg["tool"], _msg["content"])
        else:
            with st.chat_message(_msg["role"]):
                st.markdown(_msg["content"])

    # ── Chat input ─────────────────────────────────────────────────────────────
    _user_input = st.chat_input(_t("agent_placeholder"), key="agent_input")

    if _user_input:
        st.session_state["agent_display"].append(
            {"role": "user", "content": _user_input, "tool": None}
        )
        with st.chat_message("user"):
            st.markdown(_user_input)

        st.session_state["agent_messages"].append(
            {"role": "user", "content": _user_input}
        )

        with st.chat_message("assistant"):
            _text_ph = st.empty()
            try:
                _reply, _new_msgs, _tool_events = _run_agent_turn(
                    st.session_state["agent_messages"],
                    _build_spot_system(_user_input),
                    text_placeholder=_text_ph,
                )
            except Exception as _exc:
                _reply = _t("agent_error", err=str(_exc))
                _new_msgs = st.session_state["agent_messages"]
                _tool_events = []
                _text_ph.markdown(_reply)

            for _ev in _tool_events:
                st.session_state["agent_display"].append({
                    "role": "tool",
                    "content": _ev["result"],
                    "tool": _ev["tool"],
                })

            st.session_state["agent_messages"] = _new_msgs
            st.session_state["agent_display"].append(
                {"role": "assistant", "content": _reply, "tool": None}
            )

        # ── Auto-save memories + log conversation turn ─────────────────────────
        try:
            _suggestions = _extract_spot_memories(_user_input, _reply)
            for _sug in _suggestions:
                _save_spot_memory(
                    _sug["category"], _sug["subject"], _sug["content"], source="auto"
                )
            if _suggestions:
                st.toast(_t("mem_saved_ok", n=len(_suggestions)))
        except Exception:
            pass

        # ── Extract structured expert insights into kp_expert_insights ─────────
        try:
            from services.knowledge_pool.expert_memory import extract_spot_insights as _ext_insights
            _api_key_ins = _os.environ.get("ANTHROPIC_API_KEY", "")
            if _api_key_ins:
                _n_insights = _ext_insights(_user_input, _reply, _api_key_ins)
        except Exception:
            pass

        # ── Persist session display to DB ──────────────────────────────────────
        try:
            _save_spot_session(
                st.session_state["spot_strat_session_id"],
                st.session_state["agent_display"],
            )
        except Exception:
            pass

        try:
            from services.knowledge_pool.knowledge_docs import log_conversation_turn as _log_turn
            _log_turn(_user_input, _reply)
        except Exception:
            pass

    # ── Knowledge Gap Interview ────────────────────────────────────────────────
    st.divider()
    with st.expander("🎓 Teach the Strategist — Knowledge Gap Interview", expanded=False):
        for _ik in [("interview_questions", []), ("interview_idx", 0),
                    ("interview_answers", 0), ("interview_kb_queried", False),
                    ("interview_kb_results", {}), ("interview_pending_qs", [])]:
            if _ik[0] not in st.session_state:
                st.session_state[_ik[0]] = _ik[1]

        _iq  = st.session_state["interview_questions"]
        _ii  = st.session_state["interview_idx"]
        _pqs = st.session_state["interview_pending_qs"]

        # Stage 0: Generate questions
        if not _iq:
            st.markdown(
                "The agent audits its knowledge base, identifies gaps in China electricity "
                "market expertise, then searches the knowledge pool for answers before "
                "asking you only the questions it couldn't resolve."
            )
            if st.button("Generate Knowledge Gap Questions", key="gen_spot_interview"):
                with st.spinner("Auditing knowledge base and identifying gaps…"):
                    _new_qs = _generate_spot_interview_questions()
                st.session_state["interview_questions"]   = _new_qs
                st.session_state["interview_idx"]         = 0
                st.session_state["interview_answers"]     = 0
                st.session_state["interview_kb_queried"]  = False
                st.session_state["interview_kb_results"]  = {}
                st.session_state["interview_pending_qs"]  = []
                st.rerun()

        # Stage 1: KB search first-pass
        elif not st.session_state["interview_kb_queried"]:
            st.markdown("**Generated knowledge gap questions:**")
            for _qi, _qo in enumerate(_iq):
                st.markdown(f"{_qi+1}. **[{_qo['topic']}]** {_qo['question']}")
                st.caption(f"*Why: {_qo.get('why_asking', '')}*")
            st.divider()
            _col_kb, _col_skip = st.columns([2, 1])
            with _col_kb:
                if st.button("Search KB First", key="interview_kb_search", type="primary"):
                    _kbres = {}
                    with st.spinner("Searching knowledge pool for each gap question…"):
                        for _qo in _iq:
                            _ans = _answer_from_kb(_qo["question"])
                            _kbres[_qo["question"]] = _ans
                            if _ans:
                                try:
                                    _store_spot_interview_answer(
                                        _qo["question"], _ans, _qo["topic"],
                                        confidence="medium",
                                    )
                                except Exception:
                                    pass
                    st.session_state["interview_kb_results"]  = _kbres
                    st.session_state["interview_pending_qs"]  = [
                        _qo for _qo in _iq if not _kbres.get(_qo["question"])
                    ]
                    st.session_state["interview_kb_queried"]  = True
                    st.session_state["interview_idx"]         = 0
                    st.rerun()
            with _col_skip:
                if st.button("Answer Yourself (skip KB)", key="interview_skip_kb"):
                    st.session_state["interview_pending_qs"]  = list(_iq)
                    st.session_state["interview_kb_queried"]  = True
                    st.session_state["interview_idx"]         = 0
                    st.rerun()

        # Stage 3: All questions done — summary
        elif _ii >= len(_pqs):
            _kbres = st.session_state["interview_kb_results"]
            _n_kb  = sum(1 for v in _kbres.values() if v)
            _n_usr = st.session_state["interview_answers"]
            if _n_kb:
                st.success(
                    f"KB answered **{_n_kb}** gap question(s) — stored as medium-confidence insights. "
                    f"You answered **{_n_usr}** additional question(s) as high-confidence insights."
                )
                with st.expander("View KB answers", expanded=False):
                    for _qo in _iq:
                        _ans = _kbres.get(_qo["question"])
                        if _ans:
                            st.markdown(f"**Q: {_qo['question']}**")
                            st.markdown(_ans[:500] + ("…" if len(_ans) > 500 else ""))
                            st.divider()
            else:
                st.success(
                    f"Interview complete — {_n_usr} expert answer(s) stored as high-confidence insights."
                )
            if st.button("Start New Interview", key="new_spot_interview"):
                for _k2 in ["interview_questions", "interview_pending_qs"]:
                    st.session_state[_k2] = []
                st.session_state["interview_kb_results"]  = {}
                st.session_state["interview_idx"]         = 0
                st.session_state["interview_answers"]     = 0
                st.session_state["interview_kb_queried"]  = False
                st.rerun()

        # Stage 2: User Q&A for unanswered questions
        else:
            _q = _pqs[_ii]
            _kbres = st.session_state["interview_kb_results"]
            if _ii == 0 and _kbres:
                _n_auto = sum(1 for v in _kbres.values() if v)
                if _n_auto:
                    st.info(
                        f"KB answered {_n_auto} of {len(_iq)} questions. "
                        f"Please answer the remaining {len(_pqs)}."
                    )
            st.progress(_ii / max(len(_pqs), 1), text=f"Question {_ii + 1} of {len(_pqs)}")
            st.markdown(f"**[{_q['topic']}]** {_q['question']}")
            st.caption(f"*Why this matters: {_q.get('why_asking', '')}*")
            _ans = st.text_area(
                "Your answer:", key=f"spot_interview_ans_{_ii}", height=120,
                placeholder="Share what you know from hands-on experience…",
            )
            _col_sub, _col_ski = st.columns([2, 1])
            with _col_sub:
                if st.button("Submit Answer", key=f"spot_interview_submit_{_ii}", type="primary"):
                    if _ans.strip():
                        try:
                            _store_spot_interview_answer(
                                _q["question"], _ans.strip(), _q["topic"], confidence="high"
                            )
                            st.session_state["interview_idx"]     += 1
                            st.session_state["interview_answers"] += 1
                            st.rerun()
                        except Exception as _e:
                            st.error(f"Failed to store answer: {_e}")
                    else:
                        st.warning("Please enter an answer before submitting.")
            with _col_ski:
                if st.button("Skip", key=f"spot_interview_skip_{_ii}"):
                    st.session_state["interview_idx"] += 1
                    st.rerun()

    # ── Memory management ─────────────────────────────────────────────────────
    with st.expander(f"🗄️ {_t('mem_manage')}", expanded=False):
        st.caption(_t("mem_caption"))
        _mem_df = _load_spot_memories(_SPOT_MEM_KEY)
        if _mem_df.empty:
            st.info(_t("mem_empty"))
        else:
            for _row in _mem_df.itertuples():
                _c1, _c2, _c3 = st.columns([1, 5, 1])
                _c1.markdown(f"**{_row.category}**")
                _c2.markdown(f"**{_row.subject}** — {_row.content}")
                if _c3.button(_t("mem_delete"), key=f"del_spot_mem_{_row.id}"):
                    _delete_spot_memory(_row.id)
                    st.rerun()

    # ── Knowledge Base ─────────────────────────────────────────────────────────
    with st.expander(f"📚 {_t('kb_title')}", expanded=False):
        from services.knowledge_pool.knowledge_docs import (
            init_knowledge_tables as _kb_init,
            register_and_ingest as _kb_ingest,
            list_knowledge_docs as _kb_list,
            delete_knowledge_doc as _kb_delete,
            CATEGORY_LABELS as _KB_CATS,
        )
        if not st.session_state.get("_kp_tables_init_done"):
            try:
                _kb_init()
                st.session_state["_kp_tables_init_done"] = True
            except Exception as _kbi_exc:
                st.warning(f"KB table init: {_kbi_exc}")

        st.caption(_t("kb_caption"))

        # --- KB metrics ---
        try:
            _skbs = _spot_kb_stats()
            _skm1, _skm2, _skm3 = st.columns(3)
            _skm1.metric("📄 Files", f"{_skbs['total']:,}")
            _skm2.metric("🧩 Chunks", f"{_skbs['n_chunks']:,}")
            _skm3.metric("💡 Insights", f"{_skbs['n_insights']:,}")
            st.caption(
                f"**Parse success rate** — {_skbs['parsed']}/{_skbs['total']} files "
                f"parsed without error"
            )
            st.progress(_skbs["parse_pct"], text=f"{_skbs['parse_pct']:.0%}")
        except Exception:
            pass

        _kb_up_tab, _kb_url_tab = st.tabs(["📂 Upload Files", "🌐 Fetch from URL"])

        # ── Upload Files tab ───────────────────────────────────────────────────
        with _kb_up_tab:
            _kb_files = st.file_uploader(
                _t("kb_upload_label"),
                type=["pdf", "pptx", "ppt", "txt", "docx", "doc",
                      "xlsx", "xls", "png", "jpg", "jpeg", "webp"],
                accept_multiple_files=True,
                key="kb_uploader",
            )
            _cat_options = [_t("kb_category_auto")] + list(_KB_CATS.keys())
            _cat_labels  = [_t("kb_category_auto")] + list(_KB_CATS.values())
            _cat_sel_idx = st.selectbox(
                _t("kb_category_label"),
                options=range(len(_cat_options)),
                format_func=lambda i: _cat_labels[i],
                key="kb_cat_sel",
            )
            _cat_override = None if _cat_sel_idx == 0 else _cat_options[_cat_sel_idx]

            if st.button(_t("kb_upload_btn"), key="kb_upload_btn", disabled=not _kb_files):
                _api_key = _os.environ.get("ANTHROPIC_API_KEY")
                _added, _dupes, _errors = 0, [], []
                _new_doc_ids = []
                _prog = st.progress(0, text="Ingesting…")
                for _fi, _f in enumerate(_kb_files):
                    _prog.progress((_fi + 1) / len(_kb_files), text=f"Processing {_f.name}…")
                    _bytes = _f.read()
                    try:
                        _doc_id, _is_new, _cat = _kb_ingest(
                            _bytes, _f.name,
                            category_override=_cat_override,
                            api_key=_api_key,
                        )
                        if _is_new:
                            _added += 1
                            _new_doc_ids.append(_doc_id)
                        else:
                            _dupes.append(_f.name)
                    except Exception as _exc:
                        _errors.append((_f.name, str(_exc)))
                _prog.empty()

                if _added:
                    st.success(_t("kb_success", n=_added))
                    if _api_key and _new_doc_ids:
                        try:
                            from services.knowledge_pool.expert_memory import digest_spot_kb_docs as _digest
                            _n_ins = _digest(_api_key, doc_ids=_new_doc_ids)
                            if _n_ins:
                                st.toast(f"{_n_ins} insight{'s' if _n_ins != 1 else ''} extracted from uploaded document(s)")
                        except Exception:
                            pass
                for _d in _dupes:
                    st.info(_t("kb_duplicate", fname=_d))
                for _fn, _err in _errors:
                    st.error(_t("kb_failed", fname=_fn, err=_err))
                _cached_kb_docs.clear()
                st.rerun()

        # ── Fetch from URL tab ─────────────────────────────────────────────────
        with _kb_url_tab:
            st.caption("Paste a public URL (policy doc, research article, regulator notice). The page text will be extracted and added to the knowledge base.")
            _kb_url_input = st.text_input("URL", placeholder="https://...", key="kb_url_input", label_visibility="collapsed")
            if st.button("Fetch & Add to KB", key="kb_url_fetch_btn", disabled=not _kb_url_input):
                _api_key = _os.environ.get("ANTHROPIC_API_KEY")
                with st.spinner("Fetching and indexing…"):
                    try:
                        from services.knowledge_pool.knowledge_docs import register_url as _kb_register_url
                        _url_doc_id, _url_is_new, _url_cat = _kb_register_url(_kb_url_input.strip(), api_key=_api_key)
                        if _url_is_new:
                            st.success(f"Added to KB (category: {_url_cat})")
                            if _api_key:
                                try:
                                    from services.knowledge_pool.expert_memory import digest_spot_kb_docs as _digest_url
                                    _n_url_ins = _digest_url(_api_key, doc_ids=[_url_doc_id])
                                    if _n_url_ins:
                                        st.toast(f"{_n_url_ins} insight{'s' if _n_url_ins != 1 else ''} extracted")
                                except Exception:
                                    pass
                            _cached_kb_docs.clear()
                            st.rerun()
                        else:
                            st.info("This URL is already in the knowledge base.")
                    except Exception as _url_exc:
                        st.error(f"Fetch failed: {_url_exc}")

        # ── Document list ──────────────────────────────────────────────────────
        _kb_docs = _cached_kb_docs()
        _kb_total = len(_kb_docs)
        _kb_display = _kb_docs[:50]  # cap at 50 rows — rendering all 2000+ docs stalls the page
        st.markdown(
            f"**{_t('kb_doc_list_title')}** "
            + (f"({_kb_total} total, showing latest 50)" if _kb_total > 50 else f"({_kb_total})")
        )
        if not _kb_docs:
            st.info(_t("kb_doc_list_empty"))
        else:
            for _doc in _kb_display:
                _dc1, _dc2, _dc3, _dc4 = st.columns([3, 1, 1, 1])
                _dc1.markdown(
                    f"**{_doc['file_name']}**"
                    + (f"  \n_{_doc['title']}_" if _doc.get('title') and _doc['title'] != _doc['file_name'] else "")
                )
                _dc2.markdown(f"`{_KB_CATS.get(_doc['category'], _doc['category'])}`")
                _dc3.markdown(
                    _t("kb_pages", n=_doc.get('page_count', 0))
                    if _doc.get('ingest_status') == 'parsed'
                    else f"_{_doc.get('ingest_status', '—')}_"
                )
                if _dc4.button(_t("kb_delete"), key=f"kb_del_{_doc['id']}"):
                    _kb_delete(_doc['id'])
                    _cached_kb_docs.clear()
                    st.rerun()

        # ── Batch KB digest ────────────────────────────────────────────────────
        st.divider()
        _col_dig1, _col_dig2, _col_dig3 = st.columns([3, 1, 2])
        with _col_dig1:
            st.caption(
                "Extract expert insights from synthesized knowledge base documents "
                "into the insight pool (processes up to 100 undigested docs per run)."
            )
        with _col_dig2:
            if st.button("Digest KB → Insights", key="kb_digest_btn"):
                with st.spinner("Extracting insights from synthesized documents…"):
                    try:
                        from services.knowledge_pool.expert_memory import digest_spot_kb_docs as _digest_batch
                        _n_batch = _digest_batch(
                            _os.environ.get("ANTHROPIC_API_KEY", ""), limit=100
                        )
                        st.success(f"{_n_batch} new insight{'s' if _n_batch != 1 else ''} extracted.")
                    except Exception as _de:
                        st.error(f"Digest failed: {_de}")
        with _col_dig3:
            st.caption(
                "Trigger the full synthesis + digest pipeline via Hermes "
                "(runs in background — check logs for results)."
            )
            if st.button("▶ Run Digest Now", key="kb_hermes_digest_btn"):
                _hermes_url = _os.environ.get("HERMES_URL", "")
                if not _hermes_url:
                    st.warning("HERMES_URL not configured.")
                else:
                    try:
                        import requests as _req
                        _r = _req.post(
                            _hermes_url.rstrip("/") + "/hermes/knowledge/digest",
                            timeout=10,
                            verify=False,
                        )
                        if _r.status_code == 200:
                            st.success("Digest job started — insights available in ~5 min.")
                        else:
                            st.error(f"Hermes returned {_r.status_code}: {_r.text[:120]}")
                    except Exception as _he:
                        st.error(f"Could not reach Hermes: {_he}")


# ── Tab 9 helpers — module-level so @st.cache_data can hash them stably ───────
PROVINCES_MAP: dict[str, str] = {
    "山东": "Shandong", "山西": "Shanxi", "蒙西": "Mengxi", "内蒙古": "Mengxi",
    "甘肃": "Gansu", "广东": "Guangdong", "四川": "Sichuan", "云南": "Yunnan",
    "贵州": "Guizhou", "广西": "Guangxi", "湖南": "Hunan", "湖北": "Hubei",
    "安徽": "Anhui", "浙江": "Zhejiang", "江苏": "Jiangsu", "福建": "Fujian",
    "河南": "Henan", "陕西": "Shaanxi", "宁夏": "Ningxia", "新疆": "Xinjiang",
    "辽宁": "Liaoning", "吉林": "Jilin", "黑龙江": "Heilongjiang", "蒙东": "Mengdong",
    "河北": "Hebei", "冀北": "Hebei-North", "冀南": "Hebei-South",
    "河北南网": "Hebei-South", "青海": "Qinghai",
    "江西": "Jiangxi", "海南": "Hainan", "重庆": "Chongqing", "上海": "Shanghai",
    "北京": "Beijing", "天津": "Tianjin",
}

_S3_BUCKET = _os.environ.get("UPLOADS_BUCKET", "bess-uploader-data-chen-singp-2026")
_S3_PREFIX = "spot-reports"


def _parse_pdf_date_range(stem: str, year: int = 2026):
        stem = stem.strip().rstrip("）)） ")

        m = _re.fullmatch(r"(\d{2})(\d{2})(?:-(\d{2})(\d{2}))?", stem)
        if m:
            try:
                start = date(year, int(m.group(1)), int(m.group(2)))
                end   = date(year, int(m.group(3) or m.group(1)),
                             int(m.group(4) or m.group(2)))
                return start, end
            except ValueError:
                pass

        m = _re.search(r"(\d{1,2})\.(\d{1,2})(?:-(?:(\d{1,2})\.)?(\d{1,2}))?", stem)
        if m:
            try:
                m1, d1 = int(m.group(1)), int(m.group(2))
                start  = date(year, m1, d1)
                if m.group(4):
                    m2 = int(m.group(3)) if m.group(3) else m1
                    d2 = int(m.group(4))
                    end_year = year + 1 if m2 < m1 else year
                    end = date(end_year, m2, d2)
                else:
                    end = start
                return start, end
            except ValueError:
                pass

        return None


@st.cache_data(ttl=600, show_spinner=False)
def _scan_pdf_inventory(year: int = 2026):
    data_dir = _REPO / "data" / "spot reports" / str(year)
    pdfs = []

    if data_dir.exists():
        for p in sorted(data_dir.glob("*.pdf")):
            stem = p.stem
            m = _re.search(r"[（(]([^)）]+)[）)]", stem)
            if not m:
                continue
            date_range_result = _parse_pdf_date_range(m.group(1).strip(), year)
            if date_range_result:
                pdfs.append((p.name, date_range_result[0], date_range_result[1], p))
        return pdfs

    # AWS: no local data dir — scan S3
    import boto3 as _boto3
    from botocore.config import Config as _BotoCfg
    s3 = _boto3.client("s3", config=_BotoCfg(connect_timeout=5, read_timeout=15, retries={"max_attempts": 1}))
    prefix = f"{_S3_PREFIX}/{year}/"
    try:
        paginator = s3.get_paginator("list_objects_v2")
        for page in paginator.paginate(Bucket=_S3_BUCKET, Prefix=prefix):
            for obj in sorted(page.get("Contents", []), key=lambda o: o["Key"]):
                key = obj["Key"]
                fname = key.split("/")[-1]
                if not fname.lower().endswith(".pdf"):
                    continue
                stem = Path(fname).stem
                m = _re.search(r"[（(]([^)）]+)[）)]", stem)
                if not m:
                    continue
                date_range_result = _parse_pdf_date_range(m.group(1).strip(), year)
                if date_range_result:
                    pdfs.append((fname, date_range_result[0], date_range_result[1], key))
    except Exception:
        pass
    return pdfs


@st.cache_data(ttl=300, show_spinner=False)
def _db_coverage(year: int = 2026):
    cur = _conn().cursor()
    cur.execute(
        "SELECT DISTINCT report_date FROM spot_daily "
        "WHERE report_date BETWEEN %s AND %s",
        (date(year, 1, 1), date(year, 12, 31)),
    )
    return {r[0] for r in cur.fetchall()}


@st.cache_data(ttl=300, show_spinner=False)
def _db_coverage_detail(year: int = 2026):
    cur = _conn().cursor()
    cur.execute(
        """SELECT report_date::date, COUNT(da_avg), COUNT(rt_avg)
           FROM spot_daily
           WHERE report_date BETWEEN %s AND %s
           GROUP BY 1""",
        (date(year, 1, 1), date(year, 12, 31)),
    )
    return {r[0]: (r[1], r[2]) for r in cur.fetchall()}


# ── News Sources ──────────────────────────────────────────────────────────────
import psycopg2 as _ns_pg2  # top-level so it's available inside the fragment

@st.fragment
def _render_news_sources_tab():
    # Inline CRUD helpers — avoids importing services.hermes (not in this image)
    def _ns_init_db(pg_url: str) -> None:
        """Create hermes schema + news_sources table if they don't exist."""
        conn = _ns_pg2.connect(pg_url, options="-c statement_timeout=15000")
        try:
            with conn.cursor() as cur:
                cur.execute("CREATE SCHEMA IF NOT EXISTS hermes")
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS hermes.news_sources (
                        id                  SERIAL PRIMARY KEY,
                        name                TEXT NOT NULL,
                        url                 TEXT NOT NULL,
                        source_type         TEXT NOT NULL DEFAULT 'wechat',
                        biz_id              TEXT,
                        region_bucket       TEXT,
                        category_hint       TEXT,
                        scrape_config       JSONB,
                        active              BOOLEAN NOT NULL DEFAULT TRUE,
                        last_scraped_at     TIMESTAMPTZ,
                        consecutive_failures INT NOT NULL DEFAULT 0,
                        created_at          TIMESTAMPTZ DEFAULT NOW(),
                        UNIQUE(name, url)
                    )
                """)
            conn.commit()
        finally:
            conn.close()
    def _ns_get_sources(pg_url: str, active_only: bool = True) -> list:
        conn = _ns_pg2.connect(pg_url, options="-c statement_timeout=10000")
        try:
            with conn.cursor() as cur:
                sql = (
                    "SELECT id, name, url, source_type, biz_id, region_bucket, "
                    "category_hint, active, last_scraped_at, consecutive_failures "
                    "FROM hermes.news_sources"
                    + (" WHERE active = TRUE" if active_only else "")
                    + " ORDER BY name"
                )
                cur.execute(sql)
                cols = [d[0] for d in cur.description]
                return [dict(zip(cols, row)) for row in cur.fetchall()]
        finally:
            conn.close()

    def _ns_add_source(pg_url: str, name: str, url: str, source_type: str = "wechat",
                       biz_id=None, region_bucket: str = "全国", category_hint: str = "other") -> dict:
        # Auto-extract biz_id from WeChat article URL if not provided
        if source_type == "wechat" and not biz_id and "mp.weixin.qq.com/s/" in url:
            import re as _re, requests as _rq
            try:
                _r = _rq.get(url, headers={"User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15"}, timeout=20)
                _m = _re.search(r'var\s+biz\s*=\s*"([^"]+)"', _r.text) or _re.search(r'__biz=([A-Za-z0-9=+/]+)', _r.text)
                if _m:
                    biz_id = _m.group(1)
                    url = f"https://mp.weixin.qq.com/mp/profile_ext?action=home&__biz={biz_id}&scene=124"
            except Exception:
                pass
        conn = _ns_pg2.connect(pg_url, options="-c statement_timeout=10000")
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """INSERT INTO hermes.news_sources (name, url, source_type, biz_id, region_bucket, category_hint)
                       VALUES (%s, %s, %s, %s, %s, %s)
                       ON CONFLICT (name, url) DO UPDATE SET
                         source_type=EXCLUDED.source_type, biz_id=COALESCE(EXCLUDED.biz_id, hermes.news_sources.biz_id),
                         region_bucket=EXCLUDED.region_bucket, category_hint=EXCLUDED.category_hint, active=TRUE
                       RETURNING id, name, url, source_type, biz_id, region_bucket, category_hint, active""",
                    (name, url, source_type, biz_id, region_bucket, category_hint),
                )
                row = cur.fetchone(); cols = [d[0] for d in cur.description]
            conn.commit()
            return dict(zip(cols, row))
        finally:
            conn.close()

    def _ns_set_active(pg_url: str, source_id: int, active: bool) -> None:
        conn = _ns_pg2.connect(pg_url, options="-c statement_timeout=10000")
        try:
            with conn.cursor() as cur:
                cur.execute("UPDATE hermes.news_sources SET active=%s WHERE id=%s", (active, source_id))
            conn.commit()
        finally:
            conn.close()

    def _ns_delete(pg_url: str, source_id: int) -> None:
        conn = _ns_pg2.connect(pg_url, options="-c statement_timeout=10000")
        try:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM hermes.news_sources WHERE id=%s", (source_id,))
            conn.commit()
        finally:
            conn.close()

    _ns_pg_url = _os.environ.get("PGURL", "")
    _ns_hermes_url = _os.environ.get("HERMES_URL", "http://localhost:8000")

    st.subheader(_t("tab_news"))

    # ── Header row: Run Now + Backfill All + Refresh ──────────────────────────
    _ns_col1, _ns_col2, _ns_col3, _ns_col4, _ns_col5 = st.columns([1, 1.5, 1, 1, 1])
    with _ns_col1:
        _ns_run = st.button("▶ Run Now", key="ns_run_now", type="primary")
    with _ns_col2:
        _ns_backfill_all = st.button("⏮ Backfill All (2025-01-01)", key="ns_backfill_all")
    with _ns_col3:
        _ns_daily_report = st.button("📄 Daily Report", key="ns_daily_report", help="Generate today's market PDF report and send via Feishu")
    with _ns_col4:
        _ns_monthly_report = st.button("📊 Monthly Report", key="ns_monthly_report", help="Generate last month's market PDF report and send via Feishu")
    with _ns_col5:
        _ns_refresh = st.button("↻ Refresh", key="ns_refresh")

    if _ns_run:
        try:
            import requests as _ns_req
            import urllib3 as _ns_urllib3
            _ns_urllib3.disable_warnings(_ns_urllib3.exceptions.InsecureRequestWarning)
            _ns_resp = _ns_req.post(
                f"{_ns_hermes_url}/hermes/news-screener/run",
                timeout=15,
                verify=False,  # internal ALB call; cert not verifiable from container
            )
            if _ns_resp.ok:
                st.success("✅ Screener started — Feishu digest will arrive at completion.")
            else:
                st.error(f"Hermes returned {_ns_resp.status_code}: {_ns_resp.text[:200]}")
        except Exception as _ns_exc:
            st.error(f"Could not reach Hermes: {_ns_exc}")

    if _ns_backfill_all:
        try:
            import requests as _ns_req_bf2
            import urllib3 as _ns_urllib3_bf2
            _ns_urllib3_bf2.disable_warnings(_ns_urllib3_bf2.exceptions.InsecureRequestWarning)
            _ns_bf2_resp = _ns_req_bf2.post(
                f"{_ns_hermes_url}/hermes/news-screener/backfill",
                json={"start_date": "2025-01-01"},
                timeout=15,
                verify=False,
            )
            if _ns_bf2_resp.ok:
                _ns_bf2_data = _ns_bf2_resp.json()
                st.success(
                    f"⏳ Backfill started for {_ns_bf2_data.get('sources', '?')} sources from 2025-01-01 — "
                    f"Feishu notifications will arrive per-source as they complete."
                )
            else:
                st.error(f"Hermes returned {_ns_bf2_resp.status_code}: {_ns_bf2_resp.text[:200]}")
        except Exception as _ns_bf2_exc:
            st.error(f"Could not reach Hermes: {_ns_bf2_exc}")

    if _ns_daily_report:
        try:
            import requests as _ns_req_dr
            import urllib3 as _ns_urllib3_dr
            _ns_urllib3_dr.disable_warnings(_ns_urllib3_dr.exceptions.InsecureRequestWarning)
            _ns_dr_resp = _ns_req_dr.post(
                f"{_ns_hermes_url}/hermes/reports/daily",
                timeout=15, verify=False,
            )
            if _ns_dr_resp.ok:
                st.success("⏳ Daily market report generating — PDF will arrive on Feishu shortly (1-2 min).")
            else:
                st.error(f"Hermes returned {_ns_dr_resp.status_code}: {_ns_dr_resp.text[:200]}")
        except Exception as _ns_dr_exc:
            st.error(f"Could not reach Hermes: {_ns_dr_exc}")

    if _ns_monthly_report:
        try:
            import requests as _ns_req_mr
            import urllib3 as _ns_urllib3_mr
            _ns_urllib3_mr.disable_warnings(_ns_urllib3_mr.exceptions.InsecureRequestWarning)
            _ns_mr_resp = _ns_req_mr.post(
                f"{_ns_hermes_url}/hermes/reports/monthly",
                json={},
                timeout=15, verify=False,
            )
            if _ns_mr_resp.ok:
                st.success("⏳ Monthly market report generating — PDF will arrive on Feishu shortly (2-3 min).")
            else:
                st.error(f"Hermes returned {_ns_mr_resp.status_code}: {_ns_mr_resp.text[:200]}")
        except Exception as _ns_mr_exc:
            st.error(f"Could not reach Hermes: {_ns_mr_exc}")

    # Last-run timestamp
    if _ns_pg_url:
        try:
            import psycopg2 as _ns_pg
            _ns_conn = _ns_pg.connect(_ns_pg_url, options="-c statement_timeout=5000")
            with _ns_conn.cursor() as _ns_cur:
                _ns_cur.execute(
                    "SELECT MAX(last_scraped_at) FROM hermes.news_sources"
                )
                _ns_last = _ns_cur.fetchone()
            _ns_conn.close()
            if _ns_last and _ns_last[0]:
                _ns_col3.caption(f"Last run: {_ns_last[0].strftime('%Y-%m-%d %H:%M UTC')}")
        except Exception:
            pass

    # ── Sources table ─────────────────────────────────────────────────────────
    if _ns_pg_url:
        try:
            _ns_init_db(_ns_pg_url)
            _ns_sources = _ns_get_sources(_ns_pg_url, active_only=False)
        except Exception as _ns_exc:
            st.error(f"Could not load sources: {_ns_exc}")
            _ns_sources = []

        if _ns_sources:
            st.markdown("**Configured sources**")
            _NS_REGIONS = ["全国", "华北", "华东", "华南", "西北", "西南", "东北"]
            _NS_CATS = ["other", "policy", "market_rules", "market_analytics", "technology", "industry_news"]
            if "ns_editing" not in st.session_state:
                st.session_state["ns_editing"] = None
            for _ns_src in _ns_sources:
                _ns_id = _ns_src["id"]
                _ns_sc1, _ns_sc2, _ns_sc3, _ns_sc4, _ns_sc5, _ns_sc6, _ns_sc7 = st.columns(
                    [3, 1.5, 1.5, 2, 1, 0.6, 0.6]
                )
                _ns_sc1.write(_ns_src["name"])
                _ns_sc2.write(_ns_src.get("source_type", "wechat"))
                _ns_sc3.write(_ns_src.get("region_bucket") or "—")
                _ns_sc4.write(_ns_src.get("category_hint") or "other")
                _ns_active_val = bool(_ns_src.get("active", True))
                _ns_new_active = _ns_sc5.checkbox(
                    "Active", value=_ns_active_val,
                    key=f"ns_active_{_ns_id}", label_visibility="collapsed",
                )
                if _ns_new_active != _ns_active_val:
                    _ns_set_active(_ns_pg_url, _ns_id, _ns_new_active)
                    st.rerun(scope="fragment")
                if _ns_sc6.button("✏️", key=f"ns_edit_btn_{_ns_id}", help="Edit"):
                    st.session_state["ns_editing"] = (
                        None if st.session_state["ns_editing"] == _ns_id else _ns_id
                    )
                    st.rerun(scope="fragment")
                if _ns_sc7.button("🗑", key=f"ns_del_{_ns_id}"):
                    _ns_delete(_ns_pg_url, _ns_id)
                    if st.session_state.get("ns_editing") == _ns_id:
                        st.session_state["ns_editing"] = None
                    st.rerun(scope="fragment")

                # Inline edit form — shown only for the row being edited
                if st.session_state.get("ns_editing") == _ns_id:
                    with st.container():
                        _ec1, _ec2, _ec3, _ec4 = st.columns([3, 2, 2, 1])
                        _ns_e_name = _ec1.text_input(
                            "Name", value=_ns_src["name"], key=f"ns_ename_{_ns_id}"
                        )
                        _ns_e_region = _ec2.selectbox(
                            "Region", _NS_REGIONS,
                            index=_NS_REGIONS.index(_ns_src.get("region_bucket") or "全国"),
                            key=f"ns_eregion_{_ns_id}",
                        )
                        _ns_e_cat = _ec3.selectbox(
                            "Category", _NS_CATS,
                            index=_NS_CATS.index(_ns_src.get("category_hint") or "other"),
                            key=f"ns_ecat_{_ns_id}",
                        )
                        _ec4.write("")  # vertical spacer
                        _ec4.write("")
                        if _ec4.button("Save", key=f"ns_esave_{_ns_id}", type="primary"):
                            _ns_econn = _ns_pg2.connect(_ns_pg_url, options="-c statement_timeout=10000")
                            try:
                                with _ns_econn.cursor() as _ns_ecur:
                                    _ns_ecur.execute(
                                        "UPDATE hermes.news_sources SET name=%s, region_bucket=%s, category_hint=%s WHERE id=%s",
                                        (_ns_e_name, _ns_e_region, _ns_e_cat, _ns_id),
                                    )
                                _ns_econn.commit()
                            finally:
                                _ns_econn.close()
                            st.session_state["ns_editing"] = None
                            st.rerun(scope="fragment")
        else:
            st.info("No news sources configured yet. Add one below.")
    else:
        st.warning("PGURL not set — cannot manage sources.")
        _ns_sources = []

    st.divider()

    # ── Add Source ────────────────────────────────────────────────────────────
    with st.expander("➕ Add Source", expanded=len(_ns_sources) == 0):
        _ns_name = st.text_input("Display name", key="ns_name")
        _ns_type = st.selectbox(
            "Type",
            ["wechat", "web", "rss"],
            key="ns_type",
        )
        _ns_url_label = (
            "WeChat article URL (any mp.weixin.qq.com/s/… to auto-extract __biz)"
            if _ns_type == "wechat"
            else "Listing page URL"
        )
        _ns_url = st.text_input(_ns_url_label, key="ns_url")
        _ns_region = st.selectbox(
            "Region hint",
            ["全国", "华北", "华东", "华南", "西北", "西南", "东北"],
            key="ns_region",
        )
        _ns_cat = st.selectbox(
            "Category hint",
            ["other", "policy", "market_rules", "market_analytics", "technology", "industry_news"],
            key="ns_cat",
        )
        if st.button("Add Source", key="ns_add", type="primary"):
            if not _ns_name or not _ns_url:
                st.error("Name and URL are required.")
            elif not _ns_pg_url:
                st.error("PGURL not configured.")
            else:
                try:
                    _ns_new = _ns_add_source(
                        _ns_pg_url,
                        name=_ns_name,
                        url=_ns_url,
                        source_type=_ns_type,
                        region_bucket=_ns_region,
                        category_hint=_ns_cat,
                    )
                    if _ns_type == "wechat" and _ns_new.get("biz_id"):
                        st.success(
                            f"✅ Added **{_ns_name}** (biz_id: `{_ns_new['biz_id']}` auto-extracted)"
                        )
                    else:
                        st.success(f"✅ Added **{_ns_name}**")
                    # Auto-trigger backfill for new source from 2025-01-01
                    try:
                        import requests as _ns_req_bf
                        import urllib3 as _ns_urllib3_bf
                        _ns_urllib3_bf.disable_warnings(_ns_urllib3_bf.exceptions.InsecureRequestWarning)
                        _ns_bf_resp = _ns_req_bf.post(
                            f"{_ns_hermes_url}/hermes/news-screener/backfill",
                            json={"source_id": _ns_new["id"], "start_date": "2025-01-01"},
                            timeout=10,
                            verify=False,
                        )
                        if _ns_bf_resp.ok:
                            st.info("⏳ Backfilling articles since 2025-01-01 in background — you'll get a Feishu notification when done.")
                        else:
                            st.warning(f"Backfill request failed: {_ns_bf_resp.status_code}")
                    except Exception as _ns_bf_exc:
                        st.warning(f"Could not trigger backfill: {_ns_bf_exc}")
                    st.rerun()  # full-page rerun to refresh source list after add
                except Exception as _ns_exc:
                    st.error(f"Failed to add source: {_ns_exc}")

    st.divider()

    # ── Recent Ingested Articles ───────────────────────────────────────────────
    with st.expander("📋 Recent Ingested Articles", expanded=False):
        if _ns_pg_url:
            try:
                _ns_art_conn = _ns_pg2.connect(_ns_pg_url, options="-c statement_timeout=10000")
                with _ns_art_conn.cursor() as _ns_art_cur:
                    _ns_art_cur.execute("""
                        SELECT title, source_name, relevance_score, ai_summary, published_at,
                               ingest_status
                        FROM staging.spot_knowledge_docs
                        WHERE source_name IS NOT NULL
                        ORDER BY created_at DESC
                        LIMIT 40
                    """)
                    _ns_art_cols = [d[0] for d in _ns_art_cur.description]
                    _ns_articles = [dict(zip(_ns_art_cols, r)) for r in _ns_art_cur.fetchall()]
                _ns_art_conn.close()

                if _ns_articles:
                    import pandas as _ns_pd
                    _ns_art_df = _ns_pd.DataFrame(_ns_articles)
                    _ns_art_df["title"] = _ns_art_df["title"].str[:80]
                    _ns_art_df["ai_summary"] = _ns_art_df["ai_summary"].fillna("").str[:100]
                    _ns_art_df["published_at"] = _ns_art_df["published_at"].apply(
                        lambda x: x.strftime("%Y-%m-%d") if x else "—"
                    )
                    _ns_art_df["relevance_score"] = _ns_art_df["relevance_score"].fillna("—")
                    _ns_art_df.rename(columns={
                        "title": "Title", "source_name": "Source",
                        "relevance_score": "Score", "ai_summary": "Summary",
                        "published_at": "Published", "ingest_status": "Status",
                    }, inplace=True)
                    st.dataframe(
                        _ns_art_df[["Published", "Source", "Score", "Title", "Summary", "Status"]],
                        use_container_width=True, hide_index=True,
                    )
                else:
                    st.info("No articles ingested yet.")
            except Exception as _ns_art_exc:
                if "column" in str(_ns_art_exc).lower() and "does not exist" in str(_ns_art_exc).lower():
                    st.info("Extra metadata columns not yet present — run the screener once to create them.")
                else:
                    st.error(f"Could not load articles: {_ns_art_exc}")
        else:
            st.warning("PGURL not set.")

    st.divider()

    # ── Suggested Sources ─────────────────────────────────────────────────────
    _NS_SUGGESTED = [
        {"name": "国家能源局",      "url": "https://mp.weixin.qq.com/mp/profile_ext?action=home&__biz=MzA5ODM3NjYxMQ==&scene=124", "region": "全国",  "cat": "policy"},
        {"name": "中电联发布",      "url": "https://weixin.sogou.com/weixin?type=2&query=中电联发布",  "region": "全国",  "cat": "industry_news"},
        {"name": "中国储能网",      "url": "https://weixin.sogou.com/weixin?type=2&query=中国储能网",  "region": "全国",  "cat": "technology"},
        {"name": "北极星储能网",    "url": "https://weixin.sogou.com/weixin?type=2&query=北极星储能网", "region": "全国",  "cat": "industry_news"},
        {"name": "能源新媒",        "url": "https://weixin.sogou.com/weixin?type=2&query=能源新媒",    "region": "全国",  "cat": "industry_news"},
        {"name": "国网能源研究院",  "url": "https://weixin.sogou.com/weixin?type=2&query=国网能源研究院", "region": "全国", "cat": "market_analytics"},
        {"name": "华北电力交易中心","url": "https://weixin.sogou.com/weixin?type=2&query=华北电力交易中心", "region": "华北", "cat": "market_rules"},
        {"name": "电力决策与舆情研究","url": "https://weixin.sogou.com/weixin?type=2&query=电力决策与舆情研究", "region": "全国", "cat": "market_analytics"},
    ]
    # Filter out sources already in the configured list
    _ns_existing_names = {s["name"] for s in (_ns_sources if _ns_pg_url else [])}
    _ns_new_suggestions = [s for s in _NS_SUGGESTED if s["name"] not in _ns_existing_names]

    with st.expander(f"💡 Suggested Sources ({len(_ns_new_suggestions)} not yet added)", expanded=False):
        if not _ns_pg_url:
            st.warning("PGURL not set — cannot add sources.")
        elif not _ns_new_suggestions:
            st.success("All suggested sources are already configured.")
        else:
            st.caption("Click Quick Add to add a source and trigger backfill since 2025-01-01.")
            for _ns_sug in _ns_new_suggestions:
                _ns_sug_c1, _ns_sug_c2, _ns_sug_c3, _ns_sug_c4 = st.columns([2.5, 1.5, 1.5, 1])
                _ns_sug_c1.write(_ns_sug["name"])
                _ns_sug_c2.write(_ns_sug["region"])
                _ns_sug_c3.write(_ns_sug["cat"])
                if _ns_sug_c4.button("Quick Add", key=f"ns_qadd_{_ns_sug['name']}"):
                    try:
                        _ns_sug_added = _ns_add_source(
                            _ns_pg_url,
                            name=_ns_sug["name"],
                            url=_ns_sug["url"],
                            source_type="wechat",
                            region_bucket=_ns_sug["region"],
                            category_hint=_ns_sug["cat"],
                        )
                        st.success(f"✅ Added {_ns_sug['name']}")
                        try:
                            import requests as _ns_req_sug
                            import urllib3 as _ns_urllib3_sug
                            _ns_urllib3_sug.disable_warnings(_ns_urllib3_sug.exceptions.InsecureRequestWarning)
                            _ns_req_sug.post(
                                f"{_ns_hermes_url}/hermes/news-screener/backfill",
                                json={"source_id": _ns_sug_added["id"], "start_date": "2025-01-01"},
                                timeout=10, verify=False,
                            )
                            st.info("⏳ Backfill started — Feishu notification when done.")
                        except Exception:
                            pass
                        st.rerun(scope="fragment")
                    except Exception as _ns_sug_exc:
                        st.error(f"Failed: {_ns_sug_exc}")

with tab_news:
    _render_news_sources_tab()

# ── Library ───────────────────────────────────────────────────────────────────
with tab_library:
    from services.common.report_library_ui import render_library_tab
    render_library_tab("spot", "China Spot Market", "spot")

# ── 机制竞价 ──────────────────────────────────────────────────────────────────
with tab_jizhi:
    import psycopg2 as _jz_pg
    import pandas as _jz_pd
    import plotly.graph_objects as _jz_go
    from datetime import date as _jz_date

    _jz_pg_url = (
        os.environ.get("PGURL")
        or os.environ.get("DATABASE_URL")
        or "postgresql://postgres:root@127.0.0.1:5433/marketdata"
    )

    @st.cache_data(ttl=300, show_spinner=False)
    def _load_jizhi_bids(_pg: str) -> _jz_pd.DataFrame:
        try:
            conn = _jz_pg.connect(_pg)
            df = _jz_pd.read_sql("""
                SELECT id, province, year, batch, tech_type,
                       price_floor, price_cap, mechanism_type, mechanism_value,
                       supply_demand_ratio, cleared_price, cleared_volume_gwh,
                       bid_date, verified, notes, source_doc_id
                FROM staging.jizhi_bids
                ORDER BY year DESC, province, batch, tech_type
            """, conn)
            conn.close()
            return df
        except Exception:
            return _jz_pd.DataFrame()

    @st.cache_data(ttl=300, show_spinner=False)
    def _load_jizhi_upcoming(_pg: str) -> _jz_pd.DataFrame:
        try:
            conn = _jz_pg.connect(_pg)
            df = _jz_pd.read_sql("""
                SELECT id, province, year, batch, tech_type,
                       price_floor, price_cap, target_volume_gwh,
                       supply_demand_ratio, bid_open_date, bid_close_date,
                       source_url, announcement_date, verified, notes, created_at
                FROM staging.jizhi_upcoming
                ORDER BY bid_open_date ASC NULLS LAST
            """, conn)
            conn.close()
            return df
        except Exception:
            return _jz_pd.DataFrame()

    @st.cache_data(ttl=300, show_spinner=False)
    def _load_jizhi_winners(_bid_id: int, _pg: str) -> _jz_pd.DataFrame:
        try:
            conn = _jz_pg.connect(_pg)
            df = _jz_pd.read_sql("""
                SELECT project_name, operator, capacity_mw, cleared_price, tech_type
                FROM staging.jizhi_bid_winners
                WHERE bid_id = %s
                ORDER BY capacity_mw DESC NULLS LAST
            """, conn, params=[_bid_id])
            conn.close()
            return df
        except Exception:
            return _jz_pd.DataFrame()

    _jz_tab_results, _jz_tab_upcoming, _jz_tab_upload = st.tabs(
        ["📊 历史结果", "📅 即将竞价", "📂 上传 & 录入"]
    )

    # ── Sub-tab 1: 历史结果 ────────────────────────────────────────────────────
    with _jz_tab_results:
        _jz_bids_df = _load_jizhi_bids(_jz_pg_url)

        if _jz_bids_df.empty:
            st.info("暂无历史竞价数据。请在「上传 & 录入」标签中上传竞价文件。")
        else:
            # Filter row
            _jz_col_prov, _jz_col_year, _jz_col_tech = st.columns([2, 2, 2])
            with _jz_col_prov:
                _jz_provs = st.multiselect(
                    "省份", sorted(_jz_bids_df["province"].unique()), key="jz_prov_filter"
                )
            with _jz_col_year:
                _jz_years = st.multiselect(
                    "年份", sorted(_jz_bids_df["year"].unique(), reverse=True), key="jz_year_filter"
                )
            with _jz_col_tech:
                _jz_techs = st.multiselect(
                    "技术类型", sorted(_jz_bids_df["tech_type"].unique()), key="jz_tech_filter"
                )

            _jz_filtered = _jz_bids_df.copy()
            if _jz_provs:
                _jz_filtered = _jz_filtered[_jz_filtered["province"].isin(_jz_provs)]
            if _jz_years:
                _jz_filtered = _jz_filtered[_jz_filtered["year"].isin(_jz_years)]
            if _jz_techs:
                _jz_filtered = _jz_filtered[_jz_filtered["tech_type"].isin(_jz_techs)]

            # Add ⚠️ badge for unverified rows
            _jz_display = _jz_filtered.copy()
            _jz_display["verified"] = _jz_display["verified"].apply(
                lambda v: "✅" if v else "⚠️"
            )
            _jz_display = _jz_display.rename(columns={
                "province": "省份", "year": "年份", "batch": "批次",
                "tech_type": "技术", "price_floor": "价格下限", "price_cap": "价格上限",
                "mechanism_type": "机制类型", "mechanism_value": "机制量",
                "supply_demand_ratio": "供需比", "cleared_price": "中标价格",
                "cleared_volume_gwh": "中标量(GWh)", "bid_date": "竞价日期", "verified": "验证",
            })
            _jz_show_cols = [
                "省份", "年份", "批次", "技术", "价格下限", "价格上限",
                "机制类型", "机制量", "供需比", "中标价格", "中标量(GWh)", "竞价日期", "验证"
            ]
            st.dataframe(
                _jz_display[[c for c in _jz_show_cols if c in _jz_display.columns]],
                use_container_width=True, hide_index=True,
            )

            # Winner list selector
            if not _jz_filtered.empty:
                with st.expander("🏆 查看中标清单 (选择竞价记录)"):
                    _jz_opts = {
                        f"{r['province']} {r['year']} {r['batch']} {r['tech_type']}": int(r["id"])
                        for _, r in _jz_filtered.iterrows()
                    }
                    _jz_sel = st.selectbox("选择竞价记录", list(_jz_opts.keys()), key="jz_winner_sel")
                    if _jz_sel:
                        _jz_bid_id = _jz_opts[_jz_sel]
                        _jz_winners_df = _load_jizhi_winners(_jz_bid_id, _jz_pg_url)
                        if _jz_winners_df.empty:
                            st.info("该竞价暂无中标清单数据。")
                        else:
                            st.dataframe(_jz_winners_df, use_container_width=True, hide_index=True)

            # Charts
            if len(_jz_filtered) >= 2:
                st.divider()
                _jz_chart_col1, _jz_chart_col2 = st.columns(2)
                with _jz_chart_col1:
                    _jz_avg = (
                        _jz_filtered.dropna(subset=["cleared_price"])
                        .groupby(["province", "tech_type"])["cleared_price"]
                        .mean()
                        .reset_index()
                    )
                    if not _jz_avg.empty:
                        _jz_fig_bar = _jz_go.Figure()
                        for _tech in _jz_avg["tech_type"].unique():
                            _sub = _jz_avg[_jz_avg["tech_type"] == _tech]
                            _jz_fig_bar.add_trace(_jz_go.Bar(
                                x=_sub["province"], y=_sub["cleared_price"], name=_tech
                            ))
                        _jz_fig_bar.update_layout(
                            title="各省平均中标价格 (元/kWh)",
                            barmode="group", height=320,
                            margin=dict(l=0, r=0, t=36, b=0),
                            legend=dict(orientation="h", y=-0.2),
                        )
                        st.plotly_chart(_jz_fig_bar, use_container_width=True)

                with _jz_chart_col2:
                    _jz_sd = _jz_filtered.dropna(subset=["supply_demand_ratio", "year"])
                    if not _jz_sd.empty:
                        _jz_fig_sd = _jz_go.Figure()
                        for _prov in _jz_sd["province"].unique():
                            _sub = _jz_sd[_jz_sd["province"] == _prov].sort_values("year")
                            _jz_fig_sd.add_trace(_jz_go.Scatter(
                                x=_sub["year"], y=_sub["supply_demand_ratio"],
                                name=_prov, mode="lines+markers"
                            ))
                        _jz_fig_sd.update_layout(
                            title="供需比趋势",
                            height=320, margin=dict(l=0, r=0, t=36, b=0),
                            legend=dict(orientation="h", y=-0.2),
                        )
                        st.plotly_chart(_jz_fig_sd, use_container_width=True)

    # ── Sub-tab 2: 即将竞价 ────────────────────────────────────────────────────
    with _jz_tab_upcoming:
        _jz_up_df = _load_jizhi_upcoming(_jz_pg_url)

        # Last scan timestamp
        _jz_last_scan = (
            str(_jz_up_df["created_at"].max())[:16]
            if not _jz_up_df.empty and "created_at" in _jz_up_df.columns
            else "—"
        )
        st.caption(f"数据最后更新：{_jz_last_scan}  ·  每晚 18:07 (北京时间) 自动扫描")

        if _jz_up_df.empty:
            st.info("暂无即将竞价信息。数据将由 Hermes 每晚自动扫描更新。")
        else:
            _jz_up_col_prov, _jz_up_col_tech = st.columns(2)
            with _jz_up_col_prov:
                _jz_up_provs = st.multiselect(
                    "省份", sorted(_jz_up_df["province"].unique()), key="jz_up_prov"
                )
            with _jz_up_col_tech:
                _jz_up_techs = st.multiselect(
                    "技术类型", sorted(_jz_up_df["tech_type"].unique()), key="jz_up_tech"
                )

            _jz_up_filtered = _jz_up_df.copy()
            if _jz_up_provs:
                _jz_up_filtered = _jz_up_filtered[
                    _jz_up_filtered["province"].isin(_jz_up_provs)
                ]
            if _jz_up_techs:
                _jz_up_filtered = _jz_up_filtered[
                    _jz_up_filtered["tech_type"].isin(_jz_up_techs)
                ]

            # Compute days-until column
            _today = _jz_date.today()
            def _days_until(d):
                if _jz_pd.isna(d):
                    return None
                return (d.date() if hasattr(d, "date") else d) - _today

            _jz_up_filtered = _jz_up_filtered.copy()
            _jz_up_filtered["距今"] = _jz_up_filtered["bid_open_date"].apply(
                lambda d: f"{_days_until(d).days}天" if _days_until(d) is not None else "—"
            )

            _jz_up_display = _jz_up_filtered.rename(columns={
                "province": "省份", "year": "年份", "batch": "批次",
                "tech_type": "技术", "price_floor": "价格下限", "price_cap": "价格上限",
                "target_volume_gwh": "目标量(GWh)", "supply_demand_ratio": "供需比",
                "bid_open_date": "开始日期", "bid_close_date": "截止日期",
                "verified": "已验证",
            })
            _jz_up_show = [
                "省份", "年份", "批次", "技术", "价格下限", "价格上限",
                "目标量(GWh)", "供需比", "开始日期", "截止日期", "距今", "已验证"
            ]
            st.dataframe(
                _jz_up_display[[c for c in _jz_up_show if c in _jz_up_display.columns]],
                use_container_width=True, hide_index=True,
            )

    # ── Sub-tab 3: 上传 & 录入 ─────────────────────────────────────────────────
    with _jz_tab_upload:
        st.markdown("上传竞价结果文件（PPT/PDF/Excel）→ AI 自动提取结构化数据 → 预览确认 → 保存")

        _jz_up_file = st.file_uploader(
            "选择文件",
            type=["pdf", "pptx", "ppt", "xlsx", "xls", "docx", "doc", "txt", "jpg", "jpeg", "png"],
            key="jz_upload_file",
        )
        _jz_up_url = st.text_input("或输入 URL", placeholder="https://...", key="jz_upload_url")

        # Use session state to cache extraction results so clicking "保存" doesn't
        # re-trigger a slow Bedrock API call (file still in uploader on each rerun).
        _jz_ss_key  = "jz_extracted_records"
        _jz_ss_id   = "jz_extracted_doc_id"
        _jz_ss_file = "jz_extracted_filename"

        _jz_trigger_extract = False
        if _jz_up_file is not None:
            # Only re-extract when a NEW file is uploaded (filename changed)
            if st.session_state.get(_jz_ss_file) != _jz_up_file.name:
                _jz_trigger_extract = True
        elif _jz_up_url and st.button("获取 URL", key="jz_fetch_url"):
            _jz_trigger_extract = True

        if _jz_trigger_extract:
            _jz_api_key = os.environ.get("ANTHROPIC_API_KEY", "")
            with st.spinner("正在提取竞价数据…"):
                try:
                    from services.knowledge_pool.knowledge_docs import (
                        register_and_ingest, register_url, _extract_pages,
                    )
                    from services.knowledge_pool.jizhi_extractor import (
                        extract_bids, save_bids, ensure_tables,
                        _extract_pptx_text,
                    )
                    ensure_tables(_jz_pg_url)

                    if _jz_up_file is not None:
                        _jz_fbytes = _jz_up_file.read()
                        _jz_fname  = _jz_up_file.name
                        _jz_doc_id, _, _ = register_and_ingest(
                            file_bytes=_jz_fbytes, filename=_jz_fname,
                            category_override="policy_doc", app="shared",
                            api_key=_jz_api_key,
                        )
                        if _jz_fname.lower().endswith((".pptx", ".ppt")):
                            _jz_full_text = _extract_pptx_text(_jz_fbytes)
                        else:
                            _jz_pages = _extract_pages(
                                _jz_fbytes, _jz_fname, _jz_api_key
                            )
                            _jz_full_text = "\n\n".join(t for _, t in _jz_pages)
                    else:
                        _jz_doc_id, _, _ = register_url(_jz_up_url, api_key=_jz_api_key)
                        _jz_pages = _extract_pages(
                            b"", f"url_{_jz_up_url[-20:]}.txt", _jz_api_key
                        )
                        _jz_full_text = "\n\n".join(t for _, t in _jz_pages)

                    _jz_extracted = extract_bids(_jz_full_text, api_key="")
                    # Cache results in session state
                    st.session_state[_jz_ss_key]  = _jz_extracted
                    st.session_state[_jz_ss_id]   = _jz_doc_id
                    st.session_state[_jz_ss_file]  = getattr(_jz_up_file, "name", _jz_up_url)
                except Exception as _e:
                    st.error(f"提取失败：{_e}")
                    st.session_state[_jz_ss_key]  = []
                    st.session_state[_jz_ss_id]   = None
                    st.session_state[_jz_ss_file]  = None

        # Show preview from session state (persists across reruns without re-extraction)
        _jz_extracted = st.session_state.get(_jz_ss_key, [])
        _jz_doc_id    = st.session_state.get(_jz_ss_id)

        if _jz_up_file is not None or _jz_up_url:
            if _jz_extracted:
                st.success(f"提取到 {len(_jz_extracted)} 条竞价记录，请确认后保存：")
                _jz_preview_df = _jz_pd.DataFrame(_jz_extracted)
                _jz_edited = st.data_editor(
                    _jz_preview_df, use_container_width=True,
                    num_rows="dynamic", key="jz_preview_editor",
                )
                if st.button("💾 保存到数据库", key="jz_save_btn"):
                    from services.knowledge_pool.jizhi_extractor import save_bids, ensure_tables
                    ensure_tables(_jz_pg_url)
                    _jz_n = save_bids(
                        _jz_edited.to_dict("records"),
                        source_doc_id=_jz_doc_id,
                        pg_url=_jz_pg_url,
                    )
                    st.success(f"已保存 {_jz_n} 条记录（已存在且已验证的记录不会被覆盖）")
                    # Clear both the extraction cache and the query cache, then refresh
                    st.session_state[_jz_ss_key]  = []
                    st.session_state[_jz_ss_file]  = None
                    _load_jizhi_bids.clear()
                    st.rerun()
            elif not _jz_trigger_extract and st.session_state.get(_jz_ss_file) is None:
                pass  # No file processed yet, nothing to show
            else:
                st.warning("未能从文件中提取结构化竞价数据。文件已存入知识库。")

        st.divider()
        with st.expander("✏️ 手动录入单条记录"):
            _jz_m_col1, _jz_m_col2 = st.columns(2)
            with _jz_m_col1:
                _jz_m_prov  = st.text_input("省份", placeholder="广东", key="jz_m_prov")
                _jz_m_year  = st.number_input("年份", min_value=2020, max_value=2035,
                                               value=2025, step=1, key="jz_m_year")
                _jz_m_batch = st.selectbox(
                    "批次", ["存量", "增量_2025-12", "增量_2026-12", "增量_2027-12"],
                    key="jz_m_batch"
                )
                _jz_m_tech  = st.selectbox(
                    "技术类型", ["陆风", "海风", "光伏", "水电"], key="jz_m_tech"
                )
                _jz_m_pfloor = st.number_input("价格下限 (元/kWh)", min_value=0.0,
                                                step=0.001, format="%.4f", key="jz_m_pfloor")
                _jz_m_pcap   = st.number_input("价格上限 (元/kWh)", min_value=0.0,
                                                step=0.001, format="%.4f", key="jz_m_pcap")
            with _jz_m_col2:
                _jz_m_mtype  = st.selectbox("机制类型", ["小时数", "电量", "比例"], key="jz_m_mtype")
                _jz_m_mval   = st.number_input("机制量 (小时/GWh/%)", min_value=0.0,
                                                step=1.0, key="jz_m_mval")
                _jz_m_sdr    = st.number_input("供需比", min_value=0.0, step=0.01,
                                                format="%.2f", key="jz_m_sdr")
                _jz_m_cprice = st.number_input("中标价格 (元/kWh)", min_value=0.0,
                                                step=0.001, format="%.4f", key="jz_m_cprice")
                _jz_m_cvol   = st.number_input("中标量 (GWh)", min_value=0.0,
                                                step=1.0, key="jz_m_cvol")
                _jz_m_date   = st.date_input("竞价日期", key="jz_m_date")
                _jz_m_notes  = st.text_input("备注", key="jz_m_notes")

            if st.button("保存手动记录", key="jz_m_save"):
                if not _jz_m_prov:
                    st.error("省份不能为空。")
                else:
                    from services.knowledge_pool.jizhi_extractor import save_bids, ensure_tables
                    ensure_tables(_jz_pg_url)
                    _jz_manual_rec = [{
                        "province": _jz_m_prov, "year": int(_jz_m_year),
                        "batch": _jz_m_batch, "tech_type": _jz_m_tech,
                        "price_floor": _jz_m_pfloor or None, "price_cap": _jz_m_pcap or None,
                        "mechanism_type": _jz_m_mtype, "mechanism_value": _jz_m_mval or None,
                        "supply_demand_ratio": _jz_m_sdr or None,
                        "cleared_price": _jz_m_cprice or None,
                        "cleared_volume_gwh": _jz_m_cvol or None,
                        "bid_date": str(_jz_m_date), "notes": _jz_m_notes or None,
                    }]
                    _jz_n2 = save_bids(_jz_manual_rec, source_doc_id=None, pg_url=_jz_pg_url)
                    if _jz_n2:
                        st.success("已保存。（数据标记为未验证）")
                        _load_jizhi_bids.clear()
                    else:
                        st.info("记录已存在且已验证，未覆盖。")

# ── Tab: 供需结构 (Supply Structure) ─────────────────────────────────────────
with tab_supply:
    import os as _os_sup
    _sup_pg = _os_sup.environ.get("PGURL", "")

    @st.cache_data(ttl=300, show_spinner=False)
    def _load_supply_data(_pg_key: str):
        import psycopg2 as _psycopg2_sup, pandas as _pd_sup
        try:
            _conn_sup = _psycopg2_sup.connect(_pg_key)
            with _conn_sup.cursor() as _cur_sup:
                # ── Consolidated capacity from 3 sources ──────────────────────────
                # Priority: province_installed_monthly > exchange_excel_metrics > province_fundamentals
                # UNION of provinces so no source is dropped (fixes missing 广东 etc.)
                _cur_sup.execute("""
                    WITH cap_monthly AS (
                        SELECT DISTINCT ON (province)
                            province, year_month AS cap_month,
                            wind_mw, solar_mw, thermal_mw, hydro_mw, nuclear_mw
                        FROM marketdata.province_installed_monthly
                        ORDER BY province, year_month DESC
                    ),
                    cap_exchange AS (
                        SELECT DISTINCT ON (province)
                            province,
                            wind_capacity_mw    AS wind_mw,
                            solar_capacity_mw   AS solar_mw,
                            thermal_capacity_mw AS thermal_mw,
                            hydro_capacity_mw   AS hydro_mw,
                            nuclear_capacity_mw AS nuclear_mw
                        FROM staging.exchange_excel_metrics
                        ORDER BY province, report_month DESC
                    ),
                    cap_fundamentals AS (
                        SELECT DISTINCT ON (province_cn)
                            province_cn AS province,
                            wind_cap_10kw   * 10 AS wind_mw,
                            solar_cap_10kw  * 10 AS solar_mw,
                            thermal_cap_10kw* 10 AS thermal_mw,
                            hydro_cap_10kw  * 10 AS hydro_mw,
                            nuclear_cap_10kw* 10 AS nuclear_mw
                        FROM marketdata.province_fundamentals
                        ORDER BY province_cn, year DESC
                    ),
                    all_provinces AS (
                        SELECT province FROM cap_monthly
                        UNION SELECT province FROM cap_exchange
                        UNION SELECT province FROM cap_fundamentals
                    )
                    SELECT
                        ap.province,
                        cm.cap_month,
                        COALESCE(NULLIF(cm.wind_mw,    0), NULLIF(ce.wind_mw,    0), cf.wind_mw,    0) AS wind_mw,
                        COALESCE(NULLIF(cm.solar_mw,   0), NULLIF(ce.solar_mw,   0), cf.solar_mw,   0) AS solar_mw,
                        -- province_fundamentals (full NEA stats) before exchange_excel_metrics (spot participants only)
                        COALESCE(NULLIF(cm.thermal_mw, 0), NULLIF(cf.thermal_mw, 0), NULLIF(ce.thermal_mw, 0), 0) AS thermal_mw,
                        COALESCE(NULLIF(cm.hydro_mw,   0), NULLIF(cf.hydro_mw,   0), NULLIF(ce.hydro_mw,   0), 0) AS hydro_mw,
                        COALESCE(NULLIF(cm.nuclear_mw, 0), NULLIF(cf.nuclear_mw, 0), NULLIF(ce.nuclear_mw, 0), 0) AS nuclear_mw
                    FROM all_provinces ap
                    LEFT JOIN cap_monthly      cm ON ap.province = cm.province
                    LEFT JOIN cap_exchange     ce ON ap.province = ce.province
                    LEFT JOIN cap_fundamentals cf ON ap.province = cf.province
                    ORDER BY ap.province
                """)
                _cap_df = _pd_sup.DataFrame(
                    _cur_sup.fetchall(),
                    columns=[d[0] for d in _cur_sup.description],
                )
                # ── Peak load: province_fundamentals primary, spot_fundamentals_hourly fallback ──
                _cur_sup.execute("""
                    WITH peak_fund AS (
                        SELECT DISTINCT ON (province_cn)
                            province_cn AS province,
                            GREATEST(
                                COALESCE(peak_summer_mw, 0),
                                COALESCE(peak_winter_mw, 0)
                            ) AS peak_load_mw
                        FROM marketdata.province_fundamentals
                        ORDER BY province_cn, year DESC
                    ),
                    peak_hourly AS (
                        SELECT province, MAX(load_mw) AS peak_load_mw
                        FROM marketdata.spot_fundamentals_hourly
                        WHERE load_mw > 0
                          AND datetime >= NOW() - INTERVAL '18 months'
                        GROUP BY province
                    )
                    SELECT
                        COALESCE(pf.province, ph.province) AS province,
                        COALESCE(NULLIF(pf.peak_load_mw, 0), ph.peak_load_mw, 0) AS peak_load_mw
                    FROM peak_fund pf
                    FULL OUTER JOIN peak_hourly ph ON pf.province = ph.province
                """)
                _peak_df = _pd_sup.DataFrame(
                    _cur_sup.fetchall(),
                    columns=[d[0] for d in _cur_sup.description],
                )
                # ── Imports/exports from exchange excel metrics (limited provinces) ──
                _cur_sup.execute("""
                    SELECT DISTINCT ON (province)
                        province,
                        COALESCE(incoming_gwh, 0) AS incoming_gwh,
                        COALESCE(outgoing_gwh, 0) AS outgoing_gwh
                    FROM staging.exchange_excel_metrics
                    ORDER BY province, report_month DESC
                """)
                _flow_df = _pd_sup.DataFrame(
                    _cur_sup.fetchall(),
                    columns=[d[0] for d in _cur_sup.description],
                )
            _conn_sup.close()

            # Merge all sources on province
            _df_sup = _cap_df.merge(_peak_df, on='province', how='left')
            _df_sup = _df_sup.merge(_flow_df, on='province', how='left')
            for _c in ['wind_mw', 'solar_mw', 'thermal_mw', 'hydro_mw', 'nuclear_mw',
                       'peak_load_mw', 'incoming_gwh', 'outgoing_gwh']:
                _df_sup[_c] = _pd_sup.to_numeric(_df_sup.get(_c, 0), errors='coerce').fillna(0)

            # Convert monthly GWh → average MW (730 h/month)
            _df_sup['import_mw'] = _df_sup['incoming_gwh'] * 1000 / 730
            _df_sup['export_mw'] = _df_sup['outgoing_gwh'] * 1000 / 730
            # 热电缺口 = peak − (wind + solar + hydro + nuclear) − import + export
            # Negative = clean energy + imports exceed demand (structurally oversupplied)
            _df_sup['net_residual'] = (
                _df_sup['peak_load_mw']
                - _df_sup['wind_mw']
                - _df_sup['solar_mw']
                - _df_sup['hydro_mw']
                - _df_sup['nuclear_mw']
                - _df_sup['import_mw']
                + _df_sup['export_mw']
            )
            return _df_sup.dropna(subset=['province'])
        except Exception:
            return _pd_sup.DataFrame()

    _sup_df = _load_supply_data(_sup_pg)
    # Drop LingFeng metadata rows mistakenly ingested as province names
    _sup_df = _sup_df[~_sup_df['province'].str.contains('运行数据|数据披露|披露', na=False)]

    # Normalise province aliases → canonical names, then deduplicate by keeping the
    # row with the highest thermal_mw (most data) per canonical province.
    _SUP_PROV_NORM = {
        '河北南网': '冀南', '冀南网': '冀南', '河北南部': '冀南',
        '冀北电网': '冀北', '国网冀北': '冀北',
        '内蒙古东': '蒙东', '内蒙古西': '蒙西',
    }
    _DROP_ALIASES = {'中长期', '河北'}   # generic entries subsumed by 冀南/冀北
    _sup_df = _sup_df[~_sup_df['province'].isin(_DROP_ALIASES)]
    _sup_df = _sup_df.copy()
    _sup_df['province'] = _sup_df['province'].map(lambda p: _SUP_PROV_NORM.get(p, p))
    # Where both alias and canonical exist, keep the row with more thermal capacity
    _sup_df = (
        _sup_df.sort_values('thermal_mw', ascending=False)
               .drop_duplicates(subset=['province'], keep='first')
               .reset_index(drop=True)
    )

    if _sup_df.empty:
        st.info("暂无供需数据（需要 marketdata.province_installed_monthly 表）")
    else:
        import plotly.graph_objects as _pgo_sup

        _latest_sup = pd.to_datetime(_sup_df['cap_month'], errors='coerce').max()
        st.caption(
            f"数据截至 {_latest_sup.strftime('%Y-%m') if hasattr(_latest_sup, 'strftime') else _latest_sup}"
            f"  ·  省间受/送出已换算为月均功率（÷730 h）"
        )

        _sup_sorted = _sup_df.sort_values('net_residual', ascending=True).reset_index(drop=True)

        # ── Province waterfall selector ────────────────────────────────────────
        _sup_prov_sel = st.selectbox(
            "选择省份查看瀑布图分解",
            _sup_sorted['province'].tolist(),
            key="sup_prov_sel",
        )
        _sup_row = _sup_df[_sup_df['province'] == _sup_prov_sel].iloc[0]

        _col_wf, _col_th = st.columns(2)

        with _col_wf:
            st.subheader(f"{_sup_prov_sel}  供需余额分解")
            _fig_wf = _pgo_sup.Figure(_pgo_sup.Waterfall(
                orientation="v",
                measure=["absolute", "relative", "relative", "relative", "relative", "relative", "relative", "total"],
                x=["峰值负荷", "↓ 风电", "↓ 光伏", "↓ 水电", "↓ 核电", "↓ 省间受入", "↑ 省间送出", "热电缺口"],
                y=[
                    _sup_row['peak_load_mw'],
                    -_sup_row['wind_mw'],
                    -_sup_row['solar_mw'],
                    -_sup_row['hydro_mw'],
                    -_sup_row['nuclear_mw'],
                    -_sup_row['import_mw'],
                    _sup_row['export_mw'],
                    0,
                ],
                text=[
                    f"{_sup_row['peak_load_mw']:,.0f} MW",
                    f"−{_sup_row['wind_mw']:,.0f} MW",
                    f"−{_sup_row['solar_mw']:,.0f} MW",
                    f"−{_sup_row['hydro_mw']:,.0f} MW",
                    f"−{_sup_row['nuclear_mw']:,.0f} MW",
                    f"−{_sup_row['import_mw']:,.0f} MW",
                    f"+{_sup_row['export_mw']:,.0f} MW",
                    f"{_sup_row['net_residual']:,.0f} MW",
                ],
                textposition="outside",
                connector={"line": {"color": "#aaa", "width": 1, "dash": "dot"}},
                increasing={"marker": {"color": "#E74C3C"}},
                decreasing={"marker": {"color": "#27AE60"}},
                totals={"marker": {"color": "#2980B9" if _sup_row['net_residual'] > 0 else "#27AE60"}},
            ))
            _fig_wf.update_layout(
                yaxis_title="MW",
                showlegend=False,
                height=380,
                margin=dict(t=20, b=20, l=60, r=40),
            )
            st.plotly_chart(_fig_wf, use_container_width=True)

            # All-province net residual overview (horizontal bar)
            st.markdown("**所有省份热电缺口** ─ 负值=清洁能源+受入 > 峰值负荷+送出（结构性过剩）；正值=需要火电补充")
            _fig_net = _pgo_sup.Figure(_pgo_sup.Bar(
                x=_sup_sorted['net_residual'],
                y=_sup_sorted['province'],
                orientation='h',
                marker_color=[
                    '#27AE60' if v < 0 else '#E74C3C'
                    for v in _sup_sorted['net_residual']
                ],
                text=[f"{v:,.0f}" for v in _sup_sorted['net_residual']],
                textposition='outside',
            ))
            _fig_net.add_vline(x=0, line_color='black', line_width=1)
            _fig_net.update_layout(
                xaxis_title="热电缺口 MW  (负=结构性过剩)",
                height=max(380, len(_sup_sorted) * 22),
                margin=dict(t=10, b=20, l=80, r=80),
            )
            st.plotly_chart(_fig_net, use_container_width=True)

        with _col_th:
            st.subheader("火电装机容量（MW）")
            _sup_th_sorted = _sup_df.sort_values('thermal_mw', ascending=True)
            _fig_th = _pgo_sup.Figure(_pgo_sup.Bar(
                x=_sup_th_sorted['thermal_mw'],
                y=_sup_th_sorted['province'],
                orientation='h',
                marker_color='#E67E22',
                text=[f"{v:,.0f}" for v in _sup_th_sorted['thermal_mw']],
                textposition='outside',
            ))
            _fig_th.update_layout(
                xaxis_title="火电装机 MW",
                height=max(380, len(_sup_th_sorted) * 22),
                margin=dict(t=10, b=20, l=80, r=80),
            )
            st.plotly_chart(_fig_th, use_container_width=True)

            # Summary table
            _sup_tbl = _sup_sorted[[
                'province', 'peak_load_mw', 'wind_mw', 'solar_mw', 'hydro_mw', 'nuclear_mw',
                'import_mw', 'export_mw', 'net_residual', 'thermal_mw',
            ]].copy()
            _sup_tbl.columns = [
                '省份', '峰值负荷', '风电', '光伏', '水电', '核电',
                '受入均功率', '送出均功率', '热电缺口', '火电',
            ]
            for _c in _sup_tbl.columns[1:]:
                _sup_tbl[_c] = _sup_tbl[_c].round(0).astype('int', errors='ignore')
            st.dataframe(
                _sup_tbl, use_container_width=True, hide_index=True,
                column_config={
                    '热电缺口': st.column_config.NumberColumn(
                        '热电缺口 (MW)', format="%d",
                        help="峰值负荷 − (风+光+水+核) − 受入 + 送出；负值=结构性过剩，正值=需要火电/储能补充",
                    ),
                },
            )

        # ── EOH / Thermal Load Factor Analysis ──────────────────────────────────
        st.divider()
        st.subheader("热电等效利用小时数（EOH）分析")
        st.caption(
            "净剩余负荷 = 实际负荷 − 风电 − 光伏 − 水电月均 − 核电基荷（容量×90%）；"
            "水电与核电数据来自 exchange_excel_metrics。年度净剩余负荷累计 ÷ 火电装机 = 等效利用小时数（EOH）。"
        )

        @st.cache_data(ttl=3600, show_spinner=False)
        def _load_eoh_profiles(_pg_key: str, _year: int):
            import psycopg2 as _pg2, pandas as _pd2
            try:
                _c2 = _pg2.connect(_pg_key)
                with _c2.cursor() as _cr2:
                    # Join hourly load/wind/solar with monthly hydro+nuclear from exchange_excel_metrics.
                    # hydro_avg_mw = monthly generation ÷ hours in month (flat baseload proxy).
                    # nuclear_avg_mw = nuclear capacity × 90% CF (near-flat baseload).
                    # avg_net_load = thermal-only demand after removing all non-thermal generation.
                    _cr2.execute("""
                        WITH hourly AS (
                            SELECT
                                province,
                                EXTRACT(MONTH FROM datetime AT TIME ZONE 'Asia/Shanghai')::int AS month,
                                EXTRACT(HOUR  FROM datetime AT TIME ZONE 'Asia/Shanghai')::int AS hour,
                                AVG(load_mw)                  AS avg_load,
                                AVG(COALESCE(wind_mw,  0))    AS avg_wind,
                                AVG(COALESCE(solar_mw, 0))    AS avg_solar,
                                -- net_export_mw > 0 means province exports (thermal must cover exports too);
                                -- net_export_mw < 0 means province imports (reduces local thermal need).
                                AVG(GREATEST(load_mw
                                    - COALESCE(wind_mw,  0)
                                    - COALESCE(solar_mw, 0)
                                    + COALESCE(net_export_mw, 0), 0)) AS avg_raw_net
                            FROM marketdata.spot_fundamentals_hourly
                            WHERE load_mw > 0
                              AND EXTRACT(YEAR FROM datetime AT TIME ZONE 'Asia/Shanghai') = %s
                            GROUP BY province, month, hour
                        ),
                        -- Monthly hydro gen + nuclear from exchange_excel_metrics (spot market provinces)
                        mc AS (
                            SELECT DISTINCT ON (province, EXTRACT(MONTH FROM report_month)::int)
                                province,
                                EXTRACT(MONTH FROM report_month)::int AS month,
                                COALESCE(hydro_generation_gwh, 0) * 1000.0 / 730.0 AS hydro_avg_mw,
                                COALESCE(nuclear_capacity_mw, 0) * 0.90             AS nuclear_avg_mw
                            FROM staging.exchange_excel_metrics
                            WHERE EXTRACT(YEAR FROM report_month) = %s
                            ORDER BY province, EXTRACT(MONTH FROM report_month)::int, report_month DESC
                        ),
                        -- Annual generation stats from province_fundamentals (fallback for all provinces)
                        pf AS (
                            SELECT DISTINCT ON (province_cn)
                                province_cn AS province,
                                COALESCE(hydro_gen_100gwh,   0) * 100000.0 / 8760.0 AS hydro_avg_mw,
                                COALESCE(nuclear_gen_100gwh, 0) * 100000.0 / 8760.0 AS nuclear_avg_mw,
                                -- Wind/solar carried as annual GWh totals for Python-side correction
                                -- (not applied per-hour to avoid nighttime solar artefacts).
                                COALESCE(wind_gen_100gwh,  0) * 100000.0 AS pf_wind_gwh,
                                COALESCE(solar_gen_100gwh, 0) * 100000.0 AS pf_solar_gwh
                            FROM marketdata.province_fundamentals
                            ORDER BY province_cn, year DESC
                        ),
                        -- Nuclear capacity floor (MW) from NNSA 2024 report.
                        -- Used as GREATEST floor so provinces where pf under-reports nuclear
                        -- (e.g. 江苏 pf gives 5,900 MW but actual Tianwan output = 7,650 MW)
                        -- still get the correct baseload deduction.
                        nuclear_override (province, nuclear_cap_mw) AS (
                            VALUES
                                ('江苏',   8500.0),   -- Tianwan 1-6
                                ('福建',  10000.0),   -- Ningde 1-4 + Fuqing 1-6
                                ('广东',  20000.0),   -- Daya Bay + Lingao + Yangjiang + Taishan
                                ('浙江',   6600.0),   -- Qinshan 1-3 + Haiyan
                                ('辽宁',   6700.0),   -- Hongyanhe 1-6
                                ('海南',   2200.0)    -- Changjiang 1-4
                        )
                        SELECT
                            h.province, h.month, h.hour,
                            h.avg_load, h.avg_wind, h.avg_solar,
                            -- No floor at 0: negative values mean renewables exceed load (or large imports).
                            -- Python layer uses these for annual EOH; chart clips display at 0.
                            h.avg_raw_net
                                - COALESCE(NULLIF(mc.hydro_avg_mw,   0), pf.hydro_avg_mw,   0)
                                -- nuclear: use GREATEST so override acts as a floor, not just a fallback.
                                - GREATEST(
                                    COALESCE(NULLIF(mc.nuclear_avg_mw, 0), NULLIF(pf.nuclear_avg_mw, 0), 0),
                                    COALESCE(no.nuclear_cap_mw * 0.90, 0)
                                  ) AS avg_net_load,
                            COALESCE(NULLIF(mc.hydro_avg_mw,   0), pf.hydro_avg_mw,   0) AS hydro_avg_mw,
                            GREATEST(
                                COALESCE(NULLIF(mc.nuclear_avg_mw, 0), NULLIF(pf.nuclear_avg_mw, 0), 0),
                                COALESCE(no.nuclear_cap_mw * 0.90, 0)
                            ) AS nuclear_avg_mw,
                            -- Annual wind/solar generation (GWh) from pf for Python-side correction
                            -- when hourly wind_mw/solar_mw were NULL for this province.
                            COALESCE(pf.pf_wind_gwh,  0) AS pf_wind_gwh,
                            COALESCE(pf.pf_solar_gwh, 0) AS pf_solar_gwh
                        FROM hourly h
                        LEFT JOIN mc ON h.province = mc.province AND h.month = mc.month
                        LEFT JOIN pf ON h.province = pf.province
                        LEFT JOIN nuclear_override no ON h.province = no.province
                        ORDER BY h.province, h.month, h.hour
                    """, (_year, _year))
                    _eoh_raw = _pd2.DataFrame(
                        _cr2.fetchall(), columns=[d[0] for d in _cr2.description]
                    )
                _c2.close()
                return _eoh_raw
            except Exception:
                return _pd2.DataFrame()

        _eoh_year = st.selectbox("分析年份", [2025, 2024, 2026], key="eoh_year")
        _eoh_raw = _load_eoh_profiles(_sup_pg, _eoh_year)
        # Drop metadata rows that got ingested as province names (LingFeng artefact)
        _eoh_raw = _eoh_raw[~_eoh_raw['province'].str.contains('运行数据|数据披露|披露', na=False)]

        if _eoh_raw.empty:
            st.info("暂无小时级供需数据（需要 marketdata.spot_fundamentals_hourly）")
        else:
            import plotly.graph_objects as _pgo_eoh, numpy as _np_eoh
            _days_pm = {1:31, 2:28 if _eoh_year % 4 != 0 else 29, 3:31, 4:30,
                        5:31, 6:30, 7:31, 8:31, 9:30, 10:31, 11:30, 12:31}

            # EOH method selector
            _eoh_method = st.radio(
                "计算方法",
                ['实际发电量法（小时数据）', '装机利用小时法（风电2500h/光伏1800h）'],
                horizontal=True, key="eoh_method",
            )
            _use_cap_factor = '装机' in _eoh_method

            # Province → adcode mapping (for geo map)
            _PROV_ADCODE = {
                "北京": "110000", "天津": "120000", "河北": "130098", "山西": "140000",
                # 冀北=130099 (north of ~39.5°N), 冀南/河北南网=130098 (south)
                "冀北": "130099", "冀南": "130098", "河北南网": "130098",
                "蒙西": "150000", "蒙东": "150000", "内蒙古": "150000", "西藏": "540000",
                "辽宁": "210000", "吉林": "220000", "黑龙江": "230000",
                "上海": "310000", "江苏": "320000", "浙江": "330000",
                "安徽": "340000", "福建": "350000", "江西": "360000",
                "山东": "370000", "河南": "410000", "湖北": "420000", "湖南": "430000",
                "广东": "440000", "广西": "450000", "海南": "460000",
                "重庆": "500000", "四川": "510000", "贵州": "520000", "云南": "530000",
                "陕西": "610000", "甘肃": "620000", "青海": "630000",
                "宁夏": "640000", "新疆": "650000",
            }
            _PROV_CENTROIDS_EOH = {
                "110000": (39.90, 116.40), "120000": (39.13, 117.20),
                "130099": (41.20, 117.00), "130098": (38.04, 114.47),  # 冀北, 冀南
                "140000": (37.87, 112.56),
                "150000": (44.09, 113.09), "210000": (41.80, 123.43),
                "220000": (43.89, 125.32), "230000": (47.85, 127.57),
                "310000": (31.23, 121.47), "320000": (32.06, 119.59),
                "330000": (30.27, 120.15), "340000": (31.86, 117.29),
                "350000": (26.10, 118.31), "360000": (27.62, 115.70),
                "370000": (36.67, 117.02), "410000": (34.76, 113.75),
                "420000": (30.60, 114.30), "430000": (28.23, 112.94),
                "440000": (23.37, 113.50), "450000": (23.73, 108.38),
                "460000": (20.02, 110.35), "500000": (29.56, 106.54),
                "510000": (30.57, 103.99), "520000": (26.82, 106.83),
                "530000": (25.05, 101.71), "540000": (29.65,  91.11),
                "610000": (34.27, 108.95),
                "620000": (36.06, 103.83), "630000": (36.62, 101.74),
                "640000": (38.47, 106.26), "650000": (41.17,  85.29),
            }

            # Province aliases: spot_fundamentals uses 河北南网 but province_fundamentals
            # stores the same grid's data as 冀南. Map for thermal capacity lookup.
            _THERMAL_ALIAS = {'河北南网': '冀南', '冀南': '河北南网'}

            # Thermal capacity from province_fundamentals directly (most reliable source).
            # Used as fallback when province_installed_monthly has bad/missing thermal data.
            _pf_thermal_cache: dict = {}
            try:
                import psycopg2 as _pg2_th
                _conn_th = _pg2_th.connect(_sup_pg)
                with _conn_th.cursor() as _cr_th:
                    _yr_th = _eoh_year if _eoh_year in [2024, 2025] else 2025
                    _cr_th.execute(
                        "SELECT province_cn, thermal_cap_10kw * 10 FROM marketdata.province_fundamentals WHERE year = %s",
                        (_yr_th,)
                    )
                    for _pf_pn, _pf_tc in _cr_th.fetchall():
                        if _pf_tc:
                            _pf_thermal_cache[_pf_pn] = float(_pf_tc)
                _conn_th.close()
            except Exception:
                pass

            # Compute annual EOH per province.
            # EOH may be negative for provinces with large renewable surplus.
            _eoh_rows = []
            for _ep in sorted(_eoh_raw['province'].unique()):
                _pdf = _eoh_raw[_eoh_raw['province'] == _ep]
                _sup_row = _sup_df[_sup_df['province'] == _ep]
                _th_cap = float(_sup_row['thermal_mw'].iloc[0]) if not _sup_row.empty else 0
                # Use province_fundamentals thermal as authoritative source when available.
                # province_installed_monthly thermal can be wrong (bad LLM extraction).
                # Resolve alias: 河北南网 → 冀南 in province_fundamentals.
                _pf_alias = _THERMAL_ALIAS.get(_ep, _ep)
                _pf_th = _pf_thermal_cache.get(_ep) or _pf_thermal_cache.get(_pf_alias, 0)
                if _pf_th > 0:
                    _th_cap = _pf_th

                if _use_cap_factor:
                    # Method 2: annual_demand from hourly avg_load; wind/solar from installed cap × std hours
                    _annual_demand = sum(
                        _pdf[_pdf['month'] == _m]['avg_load'].sum() * _days_pm.get(_m, 30)
                        for _m in range(1, 13)
                    )  # MWh
                    _wind_cap  = float(_sup_row['wind_mw'].iloc[0])  if (not _sup_row.empty and 'wind_mw'  in _sup_row.columns) else 0
                    _solar_cap = float(_sup_row['solar_mw'].iloc[0]) if (not _sup_row.empty and 'solar_mw' in _sup_row.columns) else 0
                    _annual_wind  = _wind_cap  * 2500.0   # MWh (2500 equivalent hours)
                    _annual_solar = _solar_cap * 1800.0   # MWh (1800 equivalent hours)
                    # Hydro + nuclear from pf column (already in MWh: 亿kWh × 100000)
                    _pf_hydro  = float(_pdf['hydro_avg_mw'].iloc[0])  * 8760.0 if not _pdf.empty else 0
                    _pf_nuc    = float(_pdf['nuclear_avg_mw'].iloc[0]) * 8760.0 if not _pdf.empty else 0
                    _annual_gen = _annual_demand - _annual_wind - _annual_solar - _pf_hydro - _pf_nuc
                else:
                    # Method 1: thermal residual from hourly net load
                    _annual_gen = sum(
                        _pdf[_pdf['month'] == _m]['avg_net_load'].sum() * _days_pm.get(_m, 30)
                        for _m in range(1, 13)
                    )  # MWh
                    # Correct for missing wind/solar hourly data (all-zero provinces).
                    # pf_wind_gwh/pf_solar_gwh are in MWh (亿kWh × 100000 in SQL).
                    _has_wind  = _pdf['avg_wind'].sum()  > 0
                    _has_solar = _pdf['avg_solar'].sum() > 0
                    if not _has_wind and 'pf_wind_gwh' in _pdf.columns:
                        _annual_gen -= float(_pdf['pf_wind_gwh'].iloc[0])   # already MWh
                    if not _has_solar and 'pf_solar_gwh' in _pdf.columns:
                        _annual_gen -= float(_pdf['pf_solar_gwh'].iloc[0])  # already MWh

                _eoh_val = round(_annual_gen / _th_cap) if _th_cap > 0 else 0
                # Sanity: EOH > 8760h/yr is impossible — flag as 0 (bad hourly data)
                if _eoh_val > 8760:
                    _eoh_val = 0
                _eoh_rows.append({
                    'province': _ep,
                    'annual_gen_gwh': round(_annual_gen / 1000),
                    'thermal_mw': round(_th_cap),
                    'eoh': _eoh_val,
                    'load_factor': round(_eoh_val / 8760, 3) if _eoh_val else 0,
                    'adcode': _PROV_ADCODE.get(_ep),
                })
            # Add provinces that have province_fundamentals data but NO hourly spot data
            # (e.g. 冀北). EOH computed from annual thermal generation ÷ thermal capacity.
            _hourly_provs = set(_eoh_raw['province'].unique())
            _fund_year_sel = _eoh_year if _eoh_year in [2024, 2025] else 2025
            try:
                import psycopg2 as _pg2_fund
                _conn_fund = _pg2_fund.connect(_sup_pg)
                with _conn_fund.cursor() as _cr_fund:
                    _cr_fund.execute("""
                        SELECT province_cn,
                               COALESCE(thermal_gen_100gwh, 0) * 100000.0 AS thermal_gen_mwh,
                               COALESCE(thermal_cap_10kw,   0) * 10        AS thermal_mw
                        FROM marketdata.province_fundamentals
                        WHERE year = %s
                          AND thermal_cap_10kw > 0
                          AND thermal_gen_100gwh > 0
                    """, (_fund_year_sel,))
                    _pf_gen_rows = _cr_fund.fetchall()
                _conn_fund.close()
                for _pfp, _pfgen, _pfcap in _pf_gen_rows:
                    if _pfp in _hourly_provs:
                        continue  # already covered by hourly loop
                    # Skip if alias already covered (e.g. 冀南 covered via 河北南网)
                    if _THERMAL_ALIAS.get(_pfp) in _hourly_provs:
                        continue
                    _eoh_val = round(_pfgen / _pfcap) if _pfcap > 0 else 0
                    _eoh_rows.append({
                        'province': _pfp,
                        'annual_gen_gwh': round(_pfgen / 1000),
                        'thermal_mw': round(_pfcap),
                        'eoh': _eoh_val,
                        'load_factor': round(_eoh_val / 8760, 3) if _eoh_val else 0,
                        'adcode': _PROV_ADCODE.get(_pfp),
                    })
            except Exception:
                pass

            _eoh_df = pd.DataFrame(_eoh_rows).sort_values('eoh', ascending=False)

            _eoh_prov_list = sorted(_eoh_raw['province'].unique())
            _eoh_prov = st.selectbox("选择省份查看日内热电需求分布", _eoh_prov_list, key="eoh_prov")
            _epdf = _eoh_raw[_eoh_raw['province'] == _eoh_prov]

            _col_e1, _col_e2 = st.columns(2)

            with _col_e1:
                # Heatmap: month (x) × hour (y), value = avg_net_load
                _heat = _np_eoh.zeros((24, 12))
                for _, _r in _epdf.iterrows():
                    _h, _m = int(_r['hour']), int(_r['month']) - 1
                    if 0 <= _h < 24 and 0 <= _m < 12:
                        _heat[_h, _m] = _r['avg_net_load']
                _fig_heat = _pgo_eoh.Figure(_pgo_eoh.Heatmap(
                    z=_heat,
                    x=[f"{i}月" for i in range(1, 13)],
                    y=[f"{i:02d}:00" for i in range(24)],
                    colorscale='Reds',
                    colorbar=dict(title="MW"),
                ))
                _fig_heat.update_layout(
                    title=f"{_eoh_prov}  净剩余负荷（MW）—— 月 × 时段",
                    height=420,
                    margin=dict(t=40, b=20, l=60, r=20),
                    yaxis=dict(autorange='reversed'),
                )
                st.plotly_chart(_fig_heat, use_container_width=True)

            with _col_e2:
                # EOH bar chart — negative values allowed (surplus renewables or large net imports)
                _eoh_sorted = _eoh_df.sort_values('eoh', ascending=True)
                def _eoh_color(v):
                    if v < 0:      return '#2980B9'   # blue = net surplus
                    if v < 4000:   return '#27AE60'   # green = low utilisation
                    if v < 5500:   return '#E67E22'   # orange = moderate
                    return '#E74C3C'                   # red = high
                _fig_eoh = _pgo_eoh.Figure(_pgo_eoh.Bar(
                    x=_eoh_sorted['eoh'],
                    y=_eoh_sorted['province'],
                    orientation='h',
                    marker_color=[_eoh_color(v) for v in _eoh_sorted['eoh']],
                    text=[f"{v:,}" if v >= 0 else "" for v in _eoh_sorted['eoh']],
                    textposition='outside',
                ))
                _fig_eoh.add_vline(x=5500, line_dash='dash', line_color='gray',
                                   annotation_text='5500h')
                _fig_eoh.add_vline(x=3500, line_dash='dot', line_color='#27AE60',
                                   annotation_text='3500h')
                _fig_eoh.add_vline(x=0, line_dash='solid', line_color='black', line_width=1)
                _fig_eoh.update_layout(
                    title=f"{_eoh_year}年 各省热电等效利用小时数（负值=可再生盈余）",
                    xaxis_title="EOH（小时/年）",
                    height=max(320, len(_eoh_sorted) * 26),
                    margin=dict(t=40, b=20, l=80, r=80),
                )
                st.plotly_chart(_fig_eoh, use_container_width=True)
                # Data quality caveat for large-import provinces
                _high_eoh = _eoh_df[_eoh_df['eoh'] > 6000]['province'].tolist()
                if _high_eoh:
                    st.caption(
                        f"[!] {', '.join(_high_eoh)} 等效利用小时偏高，可能原因：省内净外送（净购入）数据缺失。"
                        "如 LingFeng 数据缺少净外送列，跨省购入电量未从热电需求中扣除。"
                    )

            # ── Geo map ──────────────────────────────────────────────────────
            st.markdown("**热电等效利用小时数 — 省级地图**")
            try:
                import json as _json_eoh, matplotlib as _mpl_eoh
                import matplotlib.pyplot as _mplt_eoh
                from matplotlib.patches import Polygon as _MplPoly
                import numpy as _np_eoh2

                _geo_path = (
                    __import__('pathlib').Path(__file__).resolve().parent / "data" / "china_provinces.geojson"
                )
                if _geo_path.exists():
                    _geojson_eoh = _json_eoh.loads(_geo_path.read_text(encoding="utf-8"))
                else:
                    _geojson_eoh = None

                if _geojson_eoh:
                    # Build adcode (int) → EOH lookup
                    # Use pd.notna to guard against both None and float NaN
                    _eoh_map = {
                        int(row['adcode']): row['eoh']
                        for _, row in _eoh_df.iterrows()
                        if pd.notna(row['adcode'])
                    }
                    # Color: blue (surplus) → green → yellow → orange → red
                    def _eoh_geo_color(v):
                        if v is None:      return "#d0d0d0"
                        if v <= 0:         return "#2980B9"
                        if v < 3000:       return "#27AE60"
                        if v < 4500:       return "#F1C40F"
                        if v < 6000:       return "#E67E22"
                        return "#E74C3C"

                    # Find CJK font for matplotlib titles
                    _cjk_font = None
                    try:
                        import matplotlib.font_manager as _fmgr
                        for _fp in _fmgr.findSystemFonts():
                            if any(k in _fp.lower() for k in ("noto", "cjk", "chinese", "wqy", "simhei")):
                                _cjk_font = _fmgr.FontProperties(fname=_fp)
                                break
                    except Exception:
                        pass

                    _fig_geo, _ax_geo = _mplt_eoh.subplots(figsize=(11, 7), facecolor="white")
                    _ax_geo.set_facecolor("#c8daf4")
                    _ax_geo.set_aspect("equal")
                    _ax_geo.axis("off")

                    for _feat in _geojson_eoh.get("features", []):
                        _adc = _feat.get("properties", {}).get("adcode")
                        _adc_int = int(_adc) if _adc is not None else None
                        _eoh_v = _eoh_map.get(_adc_int) if _adc_int is not None else None
                        _fc = _eoh_geo_color(_eoh_v)
                        _geom = _feat.get("geometry", {})
                        _rings = []
                        if _geom.get("type") == "Polygon":
                            _rings = [_geom["coordinates"][0]]
                        elif _geom.get("type") == "MultiPolygon":
                            _rings = [p[0] for p in _geom["coordinates"]]
                        for _ring in _rings:
                            _coords = _np_eoh2.array(_ring)
                            _ax_geo.add_patch(_MplPoly(
                                _coords, closed=True,
                                facecolor=_fc, edgecolor="white", linewidth=0.6,
                            ))

                    # adcode → Chinese province name
                    _ADCODE_TO_NAME = {
                        110000:"北京", 120000:"天津",
                        130099:"冀北", 130098:"冀南",  # split from Hebei 130000
                        140000:"山西",
                        150000:"内蒙古", 210000:"辽宁", 220000:"吉林", 230000:"黑龙江",
                        310000:"上海", 320000:"江苏", 330000:"浙江", 340000:"安徽",
                        350000:"福建", 360000:"江西", 370000:"山东", 410000:"河南",
                        420000:"湖北", 430000:"湖南", 440000:"广东", 450000:"广西",
                        460000:"海南", 500000:"重庆", 510000:"四川", 520000:"贵州",
                        530000:"云南", 540000:"西藏", 610000:"陕西", 620000:"甘肃",
                        630000:"青海", 640000:"宁夏", 650000:"新疆",
                    }
                    # Labels at province centroids: name + EOH
                    for _adc_str, (_lat, _lon) in _PROV_CENTROIDS_EOH.items():
                        _adc_int = int(_adc_str)
                        _eoh_v = _eoh_map.get(_adc_int)
                        _pname = _ADCODE_TO_NAME.get(_adc_int, "")
                        if _eoh_v is not None:
                            _eoh_lbl = f"{_eoh_v/1000:.1f}k" if abs(_eoh_v) >= 1000 else str(_eoh_v)
                            _lbl = f"{_pname}\n{_eoh_lbl}" if _pname else _eoh_lbl
                            _txt_kw = dict(ha="center", va="center", fontsize=5.5, fontweight="bold",
                                           color="white" if _eoh_v > 4500 else "black")
                            if _cjk_font:
                                _txt_kw["fontproperties"] = _cjk_font
                            _ax_geo.text(_lon, _lat, _lbl, **_txt_kw)

                    _ax_geo.set_xlim(72, 136)
                    _ax_geo.set_ylim(17, 54)
                    _title_txt = (f"{_eoh_year}  Thermal EOH  "
                                  "Blue=surplus | Green<3000 | Yellow3000-4500 | Orange4500-6000 | Red>6000")
                    _ax_geo.set_title(_title_txt, fontsize=8, pad=8)
                    st.pyplot(_fig_geo, use_container_width=True)
                    _mplt_eoh.close(_fig_geo)
                else:
                    st.caption("地图文件未找到（data/china_provinces.geojson）")
            except Exception as _egeo:
                st.caption(f"地图加载失败: {_egeo}")

            # Typical day by season — show full stack decomposition for one season
            st.markdown(f"**{_eoh_prov} 典型日供需结构**")
            _sel_season = st.radio(
                "季节", ['春（3-5月）', '夏（6-8月）', '秋（9-11月）', '冬（12-2月）'],
                horizontal=True, key="eoh_season",
            )
            _season_months_map = {
                '春（3-5月）': [3,4,5], '夏（6-8月）': [6,7,8],
                '秋（9-11月）': [9,10,11], '冬（12-2月）': [12,1,2],
            }
            _epdf_season = _epdf[_epdf['month'].isin(_season_months_map[_sel_season])]
            _sd = _epdf_season.groupby('hour').mean(numeric_only=True).reset_index()
            # hydro/nuclear are flat per month — take seasonal scalar average from raw data
            _hydro_scalar = float(_epdf_season['hydro_avg_mw'].mean()) if 'hydro_avg_mw' in _epdf_season.columns else 0.0
            _nuc_scalar   = float(_epdf_season['nuclear_avg_mw'].mean()) if 'nuclear_avg_mw' in _epdf_season.columns else 0.0
            if not _sd.empty:
                _hrs = _sd['hour'].tolist()
                _fig_day = _pgo_eoh.Figure()
                # Stacked area: solar, wind, hydro (flat), nuclear (flat), then load line + thermal top
                _fig_day.add_trace(_pgo_eoh.Scatter(
                    x=_hrs, y=_sd['avg_solar'].tolist(), mode='lines', name='光伏',
                    fill='tozeroy', line=dict(color='#F1C40F', width=0), fillcolor='rgba(241,196,15,0.5)',
                ))
                _fig_day.add_trace(_pgo_eoh.Scatter(
                    x=_hrs,
                    y=(_sd['avg_solar'] + _sd['avg_wind']).tolist(),
                    mode='lines', name='风电',
                    fill='tonexty', line=dict(color='#27AE60', width=0), fillcolor='rgba(39,174,96,0.5)',
                ))
                _hydro_line = _sd['avg_solar'] + _sd['avg_wind'] + _hydro_scalar
                _fig_day.add_trace(_pgo_eoh.Scatter(
                    x=_hrs, y=_hydro_line.tolist(), mode='lines', name='水电',
                    fill='tonexty', line=dict(color='#2980B9', width=0), fillcolor='rgba(41,128,185,0.4)',
                ))
                _nuc_line = _hydro_line + _nuc_scalar
                _fig_day.add_trace(_pgo_eoh.Scatter(
                    x=_hrs, y=_nuc_line.tolist(), mode='lines', name='核电',
                    fill='tonexty', line=dict(color='#8E44AD', width=0), fillcolor='rgba(142,68,173,0.4)',
                ))
                _fig_day.add_trace(_pgo_eoh.Scatter(
                    x=_hrs, y=_sd['avg_load'].tolist(), mode='lines', name='总负荷',
                    line=dict(color='black', width=2.5),
                ))
                _fig_day.add_trace(_pgo_eoh.Scatter(
                    x=_hrs, y=(_nuc_line + _sd['avg_net_load']).tolist(), mode='lines', name='火电需求上沿',
                    line=dict(color='#E74C3C', width=1.5, dash='dot'),
                ))
                _fig_day.update_layout(
                    xaxis=dict(title="时段", tickvals=list(range(0,24,2)), ticktext=[f"{h:02d}:00" for h in range(0,24,2)]),
                    yaxis_title="MW",
                    height=360, margin=dict(t=10, b=20, l=60, r=20),
                    legend=dict(orientation='h', y=-0.3),
                )
                st.plotly_chart(_fig_day, use_container_width=True)

            # Summary table
            st.dataframe(
                _eoh_df.rename(columns={
                    'province': '省份', 'annual_gen_gwh': '年净剩余发电量(GWh)',
                    'thermal_mw': '火电装机(MW)', 'eoh': 'EOH(小时)', 'load_factor': '利用率',
                }),
                use_container_width=True, hide_index=True,
                column_config={
                    'EOH(小时)': st.column_config.NumberColumn(format="%d"),
                    '利用率': st.column_config.NumberColumn(format="%.1%"),
                },
            )

# ── Tab: 价格预测 (Price Forecasting) ─────────────────────────────────────────
with tab_forecast:
    import numpy as _np_fc
    import pandas as _pd_fc
    import plotly.graph_objects as _pgo_fc

    st.subheader("价格预测 — 短中期混合模型")
    st.caption("综合 PCA时序模型 · 边际成本模型 · 贝叶斯分布模型，预测 D+1 至 M+1 日前价格")

    # ── Controls ─────────────────────────────────────────────────────────────
    _fc_c1, _fc_c2, _fc_c3, _fc_c4 = st.columns([2, 1, 1, 1])
    with _fc_c1:
        _fc_zh_provs = _load_hourly_price_provinces(_conn) or (
            sorted(df["province_cn"].dropna().unique()) if not df.empty else []
        )
        _fc_prov = st.selectbox(
            "选择省份（仅显示有小时价格数据的省份）", _fc_zh_provs,
            key="fc_province",
        )
    with _fc_c2:
        _fc_price_type = st.radio("价格类型", ["DA (日前)", "RT (实时)"], horizontal=True, key="fc_pt")
        _fc_pcol = "da_price" if "DA" in _fc_price_type else "rt_price"
    with _fc_c3:
        _fc_train_months = st.selectbox("训练数据", [6, 12, 24], index=1, key="fc_train",
                                         format_func=lambda x: f"近{x}个月")
    with _fc_c4:
        _fc_days_eoy2027 = max(1, (_pd_fc.Timestamp('2027-12-31') - _pd_fc.Timestamp.today().normalize()).days)
        _fc_horizon_opts = [1, 3, 7, 30, 90, 180, 365, _fc_days_eoy2027]
        def _fc_hlabel(x):
            if x == 1:   return "D+1"
            if x <= 7:   return f"D+{x}"
            if x <= 31:  return "M+1 (30天)"
            if x <= 92:  return "近3个月"
            if x <= 183: return "近6个月"
            if x <= 366: return "Y+1 (12个月)"
            return f"至2027年底 ({x}天)"
        _fc_horizon = st.selectbox("预测范围", _fc_horizon_opts, index=0, key="fc_horizon",
                                    format_func=_fc_hlabel)

    _fc_run = st.button("运行预测", type="primary", key="fc_run")

    if not _fc_run:
        st.info("选择省份和参数后点击「运行预测」。模型将使用历史价格数据进行训练和预测。")
    else:
        _fc_end_dt   = _pd_fc.Timestamp.today().normalize()
        _fc_start_dt = _fc_end_dt - _pd_fc.DateOffset(months=_fc_train_months)
        _fc_start    = str(_fc_start_dt.date())
        _fc_end      = str(_fc_end_dt.date())
        # Recent 30 days for Bayesian likelihood
        _fc_recent_start = str((_fc_end_dt - _pd_fc.DateOffset(days=30)).date())

        with st.spinner("加载历史价格数据…"):
            _fc_price_mat = _load_price_matrix(
                _conn, _fc_prov, _fc_start, _fc_end, _fc_pcol)
            _fc_fund_df   = _load_forecast_fundamentals(
                _conn, _fc_prov, _fc_recent_start, _fc_end)
            # Hourly price series for Bayesian
            _fc_all_hourly = _load_intraday_shape(
                _conn, (_fc_prov,), _fc_start, _fc_end, _fc_pcol)
            _fc_recent_hourly = _load_intraday_shape(
                _conn, (_fc_prov,), _fc_recent_start, _fc_end, _fc_pcol)

        if _fc_price_mat.empty:
            st.warning(f"{_fc_prov} 暂无小时价格数据（{_fc_pcol}）。")
        else:
            # ─────────────────────────────────────────────────────────────────
            # MODEL 1: PCA + ARIMA
            # ─────────────────────────────────────────────────────────────────
            _fc_tab_pca, _fc_tab_stack, _fc_tab_bayes, _fc_tab_ensemble = st.tabs([
                "PCA时序模型", "边际成本模型", "贝叶斯分布模型", "综合预测"
            ])

            # Shared: build price matrix and PCA results (used in PCA + Ensemble tabs)
            _fc_mat = _fc_price_mat.values.astype(float)   # (n_days, 24)
            _fc_mat = _np_fc.nan_to_num(_fc_mat, nan=_np_fc.nanmean(_fc_mat))
            # Normalise to ¥/kWh: spot_prices_hourly may store ¥/MWh (>5) depending on
            # ingestion pipeline version. Median > 5 ¥/kWh is physically impossible →
            # values must be in ¥/MWh; divide by 1000 so all downstream code stays in ¥/kWh.
            _fc_unit_factor = 1000.0 if float(_np_fc.nanmedian(_fc_mat)) > 5 else 1.0
            if _fc_unit_factor != 1.0:
                _fc_mat /= _fc_unit_factor
            _fc_mean_24h = _fc_mat.mean(axis=0)            # (24,), ¥/kWh
            _fc_std_24h  = _fc_mat.std(axis=0)
            _fc_std_24h  = _np_fc.where(_fc_std_24h < 1e-6, 1.0, _fc_std_24h)
            _fc_mat_c    = (_fc_mat - _fc_mean_24h) / _fc_std_24h   # centered+scaled

            # Initialise shared variables used across tabs (read in ensemble, set in inner tabs)
            _fc_fund_df2  = _pd_fc.DataFrame()
            _fc_post_mean = _np_fc.full(24, _fc_mean_24h.mean())   # fallback prior mean
            _fc_post_lo90 = _fc_post_mean - 0.05
            _fc_post_hi90 = _fc_post_mean + 0.05

            # SVD
            _fc_U, _fc_s, _fc_Vt = _np_fc.linalg.svd(_fc_mat_c, full_matrices=False)
            _fc_n_pc  = min(4, _fc_U.shape[1])
            _fc_var_exp = _fc_s**2 / (_fc_s**2).sum()
            _fc_scores  = _fc_U[:, :_fc_n_pc] * _fc_s[:_fc_n_pc]   # (n_days, n_pc)
            _fc_comps   = _fc_Vt[:_fc_n_pc, :]                       # (n_pc, 24)

            # Seasonal monthly mean per PC from training data (for long-horizon extension)
            _fc_score_months = _np_fc.array([
                int(str(d)[5:7]) for d in _fc_price_mat.index
            ])  # (n_days,) month integers
            _fc_pc_monthly_mean = _np_fc.zeros((12, _fc_n_pc))  # (12 months, n_pc)
            _fc_pc_monthly_std  = _np_fc.ones((12, _fc_n_pc))
            for _fc_m in range(1, 13):
                _fc_mask = _fc_score_months == _fc_m
                if _fc_mask.sum() > 0:
                    _fc_pc_monthly_mean[_fc_m - 1] = _fc_scores[_fc_mask].mean(axis=0)
                    _fc_pc_monthly_std[_fc_m - 1]  = _fc_scores[_fc_mask].std(axis=0).clip(min=1e-6)

            # ARIMA capped at 30 steps; long horizons extended with seasonal mean reversion
            _fc_arima_steps = min(_fc_horizon, 30)
            try:
                from statsmodels.tsa.arima.model import ARIMA as _fc_ARIMA
                _fc_pc_forecasts = []
                _fc_pc_ci_lo = []
                _fc_pc_ci_hi = []
                for _fc_i in range(_fc_n_pc):
                    _fc_series = _pd_fc.Series(_fc_scores[:, _fc_i])
                    _fc_mdl = _fc_ARIMA(_fc_series, order=(1, 0, 1)).fit()
                    _fc_res = _fc_mdl.get_forecast(steps=_fc_arima_steps)
                    _fc_arima_fcast = _fc_res.predicted_mean.values
                    _fc_arima_lo    = _fc_res.conf_int(alpha=0.1).iloc[:, 0].values
                    _fc_arima_hi    = _fc_res.conf_int(alpha=0.1).iloc[:, 1].values
                    if _fc_horizon > _fc_arima_steps:
                        # Extend to full horizon using seasonal monthly means
                        _fc_ext_fcast = []
                        _fc_ext_lo    = []
                        _fc_ext_hi    = []
                        _fc_today_m   = int(str(_fc_end_dt.date())[5:7])
                        for _fc_d in range(_fc_arima_steps, _fc_horizon):
                            _fc_future_month = ((_fc_today_m - 1 + _fc_d) % 12)
                            _fc_seas_mu  = _fc_pc_monthly_mean[_fc_future_month, _fc_i]
                            _fc_seas_sig = _fc_pc_monthly_std[_fc_future_month, _fc_i]
                            # Mean-revert toward seasonal mean with decay
                            _fc_decay = 0.97 ** (_fc_d - _fc_arima_steps)
                            _fc_last  = _fc_arima_fcast[-1]
                            _fc_ext_v = _fc_seas_mu + (_fc_last - _fc_seas_mu) * _fc_decay
                            _fc_ext_fcast.append(_fc_ext_v)
                            _fc_ext_lo.append(_fc_ext_v - 1.645 * _fc_seas_sig)
                            _fc_ext_hi.append(_fc_ext_v + 1.645 * _fc_seas_sig)
                        _fc_pc_forecasts.append(_np_fc.concatenate([_fc_arima_fcast, _fc_ext_fcast]))
                        _fc_pc_ci_lo.append(_np_fc.concatenate([_fc_arima_lo,    _fc_ext_lo]))
                        _fc_pc_ci_hi.append(_np_fc.concatenate([_fc_arima_hi,    _fc_ext_hi]))
                    else:
                        _fc_pc_forecasts.append(_fc_arima_fcast)
                        _fc_pc_ci_lo.append(_fc_arima_lo)
                        _fc_pc_ci_hi.append(_fc_arima_hi)
                _fc_pca_ok = True
            except Exception as _fc_arima_err:
                _fc_pca_ok = False
                _fc_arima_err_msg = str(_fc_arima_err)
                _fc_pc_forecasts = [_np_fc.zeros(_fc_horizon) for _ in range(_fc_n_pc)]
                _fc_pc_ci_lo = _fc_pc_forecasts
                _fc_pc_ci_hi = _fc_pc_forecasts

            # Reconstruct 24h price profiles (horizon × 24)
            _fc_pc_fcast_arr = _np_fc.array(_fc_pc_forecasts)   # (n_pc, horizon)
            _fc_pca_pred = (_fc_pc_fcast_arr.T @ _fc_comps) * _fc_std_24h + _fc_mean_24h  # (horizon, 24)
            _fc_pca_pred_lo = ((_np_fc.array(_fc_pc_ci_lo).T @ _fc_comps) * _fc_std_24h + _fc_mean_24h)
            _fc_pca_pred_hi = ((_np_fc.array(_fc_pc_ci_hi).T @ _fc_comps) * _fc_std_24h + _fc_mean_24h)

            with _fc_tab_pca:
                st.markdown("**PCA + ARIMA 时序模型**")
                st.caption(
                    "将历史每日24小时价格曲线分解为4个主成分（PC），分别对每个PC序列拟合ARIMA(1,0,1)，"
                    "再通过载荷矩阵重构预测价格曲线。"
                )

                # Show variance explained
                _fc_ve_df = _pd_fc.DataFrame({
                    '主成分': [f"PC{i+1}" for i in range(_fc_n_pc)],
                    '方差解释率': [f"{_fc_var_exp[i]:.1%}" for i in range(_fc_n_pc)],
                    '累计解释率': [f"{_fc_var_exp[:i+1].sum():.1%}" for i in range(_fc_n_pc)],
                })
                st.dataframe(_fc_ve_df, hide_index=True, use_container_width=False)

                if not _fc_pca_ok:
                    st.warning(f"ARIMA拟合失败: {_fc_arima_err_msg}")

                # Show PC shapes (loadings)
                _fc_hours = list(range(24))
                _fc_fig_pca = _pgo_fc.Figure()
                for _fc_i in range(_fc_n_pc):
                    _fc_fig_pca.add_trace(_pgo_fc.Scatter(
                        x=_fc_hours, y=_fc_comps[_fc_i, :],
                        name=f"PC{_fc_i+1} ({_fc_var_exp[_fc_i]:.1%})",
                        line=dict(width=2),
                    ))
                _fc_fig_pca.update_layout(
                    title="主成分载荷向量（24h价格曲线形态）",
                    xaxis=dict(title="时段", tickvals=list(range(0, 24, 4)),
                               ticktext=[f"{h:02d}:00" for h in range(0, 24, 4)]),
                    yaxis_title="载荷", height=300,
                    margin=dict(t=40, b=20, l=50, r=20),
                    legend=dict(orientation='h', y=-0.3),
                )
                st.plotly_chart(_fc_fig_pca, use_container_width=True)

                # Show forecast
                _fc_fig_pred = _pgo_fc.Figure()
                _fc_colors = ['#E74C3C', '#E67E22', '#F1C40F', '#2ECC71']
                for _fc_d in range(min(_fc_horizon, 7)):
                    _fc_day_lbl = f"D+{_fc_d+1}"
                    _fc_fig_pred.add_trace(_pgo_fc.Scatter(
                        x=_fc_hours, y=_fc_pca_pred[_fc_d, :],
                        name=_fc_day_lbl,
                        line=dict(width=2, color=_fc_colors[_fc_d % len(_fc_colors)]),
                    ))
                    if _fc_horizon == 1:  # show CI for single-day forecast
                        _fc_fig_pred.add_trace(_pgo_fc.Scatter(
                            x=_fc_hours + _fc_hours[::-1],
                            y=list(_fc_pca_pred_hi[_fc_d, :]) + list(_fc_pca_pred_lo[_fc_d, ::-1]),
                            fill='toself', fillcolor='rgba(231,76,60,0.15)',
                            line=dict(color='rgba(255,255,255,0)'),
                            showlegend=False, name='90% CI',
                        ))
                # Add recent actuals (last 5 days average by hour)
                if not _fc_all_hourly.empty:
                    _fc_fig_pred.add_trace(_pgo_fc.Scatter(
                        x=_fc_all_hourly['hour'].tolist(),
                        y=_fc_all_hourly['avg_price'].tolist(),
                        name='历史均值(训练期)',
                        line=dict(dash='dot', color='#7F8C8D', width=1.5),
                    ))
                _fc_fig_pred.update_layout(
                    title=f"{_fc_prov} — PCA预测日前价格曲线（¥/kWh）",
                    xaxis=dict(title="时段", tickvals=list(range(0, 24, 4)),
                               ticktext=[f"{h:02d}:00" for h in range(0, 24, 4)]),
                    yaxis_title="¥/kWh", height=380,
                    margin=dict(t=40, b=20, l=60, r=20),
                    legend=dict(orientation='h', y=-0.3),
                )
                st.plotly_chart(_fc_fig_pred, use_container_width=True)

                # Long-horizon monthly time-series view
                if _fc_horizon > 30:
                    st.markdown("**长期预测 — 月度日均价格走势**")
                    st.caption(
                        "前30天采用 ARIMA 动态预测，之后按季节性月均值均值回归延伸。"
                        "横轴为预测日期，纵轴为该日24小时均价（¥/kWh）。"
                    )
                    _fc_daily_mean     = _fc_pca_pred.mean(axis=1)    # (horizon,)
                    _fc_daily_mean_lo  = _fc_pca_pred_lo.mean(axis=1)
                    _fc_daily_mean_hi  = _fc_pca_pred_hi.mean(axis=1)
                    _fc_future_dates   = [
                        (_fc_end_dt + _pd_fc.DateOffset(days=d+1)).date()
                        for d in range(_fc_horizon)
                    ]
                    # Aggregate to monthly for cleaner chart
                    _fc_lt_df = _pd_fc.DataFrame({
                        'date':  _fc_future_dates,
                        'mean':  _fc_daily_mean,
                        'lo90':  _fc_daily_mean_lo,
                        'hi90':  _fc_daily_mean_hi,
                    })
                    _fc_lt_df['ym'] = _pd_fc.to_datetime(_fc_lt_df['date']).dt.to_period('M')
                    _fc_lt_mo = _fc_lt_df.groupby('ym').agg(
                        mean=('mean', 'mean'), lo90=('lo90', 'mean'), hi90=('hi90', 'mean')
                    ).reset_index()
                    _fc_lt_mo['ym_str'] = _fc_lt_mo['ym'].astype(str)
                    _fc_fig_lt = _pgo_fc.Figure()
                    _fc_fig_lt.add_trace(_pgo_fc.Scatter(
                        x=list(_fc_lt_mo['ym_str']) + list(_fc_lt_mo['ym_str'])[::-1],
                        y=list(_fc_lt_mo['hi90']) + list(_fc_lt_mo['lo90'])[::-1],
                        fill='toself', fillcolor='rgba(231,76,60,0.12)',
                        line=dict(color='rgba(255,255,255,0)'), name='90% 区间',
                    ))
                    _fc_fig_lt.add_trace(_pgo_fc.Scatter(
                        x=list(_fc_lt_mo['ym_str']), y=list(_fc_lt_mo['mean']),
                        name='月均预测价格', line=dict(color='#E74C3C', width=2.5),
                        mode='lines+markers',
                    ))
                    # Add vertical line at ARIMA/seasonal boundary.
                    # x-axis is categorical strings — add_vline needs the integer
                    # index, not the string itself (Plotly crashes on str+int arithmetic).
                    _fc_arima_cutoff_dt = (_fc_end_dt + _pd_fc.DateOffset(days=_fc_arima_steps)).strftime('%Y-%m')
                    _fc_ym_list = list(_fc_lt_mo['ym_str'])
                    _fc_cutoff_idx = next(
                        (i for i, s in enumerate(_fc_ym_list) if s >= _fc_arima_cutoff_dt),
                        len(_fc_ym_list) - 1,
                    )
                    _fc_fig_lt.add_vline(x=_fc_cutoff_idx, line_dash='dot', line_color='gray',
                                         annotation_text='ARIMA→季节均值')
                    _fc_fig_lt.update_layout(
                        title=f"{_fc_prov} — 长期月度均价预测（{_fc_end_dt.date()} ~ {_fc_future_dates[-1]}）",
                        xaxis_title="月份", yaxis_title="¥/kWh",
                        height=380, margin=dict(t=40, b=60, l=60, r=20),
                        xaxis_tickangle=-45,
                        legend=dict(orientation='h', y=-0.4),
                    )
                    st.plotly_chart(_fc_fig_lt, use_container_width=True)

                # PC score time series
                st.markdown("**主成分得分历史序列（最近60天）**")
                _fc_fig_scores = _pgo_fc.Figure()
                _fc_dates = _fc_price_mat.index[-60:]
                for _fc_i in range(_fc_n_pc):
                    _fc_fig_scores.add_trace(_pgo_fc.Scatter(
                        x=list(_fc_dates),
                        y=list(_fc_scores[-60:, _fc_i]),
                        name=f"PC{_fc_i+1}", line=dict(width=1.5),
                    ))
                _fc_fig_scores.update_layout(
                    xaxis_title="日期", yaxis_title="PC得分",
                    height=280, margin=dict(t=10, b=20, l=50, r=20),
                    legend=dict(orientation='h', y=-0.3),
                )
                st.plotly_chart(_fc_fig_scores, use_container_width=True)

            # ─────────────────────────────────────────────────────────────────
            # MODEL 2: Stack / Marginal Cost
            # ─────────────────────────────────────────────────────────────────
            with _fc_tab_stack:
                st.markdown("**边际成本（价格堆栈）模型**")
                st.caption(
                    "依据装机结构和发电边际成本构建供给侧报价堆栈，以残差负荷（负荷−可再生能源）"
                    "在堆栈曲线上的交叉点估算出清价格。"
                )
                _fc_s1, _fc_s2, _fc_s3 = st.columns(3)
                with _fc_s1:
                    _fc_coal_price = st.slider(
                        "动力煤价格 (元/吨SCE)", 500, 1200, 750, 50, key="fc_coal")
                with _fc_s2:
                    _fc_gas_markup = st.slider(
                        "气电溢价 (元/MWh)", 0, 200, 80, 20, key="fc_gas")
                with _fc_s3:
                    _fc_cap_premium = st.slider(
                        "尖峰容量溢价 (元/MWh)", 0, 300, 100, 25, key="fc_cap_prem")

                # Marginal costs (¥/MWh)
                _fc_mc_re      = 0.0          # wind, solar
                _fc_mc_nuclear = 25.0         # nuclear O&M only (fuel in fixed cost)
                _fc_mc_hydro   = 15.0         # water resource fee
                # Coal: heat rate ≈ 310 g SCE/kWh = 0.31 ton/MWh
                _fc_mc_coal_base = _fc_coal_price * 0.31 + 18.0  # fuel + variable O&M
                _fc_mc_coal_peak = _fc_mc_coal_base * 1.18       # peaking units, higher heat rate

                # Province installed capacity (use supply data already loaded in EOH tab)
                # Nuclear capacity override for provinces with known nuclear fleets (MW installed)
                _fc_nuclear_override = {
                    '江苏': 8500, '福建': 10000, '广东': 20000,
                    '浙江': 6600, '辽宁': 6700, '海南': 2200,
                    '山东': 2500, '广西': 2200,
                }
                _fc_sup_row = _sup_df[_sup_df['province'] == _fc_prov] if '_sup_df' in dir() else _pd_fc.DataFrame()
                if _fc_sup_row.empty:
                    st.warning("装机数据加载中 — 请先访问「供需结构」标签页。如已访问，重新点击「运行预测」。")
                    _fc_th_mw    = 30000.0
                    _fc_wind_mw  = 5000.0
                    _fc_solar_mw = 5000.0
                    _fc_hydro_mw = 0.0
                    _fc_nuc_mw   = float(_fc_nuclear_override.get(_fc_prov, 0))
                else:
                    _fc_th_mw    = float(_fc_sup_row['thermal_mw'].iloc[0])
                    _fc_wind_mw  = float(_fc_sup_row['wind_mw'].iloc[0]) if 'wind_mw' in _fc_sup_row.columns else 0.0
                    _fc_solar_mw = float(_fc_sup_row['solar_mw'].iloc[0]) if 'solar_mw' in _fc_sup_row.columns else 0.0
                    _fc_hydro_mw = float(_fc_sup_row['hydro_mw'].iloc[0]) if 'hydro_mw' in _fc_sup_row.columns else 0.0
                    _fc_nuc_db   = float(_fc_sup_row['nuclear_mw'].iloc[0]) if 'nuclear_mw' in _fc_sup_row.columns else 0.0
                    # Use GREATEST of DB value and known override (same logic as EOH tab)
                    _fc_nuc_mw   = max(_fc_nuc_db, float(_fc_nuclear_override.get(_fc_prov, 0)))

                # Merit order stack (cumulative MW → marginal cost)
                _fc_stack_blocks = [
                    ("风电",   _fc_wind_mw  * 0.28,  _fc_mc_re),       # avg capacity factor 28%
                    ("光伏",   _fc_solar_mw * 0.15,  _fc_mc_re),       # avg daytime CF 15%
                    ("核电",   _fc_nuc_mw   * 0.88,  _fc_mc_nuclear),  # 88% CF
                    ("水电",   _fc_hydro_mw * 0.40,  _fc_mc_hydro),    # dispatchable portion
                    ("基荷煤电", _fc_th_mw * 0.40,   _fc_mc_coal_base),
                    ("调峰煤电", _fc_th_mw * 0.30,   _fc_mc_coal_peak),
                    ("尖峰",   _fc_th_mw  * 0.10,   _fc_mc_coal_peak + _fc_cap_premium),
                ]
                # Draw merit order curve
                _fc_cum_mw = [0.0]
                _fc_mc_steps = []
                for _, _fc_cap, _fc_mc in _fc_stack_blocks:
                    _fc_cum_mw.append(_fc_cum_mw[-1])
                    _fc_cum_mw.append(_fc_cum_mw[-1] + _fc_cap)
                    _fc_mc_steps.append(_fc_mc)
                    _fc_mc_steps.append(_fc_mc)

                _fc_fig_stack = _pgo_fc.Figure()
                _fc_fig_stack.add_trace(_pgo_fc.Scatter(
                    x=_fc_cum_mw, y=_fc_mc_steps,
                    fill='tozeroy', fillcolor='rgba(52,152,219,0.2)',
                    line=dict(color='#2980B9', width=2),
                    name='供给曲线',
                ))
                # Plot hourly residual demand lines from fundamentals
                if not _fc_fund_df.empty:
                    _fc_peak_res  = max(0, _fc_fund_df['avg_load'].max()
                                        - _fc_fund_df['avg_wind'].min()
                                        - _fc_fund_df['avg_solar'].min())
                    _fc_off_res   = max(0, _fc_fund_df['avg_load'].min()
                                        - _fc_fund_df['avg_wind'].max()
                                        - _fc_fund_df['avg_solar'].max())
                    _fc_avg_res   = max(0, _fc_fund_df['avg_load'].mean()
                                        - _fc_fund_df['avg_wind'].mean()
                                        - _fc_fund_df['avg_solar'].mean())
                    for _fc_res_val, _fc_res_name, _fc_res_col in [
                        (_fc_peak_res,  "峰时残差负荷",  '#E74C3C'),
                        (_fc_avg_res,   "均值残差负荷",  '#E67E22'),
                        (_fc_off_res,   "谷时残差负荷",  '#27AE60'),
                    ]:
                        _fc_fig_stack.add_vline(
                            x=_fc_res_val, line_dash='dash', line_color=_fc_res_col,
                            annotation_text=_fc_res_name,
                            annotation_position="top right",
                        )
                _fc_fig_stack.update_layout(
                    title=f"{_fc_prov} — 供给堆栈曲线与残差负荷",
                    xaxis_title="累计装机容量 (MW)",
                    yaxis_title="边际成本 (¥/MWh)",
                    height=380, margin=dict(t=40, b=20, l=60, r=20),
                )
                st.plotly_chart(_fc_fig_stack, use_container_width=True)

                # Estimate hourly clearing price from fundamentals
                if not _fc_fund_df.empty:
                    def _fc_clearing_price(residual_mw):
                        _fc_cum = 0.0
                        for _, _fc_cap, _fc_mc in _fc_stack_blocks:
                            _fc_cum += _fc_cap
                            if residual_mw <= _fc_cum:
                                return _fc_mc
                        return _fc_mc_coal_peak + _fc_cap_premium

                    _fc_fund_df2 = _fc_fund_df.copy()
                    _fc_fund_df2['residual_mw'] = (
                        _fc_fund_df2['avg_load']
                        - _fc_fund_df2['avg_wind']
                        - _fc_fund_df2['avg_solar']
                    ).clip(lower=0)
                    _fc_fund_df2['stack_price_rmb_mwh'] = _fc_fund_df2['residual_mw'].apply(_fc_clearing_price)
                    _fc_fund_df2['stack_price_yuan_kwh'] = _fc_fund_df2['stack_price_rmb_mwh'] / 1000.0

                    _fc_fig_hourly_stack = _pgo_fc.Figure()
                    _fc_fig_hourly_stack.add_trace(_pgo_fc.Scatter(
                        x=_fc_fund_df2['hour'].tolist(),
                        y=_fc_fund_df2['stack_price_yuan_kwh'].tolist(),
                        name='边际成本预测价格', line=dict(color='#E74C3C', width=2),
                    ))
                    if not _fc_all_hourly.empty:
                        _fc_fig_hourly_stack.add_trace(_pgo_fc.Scatter(
                            x=_fc_all_hourly['hour'].tolist(),
                            y=_fc_all_hourly['avg_price'].tolist(),
                            name='历史均值(训练期)', line=dict(dash='dot', color='#7F8C8D', width=1.5),
                        ))
                    _fc_fig_hourly_stack.update_layout(
                        title=f"{_fc_prov} — 边际成本模型预测日内价格曲线（¥/kWh）",
                        xaxis=dict(title="时段", tickvals=list(range(0, 24, 4)),
                                   ticktext=[f"{h:02d}:00" for h in range(0, 24, 4)]),
                        yaxis_title="¥/kWh", height=320,
                        margin=dict(t=40, b=20, l=60, r=20),
                        legend=dict(orientation='h', y=-0.3),
                    )
                    st.plotly_chart(_fc_fig_hourly_stack, use_container_width=True)

                    # Summary stats
                    st.markdown("**边际成本模型预测摘要**")
                    _fc_stack_cols = st.columns(4)
                    _fc_stack_cols[0].metric("峰时预测价格",
                        f"{_fc_fund_df2['stack_price_yuan_kwh'].max():.4f} ¥/kWh")
                    _fc_stack_cols[1].metric("谷时预测价格",
                        f"{_fc_fund_df2['stack_price_yuan_kwh'].min():.4f} ¥/kWh")
                    _fc_stack_cols[2].metric("日均预测价格",
                        f"{_fc_fund_df2['stack_price_yuan_kwh'].mean():.4f} ¥/kWh")
                    _fc_stack_cols[3].metric("峰谷价差",
                        f"{((_fc_fund_df2['stack_price_yuan_kwh'].max() - _fc_fund_df2['stack_price_yuan_kwh'].min()) * 1000):.1f} ¥/MWh")
                else:
                    st.info(f"{_fc_prov} 暂无小时基本面数据，无法推算时段残差负荷。")

                # Marginal cost assumptions table
                with st.expander("边际成本假设"):
                    st.table(_pd_fc.DataFrame([
                        {"电源类型": "风电/光伏", "边际成本(¥/MWh)": f"{_fc_mc_re:.0f}",
                         "说明": "燃料为零，仅运行维护"},
                        {"电源类型": "核电", "边际成本(¥/MWh)": f"{_fc_mc_nuclear:.0f}",
                         "说明": "固定成本已摊销，可变运行约25元"},
                        {"电源类型": "水电", "边际成本(¥/MWh)": f"{_fc_mc_hydro:.0f}",
                         "说明": "水资源费+运行维护"},
                        {"电源类型": "基荷煤电", "边际成本(¥/MWh)": f"{_fc_mc_coal_base:.0f}",
                         "说明": f"煤价{_fc_coal_price}元/吨×0.31 ton/MWh + O&M"},
                        {"电源类型": "调峰煤电", "边际成本(¥/MWh)": f"{_fc_mc_coal_peak:.0f}",
                         "说明": "调峰机组效率低，热耗率+18%"},
                        {"电源类型": "尖峰", "边际成本(¥/MWh)": f"{_fc_mc_coal_peak + _fc_cap_premium:.0f}",
                         "说明": f"调峰煤电 + 尖峰容量溢价{_fc_cap_premium}元/MWh"},
                    ]))

            # ─────────────────────────────────────────────────────────────────
            # MODEL 3: Bayesian
            # ─────────────────────────────────────────────────────────────────
            with _fc_tab_bayes:
                st.markdown("**贝叶斯分布模型**")
                st.caption(
                    "以训练期小时价格分布作为先验（Prior），以近30天为似然（Likelihood）进行贝叶斯更新，"
                    "输出各时段价格的后验分布（均值 ± 置信区间）。"
                )

                # Load per-date/hour granularity for Bayesian update
                _fc_bayes_sql = f"""
                    SELECT EXTRACT(hour FROM datetime)::int AS hour,
                           {_fc_pcol} AS price
                    FROM marketdata.spot_prices_hourly
                    WHERE province = %s AND datetime BETWEEN %s AND %s
                      AND {_fc_pcol} IS NOT NULL
                """
                with st.spinner("加载小时价格分布…"):
                    _fc_prior_raw  = _pd_fc.read_sql(_fc_bayes_sql, _conn(),
                                                      params=[_fc_prov, _fc_start, _fc_end])
                    _fc_recent_raw = _pd_fc.read_sql(_fc_bayes_sql, _conn(),
                                                      params=[_fc_prov, _fc_recent_start, _fc_end])

                # Normalise to ¥/kWh (same logic as training matrix)
                if not _fc_prior_raw.empty and float(_fc_prior_raw['price'].median()) > 5:
                    _fc_prior_raw  = _fc_prior_raw.copy();  _fc_prior_raw['price']  /= 1000.0
                    _fc_recent_raw = _fc_recent_raw.copy(); _fc_recent_raw['price'] /= 1000.0

                _fc_post_mean = _np_fc.zeros(24)
                _fc_post_lo90 = _np_fc.zeros(24)
                _fc_post_hi90 = _np_fc.zeros(24)
                _fc_prior_mu  = _np_fc.zeros(24)
                _fc_prior_sig = _np_fc.zeros(24)

                for _fc_h in range(24):
                    _fc_p_h = _fc_prior_raw[_fc_prior_raw['hour'] == _fc_h]['price'].dropna()
                    _fc_r_h = _fc_recent_raw[_fc_recent_raw['hour'] == _fc_h]['price'].dropna()

                    if len(_fc_p_h) < 2:
                        continue
                    _fc_mu0 = float(_fc_p_h.mean())
                    _fc_sig0 = max(float(_fc_p_h.std()), 1e-6)
                    _fc_n0   = len(_fc_p_h)
                    _fc_prior_mu[_fc_h]  = _fc_mu0
                    _fc_prior_sig[_fc_h] = _fc_sig0

                    if len(_fc_r_h) >= 2:
                        _fc_mu1  = float(_fc_r_h.mean())
                        _fc_sig1 = max(float(_fc_r_h.std()), 1e-6)
                        _fc_n1   = len(_fc_r_h)
                        # Conjugate Gaussian update (known variance)
                        _fc_post_prec = _fc_n0 / _fc_sig0**2 + _fc_n1 / _fc_sig1**2
                        _fc_post_mu   = (_fc_n0 * _fc_mu0 / _fc_sig0**2
                                         + _fc_n1 * _fc_mu1 / _fc_sig1**2) / _fc_post_prec
                        _fc_post_std  = 1.0 / _np_fc.sqrt(_fc_post_prec)
                    else:
                        _fc_post_mu  = _fc_mu0
                        _fc_post_std = _fc_sig0

                    _fc_post_mean[_fc_h] = _fc_post_mu
                    _fc_post_lo90[_fc_h] = _fc_post_mu - 1.645 * _fc_post_std
                    _fc_post_hi90[_fc_h] = _fc_post_mu + 1.645 * _fc_post_std

                _fc_hours = list(range(24))
                _fc_fig_bayes = _pgo_fc.Figure()
                # 90% CI band
                _fc_fig_bayes.add_trace(_pgo_fc.Scatter(
                    x=_fc_hours + _fc_hours[::-1],
                    y=list(_fc_post_hi90) + list(_fc_post_lo90[::-1]),
                    fill='toself', fillcolor='rgba(231,76,60,0.15)',
                    line=dict(color='rgba(255,255,255,0)'),
                    name='90% 置信区间',
                ))
                # Posterior mean
                _fc_fig_bayes.add_trace(_pgo_fc.Scatter(
                    x=_fc_hours, y=list(_fc_post_mean),
                    name='后验均值', line=dict(color='#E74C3C', width=2.5),
                ))
                # Prior mean
                _fc_fig_bayes.add_trace(_pgo_fc.Scatter(
                    x=_fc_hours, y=list(_fc_prior_mu),
                    name=f'先验均值（训练期{_fc_train_months}个月）',
                    line=dict(dash='dot', color='#7F8C8D', width=1.5),
                ))
                if not _fc_recent_hourly.empty:
                    _fc_fig_bayes.add_trace(_pgo_fc.Scatter(
                        x=_fc_recent_hourly['hour'].tolist(),
                        y=_fc_recent_hourly['avg_price'].tolist(),
                        name='近30天均值（似然）',
                        line=dict(dash='dash', color='#2980B9', width=1.5),
                    ))
                _fc_fig_bayes.update_layout(
                    title=f"{_fc_prov} — 贝叶斯后验价格分布（¥/kWh，90%置信区间）",
                    xaxis=dict(title="时段", tickvals=list(range(0, 24, 4)),
                               ticktext=[f"{h:02d}:00" for h in range(0, 24, 4)]),
                    yaxis_title="¥/kWh", height=400,
                    margin=dict(t=40, b=20, l=60, r=20),
                    legend=dict(orientation='h', y=-0.3),
                )
                st.plotly_chart(_fc_fig_bayes, use_container_width=True)

                # Price distribution violin for peak/off-peak hours
                st.markdown("**典型时段价格分布（先验 vs 近30天）**")
                _fc_b1, _fc_b2 = st.columns(2)
                for _fc_col_widget, _fc_h_label, _fc_h_range in [
                    (_fc_b1, "峰时 (08-12 & 18-22)", list(range(8, 12)) + list(range(18, 22))),
                    (_fc_b2, "谷时 (00-06)",          list(range(0, 6))),
                ]:
                    _fc_p_sel  = _fc_prior_raw[_fc_prior_raw['hour'].isin(_fc_h_range)]['price'].dropna()
                    _fc_r_sel  = _fc_recent_raw[_fc_recent_raw['hour'].isin(_fc_h_range)]['price'].dropna()
                    _fc_fig_vio = _pgo_fc.Figure()
                    if len(_fc_p_sel) > 0:
                        _fc_fig_vio.add_trace(_pgo_fc.Violin(
                            y=_fc_p_sel.tolist(), name=f'先验({_fc_train_months}m)',
                            box_visible=True, meanline_visible=True,
                            fillcolor='rgba(127,140,141,0.4)', line_color='#7F8C8D',
                        ))
                    if len(_fc_r_sel) > 0:
                        _fc_fig_vio.add_trace(_pgo_fc.Violin(
                            y=_fc_r_sel.tolist(), name='近30天',
                            box_visible=True, meanline_visible=True,
                            fillcolor='rgba(41,128,185,0.4)', line_color='#2980B9',
                        ))
                    _fc_fig_vio.update_layout(
                        title=_fc_h_label, yaxis_title="¥/kWh",
                        height=280, margin=dict(t=40, b=10, l=50, r=10),
                        showlegend=True,
                    )
                    with _fc_col_widget:
                        st.plotly_chart(_fc_fig_vio, use_container_width=True)

            # ─────────────────────────────────────────────────────────────────
            # ENSEMBLE
            # ─────────────────────────────────────────────────────────────────
            with _fc_tab_ensemble:
                st.markdown("**综合预测 — 三模型加权集成**")
                st.caption("PCA × 权重₁ + 边际成本 × 权重₂ + 贝叶斯 × 权重₃，调整各模型权重以反映对不同信号的置信度。")

                _fc_e1, _fc_e2, _fc_e3 = st.columns(3)
                with _fc_e1:
                    _fc_w_pca   = st.slider("PCA权重",   0.0, 1.0, 0.40, 0.05, key="fc_w_pca")
                with _fc_e2:
                    _fc_w_stack = st.slider("边际成本权重", 0.0, 1.0, 0.30, 0.05, key="fc_w_stack")
                with _fc_e3:
                    _fc_w_bayes = st.slider("贝叶斯权重", 0.0, 1.0, 0.30, 0.05, key="fc_w_bayes")
                _fc_w_total = _fc_w_pca + _fc_w_stack + _fc_w_bayes
                if abs(_fc_w_total - 1.0) > 0.01:
                    st.warning(f"权重之和 = {_fc_w_total:.2f}，建议调整至1.0。当前将自动归一化。")
                _fc_w_total = max(_fc_w_total, 1e-6)

                _fc_hours = list(range(24))
                # PCA D+1 forecast (first day)
                _fc_pca_d1 = _fc_pca_pred[0, :]  # (24,), already in ¥/kWh

                # Stack price (24h)
                if not _fc_fund_df.empty and 'stack_price_yuan_kwh' in _fc_fund_df2.columns:
                    _fc_stack_h24 = _np_fc.array([
                        float(_fc_fund_df2[_fc_fund_df2['hour'] == h]['stack_price_yuan_kwh'].iloc[0])
                        if not _fc_fund_df2[_fc_fund_df2['hour'] == h].empty else float(_fc_post_mean[h])
                        for h in _fc_hours
                    ])
                else:
                    # Fallback: use historical mean
                    _fc_stack_h24 = _np_fc.array([
                        float(_fc_all_hourly[_fc_all_hourly['hour'] == h]['avg_price'].iloc[0])
                        if not _fc_all_hourly[_fc_all_hourly['hour'] == h].empty else 0.3
                        for h in _fc_hours
                    ])

                # Ensemble (weighted)
                _fc_ensemble = (
                    _fc_w_pca   / _fc_w_total * _fc_pca_d1
                    + _fc_w_stack / _fc_w_total * _fc_stack_h24
                    + _fc_w_bayes / _fc_w_total * _fc_post_mean
                )
                # Uncertainty: weighted CI from PCA and Bayesian
                _fc_ens_lo = (
                    _fc_w_pca   / _fc_w_total * _fc_pca_pred_lo[0, :]
                    + _fc_w_stack / _fc_w_total * _fc_stack_h24 * 0.93
                    + _fc_w_bayes / _fc_w_total * _fc_post_lo90
                )
                _fc_ens_hi = (
                    _fc_w_pca   / _fc_w_total * _fc_pca_pred_hi[0, :]
                    + _fc_w_stack / _fc_w_total * _fc_stack_h24 * 1.07
                    + _fc_w_bayes / _fc_w_total * _fc_post_hi90
                )

                _fc_fig_ens = _pgo_fc.Figure()
                _fc_fig_ens.add_trace(_pgo_fc.Scatter(
                    x=_fc_hours + _fc_hours[::-1],
                    y=list(_fc_ens_hi) + list(_fc_ens_lo[::-1]),
                    fill='toself', fillcolor='rgba(231,76,60,0.15)',
                    line=dict(color='rgba(255,255,255,0)'),
                    name='90% 置信区间',
                ))
                _fc_fig_ens.add_trace(_pgo_fc.Scatter(
                    x=_fc_hours, y=list(_fc_ensemble),
                    name='综合预测', line=dict(color='#E74C3C', width=3),
                ))
                _fc_fig_ens.add_trace(_pgo_fc.Scatter(
                    x=_fc_hours, y=list(_fc_pca_d1),
                    name=f'PCA(w={_fc_w_pca:.2f})', line=dict(dash='dot', color='#9B59B6', width=1.5),
                ))
                _fc_fig_ens.add_trace(_pgo_fc.Scatter(
                    x=_fc_hours, y=list(_fc_stack_h24),
                    name=f'边际成本(w={_fc_w_stack:.2f})', line=dict(dash='dot', color='#2980B9', width=1.5),
                ))
                _fc_fig_ens.add_trace(_pgo_fc.Scatter(
                    x=_fc_hours, y=list(_fc_post_mean),
                    name=f'贝叶斯(w={_fc_w_bayes:.2f})', line=dict(dash='dot', color='#27AE60', width=1.5),
                ))
                if not _fc_all_hourly.empty:
                    _fc_fig_ens.add_trace(_pgo_fc.Scatter(
                        x=_fc_all_hourly['hour'].tolist(),
                        y=_fc_all_hourly['avg_price'].tolist(),
                        name='历史均值', line=dict(dash='dash', color='#7F8C8D', width=1),
                    ))
                _fc_fig_ens.update_layout(
                    title=f"{_fc_prov} — D+1 综合价格预测（¥/kWh）",
                    xaxis=dict(title="时段", tickvals=list(range(0, 24, 4)),
                               ticktext=[f"{h:02d}:00" for h in range(0, 24, 4)]),
                    yaxis_title="¥/kWh", height=440,
                    margin=dict(t=40, b=20, l=60, r=20),
                    legend=dict(orientation='h', y=-0.32),
                )
                st.plotly_chart(_fc_fig_ens, use_container_width=True)

                # Summary metrics
                st.markdown("**综合预测关键指标**")
                _fc_m1, _fc_m2, _fc_m3, _fc_m4, _fc_m5 = st.columns(5)
                _fc_m1.metric("日均预测",   f"{_fc_ensemble.mean():.4f} ¥/kWh")
                _fc_m2.metric("峰时均值",   f"{_fc_ensemble[list(range(8,12))+list(range(18,22))].mean():.4f} ¥/kWh")
                _fc_m3.metric("谷时均值",   f"{_fc_ensemble[list(range(0,6))].mean():.4f} ¥/kWh")
                _fc_m4.metric("峰谷价差",   f"{((_fc_ens_hi.max() - _fc_ens_lo.min()) * 1000):.1f} ¥/MWh")
                _fc_m5.metric("预测不确定性(σ)",
                              f"{((_fc_ens_hi - _fc_ens_lo).mean() / 2 / 1.645 * 1000):.1f} ¥/MWh")

                # Downloadable forecast table
                _fc_out_df = _pd_fc.DataFrame({
                    '时段': [f"{h:02d}:00" for h in _fc_hours],
                    'PCA预测(¥/kWh)':     [round(float(v), 4) for v in _fc_pca_d1],
                    '边际成本(¥/kWh)':   [round(float(v), 4) for v in _fc_stack_h24],
                    '贝叶斯均值(¥/kWh)': [round(float(v), 4) for v in _fc_post_mean],
                    '综合预测(¥/kWh)':   [round(float(v), 4) for v in _fc_ensemble],
                    '90%区间下限(¥/kWh)':[round(float(v), 4) for v in _fc_ens_lo],
                    '90%区间上限(¥/kWh)':[round(float(v), 4) for v in _fc_ens_hi],
                })
                st.dataframe(_fc_out_df, hide_index=True, use_container_width=True)
                st.download_button(
                    "下载预测结果 CSV",
                    _fc_out_df.to_csv(index=False, encoding='utf-8-sig'),
                    file_name=f"price_forecast_{_fc_prov}_D+1.csv",
                    mime="text/csv",
                    key="fc_download",
                )

            # ─────────────────────────────────────────────────────────────────
            # BACKTEST — evaluate model accuracy on last 14-day holdout
            # ─────────────────────────────────────────────────────────────────
            st.divider()
            st.markdown("### 模型回测验证（样本外14天）")
            st.caption(
                "以训练期之前14天作为验证集（holdout），用相同训练数据拟合的贝叶斯后验均值 "
                "与 PCA 历史均值和 Bayesian 后验均值进行比较，计算 MAE / RMSE / MAPE。"
            )

            _fc_holdout_end   = _fc_start_dt - _pd_fc.DateOffset(days=1)
            _fc_holdout_start = _fc_holdout_end - _pd_fc.DateOffset(days=13)
            with st.spinner("加载回测数据…"):
                _fc_holdout_df = _load_price_holdout(
                    _conn, _fc_prov,
                    str(_fc_holdout_start.date()), str(_fc_holdout_end.date()),
                    _fc_pcol,
                )

            if _fc_holdout_df.empty:
                st.caption("无法获取验证期数据（可能超出历史范围）。")
            else:
                # Group holdout by hour, compute actual average per hour
                _fc_ho_hourly = _fc_holdout_df.groupby('hour')['actual_price'].mean().reset_index()
                _fc_ho_mean   = _np_fc.zeros(24)
                for _, _hrow in _fc_ho_hourly.iterrows():
                    _fc_ho_mean[int(_hrow['hour'])] = float(_hrow['actual_price'])
                # Normalise holdout to ¥/kWh (holdout period may have been ingested
                # in a different unit than the training period)
                if _fc_ho_mean.max() > 5:
                    _fc_ho_mean /= 1000.0

                # Treat holdout as invalid if max price < 0.05 ¥/kWh (= 50 ¥/MWh).
                # Old ingestion pipeline stored tiny non-zero garbage values;
                # the > 0 SQL filter is not sufficient to exclude them.
                _fc_ho_valid = _fc_ho_mean.max() >= 0.05

                # PCA backtest: use historical mean from training window
                _fc_bt_pca    = _fc_mean_24h   # training period hourly mean, ¥/kWh
                _fc_bt_bayes  = _fc_post_mean  # Bayesian posterior mean
                _fc_bt_ens    = (_fc_w_pca / _fc_w_total * _fc_bt_pca
                                 + _fc_w_bayes / _fc_w_total * _fc_bt_bayes
                                 + _fc_w_stack / _fc_w_total * _fc_stack_h24)

                def _fc_metrics(pred, actual):
                    mask = actual > 0
                    if mask.sum() == 0:
                        return 0, 0, 0
                    mae  = _np_fc.abs(pred[mask] - actual[mask]).mean()
                    rmse = _np_fc.sqrt(((pred[mask] - actual[mask])**2).mean())
                    mape = (_np_fc.abs((pred[mask] - actual[mask]) / actual[mask])).mean() * 100
                    return float(mae), float(rmse), float(mape)

                if not _fc_ho_valid:
                    st.caption("验证期无有效价格数据（数据库存储为零值占位），跳过回测。")
                else:
                    _fc_bt_cols = st.columns(3)
                    for _fc_bt_lbl, _fc_bt_pred, _fc_bt_col in [
                        ("历史均值(PCA基准)", _fc_bt_pca,   _fc_bt_cols[0]),
                        ("贝叶斯后验均值",   _fc_bt_bayes, _fc_bt_cols[1]),
                        ("综合预测",         _fc_bt_ens,   _fc_bt_cols[2]),
                    ]:
                        _fc_mae, _fc_rmse, _fc_mape = _fc_metrics(_fc_bt_pred, _fc_ho_mean)
                        with _fc_bt_col:
                            st.markdown(f"**{_fc_bt_lbl}**")
                            st.metric("MAE",  f"{_fc_mae*1000:.2f} ¥/MWh")
                            st.metric("RMSE", f"{_fc_rmse*1000:.2f} ¥/MWh")
                            st.metric("MAPE", f"{_fc_mape:.1f}%")

                    _fc_fig_bt = _pgo_fc.Figure()
                    _fc_fig_bt.add_trace(_pgo_fc.Scatter(
                        x=_fc_hours, y=list(_fc_ho_mean),
                        name="验证期实际均值", line=dict(color='#2C3E50', width=2.5),
                    ))
                    _fc_fig_bt.add_trace(_pgo_fc.Scatter(
                        x=_fc_hours, y=list(_fc_bt_bayes),
                        name="贝叶斯后验均值", line=dict(color='#E74C3C', width=2, dash='dot'),
                    ))
                    _fc_fig_bt.add_trace(_pgo_fc.Scatter(
                        x=_fc_hours, y=list(_fc_bt_pca),
                        name="PCA历史均值", line=dict(color='#9B59B6', width=2, dash='dot'),
                    ))
                    _fc_fig_bt.update_layout(
                        title=f"{_fc_prov} — 模型回测（验证期 {_fc_holdout_start.date()} ~ {_fc_holdout_end.date()}）",
                        xaxis=dict(title="时段", tickvals=list(range(0, 24, 4)),
                                   ticktext=[f"{h:02d}:00" for h in range(0, 24, 4)]),
                        yaxis_title="¥/kWh", height=320,
                        margin=dict(t=40, b=20, l=60, r=20),
                        legend=dict(orientation='h', y=-0.3),
                    )
                    st.plotly_chart(_fc_fig_bt, use_container_width=True)

            # ─────────────────────────────────────────────────────────────────
            # MONTHLY AGGREGATE VIEW (horizon >= 30)
            # ─────────────────────────────────────────────────────────────────
            if _fc_horizon >= 30:
                st.divider()
                _fc_mo_title = "### M+1 月度价格分布预测" if _fc_horizon <= 31 else f"### 月度价格分布预测（未来{_fc_horizon}天）"
                st.markdown(_fc_mo_title)
                st.caption(
                    "基于贝叶斯后验分布，展示下月各时段价格的期望区间。"
                    "箱线图为后验分布（均值 ± 1σ / 2σ）在月内的变化范围。"
                )
                # Seasonal adjustment: split training data by month to estimate seasonal σ
                _fc_monthly_sql = f"""
                    SELECT EXTRACT(month FROM datetime)::int AS month,
                           EXTRACT(hour  FROM datetime)::int AS hour,
                           AVG({_fc_pcol}) AS avg_price,
                           STDDEV({_fc_pcol}) AS std_price
                    FROM marketdata.spot_prices_hourly
                    WHERE province = %s AND datetime BETWEEN %s AND %s
                      AND {_fc_pcol} IS NOT NULL
                    GROUP BY month, hour
                    ORDER BY month, hour
                """
                with st.spinner("加载月度分布数据…"):
                    _fc_monthly_df = _pd_fc.read_sql(
                        _fc_monthly_sql, _conn(),
                        params=[_fc_prov, _fc_start, _fc_end]
                    )

                # Next calendar month
                _fc_next_month = (_fc_end_dt + _pd_fc.DateOffset(months=1)).month
                _fc_month_data = _fc_monthly_df[_fc_monthly_df['month'] == _fc_next_month]
                if _fc_month_data.empty:
                    # Fallback: use all months average
                    _fc_month_data = _fc_monthly_df.groupby('hour').agg(
                        avg_price=('avg_price', 'mean'),
                        std_price=('std_price', 'mean')
                    ).reset_index()
                    _fc_month_data['month'] = _fc_next_month

                _fc_month_data = _fc_month_data.sort_values('hour')
                _fc_m_hours = _fc_month_data['hour'].tolist()
                _fc_m_mean  = _fc_month_data['avg_price'].fillna(0).tolist()
                _fc_m_std   = _fc_month_data['std_price'].fillna(0.02).tolist()

                _fc_fig_monthly = _pgo_fc.Figure()
                _fc_m_mean_arr = _np_fc.array(_fc_m_mean)
                _fc_m_std_arr  = _np_fc.array(_fc_m_std)
                # 2σ band
                _fc_fig_monthly.add_trace(_pgo_fc.Scatter(
                    x=_fc_m_hours + _fc_m_hours[::-1],
                    y=list(_fc_m_mean_arr + 2*_fc_m_std_arr) + list((_fc_m_mean_arr - 2*_fc_m_std_arr)[::-1]),
                    fill='toself', fillcolor='rgba(231,76,60,0.08)',
                    line=dict(color='rgba(255,255,255,0)'), name='±2σ (95%)',
                ))
                # 1σ band
                _fc_fig_monthly.add_trace(_pgo_fc.Scatter(
                    x=_fc_m_hours + _fc_m_hours[::-1],
                    y=list(_fc_m_mean_arr + _fc_m_std_arr) + list((_fc_m_mean_arr - _fc_m_std_arr)[::-1]),
                    fill='toself', fillcolor='rgba(231,76,60,0.20)',
                    line=dict(color='rgba(255,255,255,0)'), name='±1σ (68%)',
                ))
                _fc_fig_monthly.add_trace(_pgo_fc.Scatter(
                    x=_fc_m_hours, y=_fc_m_mean,
                    name=f'{_fc_next_month}月 历史同期均值',
                    line=dict(color='#E74C3C', width=2.5),
                ))
                # Overlay Bayesian posterior mean
                _fc_fig_monthly.add_trace(_pgo_fc.Scatter(
                    x=list(range(24)), y=list(_fc_post_mean),
                    name='贝叶斯后验均值(综合调整)',
                    line=dict(color='#2980B9', width=2, dash='dot'),
                ))
                _fc_fig_monthly.update_layout(
                    title=f"{_fc_prov} — {_fc_next_month}月 日前价格分布预测（历史同期 ± σ）",
                    xaxis=dict(title="时段", tickvals=list(range(0, 24, 4)),
                               ticktext=[f"{h:02d}:00" for h in range(0, 24, 4)]),
                    yaxis_title="¥/kWh", height=380,
                    margin=dict(t=40, b=20, l=60, r=20),
                    legend=dict(orientation='h', y=-0.3),
                )
                st.plotly_chart(_fc_fig_monthly, use_container_width=True)

                # Monthly summary table
                _fc_monthly_summary = _pd_fc.DataFrame({
                    '指标': ['月均价 (¥/kWh)', '峰时均价 (¥/kWh)', '谷时均价 (¥/kWh)',
                             '峰谷差 (¥/MWh)', '价格波动率 (1σ, ¥/MWh)'],
                    '预测值': [
                        f"{_fc_m_mean_arr.mean():.4f}",
                        f"{_fc_m_mean_arr[list(range(8,12))+list(range(18,22))].mean():.4f}",
                        f"{_fc_m_mean_arr[list(range(0,6))].mean():.4f}",
                        f"{(_fc_m_mean_arr.max() - _fc_m_mean_arr.min()) * 1000:.1f}",
                        f"{_fc_m_std_arr.mean() * 1000:.1f}",
                    ],
                })
                st.dataframe(_fc_monthly_summary, hide_index=True, use_container_width=False)


# ── Tab 9: Data Management ────────────────────────────────────────────────────
with tab_mgmt:
    # ── Layout ────────────────────────────────────────────────────────────────
    st.subheader(_t("data_mgmt_title"))

    col_yr, _, _ = st.columns([1, 2, 1])
    with col_yr:
        sel_year = st.selectbox(_t("report_year"), [2026, 2025, 2024], key="mgmt_year")

    c_left, c_right = st.columns([2, 1])

    with c_left:
        mgmt_mode = st.radio(
            _t("mode_label"),
            [_t("mode_fill_gaps"), _t("mode_backfill")],
            horizontal=False,
            key="mgmt_mode",
        )
        st.caption(_t("additional_steps"))
        run_interprov = st.checkbox(
            _t("chk_interprov"),
            value=True,
            key="mgmt_interprov",
            help=_t("chk_interprov_help"),
        )
        run_ai = st.checkbox(
            _t("chk_ai"),
            value=False,
            key="mgmt_ai",
            help=_t("chk_ai_help"),
        )

    with c_right:
        _yr_end = date(sel_year, 12, 31) if sel_year < date.today().year else date.today() - timedelta(days=1)
        bf_start = st.date_input(_t("start_date"), date(sel_year, 1, 1), key=f"bf_start_{sel_year}")
        bf_end   = st.date_input(_t("end_date"),   _yr_end,              key=f"bf_end_{sel_year}")

    # ── S3 uploader (AWS only — when no local data folder) ────────────────────
    _local_data = _REPO / "data" / "spot reports" / str(sel_year)
    if not _local_data.exists():
        with st.expander(_t("upload_pdf"), expanded=True):
            st.caption(_t("upload_help"))
            _uploaded = st.file_uploader(
                _t("upload_pdf"),
                type=["pdf"],
                accept_multiple_files=True,
                key=f"mgmt_upload_{sel_year}",
                label_visibility="collapsed",
            )
            if _uploaded:
                if st.button(
                    _t("upload_btn", n=len(_uploaded)),
                    key=f"mgmt_do_upload_{sel_year}",
                    type="primary",
                ):
                    import boto3 as _boto3
                    _s3 = _boto3.client("s3")
                    for _uf in _uploaded:
                        _s3.put_object(
                            Bucket=_S3_BUCKET,
                            Key=f"{_S3_PREFIX}/{sel_year}/{_uf.name}",
                            Body=_uf.read(),
                        )
                    st.success(_t("upload_success", n=len(_uploaded)))
                    _scan_pdf_inventory.clear()
                    st.session_state.pop(f"mgmt_upload_{sel_year}", None)
                    st.rerun()

    st.divider()

    # ── DA / RT Price Data Export ──────────────────────────────────────────────
    with st.expander(f"📥 {_t('export_title')}", expanded=True):
        st.caption(_t("export_caption"))
        _ex_col1, _ex_col2 = st.columns([1, 1])
        _ex_start = _ex_col1.date_input("开始日期 Start", date(sel_year, 1, 1), key="export_start")
        _ex_end   = _ex_col2.date_input("结束日期 End",   min(date(sel_year, 12, 31), date.today()), key="export_end")

        @st.cache_data(ttl=300, show_spinner=False)
        def _build_export_xlsx(_start: str, _end: str) -> bytes:
            import io
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            df_exp = pd.read_sql(
                """SELECT report_date, province_cn, da_avg, rt_avg
                   FROM spot_daily
                   WHERE report_date BETWEEN %s AND %s AND da_avg IS NOT NULL
                   ORDER BY report_date, province_cn""",
                _conn(),
                params=(_start, _end),
            )

            if df_exp.empty:
                # return empty workbook
                wb = openpyxl.Workbook()
                wb.active.title = "无数据"
                buf = io.BytesIO()
                wb.save(buf)
                return buf.getvalue()

            # Pivot: rows = date, columns = province_cn
            pivot_da = df_exp.pivot_table(index="report_date", columns="province_cn", values="da_avg")
            pivot_rt = df_exp.pivot_table(index="report_date", columns="province_cn", values="rt_avg")

            # Province column order: sort by name
            provinces = sorted(set(df_exp["province_cn"].dropna()))
            pivot_da = pivot_da.reindex(columns=provinces)
            pivot_rt = pivot_rt.reindex(columns=provinces)

            # Helper: write one sheet
            header_fill = PatternFill("solid", fgColor="1F4E79")
            header_font = Font(bold=True, color="FFFFFF")
            date_fill   = PatternFill("solid", fgColor="D6E4F0")
            thin        = Side(style="thin", color="CCCCCC")
            border      = Border(left=thin, right=thin, top=thin, bottom=thin)

            def _write_sheet(ws, pivot, sheet_title: str, unit_label: str):
                ws.title = sheet_title
                cols = list(pivot.columns)
                # Row 1: title
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols) + 1)
                title_cell = ws.cell(1, 1, f"{sheet_title}（{unit_label}）")
                title_cell.font = Font(bold=True, size=13)
                title_cell.alignment = Alignment(horizontal="center")
                # Row 2: headers
                ws.cell(2, 1, "日期").font = header_font
                ws.cell(2, 1).fill = header_fill
                ws.cell(2, 1).alignment = Alignment(horizontal="center")
                for ci, prov in enumerate(cols, start=2):
                    c = ws.cell(2, ci, prov)
                    c.font = header_font
                    c.fill = header_fill
                    c.alignment = Alignment(horizontal="center")
                # Data rows
                for ri, (dt_idx, row) in enumerate(pivot.iterrows(), start=3):
                    dt_str = dt_idx.strftime("%Y-%m-%d") if hasattr(dt_idx, "strftime") else str(dt_idx)
                    dc = ws.cell(ri, 1, dt_str)
                    dc.fill = date_fill
                    dc.font = Font(bold=True)
                    dc.alignment = Alignment(horizontal="center")
                    for ci, prov in enumerate(cols, start=2):
                        val = row.get(prov)
                        c = ws.cell(ri, ci)
                        if val is not None and not (isinstance(val, float) and val != val):
                            c.value = round(float(val), 4)
                            c.number_format = "0.0000"
                        c.border = border
                        c.alignment = Alignment(horizontal="right")
                # Column widths
                ws.column_dimensions["A"].width = 13
                for ci in range(2, len(cols) + 2):
                    ws.column_dimensions[get_column_letter(ci)].width = 9
                # Freeze header row
                ws.freeze_panes = "B3"

            wb = openpyxl.Workbook()
            _write_sheet(wb.active,    pivot_da, "日前均价",  "元/千瓦时")
            _write_sheet(wb.create_sheet(), pivot_rt, "实时均价", "元/千瓦时")

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        if st.button("生成 Excel / Generate", key="export_generate"):
            with st.spinner("查询数据库…"):
                try:
                    _xlsx_bytes = _build_export_xlsx(str(_ex_start), str(_ex_end))
                    st.download_button(
                        label=f"📥 {_t('export_btn')}",
                        data=_xlsx_bytes,
                        file_name=_t("export_filename"),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="export_dl",
                    )
                    st.success("点击上方按钮下载 / Click button above to download")
                except Exception as _ex_exc:
                    st.error(f"Export failed: {_ex_exc}")

    st.divider()

    # ── PDF inventory + gap analysis ──────────────────────────────────────────
    inventory = _scan_pdf_inventory(sel_year)
    try:
        coverage = _db_coverage_detail(sel_year)
    except Exception as _cov_exc:
        st.warning(f"Could not load coverage data: {_cov_exc}")
        coverage = {}
    existing_dates = set(coverage.keys())

    relevant_pdfs = [
        (fname, s, e, path)
        for fname, s, e, path in inventory
        if s <= bf_end and e >= bf_start
    ]

    inv_rows = []
    for fname, s, e, path in relevant_pdfs:
        dates_in_range = [
            s + timedelta(days=i)
            for i in range((e - s).days + 1)
            if bf_start <= s + timedelta(days=i) <= bf_end
        ]
        missing = [d for d in dates_in_range if d not in existing_dates]
        partial = [
            d for d in dates_in_range
            if d in existing_dates and (coverage[d][0] == 0 or coverage[d][1] == 0)
        ]
        inv_rows.append({
            _t("col_pdf"):          fname,
            _t("col_covers"):       f"{s} → {e}",
            _t("col_dates_range"):  len(dates_in_range),
            _t("col_missing"):      len(missing),
            _t("col_partial"):      len(partial),
            _t("col_status"):       _t("status_missing") if missing else (
                                        _t("status_partial") if partial else _t("status_ok")
                                    ),
        })

    if inv_rows:
        inv_df = pd.DataFrame(inv_rows)
        status_col = _t("col_status")
        st.dataframe(
            inv_df.style.apply(
                lambda col: [
                    "background-color: #ffe0e0" if v == _t("status_missing")
                    else "background-color: #fff3cd" if v == _t("status_partial")
                    else "background-color: #d4edda"
                    for v in col
                ],
                subset=[status_col],
            ),
            use_container_width=True, hide_index=True,
        )
    else:
        st.info(_t("no_pdfs"))

    if relevant_pdfs:
        needs_work = [
            (fname, s, e, path)
            for fname, s, e, path in relevant_pdfs
            if any(
                s + timedelta(days=i) not in existing_dates
                for i in range((e - s).days + 1)
                if bf_start <= s + timedelta(days=i) <= bf_end
            )
        ]
        partial_pdfs = [
            (fname, s, e, path)
            for fname, s, e, path in relevant_pdfs
            if any(
                (s + timedelta(days=i)) in existing_dates
                and (coverage[s + timedelta(days=i)][0] == 0
                     or coverage[s + timedelta(days=i)][1] == 0)
                for i in range((e - s).days + 1)
                if bf_start <= s + timedelta(days=i) <= bf_end
            )
        ]

        if mgmt_mode == _t("mode_fill_gaps"):
            pdfs_to_run = needs_work
            btn_label = _t("btn_fill_gaps", n=len(pdfs_to_run))
        else:
            pdfs_to_run = relevant_pdfs
            btn_label = _t("btn_reingest", n=len(pdfs_to_run))

        col_btn, col_info = st.columns([1, 3])
        with col_btn:
            run_backfill = st.button(btn_label, type="primary", disabled=len(pdfs_to_run) == 0)
        with col_info:
            if mgmt_mode == _t("mode_fill_gaps") and not needs_work and partial_pdfs:
                st.warning(_t("warn_partial", n=len(partial_pdfs)))
            elif not pdfs_to_run:
                st.success(_t("all_present"))

        if run_backfill:
            from services.spot_ingest.pdf_parser import parse_pdf as _parse_pdf
            from services.spot_ingest.db_upsert import upsert_rows as _upsert_rows
            if run_interprov:
                from services.spot_ingest.interprov_parser import parse_interprov as _parse_interprov
                from services.spot_ingest.interprov_upsert import upsert_interprov_rows as _upsert_interprov_rows
            if run_ai:
                from services.spot_ingest.ai_summary import generate_summary as _gen_summary
                from services.spot_ingest.interprov_upsert import upsert_summary as _upsert_summary

            provinces_cn = list(PROVINCES_MAP.keys())
            total = len(pdfs_to_run)
            progress = st.progress(0, text=_t("prog_starting"))
            results = []

            for i, (fname, s, e, path) in enumerate(pdfs_to_run):
                progress.progress(i / total, text=_t("prog_parsing", fname=fname))
                # S3 path is a str key; local path is a _Path object
                if isinstance(path, str):
                    import boto3 as _boto3
                    _tmp_path = Path(f"/tmp/{fname}")
                    _boto3.client("s3").download_file(_S3_BUCKET, path, str(_tmp_path))
                    actual_path = _tmp_path
                    _key_parts = path.split("/")
                    pdf_year = int(_key_parts[1]) if len(_key_parts) > 1 and _key_parts[1].isdigit() else sel_year
                else:
                    actual_path = path
                    pdf_year = int(path.parent.name) if path.parent.name.isdigit() else sel_year
                interprov_count = 0
                ai_count = 0
                try:
                    parsed = _parse_pdf(actual_path, pdf_year, provinces_cn)
                    rows = []
                    for rdate, provs in parsed.items():
                        for pcn, vals in provs.items():
                            rows.append({
                                "report_date": rdate,
                                "province_cn": pcn,
                                "province_en": PROVINCES_MAP.get(pcn, pcn),
                                **vals,
                            })
                    n = _upsert_rows(rows)

                    interprov_rows: list = []
                    if run_interprov:
                        progress.progress((i + 0.5) / total,
                                          text=_t("prog_interprov", fname=fname))
                        interprov_rows = _parse_interprov(actual_path, pdf_year)
                        if interprov_rows:
                            interprov_count = _upsert_interprov_rows(interprov_rows)

                    if run_ai:
                        for rdate in sorted(parsed.keys()):
                            progress.progress((i + 0.7) / total,
                                              text=_t("prog_ai", rdate=rdate))
                            day_prices = [
                                {
                                    "province_en": r.get("province_en", r.get("province_cn", "")),
                                    "da_avg": r.get("da_avg"),
                                    "rt_avg": r.get("rt_avg"),
                                }
                                for r in rows
                                if r.get("report_date") == rdate
                            ]
                            day_interprov = [r for r in interprov_rows if r["report_date"] == rdate]
                            summary = _gen_summary(rdate, day_prices, day_interprov, fname)
                            if summary:
                                _upsert_summary(summary)
                                ai_count += 1

                    results.append({
                        _t("col_pdf"):       fname,
                        _t("col_dates"):     str(sorted(parsed.keys())),
                        _t("col_rows"):      n,
                        _t("col_interprov"): interprov_count,
                        _t("col_ai"):        ai_count,
                        _t("col_error"):     "",
                    })
                except Exception as exc:
                    results.append({
                        _t("col_pdf"):       fname,
                        _t("col_dates"):     "",
                        _t("col_rows"):      0,
                        _t("col_interprov"): 0,
                        _t("col_ai"):        0,
                        _t("col_error"):     str(exc)[:120],
                    })

            progress.progress(1.0, text=_t("prog_done"))
            st.success(_t("backfill_complete", n=total))
            st.dataframe(pd.DataFrame(results), use_container_width=True, hide_index=True)

            load_all.clear()
            load_kpis.clear()
            _db_coverage.clear()
            _db_coverage_detail.clear()
            st.rerun()

    # ── Knowledge Base Sync ────────────────────────────────────────────────────
    st.divider()
    with st.expander(f"📚 {_t('kb_sync_title')}", expanded=False):
        _KB_SYNC_DIR = _REPO / "data" / "market-fundamentals"
        _KB_SYNC_EXCLUDE = "各省现货价格及边界数据"
        _KB_SUPPORTED_EXTS = {
            ".pdf", ".pptx", ".ppt", ".docx", ".doc",
            ".xlsx", ".xls", ".txt",
            ".png", ".jpg", ".jpeg", ".webp",
        }

        st.caption(_t("kb_sync_caption"))

        if not _KB_SYNC_DIR.exists():
            st.info(_t("kb_sync_unavailable"))
        else:
            # Resolve app scope: trader for trading/settlement/frequency-result folders, else shared
            _KBS_TRADER_MARKERS = ("5-交易数据", "交易数据", "电力市场结算情况", "调频结果数据")
            def _kbs_resolve_app(p: Path) -> str:
                p_str = str(p)
                return "trader" if any(m in p_str for m in _KBS_TRADER_MARKERS) else "shared"

            _api_key = _os.environ.get("ANTHROPIC_API_KEY")

            # Scan is deferred to avoid freezing the tab on page load (3000+ files on network drive)
            _kbs_scan_key = "kbs_scanned_files"
            col_check, col_sync = st.columns([1, 1])

            if col_check.button("🔍 Scan for new files", key="kb_sync_scan"):
                _results = []
                for _p in sorted(_KB_SYNC_DIR.rglob("*")):
                    if not _p.is_file(): continue
                    if _p.name.endswith("_Error.txt"): continue
                    if _p.suffix.lower() not in _KB_SUPPORTED_EXTS: continue
                    if _KB_SYNC_EXCLUDE in str(_p): continue
                    _results.append(str(_p))
                st.session_state[_kbs_scan_key] = _results

            _kbs_paths = [Path(p) for p in st.session_state.get(_kbs_scan_key, [])]

            if _kbs_paths:
                st.markdown(f"**{len(_kbs_paths)}** file(s) found under `data/market-fundamentals/`")
                _run_sync = col_sync.button(
                    _t("kb_sync_btn", n=len(_kbs_paths)),
                    type="primary",
                    key="kb_sync_run",
                )
                if _run_sync:
                    from services.knowledge_pool.knowledge_docs import (
                        init_knowledge_tables as _kbs_init,
                        register_and_ingest as _kbs_ingest,
                    )
                    _kbs_init()
                    _kbs_added, _kbs_skipped, _kbs_errors = 0, 0, 0
                    _kbs_total = len(_kbs_paths)
                    _kbs_prog = st.progress(0, text=_t("prog_starting"))
                    for _ki, _kp in enumerate(_kbs_paths):
                        _kbs_prog.progress(
                            _ki / _kbs_total,
                            text=_t("kb_sync_progress", i=_ki + 1, n=_kbs_total, fname=_kp.name),
                        )
                        try:
                            _kbs_app = _kbs_resolve_app(_kp)
                            _, _is_new, _ = _kbs_ingest(
                                file_bytes=_kp.read_bytes(),
                                filename=_kp.name,
                                app=_kbs_app,
                                api_key=_api_key,
                            )
                            if _is_new:
                                _kbs_added += 1
                            else:
                                _kbs_skipped += 1
                        except Exception as _kbs_exc:
                            _kbs_errors += 1
                            st.warning(f"{_kp.name}: {_kbs_exc}")
                    _kbs_prog.progress(1.0, text=_t("prog_done"))
                    st.success(_t("kb_sync_done",
                                   added=_kbs_added, skipped=_kbs_skipped, errors=_kbs_errors))
                    st.session_state.pop(_kbs_scan_key, None)
            else:
                st.caption("Click 'Scan' to check for new files in data/market-fundamentals/")

    # ── WeChat Article Batch Import ───────────────────────────────────────────
    st.divider()
    with st.expander(f"💬 {_t('wechat_title')}", expanded=False):
        st.caption(_t("wechat_caption"))

        _wc_urls_raw = st.text_area(
            _t("wechat_url_label"),
            height=140,
            placeholder=_t("wechat_url_placeholder"),
            key="wechat_urls_input",
        )
        _wc_urls = [u.strip() for u in _wc_urls_raw.splitlines() if u.strip().startswith("http")]

        _wc_col1, _wc_col2 = st.columns([2, 1])
        _wc_category = _wc_col2.selectbox(
            "Category",
            options=["research_report", "market_rules", "policy_doc", "technical_spec", "annual_report", "other"],
            index=0,
            key="wechat_category",
        )
        _wc_app = _wc_col2.selectbox(
            "App scope",
            options=["shared", "strategist", "trader"],
            index=0,
            key="wechat_app_scope",
        )

        _wc_run = _wc_col1.button(
            _t("wechat_run_btn", n=len(_wc_urls)) if _wc_urls else _t("wechat_no_urls"),
            disabled=not _wc_urls,
            type="primary",
            key="wechat_run",
        )

        if _wc_run and _wc_urls:
            import requests as _requests

            def _fetch_wechat(url: str) -> tuple[bytes, str]:
                """Fetch a WeChat article and return (text_bytes, title)."""
                headers = {
                    "User-Agent": (
                        "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) "
                        "AppleWebKit/605.1.15 (KHTML, like Gecko) "
                        "Version/17.0 Mobile/15E148 Safari/604.1"
                    ),
                    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                    "Referer": "https://mp.weixin.qq.com/",
                }
                resp = _requests.get(url, headers=headers, timeout=30)
                resp.raise_for_status()
                from bs4 import BeautifulSoup
                soup = BeautifulSoup(resp.content, "html.parser")
                # Extract title
                title_tag = (
                    soup.find("h1", id="activity-name")
                    or soup.find("h2", class_="rich_media_title")
                    or soup.find("title")
                )
                title = (title_tag.get_text(strip=True) if title_tag else "") or url.split("/s/")[-1][:40]
                # Remove non-content elements
                for _tag in soup(["script", "style", "nav", "footer", "header",
                                   "iframe", "img", "svg"]):
                    _tag.decompose()
                # Try article body first, fall back to full page text
                content_div = (
                    soup.find("div", id="js_content")
                    or soup.find("div", class_="rich_media_content")
                )
                if content_div:
                    text = content_div.get_text(separator="\n")
                else:
                    text = soup.get_text(separator="\n")
                lines = [l.strip() for l in text.splitlines() if l.strip()]
                return "\n".join(lines).encode("utf-8"), title

            from services.knowledge_pool.knowledge_docs import (
                init_knowledge_tables as _wc_init,
                register_and_ingest as _wc_ingest,
            )
            _wc_init()

            _wc_added, _wc_skipped, _wc_errors = 0, 0, 0
            _wc_total = len(_wc_urls)
            _wc_prog = st.progress(0, text=_t("prog_starting"))
            _wc_api_key = _os.environ.get("ANTHROPIC_API_KEY")
            _wc_status_lines = []

            for _wi, _wu in enumerate(_wc_urls):
                _wc_prog.progress(
                    _wi / _wc_total,
                    text=_t("wechat_fetching", i=_wi + 1, n=_wc_total, url=_wu[:60]),
                )
                try:
                    _wc_bytes, _wc_title = _fetch_wechat(_wu)
                    _safe_title = _wc_title[:120].replace("/", "_").replace("\\", "_")
                    _wc_fname = f"{_safe_title}.txt"
                    _, _wc_is_new, _wc_cat = _wc_ingest(
                        file_bytes=_wc_bytes,
                        filename=_wc_fname,
                        category_override=_wc_category,
                        app=_wc_app,
                        api_key=_wc_api_key,
                        synthesize=False,
                    )
                    if _wc_is_new:
                        _wc_added += 1
                        _wc_status_lines.append(f"✅ {_wc_title[:80]}")
                    else:
                        _wc_skipped += 1
                        _wc_status_lines.append(f"⏭️ already indexed: {_wc_title[:80]}")
                except Exception as _wc_exc:
                    _wc_errors += 1
                    _wc_status_lines.append(f"❌ {_wu[:60]} — {_wc_exc}")

            _wc_prog.progress(1.0, text=_t("prog_done"))
            st.success(_t("wechat_done", added=_wc_added, skipped=_wc_skipped, errors=_wc_errors))
            for _sl in _wc_status_lines:
                st.markdown(_sl)

            # Offer immediate KB digest
            if _wc_added > 0:
                if st.button(_t("wechat_digest_btn"), key="wechat_digest_now"):
                    from services.knowledge_pool.expert_memory import digest_spot_kb_docs as _wc_digest
                    _wc_api_key2 = _os.environ.get("ANTHROPIC_API_KEY")
                    if _wc_api_key2:
                        with st.spinner("Digesting new articles into insights…"):
                            _wc_n = _wc_digest(api_key=_wc_api_key2)
                        st.toast(f"Extracted {_wc_n} insight(s) from new articles.")
                    else:
                        st.warning("ANTHROPIC_API_KEY not set — cannot digest.")

    # ── Daily Market Report ───────────────────────────────────────────────────
    st.divider()
    with st.expander(f"📧 {_t('report_section_title')}", expanded=False):
        st.caption(_t("report_section_caption"))

        # Scheduler status
        try:
            _rpt_sched = _start_spot_scheduler()
            if _rpt_sched is not None:
                _rpt_jobs = _rpt_sched.get_jobs()
                _rpt_job  = next((j for j in _rpt_jobs if j.id == "spot_daily_report"), None)
                if _rpt_job and _rpt_job.next_run_time:
                    st.info(
                        f"⏰ {_t('report_schedule_status')}: running · "
                        f"{_t('report_next_run')}: "
                        f"{_rpt_job.next_run_time.strftime('%Y-%m-%d %H:%M %Z')}"
                    )
                else:
                    st.warning("Scheduler running but no next run time found.")
            else:
                st.warning("APScheduler not available — install `apscheduler`.")
        except Exception as _sched_exc:
            st.warning(f"Scheduler status unavailable: {_sched_exc}")

        st.markdown(f"**{_t('report_webhook_title')}**")
        st.caption(_t("report_webhook_caption"))

        # Module is loaded once per process via @st.cache_resource
        try:
            _rpt_mod = _get_spot_report_mod()
        except Exception as _wt_exc:
            st.warning(f"Could not load spot_report module: {_wt_exc}")
            _rpt_mod = None

        # ── Existing webhooks (cached — no DB hit on every rerun) ─────────────
        _wh_rows = _cached_webhooks()
        if _wh_rows:
            for _wh in _wh_rows:
                _wh_col1, _wh_col2, _wh_col3, _wh_col4 = st.columns([3, 2, 1, 1])
                _wh_col1.text(_wh.get("webhook_url", "")[:60])
                _wh_col2.text(_wh.get("label", ""))
                _enabled_now = _wh_col3.checkbox(
                    _t("report_webhook_enabled"),
                    value=bool(_wh.get("enabled", True)),
                    key=f"wh_enabled_{_wh['id']}",
                    label_visibility="collapsed",
                )
                if _enabled_now != bool(_wh.get("enabled", True)):
                    if _rpt_mod:
                        _rpt_mod.upsert_webhook(
                            _wh["webhook_url"], _wh.get("label", ""), _enabled_now
                        )
                    _cached_webhooks.clear()
                    st.rerun()
                if _wh_col4.button(_t("report_webhook_delete"),
                                   key=f"wh_del_{_wh['id']}"):
                    if _rpt_mod:
                        _rpt_mod.delete_webhook(_wh["id"])
                    _cached_webhooks.clear()
                    st.rerun()
        else:
            st.caption(_t("report_webhook_empty"))

        # ── Add new webhook ───────────────────────────────────────────────────
        with st.form("add_webhook_form", clear_on_submit=True):
            _new_wh_url = st.text_input(
                _t("report_webhook_add_label"),
                placeholder="https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=...",
                type="password",
            )
            _new_wh_label = st.text_input(
                _t("report_webhook_add_label_label"),
                placeholder="e.g. Trading Team",
            )
            if st.form_submit_button(_t("report_webhook_add_btn"), type="primary"):
                if _new_wh_url and _new_wh_url.startswith("http") and _rpt_mod:
                    _rpt_mod.upsert_webhook(_new_wh_url, _new_wh_label, True)
                    _cached_webhooks.clear()
                    st.success(_t("report_webhook_added"))
                    st.rerun()
                else:
                    st.error("Please enter a valid webhook URL.")

        st.divider()

        # ── Send Now ──────────────────────────────────────────────────────────
        @st.cache_data(ttl=600, show_spinner=False)
        def _last_rt_date():
            try:
                import pandas as _pd2
                _df = _pd2.read_sql(
                    "SELECT MAX(report_date) AS d FROM spot_daily WHERE rt_avg IS NOT NULL",
                    _conn(),
                )
                _v = _df.iloc[0]["d"]
                if _v is not None and not _pd2.isna(_v):
                    return _pd2.Timestamp(_v).date()
            except Exception:
                pass
            import datetime as _dt2
            return _dt2.date.today() - _dt2.timedelta(days=1)

        import datetime as _rdt
        _rpt_default_date = _last_rt_date()
        _rpt_date = st.date_input(
            "Report date",
            value=_rpt_default_date,
            max_value=_rdt.date.today(),
            key="report_date_pick",
        )
        _rpt_email_default = _os.environ.get("REPORT_TO_EMAIL", _DEFAULT_RECIPIENT)
        _rpt_email = st.text_input(
            _t("report_email_label"),
            value=_rpt_email_default,
            help=_t("report_email_help"),
            key="report_email_override",
        )
        if st.button(_t("report_send_now"), type="primary", key="report_send_now_btn"):
            if _rpt_mod is not None:
                with st.spinner("Generating and sending report…"):
                    try:
                        _rpt_result = _rpt_mod.run_daily_report(
                            to_email=_rpt_email or None,
                            report_date=_rpt_date,
                        )
                        if _rpt_result["status"] == "success":
                            st.success(
                                _t("report_send_success",
                                   size=_rpt_result.get("size_bytes", 0),
                                   rdate=_rpt_result.get("date", "?"))
                            )
                            st.info(_t("report_send_wecom", result=_rpt_result.get("wecom", "skipped")))
                        else:
                            st.error(_t("report_send_error", err=_rpt_result.get("error", "?")))
                    except Exception as _rpt_exc:
                        st.error(_t("report_send_error", err=str(_rpt_exc)))
            else:
                st.error("spot_report module not available.")

    # ── Exchange Monthly Reports ───────────────────────────────────────────────
    st.divider()
    with st.expander("📋 交易所月报管理 Exchange Monthly Reports", expanded=False):
        st.caption(
            "Upload and manage provincial power exchange monthly/quarterly reports. "
            "Files are ingested into the shared knowledge base and searchable by the Strategist agent."
        )

        # ── Upload new reports ─────────────────────────────────────────────────
        _emr_uploaded = st.file_uploader(
            "Upload report (PDF or DOCX)",
            type=["pdf", "doc", "docx"],
            accept_multiple_files=True,
            key="emr_upload",
        )
        if _emr_uploaded:
            _emr_province_options = [
                "自动识别 Auto-detect",
                "上海", "冀南", "安徽", "山东", "广东", "江苏", "浙江", "福建", "蒙西",
            ]
            _emr_province_sel = st.selectbox(
                "Province (leave auto to infer from filename)",
                _emr_province_options,
                key="emr_province_sel",
            )
            if st.button(
                f"Ingest {len(_emr_uploaded)} file(s) into KB",
                type="primary",
                key="emr_do_ingest",
            ):
                import sys as _sys
                _sys.path.insert(0, str(_REPO))
                try:
                    from services.exchange_reports.ingestor import ingest_report
                    _emr_ok, _emr_dup, _emr_fail = 0, 0, 0
                    for _emr_f in _emr_uploaded:
                        _emr_bytes = _emr_f.read()
                        _emr_prov = (
                            None if _emr_province_sel == "自动识别 Auto-detect"
                            else _emr_province_sel
                        )
                        try:
                            _emr_res = ingest_report(
                                file_bytes=_emr_bytes,
                                filename=_emr_f.name,
                                province=_emr_prov,
                                pg_url=_os.environ.get("PGURL"),
                                anthropic_api_key=(
                                    _os.environ.get("ANTHROPIC_API_KEY")
                                    or _os.environ.get("DEEPSEEK_API_KEY")
                                ),
                            )
                            if _emr_res["status"] == "ingested":
                                _emr_ok += 1
                            elif _emr_res["status"] == "duplicate":
                                _emr_dup += 1
                            else:
                                _emr_fail += 1
                        except Exception as _emr_exc:
                            st.error(f"Failed {_emr_f.name}: {_emr_exc}")
                            _emr_fail += 1
                    if _emr_ok:
                        st.success(f"✅ Ingested {_emr_ok} new report(s).")
                    if _emr_dup:
                        st.info(f"ℹ️ {_emr_dup} file(s) already in KB (skipped).")
                    if _emr_fail:
                        st.warning(f"⚠️ {_emr_fail} file(s) failed — check logs.")
                    st.rerun()
                except ImportError as _emr_ie:
                    st.error(f"exchange_reports service not found: {_emr_ie}")

        st.divider()

        # ── Report inventory ───────────────────────────────────────────────────
        _emr_filter_prov = st.selectbox(
            "Filter by province",
            ["All", "上海", "冀南", "安徽", "山东", "广东", "广西", "江苏", "浙江", "福建", "蒙西"],
            key="emr_filter_prov",
        )

        # ── Metrics tab switcher ───────────────────────────────────────────────
        _emr_view = st.radio(
            "View",
            ["📊 数据汇总表 Metrics", "📈 趋势分析 Trends", "📂 文件清单 File List", "🗺️ 省级Excel数据 Province Data"],
            horizontal=True, key="emr_view",
        )

        if _emr_view == "📊 数据汇总表 Metrics":
            # Month selector
            @st.cache_data(ttl=120, show_spinner=False)
            def _emr_avail_months() -> list:
                try:
                    from services.exchange_reports.metrics_extractor import get_available_months
                    return get_available_months(pg_url=_os.environ.get("PGURL"))
                except Exception:
                    return []

            _emr_months = _emr_avail_months()
            if not _emr_months:
                st.info("No structured metrics yet. Run the backfill script with ANTHROPIC_API_KEY set, or upload reports via Feishu.")
            else:
                _emr_sel_month = st.selectbox(
                    "Report month", _emr_months, key="emr_sel_month",
                )
                _emr_yr, _emr_mo = int(_emr_sel_month[:4]), int(_emr_sel_month[5:])

                @st.cache_data(ttl=120, show_spinner=False)
                def _load_emr_metrics(_yr: int, _mo: int) -> list:
                    try:
                        from services.exchange_reports.metrics_extractor import get_metrics_table
                        return get_metrics_table(year=_yr, month=_mo, pg_url=_os.environ.get("PGURL"))
                    except Exception:
                        return []

                _emr_mrows = _load_emr_metrics(_emr_yr, _emr_mo)
                if not _emr_mrows:
                    st.info(f"No metrics extracted for {_emr_sel_month}.")
                else:
                    import pandas as _pd_emr2
                    _emr_mdf = _pd_emr2.DataFrame(_emr_mrows)

                    # ── Sub-tabs ───────────────────────────────────────────────
                    _emr_t1, _emr_t2, _emr_t3, _emr_t4, _emr_t5 = st.tabs(
                        ["概览", "结算价格", "发电量", "装机容量", "零售市场"]
                    )

                    with _emr_t1:
                        _emr_c1 = {
                            "province": "省份",
                            "total_volume_gwh": "总成交量(亿kWh)",
                            "volume_yoy_pct": "同比(%)",
                            "avg_price_yuan_mwh": "结算均价(元/MWh)",
                            "contract_avg_price_yuan_mwh": "合约均价",
                            "spot_volume_gwh": "现货量(亿kWh)",
                            "spot_avg_price_yuan_mwh": "现货均价",
                            "peak_price_yuan_mwh": "峰段价",
                            "valley_price_yuan_mwh": "谷段价",
                            "renewable_pct": "新能源占比(%)",
                            "installed_capacity_gw": "总装机(GW)",
                            "max_load_gw": "最大负荷(GW)",
                            "market_participants_total": "市场主体(户)",
                        }
                        _a1 = [c for c in _emr_c1 if c in _emr_mdf.columns]
                        st.dataframe(
                            _emr_mdf[_a1].rename(columns=_emr_c1).set_index("省份"),
                            use_container_width=True,
                        )
                        st.caption(f"{len(_emr_mrows)} provinces · {_emr_sel_month}")

                    with _emr_t2:
                        _emr_c2 = {
                            "province": "省份",
                            "contract_avg_price_yuan_mwh": "合约均价",
                            "avg_price_yuan_mwh": "结算均价",
                            "thermal_settlement_price_yuan_mwh": "火电",
                            "wind_settlement_price_yuan_mwh": "风电",
                            "solar_settlement_price_yuan_mwh": "光伏",
                            "nuclear_settlement_price_yuan_mwh": "核电",
                            "bess_settlement_price_yuan_mwh": "储能",
                            "retailer_settlement_price_yuan_mwh": "售电公司均价",
                            "spot_avg_price_yuan_mwh": "现货均价",
                        }
                        _a2 = [c for c in _emr_c2 if c in _emr_mdf.columns]
                        st.dataframe(
                            _emr_mdf[_a2].rename(columns=_emr_c2).set_index("省份"),
                            use_container_width=True,
                        )
                        st.caption("单位: 元/MWh")

                    with _emr_t3:
                        _emr_c3 = {
                            "province": "省份",
                            "total_volume_gwh": "总成交量",
                            "thermal_volume_gwh": "火电",
                            "wind_volume_gwh": "风电",
                            "solar_volume_gwh": "光伏",
                            "hydro_volume_gwh": "水电",
                            "nuclear_volume_gwh": "核电",
                            "bess_traded_volume_gwh": "储能",
                            "spot_volume_gwh": "现货量",
                            "incoming_volume_gwh": "外来电",
                            "outgoing_volume_gwh": "外送电",
                        }
                        _a3 = [c for c in _emr_c3 if c in _emr_mdf.columns]
                        st.dataframe(
                            _emr_mdf[_a3].rename(columns=_emr_c3).set_index("省份"),
                            use_container_width=True,
                        )
                        st.caption("单位: 亿kWh")

                    with _emr_t4:
                        _emr_c4 = {
                            "province": "省份",
                            "installed_capacity_gw": "总装机(GW)",
                            "wind_capacity_gw": "风电",
                            "solar_capacity_gw": "光伏",
                            "thermal_capacity_gw": "火电",
                            "nuclear_capacity_gw": "核电",
                            "bess_capacity_gw": "储能",
                        }
                        _a4 = [c for c in _emr_c4 if c in _emr_mdf.columns]
                        st.dataframe(
                            _emr_mdf[_a4].rename(columns=_emr_c4).set_index("省份"),
                            use_container_width=True,
                        )
                        st.caption("单位: GW")

                    with _emr_t5:
                        _emr_c5 = {
                            "province": "省份",
                            "retailer_volume_gwh": "零售交易量(亿kWh)",
                            "retailer_settlement_price_yuan_mwh": "零售结算均价(元/MWh)",
                            "retailer_service_fee_million_yuan": "代理服务费(百万元)",
                        }
                        _a5 = [c for c in _emr_c5 if c in _emr_mdf.columns]
                        st.dataframe(
                            _emr_mdf[_a5].rename(columns=_emr_c5).set_index("省份"),
                            use_container_width=True,
                        )

                    # Highlights
                    _emr_hl = [(r["province"], r["key_highlights"]) for r in _emr_mrows if r.get("key_highlights")]
                    if _emr_hl:
                        with st.expander("📝 各省市场要点", expanded=False):
                            for _p, _h in sorted(_emr_hl):
                                st.markdown(f"**{_p}**：{_h}")

                    # PDF download
                    if st.button("📄 下载汇总PDF", key="emr_dl_pdf"):
                        try:
                            from services.exchange_reports.summary_pdf import build_summary_pdf
                            _emr_pdf_bytes = build_summary_pdf(
                                _emr_mrows,
                                month_label=f"{_emr_yr}年{_emr_mo}月",
                            )
                            st.download_button(
                                "⬇️ 保存PDF",
                                data=_emr_pdf_bytes,
                                file_name=f"交易所月报汇总_{_emr_yr}-{_emr_mo:02d}.pdf",
                                mime="application/pdf",
                                key="emr_pdf_dl_btn",
                            )
                        except Exception as _emr_pe:
                            st.error(f"PDF生成失败：{_emr_pe}")

        elif _emr_view == "📈 趋势分析 Trends":
            @st.cache_data(ttl=300, show_spinner=False)
            def _emr_avail_provinces() -> list:
                try:
                    from services.exchange_reports.metrics_extractor import get_available_provinces
                    return get_available_provinces(pg_url=_os.environ.get("PGURL"))
                except Exception:
                    return ["上海", "冀南", "安徽", "山东", "广东", "广西", "江苏", "浙江", "福建", "蒙西"]

            _emr_trend_provs = _emr_avail_provinces()
            _emr_trend_prov = st.selectbox(
                "选择省份", _emr_trend_provs, key="emr_trend_prov",
            )

            @st.cache_data(ttl=300, show_spinner=False)
            def _load_emr_ts(_prov: str) -> list:
                try:
                    from services.exchange_reports.metrics_extractor import get_metrics_timeseries
                    return get_metrics_timeseries(_prov, report_type="monthly", pg_url=_os.environ.get("PGURL"))
                except Exception:
                    return []

            _emr_ts_rows = _load_emr_ts(_emr_trend_prov)
            if not _emr_ts_rows:
                st.info(f"No monthly data for {_emr_trend_prov}.")
            else:
                import pandas as _pd_ts
                _emr_tsdf = _pd_ts.DataFrame(_emr_ts_rows)
                _emr_tsdf["month"] = _pd_ts.to_datetime(_emr_tsdf["report_month"]).dt.strftime("%Y-%m")
                _emr_tsdf = _emr_tsdf.set_index("month")

                _emr_trend_cat = st.radio(
                    "指标类别",
                    ["价格 Prices", "电量 Volumes", "装机 Capacity"],
                    horizontal=True, key="emr_trend_cat",
                )

                if _emr_trend_cat == "价格 Prices":
                    _px_cols = {
                        "avg_price_yuan_mwh": "结算均价",
                        "contract_avg_price_yuan_mwh": "合约均价",
                        "spot_avg_price_yuan_mwh": "现货均价",
                        "thermal_settlement_price_yuan_mwh": "火电",
                        "wind_settlement_price_yuan_mwh": "风电",
                        "solar_settlement_price_yuan_mwh": "光伏",
                        "bess_settlement_price_yuan_mwh": "储能",
                        "retailer_settlement_price_yuan_mwh": "售电公司",
                    }
                    _px_avail = {k: v for k, v in _px_cols.items() if k in _emr_tsdf.columns and _emr_tsdf[k].notna().any()}
                    if _px_avail:
                        _px_plot = _emr_tsdf[list(_px_avail.keys())].rename(columns=_px_avail)
                        st.line_chart(_px_plot, use_container_width=True)
                        st.caption(f"{_emr_trend_prov} 各类型结算价格趋势 (元/MWh)")
                    else:
                        st.info("No price data available.")

                elif _emr_trend_cat == "电量 Volumes":
                    _vol_cols = {
                        "total_volume_gwh": "总成交量",
                        "wind_volume_gwh": "风电",
                        "solar_volume_gwh": "光伏",
                        "thermal_volume_gwh": "火电",
                        "hydro_volume_gwh": "水电",
                        "nuclear_volume_gwh": "核电",
                        "bess_traded_volume_gwh": "储能",
                        "spot_volume_gwh": "现货量",
                        "incoming_volume_gwh": "外来电",
                    }
                    _vol_avail = {k: v for k, v in _vol_cols.items() if k in _emr_tsdf.columns and _emr_tsdf[k].notna().any()}
                    if _vol_avail:
                        _vol_plot = _emr_tsdf[list(_vol_avail.keys())].rename(columns=_vol_avail)
                        st.line_chart(_vol_plot, use_container_width=True)
                        st.caption(f"{_emr_trend_prov} 各类型电量趋势 (亿kWh)")
                    else:
                        st.info("No volume data available.")

                else:  # Capacity
                    _cap_cols = {
                        "installed_capacity_gw": "总装机",
                        "wind_capacity_gw": "风电",
                        "solar_capacity_gw": "光伏",
                        "thermal_capacity_gw": "火电",
                        "bess_capacity_gw": "储能",
                        "nuclear_capacity_gw": "核电",
                    }
                    _cap_avail = {k: v for k, v in _cap_cols.items() if k in _emr_tsdf.columns and _emr_tsdf[k].notna().any()}
                    if _cap_avail:
                        _cap_plot = _emr_tsdf[list(_cap_avail.keys())].rename(columns=_cap_avail)
                        st.line_chart(_cap_plot, use_container_width=True)
                        st.caption(f"{_emr_trend_prov} 装机容量趋势 (GW)")
                    else:
                        st.info("No capacity data available.")

        else:
            # File list view
            @st.cache_data(ttl=120, show_spinner=False)
            def _load_exchange_reports(_prov: str) -> list:
                try:
                    from services.exchange_reports.ingestor import list_reports
                    return list_reports(
                        province=None if _prov == "All" else _prov,
                        pg_url=_os.environ.get("PGURL"),
                    )
                except Exception:
                    return []

            _emr_rows = _load_exchange_reports(_emr_filter_prov)
            if _emr_rows:
                import pandas as _pd_emr
                _emr_df = _pd_emr.DataFrame(_emr_rows)
                _emr_df["report_month"] = _pd_emr.to_datetime(_emr_df["report_month"]).dt.strftime("%Y-%m")
                _emr_df["created_at"] = _pd_emr.to_datetime(_emr_df["created_at"]).dt.strftime("%Y-%m-%d")
                st.dataframe(
                    _emr_df[["province", "report_month", "report_type", "file_name", "created_at"]],
                    use_container_width=True,
                    hide_index=True,
                )
                st.caption(f"{len(_emr_rows)} report(s) in KB.")
            else:
                st.info("No exchange monthly reports ingested yet. Upload files above or run the backfill script.")

        # ── Province Excel Data tab ────────────────────────────────────────────
        if _emr_view == "🗺️ 省级Excel数据 Province Data":
            st.caption("来自各省信息披露月报Excel数据库的结构化数据 | Structured data from vendor-curated Excel monthly reports")

            # ── Helper ─────────────────────────────────────────────────────────
            @st.cache_data(ttl=300, show_spinner=False)
            def _excel_all_provinces():
                try:
                    import psycopg2 as _pg2
                    conn = _pg2.connect(_os.environ.get("PGURL", ""))
                    with conn.cursor() as _cur:
                        _cur.execute(
                            "SELECT DISTINCT province FROM staging.exchange_excel_metrics ORDER BY province"
                        )
                        return [r[0] for r in _cur.fetchall()]
                except Exception:
                    return []

            @st.cache_data(ttl=300, show_spinner=False)
            def _excel_latest_all():
                try:
                    import psycopg2 as _pg2
                    conn = _pg2.connect(_os.environ.get("PGURL", ""))
                    with conn.cursor() as _cur:
                        _cur.execute("""
                            SELECT DISTINCT ON (province)
                                province, report_month, source_file,
                                total_capacity_mw, wind_capacity_mw, solar_capacity_mw,
                                thermal_capacity_mw, bess_capacity_mw,
                                total_traded_gwh, spot_traded_gwh,
                                avg_settlement_price, spot_avg_price,
                                max_load_mw, market_participants_total, retailers,
                                incoming_gwh, outgoing_gwh,
                                fr_pool_million_yuan, total_ancillary_million_yuan
                            FROM staging.exchange_excel_metrics
                            ORDER BY province, report_month DESC
                        """)
                        cols = [d[0] for d in _cur.description]
                        import pandas as _pd2
                        return _pd2.DataFrame(_cur.fetchall(), columns=cols)
                except Exception:
                    import pandas as _pd2
                    return _pd2.DataFrame()

            @st.cache_data(ttl=300, show_spinner=False)
            def _excel_timeseries_prov(_prov: str):
                try:
                    import psycopg2 as _pg2
                    conn = _pg2.connect(_os.environ.get("PGURL", ""))
                    with conn.cursor() as _cur:
                        _cur.execute("""
                            SELECT report_month, source_file,
                                total_capacity_mw, wind_capacity_mw, solar_capacity_mw,
                                thermal_capacity_mw, bess_capacity_mw,
                                total_traded_gwh, spot_traded_gwh, contract_traded_gwh,
                                avg_settlement_price, spot_avg_price, contract_avg_price,
                                thermal_settlement_price, wind_settlement_price,
                                solar_settlement_price, bess_settlement_price,
                                max_load_mw, market_participants_total, retailers,
                                incoming_gwh, outgoing_gwh,
                                fr_pool_million_yuan, peak_shaving_million_yuan,
                                renewable_deviation_million_yuan, total_ancillary_million_yuan,
                                retailer_traded_gwh, retailer_settlement_price,
                                retailer_service_fee_million_yuan
                            FROM staging.exchange_excel_metrics
                            WHERE province = %s
                            ORDER BY report_month
                        """, (_prov,))
                        cols = [d[0] for d in _cur.description]
                        import pandas as _pd2
                        return _pd2.DataFrame(_cur.fetchall(), columns=cols)
                except Exception:
                    import pandas as _pd2
                    return _pd2.DataFrame()

            _ex_provinces = _excel_all_provinces()
            _ex_latest_df = _excel_latest_all()

            # ── Sub-tabs ────────────────────────────────────────────────────────
            _ex_subtab_all, _ex_subtab_prov, _ex_subtab_bess, _ex_subtab_fr = st.tabs([
                "📋 全省概览 All Provinces",
                "📈 省份详情 Province Detail",
                "🔋 储能装机 BESS Capacity",
                "⚡ 辅助服务 Ancillary",
            ])

            # ── All Provinces overview ─────────────────────────────────────────
            with _ex_subtab_all:
                if _ex_latest_df.empty:
                    st.info("No Excel data ingested. Run scripts/ingest_excel_reports.py.")
                else:
                    import pandas as _pd_ex
                    _show_df = _ex_latest_df.copy()
                    _show_df["report_month"] = _pd_ex.to_datetime(_show_df["report_month"]).dt.strftime("%Y-%m")
                    _show_df["total_capacity_gw"] = (_show_df["total_capacity_mw"] / 1000).round(1)
                    _show_df["bess_capacity_gw"]  = (_show_df["bess_capacity_mw"]  / 1000).round(2)
                    _show_df["wind_capacity_gw"]  = (_show_df["wind_capacity_mw"]  / 1000).round(1)
                    _show_df["solar_capacity_gw"] = (_show_df["solar_capacity_mw"] / 1000).round(1)

                    _disp_cols = {
                        "province":            "省份",
                        "report_month":        "最新数据月份",
                        "total_capacity_gw":   "总装机(GW)",
                        "wind_capacity_gw":    "风电(GW)",
                        "solar_capacity_gw":   "光伏(GW)",
                        "bess_capacity_gw":    "储能(GW)",
                        "total_traded_gwh":    "成交量(亿kWh)",
                        "avg_settlement_price":"均价(元/MWh)",
                        "max_load_mw":         "最大负荷(MW)",
                        "market_participants_total": "市场主体(家)",
                    }
                    _disp_df = _show_df.rename(columns=_disp_cols)[list(_disp_cols.values())]
                    st.dataframe(_disp_df, use_container_width=True, hide_index=True)
                    st.caption(f"各省最新月度数据 — {len(_show_df)} 个省份")

                    # Capacity bar chart comparing all provinces (plotly grouped bar)
                    if "bess_capacity_gw" in _show_df.columns:
                        _cap_chart_df = _show_df[
                            _show_df["bess_capacity_gw"].notna()
                        ][["province", "wind_capacity_gw", "solar_capacity_gw", "bess_capacity_gw"]].copy()
                        _cap_chart_df = _cap_chart_df.sort_values("bess_capacity_gw", ascending=True)
                        _cap_chart_df = _cap_chart_df.fillna(0)
                        if not _cap_chart_df.empty:
                            st.subheader("各省储能与新能源装机对比")
                            try:
                                import plotly.graph_objects as _go
                                _fig = _go.Figure()
                                for _col, _lbl, _color in [
                                    ("wind_capacity_gw",  "风电",  "#4db8ff"),
                                    ("solar_capacity_gw", "光伏",  "#ffd700"),
                                    ("bess_capacity_gw",  "储能",  "#ff6b35"),
                                ]:
                                    _fig.add_trace(_go.Bar(
                                        name=_lbl,
                                        x=_cap_chart_df[_col],
                                        y=_cap_chart_df["province"],
                                        orientation="h",
                                        marker_color=_color,
                                    ))
                                _fig.update_layout(
                                    barmode="group",
                                    height=max(300, len(_cap_chart_df) * 28),
                                    xaxis_title="装机容量 (GW)",
                                    margin=dict(l=60, r=20, t=20, b=40),
                                    legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="right", x=1),
                                )
                                st.plotly_chart(_fig, use_container_width=True)
                            except ImportError:
                                st.bar_chart(_cap_chart_df.set_index("province")[
                                    ["wind_capacity_gw", "solar_capacity_gw", "bess_capacity_gw"]
                                ])

            # ── Province detail ────────────────────────────────────────────────
            with _ex_subtab_prov:
                if not _ex_provinces:
                    st.info("No data.")
                else:
                    _ex_sel_prov = st.selectbox(
                        "选择省份", _ex_provinces, key="ex_sel_prov",
                    )
                    _ex_ts = _excel_timeseries_prov(_ex_sel_prov)
                    if _ex_ts.empty:
                        st.info(f"No data for {_ex_sel_prov}.")
                    else:
                        import pandas as _pd_ex2
                        _ex_ts["month"] = _pd_ex2.to_datetime(_ex_ts["report_month"]).dt.strftime("%Y-%m")
                        _ex_ts_idx = _ex_ts.set_index("month")

                        _ex_cat = st.radio(
                            "指标", ["价格 Prices", "电量 Volumes", "装机 Capacity", "零售 Retail"],
                            horizontal=True, key="ex_cat",
                        )

                        if _ex_cat == "价格 Prices":
                            _px = {
                                "avg_settlement_price":     "结算均价",
                                "spot_avg_price":           "现货均价",
                                "contract_avg_price":       "合约均价",
                                "thermal_settlement_price": "火电结算",
                                "wind_settlement_price":    "风电结算",
                                "solar_settlement_price":   "光伏结算",
                                "bess_settlement_price":    "储能结算",
                                "retailer_settlement_price":"售电公司",
                            }
                            _px_av = {k: v for k, v in _px.items()
                                      if k in _ex_ts_idx.columns and _ex_ts_idx[k].notna().any()}
                            if _px_av:
                                st.line_chart(_ex_ts_idx[list(_px_av.keys())].rename(columns=_px_av))
                                st.caption(f"{_ex_sel_prov} 结算价格趋势 (元/MWh)")
                            else:
                                st.info("该省暂无价格数据。")

                        elif _ex_cat == "电量 Volumes":
                            _vl = {
                                "total_traded_gwh":    "总成交",
                                "spot_traded_gwh":     "现货",
                                "contract_traded_gwh": "合约",
                                "incoming_gwh":        "外来电",
                                "outgoing_gwh":        "外送电",
                            }
                            _vl_av = {k: v for k, v in _vl.items()
                                      if k in _ex_ts_idx.columns and _ex_ts_idx[k].notna().any()}
                            if _vl_av:
                                st.line_chart(_ex_ts_idx[list(_vl_av.keys())].rename(columns=_vl_av))
                                st.caption(f"{_ex_sel_prov} 月度电量趋势 (亿kWh)")
                            else:
                                st.info("该省暂无电量数据。")

                        elif _ex_cat == "装机 Capacity":
                            _cp = {
                                "total_capacity_mw":   "总装机",
                                "wind_capacity_mw":    "风电",
                                "solar_capacity_mw":   "光伏",
                                "thermal_capacity_mw": "火电",
                                "bess_capacity_mw":    "储能",
                            }
                            _cp_av = {k: v for k, v in _cp.items()
                                      if k in _ex_ts_idx.columns and _ex_ts_idx[k].notna().any()}
                            if _cp_av:
                                st.line_chart(_ex_ts_idx[list(_cp_av.keys())].rename(columns=_cp_av))
                                st.caption(f"{_ex_sel_prov} 装机容量趋势 (MW)")
                            else:
                                st.info("该省暂无装机数据。")

                        else:  # Retail
                            _rt = {
                                "retailers":                        "售电公司数",
                                "retailer_traded_gwh":              "售电量(亿kWh)",
                                "retailer_settlement_price":        "售电均价(元/MWh)",
                                "retailer_service_fee_million_yuan":"服务费(百万元)",
                            }
                            _rt_av = {k: v for k, v in _rt.items()
                                      if k in _ex_ts_idx.columns and _ex_ts_idx[k].notna().any()}
                            if _rt_av:
                                # Show retailer count separately from price/volume
                                _cnt_cols = [k for k in _rt_av if "数" in _rt_av[k] or "家" in _rt_av[k]]
                                _val_cols = [k for k in _rt_av if k not in _cnt_cols]
                                if _cnt_cols:
                                    st.line_chart(_ex_ts_idx[_cnt_cols].rename(columns=_rt_av))
                                if _val_cols:
                                    st.line_chart(_ex_ts_idx[_val_cols].rename(columns=_rt_av))
                                st.caption(f"{_ex_sel_prov} 零售市场趋势")
                            else:
                                st.info("该省暂无零售市场数据。")

            # ── BESS capacity trend ────────────────────────────────────────────
            with _ex_subtab_bess:
                st.caption("各省储能装机容量历史走势（来自Excel数据库）")
                if not _ex_provinces:
                    st.info("No data.")
                else:
                    _bess_provs = st.multiselect(
                        "选择省份（可多选）",
                        _ex_provinces,
                        default=[p for p in ["蒙西", "安徽", "广西", "山西", "宁夏"] if p in _ex_provinces],
                        key="ex_bess_provs",
                    )
                    if _bess_provs:
                        import pandas as _pd_bess
                        _bess_frames = []
                        for _bp in _bess_provs:
                            _bt = _excel_timeseries_prov(_bp)
                            if not _bt.empty and "bess_capacity_mw" in _bt.columns:
                                _bt_filt = _bt[["report_month", "bess_capacity_mw"]].dropna()
                                _bt_filt = _bt_filt.rename(columns={"bess_capacity_mw": _bp})
                                _bt_filt = _bt_filt.set_index("report_month")
                                _bess_frames.append(_bt_filt)
                        if _bess_frames:
                            _bess_merged = _pd_bess.concat(_bess_frames, axis=1).sort_index()
                            _bess_merged.index = _pd_bess.to_datetime(_bess_merged.index).strftime("%Y-%m")
                            _bess_merged /= 1000  # MW → GW
                            st.line_chart(_bess_merged, use_container_width=True)
                            st.caption("储能装机 (GW)")
                        else:
                            st.info("所选省份暂无储能装机数据。")

            # ── Ancillary / FR costs ───────────────────────────────────────────
            with _ex_subtab_fr:
                st.caption("山东辅助服务费用月度数据（来自山东电力市场信息披露月报）")
                _fr_ts = _excel_timeseries_prov("山东")
                if _fr_ts.empty or "fr_pool_million_yuan" not in _fr_ts.columns:
                    st.info("No ancillary cost data found.")
                else:
                    import pandas as _pd_fr
                    _fr_ts["month"] = _pd_fr.to_datetime(_fr_ts["report_month"]).dt.strftime("%Y-%m")
                    _fr_ts_idx = _fr_ts.set_index("month")
                    _fr_cols = {
                        "fr_pool_million_yuan":              "调频费用",
                        "peak_shaving_million_yuan":         "调峰补贴",
                        "renewable_deviation_million_yuan":  "新能源偏差",
                        "total_ancillary_million_yuan":      "辅助服务合计",
                    }
                    _fr_av = {k: v for k, v in _fr_cols.items()
                              if k in _fr_ts_idx.columns and _fr_ts_idx[k].notna().any()}
                    if _fr_av:
                        st.line_chart(_fr_ts_idx[list(_fr_av.keys())].rename(columns=_fr_av))
                        st.caption("山东辅助服务费用趋势 (百万元/月)")
                    # Summary table
                    _fr_disp = _fr_ts[["month"] + [k for k in _fr_cols if k in _fr_ts.columns]].rename(
                        columns={**{"month": "月份"}, **_fr_cols}
                    )
                    st.dataframe(_fr_disp.dropna(subset=["调频费用"]), use_container_width=True, hide_index=True)
