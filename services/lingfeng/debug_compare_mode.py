"""
Debug: investigates the 数据对比方案 comparison mode and finds the raw export.

Run:
    py services/lingfeng/debug_compare_mode.py
"""
import asyncio, os
from pathlib import Path

_REPO = Path(__file__).resolve().parent.parent.parent
_DEBUG_DIR = _REPO / "debug" / "lingfeng"
_DEBUG_DIR.mkdir(parents=True, exist_ok=True)

_env_file = _REPO / "config" / ".env"
if _env_file.exists():
    try:
        from dotenv import load_dotenv; load_dotenv(str(_env_file))
    except ImportError: pass

USERNAME = os.environ.get("LINGFENG_USERNAME", "")
PASSWORD = os.environ.get("LINGFENG_PASSWORD", "")
_LOGIN_URL = "https://lingfeng-saas.tradingthink.cn/#/login"
_DATA_URL  = "https://lingfeng-saas.tradingthink.cn/#/powerTrading/market"
MARKET = "山西"


async def run():
    from playwright.async_api import async_playwright

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(accept_downloads=True)
        page    = await context.new_page()
        page.set_default_timeout(20_000)

        # Login
        print("[1] Logging in …")
        await page.goto(_LOGIN_URL)
        await page.wait_for_load_state("networkidle")
        try:
            tab = page.locator("div.login-tab-item", has_text="账号登录")
            if await tab.count() > 0:
                await tab.first.click(); await page.wait_for_timeout(300)
        except Exception: pass
        inputs = page.locator("input")
        await inputs.nth(0).fill(USERNAME)
        await inputs.nth(1).fill(PASSWORD)
        await page.locator("button", has_text="登录").first.click()
        await page.wait_for_url(lambda u: "/login" not in u, timeout=15_000)

        # Navigate
        await page.goto(_DATA_URL)
        await page.wait_for_load_state("networkidle")
        await page.wait_for_selector(".ant-select-selector", timeout=30_000)
        await page.wait_for_timeout(1000)

        # Screenshot initial state
        await page.screenshot(path=str(_DEBUG_DIR / "cmp_01_initial.png"))
        print("[2] Initial state screenshot → cmp_01_initial.png")

        # Print all .ant-tabs-tab-btn to understand tab structure
        print("\n[3] All tab buttons:")
        tabs = await page.locator(".ant-tabs-tab-btn").all()
        for i, t in enumerate(tabs):
            try:
                txt = (await t.inner_text()).strip()
                active = await t.evaluate("el => el.closest('.ant-tabs-tab')?.classList.contains('ant-tabs-tab-active') ?? false")
                print(f"    [{i}] active={active} text='{txt}'")
            except Exception: pass

        # Print the actions-btn / 数据对比 area
        print("\n[4] Actions / 数据对比 buttons:")
        btns = await page.locator(".actions-btn, [class*='actions']").all()
        for i, b in enumerate(btns):
            try:
                txt = (await b.inner_text()).strip()[:80]
                cls = (await b.get_attribute("class") or "")[:60]
                vis = await b.is_visible()
                if vis:
                    print(f"    [{i}] cls='{cls}' text='{txt}'")
            except Exception: pass

        # Check state of 数据对比方案 - look for any "+ 0" or comparison indicator
        print("\n[5] Comparison scheme indicators:")
        for sel in [
            "[class*='scheme']", "[class*='compare']", "[class*='contrast']",
            ".ant-tag", ".ant-badge", "[class*='tag']",
        ]:
            cnt = await page.locator(sel).count()
            if cnt > 0:
                elems = await page.locator(sel).all()
                for el in elems:
                    try:
                        txt = (await el.inner_text()).strip()[:60]
                        vis = await el.is_visible()
                        cls = (await el.get_attribute("class") or "")[:50]
                        if vis and txt:
                            print(f"    {sel}: cls='{cls}' text='{txt}'")
                    except Exception: pass

        # Select market + date range
        print("\n[6] Setting up query …")
        current = (await page.locator(".ant-select-selection-item").nth(0).inner_text()).strip()
        if current != MARKET:
            await page.locator(".ant-select-selector").nth(0).click()
            await page.wait_for_selector(".ant-select-dropdown:not(.ant-select-dropdown-hidden)", timeout=10_000)
            await page.wait_for_timeout(300)
            _list = page.locator(".rc-virtual-list-holder").first
            for _ in range(50):
                _opt = page.locator(".ant-select-item-option-content").filter(has_text=MARKET)
                if await _opt.count() > 0:
                    await _opt.first.click(); break
                await _list.evaluate("el => { el.scrollTop += 120; }")
                await page.wait_for_timeout(80)
            await page.wait_for_timeout(400)

        start_input = page.locator("input[date-range='start']").first
        end_input   = page.locator("input[date-range='end']").first
        await start_input.click(); await page.wait_for_timeout(300)
        await start_input.fill("2026-07-05"); await start_input.press("Tab")
        await page.wait_for_timeout(300)
        await end_input.click(); await end_input.fill("2026-07-06"); await end_input.press("Enter")
        await page.wait_for_timeout(400); await page.keyboard.press("Escape"); await page.wait_for_timeout(300)

        # Click 查询
        await page.locator("button.ant-btn-primary").first.click()
        await page.wait_for_timeout(3000)
        await page.screenshot(path=str(_DEBUG_DIR / "cmp_02_after_query.png"))
        print("    Screenshot → cmp_02_after_query.png")

        # Count ALL download spans and their locations
        print("\n[7] ALL span.down-load-container elements (before any click):")
        spans = await page.locator("span.down-load-container").all()
        for i, sp in enumerate(spans):
            try:
                vis = await sp.is_visible()
                cls = (await sp.get_attribute("class") or "")
                txt = (await sp.inner_text()).strip()[:40]
                # Get parent container info
                parent_cls = await sp.evaluate("el => el.parentElement?.className || ''")
                print(f"    [{i}] vis={vis} cls='{cls}' text='{txt}' parent='{parent_cls[:60]}'")
            except Exception as e:
                print(f"    [{i}] error: {e}")

        # Try clicking each visible download span and check format
        print("\n[8] Testing each visible download span:")
        spans = await page.locator("span.down-load-container").all()
        visible_spans = []
        for i, sp in enumerate(spans):
            try:
                if await sp.is_visible():
                    visible_spans.append((i, sp))
            except Exception: pass

        print(f"    {len(visible_spans)} visible spans found")

        for idx, (orig_i, sp) in enumerate(visible_spans):
            cls = (await sp.get_attribute("class") or "")
            txt = (await sp.inner_text()).strip()[:40]
            print(f"\n    Testing span [{orig_i}] cls='{cls}' text='{txt}'")
            try:
                dl_ctx = page.expect_download(timeout=8_000)
                async with dl_ctx as dl_info:
                    await sp.click()
                dl = await dl_info.value
                fname = dl.suggested_filename
                tmp = _DEBUG_DIR / f"test_span_{orig_i}_{fname}"
                await dl.save_as(str(tmp))
                # Check columns
                import openpyxl
                wb = openpyxl.load_workbook(str(tmp), data_only=True, read_only=True)
                ws = wb.active
                hdrs = list(ws.iter_rows(max_row=1, values_only=True))[0]
                row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
                print(f"      Downloaded: {fname} ({len(hdrs)} cols)")
                print(f"      Col[2]={hdrs[2]}")
                print(f"      Col[4]={hdrs[4]}")
                if len(hdrs) > 6:
                    print(f"      Col[6]={hdrs[6]}")
                print(f"      Row2 col[4]={row2[4]}, col[6]={row2[6] if len(row2)>6 else 'N/A'}")
                wb.close()
            except Exception as e:
                print(f"      No download (timeout or error): {e}")
                await page.screenshot(path=str(_DEBUG_DIR / f"cmp_span_{orig_i}_click.png"))
                print(f"      Screenshot → cmp_span_{orig_i}_click.png")
                # Close any drawer/modal that opened
                await page.keyboard.press("Escape")
                await page.wait_for_timeout(500)

        print("\n[DONE] Close browser or wait 30s.")
        await page.wait_for_timeout(30_000)
        await browser.close()


if __name__ == "__main__":
    asyncio.run(run())
