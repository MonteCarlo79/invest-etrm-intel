"""
Province System Operation Fee ETL
==================================
Parses 各省市电网系统运行费用 Excel files and upserts into
province_sysopfee_monthly, which feeds the bess-map 系统运行费 tab.

Standard format (202501-202605 layout):
  Row 0: title
  Row 1: year headers (省市 | 2025 | ... | 2026 | ...)
  Row 2: month headers (None | 1月 | 2月 | ... | 12月 | 均值 | 1月 | ... | 均值)
  Rows 3+: province data

Entry points:
  upsert_sysopfee(file_bytes, filename, pg_url, api_key)  — from Feishu upload
  upsert_sysopfee_rows(rows, pg_url, source_file)          — from web screener
  is_sysopfee_file(filename)                               — file detection
"""
from __future__ import annotations

import io
import json
import logging
import re
from datetime import date
from typing import Optional

import openpyxl
import psycopg2

logger = logging.getLogger(__name__)

# ── DB ─────────────────────────────────────────────────────────────────────────

_ENSURE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS province_sysopfee_monthly (
    id           SERIAL PRIMARY KEY,
    province     TEXT    NOT NULL,
    year_month   DATE    NOT NULL,
    fee_yuan_kwh NUMERIC,
    source_file  TEXT,
    ingested_at  TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    UNIQUE (province, year_month)
);
CREATE INDEX IF NOT EXISTS idx_psm_province_ym
    ON province_sysopfee_monthly (province, year_month DESC);
"""

_UPSERT_SQL = """
INSERT INTO province_sysopfee_monthly (province, year_month, fee_yuan_kwh, source_file)
VALUES (%s, %s, %s, %s)
ON CONFLICT (province, year_month) DO UPDATE SET
    fee_yuan_kwh = EXCLUDED.fee_yuan_kwh,
    source_file  = EXCLUDED.source_file,
    ingested_at  = NOW()
"""

# ── Known column layout ────────────────────────────────────────────────────────
# Standard 202501-202605 format:
#   col 1-12  → 2025 Jan-Dec
#   col 13    → 2025 annual average (skip)
#   col 14-18 → 2026 Jan-May
#   col 19    → 2026 partial average (skip)

_SKIP_KEYWORDS = {"合计", "total", "均值", "小计", "全国", "average", "avg", "省市", "汇总"}


def _build_col_map(header_row2: tuple, header_row3: tuple) -> dict[int, date]:
    """
    Infer column → year_month from the two header rows.
    header_row2: (None/省市, year1, None, ..., year2, ...)
    header_row3: (None, 1月, 2月, ..., 12月, 均值, 1月, ...)
    Falls back to the hard-coded standard layout if detection fails.
    """
    col_map: dict[int, date] = {}
    current_year = None
    for i, (y_cell, m_cell) in enumerate(zip(header_row2, header_row3)):
        if isinstance(y_cell, (int, float)) and 2000 <= int(y_cell) <= 2100:
            current_year = int(y_cell)
        if current_year and isinstance(m_cell, str):
            m_str = m_cell.strip()
            mo_m = re.match(r'^(\d{1,2})月$', m_str)
            if mo_m:
                mo = int(mo_m.group(1))
                try:
                    col_map[i] = date(current_year, mo, 1)
                except ValueError:
                    pass
    if not col_map:
        # Fall back to hard-coded standard layout
        standard: list[tuple[int, int]] = [
            (1,2025,1),(2,2025,2),(3,2025,3),(4,2025,4),(5,2025,5),(6,2025,6),
            (7,2025,7),(8,2025,8),(9,2025,9),(10,2025,10),(11,2025,11),(12,2025,12),
            (14,2026,1),(15,2026,2),(16,2026,3),(17,2026,4),(18,2026,5),
        ]
        col_map = {c: date(y, m, 1) for c, y, m in standard}
    return col_map


def _parse_excel(file_bytes: bytes) -> list[dict]:
    """Parse wide-format sysopfee Excel → list of {province, year_month, fee_yuan_kwh}."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    all_rows = [row for row in ws.iter_rows(values_only=True)]

    # Find header rows: look for the row containing month markers (1月, 2月, ...)
    month_row_idx = None
    for idx, row in enumerate(all_rows):
        if sum(1 for c in row if isinstance(c, str) and re.match(r'^\d{1,2}月$', c.strip())) >= 3:
            month_row_idx = idx
            break

    if month_row_idx is None or month_row_idx < 1:
        logger.warning("Could not find month header row — trying default col map")
        col_map = _build_col_map((), ())
        data_start = 3
    else:
        year_row = all_rows[month_row_idx - 1] if month_row_idx >= 1 else ()
        col_map = _build_col_map(year_row, all_rows[month_row_idx])
        data_start = month_row_idx + 1

    rows: list[dict] = []
    for row in all_rows[data_start:]:
        province = row[0] if row else None
        if not isinstance(province, str) or not province.strip():
            continue
        province = province.strip()
        if any(kw in province for kw in _SKIP_KEYWORDS):
            continue
        for col_idx, ym in col_map.items():
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None or val == "":
                continue
            try:
                fee = float(val)
            except (TypeError, ValueError):
                continue
            rows.append({"province": province, "year_month": ym, "fee_yuan_kwh": fee})
    return rows


def _excel_to_text(file_bytes: bytes, max_rows: int = 80) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                sheet_rows.append("\t".join(cells))
            if len(sheet_rows) >= max_rows:
                break
        if sheet_rows:
            parts.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(sheet_rows))
    return "\n\n".join(parts)


def _parse_with_llm(text: str, filename: str, api_key: str) -> list[dict]:
    from anthropic import Anthropic
    client = Anthropic(api_key=api_key)
    resp = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=1500,
        system="""You are a data extraction assistant for China electricity market data.
Extract province-level 系统运行费 (grid system operation fee) in yuan/kWh.

Return ONLY a valid JSON array:
[{"province": "省份", "year_month": "YYYY-MM-01", "fee_yuan_kwh": 数值}]

Rules:
- province: Chinese province/grid name as in the file
- year_month: first day of the month (e.g. "2025-03-01")
- fee_yuan_kwh: numeric value in yuan/kWh (typical range 0.001–0.20)
- SKIP rows that are averages (均值), totals (合计/小计), or headers
- Return [] if no valid data found""",
        messages=[{"role": "user", "content": f"File: {filename}\n\nContent:\n{text[:5000]}"}],
    )
    raw = resp.content[0].text.strip()
    m = re.search(r'\[.*\]', raw, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(0))
        except json.JSONDecodeError:
            logger.warning("LLM returned invalid JSON for sysopfee ETL: %s", raw[:200])
    return []


# ── Public upsert helpers ──────────────────────────────────────────────────────

def upsert_sysopfee_rows(rows: list[dict], pg_url: str, source_file: str) -> dict:
    """Upsert pre-parsed rows directly (no LLM step).

    Each row: {"province": str, "year_month": date | "YYYY-MM-DD", "fee_yuan_kwh": float}
    Returns {"upserted": int, "rows": list[str], "errors": list[str]}
    """
    errors: list[str] = []
    upserted: list[str] = []

    conn = psycopg2.connect(pg_url)
    try:
        with conn.cursor() as cur:
            cur.execute(_ENSURE_TABLE_SQL)
            for row in rows:
                province = str(row.get("province", "")).strip()
                if not province or any(kw in province for kw in _SKIP_KEYWORDS):
                    continue
                ym = row.get("year_month")
                if isinstance(ym, str):
                    try:
                        ym = date.fromisoformat(ym[:10])
                    except ValueError:
                        continue
                if not isinstance(ym, date):
                    continue
                fee = row.get("fee_yuan_kwh")
                if fee is None:
                    continue
                try:
                    fee = float(fee)
                except (TypeError, ValueError):
                    continue
                try:
                    cur.execute(_UPSERT_SQL, (province, ym, fee, source_file))
                    upserted.append(f"{province}/{ym}")
                except Exception as exc:
                    errors.append(f"{province}/{ym}: {exc}")
                    logger.error("sysopfee upsert failed for %s/%s: %s", province, ym, exc)
        conn.commit()
    finally:
        conn.close()

    return {"upserted": len(upserted), "rows": upserted, "errors": errors}


def upsert_sysopfee(
    file_bytes: bytes,
    filename: str,
    pg_url: str,
    api_key: str,
) -> dict:
    """Parse sysopfee Excel and upsert to province_sysopfee_monthly.

    Returns {"upserted": int, "rows": list[str], "errors": list[str]}
    """
    try:
        rows = _parse_excel(file_bytes)
    except Exception as exc:
        logger.error("sysopfee Excel parse failed for %s: %s", filename, exc)
        rows = []

    if not rows:
        logger.info("Standard parse yielded no rows for %s — trying LLM fallback", filename)
        try:
            text = _excel_to_text(file_bytes)
            llm_rows = _parse_with_llm(text, filename, api_key)
            for r in llm_rows:
                ym_str = r.get("year_month", "")
                try:
                    r["year_month"] = date.fromisoformat(ym_str[:10])
                    rows.append(r)
                except (ValueError, TypeError):
                    continue
        except Exception as exc:
            logger.error("sysopfee LLM fallback failed for %s: %s", filename, exc)

    if not rows:
        return {"upserted": 0, "rows": [], "errors": ["No data extracted from file"]}

    return upsert_sysopfee_rows(rows, pg_url, filename)


def is_sysopfee_file(filename: str) -> bool:
    """Return True if filename looks like a 系统运行费 Excel file."""
    name = filename.lower()
    if not (name.endswith(".xlsx") or name.endswith(".xls")):
        return False
    keywords = ["系统运行费", "sysopfee", "system_op_fee", "运行费用", "system_operation"]
    return any(kw in name for kw in keywords)
