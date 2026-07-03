"""
Manual 各省装机容量 Data Entry
==============================
Extracts per-province installed capacity (wind/solar/thermal/hydro/nuclear/BESS)
from free-form text, PDF/Excel/TXT files, or URLs, and upserts into
province_installed_monthly.

Entry points:
  extract_capacity_from_text(text, api_key, pg_url, year_month?)  → summary dict
  extract_capacity_from_file(filename, file_bytes, api_key, pg_url) → summary dict
  extract_capacity_from_url(url, api_key, pg_url) → summary dict
  is_capacity_file_extended(filename) → bool  (widens is_capacity_file to PDF/TXT)
"""
from __future__ import annotations

import io
import json
import logging
import re
from datetime import date
from typing import Optional

import anthropic

logger = logging.getLogger(__name__)

# ── Extended filename detection ────────────────────────────────────────────────

_CAPACITY_KEYWORDS = [
    "装机容量", "各省装机", "储能装机", "installed_cap", "installed_capacity",
    "province_cap", "capacity_scan",
    "装机数据", "装机情况", "装机统计",
]


def is_capacity_file_extended(filename: str) -> bool:
    """
    Return True if filename suggests province capacity data.
    Widens the existing is_capacity_file() to also match PDF/TXT (not just Excel).
    """
    fn = filename.lower()
    ext = fn.rsplit(".", 1)[-1] if "." in fn else ""
    if ext not in ("xlsx", "xls", "xlsm", "pdf", "txt", "csv"):
        return False
    return any(kw.lower() in fn for kw in _CAPACITY_KEYWORDS)


# ── Text extraction from bytes ─────────────────────────────────────────────────

def _text_from_pdf(file_bytes: bytes) -> str:
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            return "\n\n".join(
                (page.extract_text() or "") for page in pdf.pages[:30]
            )
    except Exception as exc:
        logger.warning("PDF text extraction failed: %s", exc)
        return ""


def _text_from_excel(file_bytes: bytes) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    rows.append("\t".join(cells))
                if len(rows) >= 150:
                    break
            if rows:
                parts.append(f"=== {sheet_name} ===\n" + "\n".join(rows))
        return "\n\n".join(parts)
    except Exception as exc:
        logger.warning("Excel text extraction failed: %s", exc)
        return ""


def _text_from_bytes(filename: str, file_bytes: bytes) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "pdf":
        return _text_from_pdf(file_bytes)
    if ext in ("xlsx", "xls", "xlsm"):
        return _text_from_excel(file_bytes)
    for enc in ("utf-8", "gbk", "gb2312"):
        try:
            return file_bytes.decode(enc)
        except Exception:
            pass
    return file_bytes.decode("utf-8", errors="replace")


# ── Claude extraction ──────────────────────────────────────────────────────────

_SYSTEM = (
    "你是中国电力市场数据提取专家，专注于各省装机容量数据。"
    "从提供的文本中准确提取各省每个月的装机容量数据。"
    "Respond ONLY with valid JSON, no other text."
)

_PROMPT = """\
今天日期：{today}

以下是待提取的文本：

{context}

请提取所有省份的装机容量数据，以JSON格式回答：
{{
  "records": [
    {{
      "province": "省份名称",
      "year_month": "YYYY-MM",
      "wind_mw": <风电装机，MW，数字或null>,
      "solar_mw": <光伏装机，MW，数字或null>,
      "thermal_mw": <火电装机，MW，数字或null>,
      "hydro_mw": <水电装机，MW，数字或null>,
      "nuclear_mw": <核电装机，MW，数字或null>,
      "bess_mw": <储能装机，MW，数字或null>,
      "total_mw": <合计装机，MW，数字或null>
    }}
  ]
}}

规则：
- province：使用标准省份名称（如广东、山西、内蒙古、蒙西、蒙东、冀北、河北等）
- year_month：从文本中推断年月，格式YYYY-MM（如2026-05）。若文本未明确月份，使用今天日期的当月（{today_ym}）
- 单位转换：万kW×10=MW，万MW极少见勿用；GW×1000=MW
- 跳过合计、小计、全国、total等汇总行
- 只填写文本中实际出现的数值，无数据填null
- 若文本含多省多月数据，每个省+月组合单独一条记录
- 若文本无有效装机数据，records返回空列表[]
"""


def _call_claude(context: str, api_key: str) -> Optional[dict]:
    """Call Claude Haiku to extract province capacity data."""
    from datetime import date as _date
    today = _date.today()
    context = context[:12000]
    prompt = _PROMPT.format(
        context=context,
        today=today.strftime("%Y-%m-%d"),
        today_ym=today.strftime("%Y-%m"),
    )
    try:
        client = anthropic.Anthropic(api_key=api_key)
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=2048,
            system=_SYSTEM,
            messages=[{"role": "user", "content": prompt}],
        )
        text = resp.content[0].text.strip()
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
        return json.loads(text)
    except Exception as exc:
        logger.error("Claude capacity extraction failed: %s", exc)
        return None


# ── Row validation + upsert ────────────────────────────────────────────────────

def _parse_year_month(val: str) -> Optional[date]:
    """Parse 'YYYY-MM' or 'YYYY-M' string to date(YYYY, M, 1)."""
    m = re.match(r"(\d{4})-(\d{1,2})$", str(val).strip())
    if m:
        yr, mo = int(m.group(1)), int(m.group(2))
        if 2015 <= yr <= 2035 and 1 <= mo <= 12:
            return date(yr, mo, 1)
    return None


def _safe_mw(val) -> Optional[float]:
    if val is None or val == "":
        return None
    try:
        f = float(val)
        return f if f > 0 else None
    except (TypeError, ValueError):
        return None


_UPSERT_SQL = """
INSERT INTO marketdata.province_installed_monthly
    (province, year_month, wind_mw, solar_mw, thermal_mw, hydro_mw,
     nuclear_mw, bess_mw, total_mw, source_file)
VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
ON CONFLICT (province, year_month) DO UPDATE SET
    wind_mw     = COALESCE(EXCLUDED.wind_mw,    marketdata.province_installed_monthly.wind_mw),
    solar_mw    = COALESCE(EXCLUDED.solar_mw,   marketdata.province_installed_monthly.solar_mw),
    thermal_mw  = COALESCE(EXCLUDED.thermal_mw, marketdata.province_installed_monthly.thermal_mw),
    hydro_mw    = COALESCE(EXCLUDED.hydro_mw,   marketdata.province_installed_monthly.hydro_mw),
    nuclear_mw  = COALESCE(EXCLUDED.nuclear_mw, marketdata.province_installed_monthly.nuclear_mw),
    bess_mw     = COALESCE(EXCLUDED.bess_mw,    marketdata.province_installed_monthly.bess_mw),
    total_mw    = COALESCE(EXCLUDED.total_mw,   marketdata.province_installed_monthly.total_mw),
    source_file = EXCLUDED.source_file,
    ingested_at = NOW()
"""


def _upsert_records(records: list[dict], pg_url: str, source_name: str) -> dict:
    import psycopg2
    upserted = []
    errors = []
    conn = psycopg2.connect(pg_url)
    try:
        with conn.cursor() as cur:
            for rec in records:
                province = str(rec.get("province", "")).strip()
                if not province:
                    continue
                ym = _parse_year_month(rec.get("year_month", ""))
                if ym is None:
                    errors.append(f"{province}: invalid year_month '{rec.get('year_month')}'")
                    continue
                # Skip totals
                if any(kw in province for kw in ("合计", "小计", "全国", "total", "汇总")):
                    continue
                try:
                    cur.execute(_UPSERT_SQL, (
                        province, ym,
                        _safe_mw(rec.get("wind_mw")),
                        _safe_mw(rec.get("solar_mw")),
                        _safe_mw(rec.get("thermal_mw")),
                        _safe_mw(rec.get("hydro_mw")),
                        _safe_mw(rec.get("nuclear_mw")),
                        _safe_mw(rec.get("bess_mw")),
                        _safe_mw(rec.get("total_mw")),
                        source_name[:200],
                    ))
                    upserted.append(f"{province}({ym.strftime('%Y-%m')})")
                except Exception as exc:
                    errors.append(f"{province}: {exc}")
        conn.commit()
    finally:
        conn.close()
    return {"upserted": len(upserted), "provinces": upserted, "errors": errors}


def _process(data: Optional[dict], pg_url: str, source_name: str) -> dict:
    if not data:
        return {"upserted": 0, "provinces": [], "errors": ["Claude未返回有效JSON"]}
    records = data.get("records") or []
    if not records:
        return {"upserted": 0, "provinces": [], "errors": ["未找到有效装机数据"]}
    return _upsert_records(records, pg_url, source_name)


# ── Public entry points ────────────────────────────────────────────────────────

def extract_capacity_from_text(
    text: str,
    api_key: str,
    pg_url: str,
    year_month: Optional[date] = None,
) -> dict:
    """Extract province capacity data from a free-form text message and upsert."""
    # Prepend year_month hint if caller supplies it
    if year_month:
        text = f"数据月份：{year_month.strftime('%Y-%m')}\n\n{text}"
    data = _call_claude(text, api_key)
    return _process(data, pg_url, f"manual_text")


def extract_capacity_from_file(
    filename: str,
    file_bytes: bytes,
    api_key: str,
    pg_url: str,
) -> dict:
    """Extract province capacity data from an uploaded file and upsert."""
    text = _text_from_bytes(filename, file_bytes)
    if not text.strip():
        return {"upserted": 0, "provinces": [], "errors": ["文件无法提取文本内容"]}
    # Prepend filename as context hint for year_month inference
    text = f"文件名：{filename}\n\n{text}"
    data = _call_claude(text, api_key)
    return _process(data, pg_url, f"manual_file:{filename[:80]}")


def extract_capacity_from_url(
    url: str,
    api_key: str,
    pg_url: str,
) -> dict:
    """Fetch a URL and extract province capacity data."""
    import requests
    try:
        resp = requests.get(url, timeout=20, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        content_type = resp.headers.get("content-type", "")
        if "pdf" in content_type or url.lower().endswith(".pdf"):
            text = _text_from_pdf(resp.content)
        elif "html" in content_type:
            text = re.sub(r"<[^>]+>", " ", resp.text)
            text = re.sub(r"\s+", " ", text)
        else:
            text = resp.text
    except Exception as exc:
        return {"upserted": 0, "provinces": [], "errors": [f"URL获取失败：{exc}"]}
    text = f"来源URL：{url}\n\n{text}"
    data = _call_claude(text, api_key)
    return _process(data, pg_url, f"manual_url:{url[:100]}")


def format_result_message(result: dict) -> str:
    """Format extraction result as a human-readable message."""
    lines = []
    if result["upserted"] > 0:
        # Group by province name (strip year_month suffix for display)
        provs = list(dict.fromkeys(p.split("(")[0] for p in result["provinces"]))
        prov_str = "、".join(provs[:10])
        if len(provs) > 10:
            prov_str += f"等{len(provs)}省"
        lines.append(
            f"✅ 装机数据已入库：{result['upserted']} 条\n"
            f"涉及省份：{prov_str}\n"
            f"bess-map 储能需求Tab已自动更新。"
        )
    else:
        lines.append("⚠️ 未找到可入库的装机容量数据")
        lines.append(
            "支持格式：「广东 2026-05 储能2000MW 风电5000MW 光伏8000MW」\n"
            "或上传含装机数据的PDF/Excel/TXT文件"
        )
    if result["errors"]:
        lines.append(f"错误：{result['errors'][0]}")
    return "\n".join(lines)
