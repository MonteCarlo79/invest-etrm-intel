"""
Daily BESS screener for Mengxi (Inner Mongolia) market.

Checks the first 15-min interval (datetime = data_date 00:00:00) of the latest data_date.
Finds plant_names NOT already listed in OneDrive data/电站.xlsx.
Confirms BESS signature (both charge + discharge) over the full day.
Appends new plants to 电站.xlsx and notifies via Feishu.

Schedule: daily at 06:30 UTC (14:30 Beijing UTC+8), after market data typically arrives.
"""
from __future__ import annotations

import io
import logging
from datetime import date, timedelta
from typing import Optional

import psycopg2

logger = logging.getLogger(__name__)


_BESS_EXCEL_PATH = "bess-platform/data/电站.xlsx"


def _load_known_plant_names(onedrive_client) -> tuple[set[str], object, bytes]:
    """Download 电站.xlsx, return (known_plant_names, workbook, raw_bytes)."""
    import openpyxl

    raw = onedrive_client.read_file_by_path(_BESS_EXCEL_PATH)

    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    known: set[str] = set()
    first_row = True
    for row in ws.iter_rows(values_only=True):
        if first_row:
            first_row = False
            continue
        plant_name = row[1] if len(row) > 1 else None
        if plant_name:
            known.add(str(plant_name).strip())
    return known, wb, raw


def _get_latest_data_date(pg_url: str) -> Optional[date]:
    conn = psycopg2.connect(pg_url, options="-c statement_timeout=30000")
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT MAX(data_date) FROM marketdata.md_id_cleared_energy")
            row = cur.fetchone()
            return row[0] if row and row[0] else None
    finally:
        conn.close()


def _find_new_bess_candidates(pg_url: str, screen_date: date, known_names: set[str]) -> list[tuple[str, int]]:
    """
    Step 1: Plants active at 00:00 on screen_date, not in known_names.
    Step 2: Confirm BESS (both charge + discharge over full screen_date).
    Returns list of (plant_name, inferred_mw).
    """
    from datetime import datetime as dt

    first_interval = dt.combine(screen_date, __import__("datetime").time(0, 0))
    screen_date_end = screen_date + timedelta(days=1)

    conn = psycopg2.connect(pg_url, options="-c statement_timeout=60000")
    try:
        with conn.cursor() as cur:
            # Step 1: plants at 00:00 on screen_date not in known list
            cur.execute(
                """
                SELECT DISTINCT plant_name
                FROM marketdata.md_id_cleared_energy
                WHERE data_date = %s
                  AND datetime = %s
                """,
                (screen_date, first_interval),
            )
            all_plants = {r[0] for r in cur.fetchall()}
            candidates = [p for p in all_plants if p not in known_names]

            if not candidates:
                return []

            logger.info("Screener: %d candidate(s) not in 电站.xlsx: %s", len(candidates), candidates)

            # Step 2: confirm BESS — must have both positive and negative cleared_energy
            cur.execute(
                """
                SELECT plant_name, ROUND(MAX(cleared_energy_mwh)::numeric, 0) AS inferred_mw
                FROM marketdata.md_id_cleared_energy
                WHERE data_date = %s
                  AND datetime >= %s AND datetime < %s
                  AND plant_name = ANY(%s)
                GROUP BY plant_name
                HAVING MAX(cleared_energy_mwh) > 0 AND MIN(cleared_energy_mwh) < 0
                """,
                (screen_date, screen_date, screen_date_end, candidates),
            )
            confirmed = cur.fetchall()
    finally:
        conn.close()

    return [(row[0], int(row[1])) for row in confirmed]


def _append_to_excel(wb, new_plants: list[tuple[str, int]], screen_date: date) -> bytes:
    """Append new plant rows to the workbook and return updated bytes."""
    import openpyxl

    ws = wb.active

    # Determine next ID
    max_id = 0
    first_row = True
    for row in ws.iter_rows(values_only=True):
        if first_row:
            first_row = False
            continue
        id_val = row[0]
        if id_val and isinstance(id_val, (int, float)):
            max_id = max(max_id, int(id_val))

    for plant_name, inferred_mw in new_plants:
        next_id = max_id + 1
        max_id = next_id
        # Columns: ID, 储能电站名称, (col3), 省份, 位置, 业主, 运营方, 装机MW, 装机容量
        ws.append([
            next_id,        # ID
            plant_name,     # 储能电站名称
            None,           # col3 (alias/pinyin — unknown)
            "内蒙古西部",    # 省份 (screener only covers Mengxi)
            None,           # 位置
            None,           # 业主 (unknown, to be filled manually)
            None,           # 运营方
            inferred_mw,    # 装机MW
            inferred_mw * 4,  # 装机容量MWh (assume 4h duration)
        ])
        logger.info("Screener: appending new BESS to 电站.xlsx: %s (%d MW)", plant_name, inferred_mw)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def screen_new_bess(
    pg_url: str,
    onedrive_client,
    feishu=None,
    owner_open_id: str = "",
) -> list[str]:
    """
    Main entry point. Returns list of newly detected BESS plant_names (empty on most days).
    Appends new plants to OneDrive 电站.xlsx and notifies via Feishu.
    """
    # ── Load known plants ──────────────────────────────────────────────────────
    try:
        known_names, wb, _raw = _load_known_plant_names(onedrive_client)
    except Exception as exc:
        logger.error("Screener: failed to load 电站.xlsx: %s", exc)
        if feishu and owner_open_id:
            feishu.send_text(owner_open_id, f"⚠️ 蒙西BESS筛查失败（无法读取电站.xlsx）：{exc}")
        return []

    logger.info("Screener: %d plants currently in 电站.xlsx", len(known_names))

    # ── Get latest data date ───────────────────────────────────────────────────
    try:
        screen_date = _get_latest_data_date(pg_url)
    except Exception as exc:
        logger.error("Screener: failed to get latest data date: %s", exc)
        return []

    if screen_date is None:
        logger.warning("Screener: no data in md_id_cleared_energy — skipping")
        return []

    logger.info("Screener: screening date %s", screen_date)

    # ── Find new BESS ─────────────────────────────────────────────────────────
    try:
        new_plants = _find_new_bess_candidates(pg_url, screen_date, known_names)
    except Exception as exc:
        logger.error("Screener: DB query failed: %s", exc)
        if feishu and owner_open_id:
            feishu.send_text(owner_open_id, f"⚠️ 蒙西BESS筛查DB查询失败：{exc}")
        return []

    if not new_plants:
        logger.info("Screener: no new BESS found on %s", screen_date)
        return []

    # ── Update 电站.xlsx on OneDrive ──────────────────────────────────────────
    try:
        updated_bytes = _append_to_excel(wb, new_plants, screen_date)
        _folder, _fname = _BESS_EXCEL_PATH.rsplit("/", 1)
        onedrive_client.upload_file(
            folder_path=_folder,
            filename=_fname,
            content=updated_bytes,
            conflict_behavior="replace",
        )
        logger.info("Screener: 电站.xlsx updated with %d new plant(s)", len(new_plants))
    except Exception as exc:
        logger.error("Screener: failed to update 电站.xlsx on OneDrive: %s", exc)
        if feishu and owner_open_id:
            feishu.send_text(owner_open_id, f"⚠️ 蒙西BESS筛查：检测到新电站但更新电站.xlsx失败：{exc}")
        return []

    # ── Notify via Feishu ─────────────────────────────────────────────────────
    new_names = [p[0] for p in new_plants]
    if feishu and owner_open_id:
        lines = [f"📡 蒙西BESS筛查：发现 {len(new_names)} 个新储能电站（{screen_date}）"]
        for name, mw in new_plants:
            lines.append(f"  • {name}（推算装机：{mw} MW）")
        lines.append("已自动添加至 OneDrive data/电站.xlsx，请补充业主及运营方信息。")
        feishu.send_text(owner_open_id, "\n".join(lines))

    return new_names
