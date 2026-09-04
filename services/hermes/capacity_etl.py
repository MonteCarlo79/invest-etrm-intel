"""
BESS / Generation Installed Capacity ETL
=========================================
Parses 各省储能装机容量 Excel files (any format) and upserts into
province_installed_monthly, which feeds the bess-map BESS Demand tab.

Uses Claude Haiku to extract structured data — handles format variations
without brittle column-index logic.
"""
from __future__ import annotations

import io
import json
import logging
import re
from datetime import date
from typing import Optional

import openpyxl
import psycopg2

logger = logging.getLogger(__name__)

# ── Province name normalisation ───────────────────────────────────────────────
# Maps common aliases → canonical name stored in province_installed_monthly
_PROVINCE_NORM: dict[str, str] = {
    # Direct matches
    "山西": "山西", "吉林": "吉林", "山东": "山东", "青海": "青海",
    "新疆": "新疆", "海南": "海南", "内蒙古": "内蒙古", "广东": "广东",
    "浙江": "浙江", "湖南": "湖南", "安徽": "安徽", "福建": "福建",
    "江苏": "江苏", "四川": "四川", "湖北": "湖北", "广西": "广西",
    "贵州": "贵州", "云南": "云南", "河南": "河南", "江西": "江西",
    "陕西": "陕西", "北京": "北京", "上海": "上海", "天津": "天津",
    "重庆": "重庆", "辽宁": "辽宁", "黑龙江": "黑龙江", "甘肃": "甘肃",
    "宁夏": "宁夏", "西藏": "西藏", "吉林": "吉林", "河北": "河北",
    # Aliases
    "新疆（疆内）": "新疆", "疆内": "新疆", "新疆疆内": "新疆",
    "河北南网": "河北南网", "河北南部": "河北南网", "冀南": "河北南网",
    "冀北": "冀北", "冀北电网": "冀北", "国网冀北": "冀北",
    "蒙西": "蒙西", "蒙东": "蒙东",
    "南方电网": "南方电网", "华北": "华北",
}

# Keywords that suggest a row is a total/header (should be skipped)
_SKIP_KEYWORDS = {"合计", "total", "小计", "全国", "合", "sum", "汇总", "统计"}


def _normalise_province(raw: str) -> Optional[str]:
    raw = raw.strip()
    if raw in _PROVINCE_NORM:
        return _PROVINCE_NORM[raw]
    # Partial match for 2–3 char prefixes
    for key, val in _PROVINCE_NORM.items():
        if raw.startswith(key) or key.startswith(raw):
            return val
    return raw  # Keep as-is if unknown — DB will store it


def _extract_year_month_from_filename(filename: str) -> date:
    """Extract year-month from filename. Falls back to current month.

    Handles formats:
    - YYYYMM or YYYYMMDD digits (e.g. 202605, 20260501)
    - YYYY年M月 / YYYY年MM月 (e.g. 2026年5月, 2026年05月)
    - YYYY-MM or YYYY/MM
    """
    # YYYY年M月 / YYYY年MM月 — must check before generic digit scan
    m = re.search(r'(\d{4})[年/-](\d{1,2})[月/-]?', filename)
    if m:
        yr, mo = int(m.group(1)), int(m.group(2))
        if 2000 <= yr <= 2100 and 1 <= mo <= 12:
            return date(yr, mo, 1)
    # 6-digit YYYYMM or 8-digit YYYYMMDD
    m = re.search(r'(\d{4})(\d{2})(?:\d{2})?', filename)
    if m:
        yr, mo = int(m.group(1)), int(m.group(2))
        if 2000 <= yr <= 2100 and 1 <= mo <= 12:
            return date(yr, mo, 1)
    today = date.today()
    return date(today.year, today.month, 1)


# Province name patterns for detecting single-province files
_PROVINCE_NAMES = [
    "黑龙江", "内蒙古东", "内蒙古西", "蒙东", "蒙西",
    "北京", "天津", "山西", "山东", "辽宁", "吉林", "上海",
    "江苏", "浙江", "安徽", "福建", "江西", "河南", "湖北",
    "湖南", "广东", "广西", "海南", "重庆", "四川", "贵州",
    "云南", "陕西", "甘肃", "青海", "宁夏", "新疆", "西藏",
    "冀北", "冀南", "河北",
]


def province_from_filename(filename: str) -> Optional[str]:
    """Return the Chinese province name found in the filename, or None."""
    for prov in _PROVINCE_NAMES:
        if prov in filename:
            return prov
    return None


def _excel_to_text(file_bytes: bytes, max_rows: int = 120) -> str:
    """Convert all sheets to tab-separated text for LLM parsing."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append("\t".join(cells))
            if len(rows) >= max_rows:
                break
        if rows:
            parts.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows))
    return "\n\n".join(parts)


def _parse_with_llm(text: str, filename: str, api_key: str,
                    province_hint: Optional[str] = None,
                    year_month_hint: Optional[date] = None) -> list[dict]:
    """Use Claude Haiku to extract structured capacity data from raw Excel text."""
    from shared.anthropic_client import make_client as _make_anthropic_client

    client = _make_anthropic_client(api_key)

    # Build context hints so LLM can resolve "5月" → correct year
    hint_lines = []
    if province_hint:
        hint_lines.append(f"Province: {province_hint} (this is a single-province file)")
    if year_month_hint:
        hint_lines.append(
            f"Target month from filename: {year_month_hint.strftime('%Y年%m月')} "
            f"({year_month_hint.year}-{year_month_hint.month:02d}). "
            f"Use this year when the file only shows a month number (e.g. '5月' means {year_month_hint.year}-{year_month_hint.month:02d})."
        )
    hint_block = "\n".join(hint_lines)

    resp = client.messages.create(
        model="claude-sonnet-4-6",  # haiku-4-5 requires use-case form on this Bedrock account
        max_tokens=1000,
        system="""You are a data extraction assistant for China electricity market data.
Extract province-level installed capacity in MW from the Excel content provided.

Return ONLY a valid JSON array with this structure:
[{"province": "省份", "bess_mw": 数值, "wind_mw": 数值或null, "solar_mw": 数值或null, "thermal_mw": 数值或null, "hydro_mw": 数值或null, "nuclear_mw": 数值或null, "total_mw": 数值或null}]

Rules:
- province: Chinese province/grid name as shown in the file (or from the hint if single-province)
- All MW values: numeric. Convert units if needed: 万kW×10=MW, GW×1000=MW
- Set missing columns to null (not 0)
- SKIP rows that are subtotals, grand totals, headers (合计/小计/全国/total etc.)
- SKIP rows where all capacity values are 0 or null
- For single-province transposed format (rows = fuel types, one data column): extract all fuel types into one record for the target month
- Return [] if no valid province data found""",
        messages=[{
            "role": "user",
            "content": (
                f"File: {filename}\n"
                + (f"{hint_block}\n" if hint_block else "")
                + f"\nContent:\n{text[:5000]}"
            ),
        }],
    )

    raw = resp.content[0].text.strip()
    m = re.search(r'\[.*\]', raw, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(0))
        except json.JSONDecodeError:
            logger.warning("LLM returned invalid JSON for capacity ETL: %s", raw[:200])
    return []


# ── DB upsert ─────────────────────────────────────────────────────────────────

_UPSERT_SQL = """
INSERT INTO marketdata.province_installed_monthly
    (province, year_month, wind_mw, solar_mw, thermal_mw, hydro_mw,
     nuclear_mw, bess_mw, total_mw, source_file)
VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
ON CONFLICT (province, year_month) DO UPDATE SET
    wind_mw     = COALESCE(EXCLUDED.wind_mw,    marketdata.province_installed_monthly.wind_mw),
    solar_mw    = COALESCE(EXCLUDED.solar_mw,   marketdata.province_installed_monthly.solar_mw),
    thermal_mw  = COALESCE(EXCLUDED.thermal_mw, marketdata.province_installed_monthly.thermal_mw),
    hydro_mw    = COALESCE(EXCLUDED.hydro_mw,   marketdata.province_installed_monthly.hydro_mw),
    nuclear_mw  = COALESCE(EXCLUDED.nuclear_mw, marketdata.province_installed_monthly.nuclear_mw),
    bess_mw     = COALESCE(EXCLUDED.bess_mw,    marketdata.province_installed_monthly.bess_mw),
    total_mw    = COALESCE(EXCLUDED.total_mw,   marketdata.province_installed_monthly.total_mw),
    source_file = EXCLUDED.source_file,
    ingested_at = NOW()
"""

_ENSURE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS marketdata.province_installed_monthly (
    id          SERIAL PRIMARY KEY,
    province    TEXT    NOT NULL,
    year_month  DATE    NOT NULL,
    wind_mw     NUMERIC,
    solar_mw    NUMERIC,
    thermal_mw  NUMERIC,
    hydro_mw    NUMERIC,
    nuclear_mw  NUMERIC,
    bess_mw     NUMERIC,
    total_mw    NUMERIC,
    source_file TEXT,
    ingested_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    UNIQUE (province, year_month)
);
CREATE INDEX IF NOT EXISTS idx_pim_province_ym
    ON marketdata.province_installed_monthly (province, year_month DESC);
"""


def upsert_capacity(
    file_bytes: bytes,
    filename: str,
    pg_url: str,
    api_key: str,
    year_month: Optional[date] = None,
) -> dict:
    """Parse capacity Excel and upsert into province_installed_monthly.

    Returns:
        {"upserted": int, "provinces": list[str], "year_month": str, "errors": list[str]}
    """
    if year_month is None:
        year_month = _extract_year_month_from_filename(filename)

    province_hint = province_from_filename(filename)

    text = _excel_to_text(file_bytes)
    if not text.strip():
        return {"upserted": 0, "provinces": [], "year_month": str(year_month),
                "errors": ["Could not read Excel content"]}

    rows = _parse_with_llm(text, filename, api_key,
                           province_hint=province_hint,
                           year_month_hint=year_month)
    if not rows:
        return {"upserted": 0, "provinces": [], "year_month": str(year_month),
                "errors": ["LLM extracted no province data from file"]}

    errors: list[str] = []
    upserted: list[str] = []

    conn = psycopg2.connect(pg_url)
    try:
        with conn.cursor() as cur:
            cur.execute(_ENSURE_TABLE_SQL)

            for row in rows:
                province_raw = str(row.get("province", "")).strip()
                if not province_raw:
                    continue
                # Skip totals/headers
                if any(kw in province_raw.lower() for kw in _SKIP_KEYWORDS):
                    continue

                province = _normalise_province(province_raw)

                def _mw(key: str) -> Optional[float]:
                    v = row.get(key)
                    if v is None or v == "":
                        return None
                    try:
                        return float(v)
                    except (TypeError, ValueError):
                        return None

                try:
                    cur.execute(_UPSERT_SQL, (
                        province, year_month,
                        _mw("wind_mw"), _mw("solar_mw"), _mw("thermal_mw"),
                        _mw("hydro_mw"), _mw("nuclear_mw"), _mw("bess_mw"),
                        _mw("total_mw"), filename,
                    ))
                    upserted.append(province)
                    logger.info(
                        "Upserted capacity for %s %s (bess=%.0f MW)",
                        province, year_month, _mw("bess_mw") or 0,
                    )
                except Exception as exc:
                    errors.append(f"{province}: {exc}")
                    logger.error("Capacity upsert failed for %s: %s", province, exc)

        conn.commit()
    finally:
        conn.close()

    return {
        "upserted": len(upserted),
        "provinces": upserted,
        "year_month": str(year_month),
        "errors": errors,
    }


def upsert_capacity_rows(
    rows: list[dict],
    pg_url: str,
    source_name: str,
    year_month: date,
) -> dict:
    """Upsert pre-extracted capacity rows (no LLM step).

    Each row: {"province": str, "bess_mw": float|None, "hydro_mw": float|None,
               "wind_mw": float|None, "solar_mw": float|None,
               "thermal_mw": float|None, "nuclear_mw": float|None, "total_mw": float|None}
    Returns {"upserted": int, "provinces": list[str], "year_month": str, "errors": list[str]}
    """
    errors: list[str] = []
    upserted: list[str] = []

    conn = psycopg2.connect(pg_url)
    try:
        with conn.cursor() as cur:
            cur.execute(_ENSURE_TABLE_SQL)
            for row in rows:
                province_raw = str(row.get("province", "")).strip()
                if not province_raw:
                    continue
                if any(kw in province_raw.lower() for kw in _SKIP_KEYWORDS):
                    continue
                province = _normalise_province(province_raw)

                def _mw(key: str) -> Optional[float]:
                    v = row.get(key)
                    if v is None or v == "":
                        return None
                    try:
                        return float(v)
                    except (TypeError, ValueError):
                        return None

                try:
                    cur.execute(_UPSERT_SQL, (
                        province, year_month,
                        _mw("wind_mw"), _mw("solar_mw"), _mw("thermal_mw"),
                        _mw("hydro_mw"), _mw("nuclear_mw"), _mw("bess_mw"),
                        _mw("total_mw"), source_name,
                    ))
                    upserted.append(province)
                except Exception as exc:
                    errors.append(f"{province}: {exc}")
                    logger.error("Capacity row upsert failed for %s: %s", province, exc)
        conn.commit()
    finally:
        conn.close()

    return {
        "upserted": len(upserted),
        "provinces": upserted,
        "year_month": str(year_month),
        "errors": errors,
    }


def is_capacity_file(filename: str) -> bool:
    """Return True if filename looks like a 各省储能/装机 capacity Excel file."""
    name_lower = filename.lower()
    # Must be an Excel file — PDFs/Word docs with "装机" in name should not trigger ETL
    if not (name_lower.endswith(".xlsx") or name_lower.endswith(".xls")):
        return False
    keywords = ["储能装机", "装机容量", "装机数据", "installed_cap", "installed_capacity",
                "各省装机", "province_cap", "capacity_scan"]
    if any(kw in name_lower for kw in keywords):
        return True
    # Also match single-province files like "重庆装机-gpt-20260716.xlsx"
    if "装机" in filename and province_from_filename(filename) is not None:
        return True
    return False
