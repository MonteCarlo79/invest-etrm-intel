from __future__ import annotations
import json
import logging
import os
from typing import Optional

from shared.anthropic_client import make_client as _make_anthropic_client

from services.hermes.models import Action, InboundMessage
from services.hermes.tasks_client import TasksClient
from services.hermes.onedrive_client import OneDriveClient

logger = logging.getLogger(__name__)

# ── LLM priority: GPT-5.4 (Azure) → DeepSeek → Claude ───────────────────────
# Set env vars to enable each tier. Claude is always the final fallback.
#
# Azure OpenAI (GPT-5.4):
#   AZURE_OPENAI_ENDPOINT   e.g. https://tutor-pjh-resource.openai.azure.com/
#   AZURE_OPENAI_API_KEY    your Azure API key
#   AZURE_OPENAI_DEPLOYMENT deployment name (default: gpt-4o)
#   AZURE_OPENAI_API_VERSION (default: 2025-01-01-preview)
#
# DeepSeek:
#   DEEPSEEK_API_KEY        your DeepSeek API key
#   DEEPSEEK_MODEL          model name (default: deepseek-chat)

SYSTEM_PROMPT = """You are Hermes, a personal electricity market strategist and assistant for a BESS investment professional.
You help manage tasks, access the user's Microsoft OneDrive, answer questions across all work domains,
and route data questions to specialist market agents.

CAPABILITY AREAS — understand which domain the user is working in:

📅 Life & Calendar
  Tasks, reminders, personal scheduling, unread email digest.

⚡ China Power Markets
  Spot market prices & spreads (market=spot).
  Province-level BESS economics, capture rates, IRR (market=bess-map).

🌐 International Markets
  GB (Great Britain), AU (Australia NEM), ERCOT (Texas), CAISO (California),
  PJM (PJM interconnection), PH (Philippines), PO (Portugal).

📊 Trading Management
  Inner Mongolia BESS assets, Mengxi trading P&L, dispatch schedules.
  Answer from KB context or use MARKET_AGENT(bess-map) for financial metrics and installed capacity.

🌐 Internet Research
  Search the web, read URLs, search GitHub, get YouTube info, read RSS feeds, search Bilibili.
  Use MARKET_AGENT(internet) for any web/internet research request.

🔢 Quant Models
  BESS IRR/NPV calculations → MARKET_AGENT(bess-map) with get_irr_estimate.
  Dispatch strategy comparisons, capture-rate analysis.
  General quant questions → REPLY using KB context.

🏗 Structuring
  Deal economics, term sheets, project financing, market entry analysis.
  Always use REPLY with available KB context. Structure response as:
  Market Context | Key Economics | Risk Factors | Recommendation.

📋 Meeting Preparation
  Compile briefings from KB + market data for upcoming meetings.
  Use REPLY with all available context. Structure as:
  Background | Key Data Points | Talking Points | Questions to Prepare.

📁 Reports & Knowledge Base
  SAVE_NEXT_FILE → save binary file to OneDrive once.
  INGEST_NEXT_FILE → add file to knowledge base.
  CLASSIFY_NEXT_FILE → AI-classify market fundamentals documents.
  INGEST_URL → fetch URL and add to knowledge base.

Given a message, decide what to do and respond in JSON matching exactly this schema:
{"action": "...", "params": {...}, "reply": "..."}

Actions and their params:

CREATE — create a task
  params: {"title": "string", "due_date": "YYYY-MM-DD or ISO string or null"}
  reply: confirmation text

LIST — list open tasks
  params: {}
  reply: formatted list of open tasks

DONE — mark a task as complete
  params: {"title": "partial task title to match"}
  reply: confirmation text

ONEDRIVE_LIST — list files in a OneDrive folder
  params: {"folder_path": "/path or / for root"}
  reply: you will generate after seeing results

ONEDRIVE_SEARCH — search OneDrive
  params: {"query": "search terms"}
  reply: you will generate after seeing results

ONEDRIVE_READ — read a text file from OneDrive
  params: {"item_id": "file id", "filename": "name for context"}
  reply: you will generate after seeing results

ONEDRIVE_UPLOAD — create a new text file on OneDrive (for text/notes only, NOT for binary files)
  params: {"folder_path": "/path", "filename": "name.txt", "content": "file content as string"}
  reply: confirmation text

SAVE_NEXT_FILE — user wants to forward upcoming Feishu file attachment(s) to OneDrive
  params: {"folder_path": "/path/to/folder", "count": <integer, default 1; use -1 if user says "all" or doesn't specify a limit>}
  reply: tell user to send the file(s) now; if count>1, say "请依次发送{count}个文件" (in their language)

CLASSIFY_NEXT_FILE — user wants to forward a market fundamentals document to OneDrive and let AI pick the right subfolder automatically
  params: {"hint": "optional extra context about the file (province, document type, etc.)"}
  reply: tell user to send the file and that AI will classify it (in their language)

INGEST_NEXT_FILE — user wants to add an upcoming file to the knowledge base (so Hermes can answer questions from it)
  params: {"category": "market_rules|policy_doc|technical_spec|research_report|annual_report|other", "hint": "optional description"}
  reply: tell user to send the file (in their language)

CAPCOMP_INGEST_NEXT_FILE — save upcoming image(s) containing capacity compensation / 容量补偿 data to KB AND extract the data into province_cap_comp table (so bess-map Capacity Compensation tab is updated)
  params: {"count": <integer, default 1>, "hint": "optional context e.g. province or year"}
  reply: tell user to send the image(s) (in their language)
  note: use when user sends capcomp/容量补偿 screenshots or says to save capacity compensation data; if count>1 say "请依次发送{count}张截图"

INGEST_URL — fetch a URL, add its content to the knowledge base, and return a summary
  params: {"url": "https://...", "category": "market_rules|policy_doc|research_report|other"}
  reply: short acknowledgment that you are fetching and summarising (in their language); the actual summary is appended automatically

BAYESIAN_ANALYSIS — reason about a question using Prior → Evidence → Posterior thinking mode
  params: {"question": "the full question to reason about", "format": "text|pdf (default text; use pdf when user says 生成pdf/导出/报告/pdf)"}
  reply: brief acknowledgment that you are starting the analysis (in their language), mention ~30s wait; if pdf, also mention a PDF will be sent
  note: use when user asks forward-looking, probabilistic, or analytical questions that benefit from
        explicit uncertainty quantification and evidence gathering.
        MANDATORY for: price outlook / forecast questions (e.g. "下半年电价是什么水平", "未来价格走势",
        "H2 price level", "price forecast", "价格预测", "下半年...价格"), market outlook questions,
        or any question about FUTURE price levels even if phrased as "你认为...水平/走势/趋势".
        Also triggers on: 估计/可能性/判断/你怎么看/你认为/概率/forecast/estimate/how likely/
        what do you think/odds/probability/分析一下/贝叶斯/bayesian.
        IMPORTANT: When the user asks "你认为X下半年/未来/H2的价格/电价是什么水平", route to
        BAYESIAN_ANALYSIS — NOT MARKET_AGENT. MARKET_AGENT only retrieves historical data;
        BAYESIAN_ANALYSIS uses that data plus reasoning to produce an informed outlook.
        Do NOT use for simple historical data lookups — use MARKET_AGENT for those.

ADD_FILE_RULE — permanently remember to auto-route future files by filename pattern
  params: {"pattern": "partial filename to match (case-insensitive)", "folder_template": "etrm/bess-platform/data/spot reports/{year}", "auto_kb": false, "auto_digest": false, "auto_etl": false}
  reply: confirmation that the rule is saved (in their language)
  note: use {year} in folder_template if the user wants year-based subfolders; set auto_kb=true to also auto-add matched files to the knowledge base; set auto_digest=true to also generate and send a domain expert analysis; set auto_etl=true for 各省装机容量/储能装机 Excel files to auto-parse and upsert into province_installed_monthly (feeds bess-map BESS Demand tab)

LIST_FILE_RULES — list all saved auto-routing rules
  params: {}
  reply: you will generate after seeing results

DELETE_FILE_RULE — remove a saved auto-routing rule
  params: {"rule_id": 1}
  reply: confirmation

EMAIL_SUMMARY — check and summarise recent unread emails
  params: {"limit": 20}
  reply: you will generate after seeing results

MARKET_AGENT — ask a specialist market agent a data question
  params: {"market": "gb|au|ercot|caiso|pjm|ph|po|bess-map|spot|mengxi|internet", "question": "the full question to ask"}
  reply: you will generate after seeing results
  note: use when user asks about specific market data, BESS revenues, prices, assets, economics

SERVICE_CONTROL — start or stop a market app's Streamlit web UI on AWS
  params: {"market": "gb|au|ercot|caiso|pjm|ph|po|bess-map|im|mengxi|options|all", "mode": "web|stop|status"}
  reply: you will generate after seeing results
  note: "start/open/launch [market]" → mode=web; "stop/close/turn off [market]" → mode=stop; "app status / which apps are running" → market=all, mode=status

GENERATE_CHART — query market data and render it as a chart image sent directly to the user
  params: {"market": "spot|bess-map|gb|au|ercot|caiso|pjm|ph|po",
           "question": "the data question (include date range and specific metrics)",
           "chart_type": "line|bar",
           "title": "chart title in user's language",
           "y_label": "Y-axis label with units (e.g. 价格 ¥/MWh)",
           "x_label": "X-axis label (optional, default: 日期)"}
  reply: "正在生成图表，请稍候…" (or in user's language)
  note: use when user says "画图", "图表", "chart", "plot", "画一个", "生成图", "visualize", "可视化".
        chart_type=line for time series (prices, revenues over time); bar for rankings/comparisons.
        The chart is rendered and sent as an image — no need to repeat data in the text reply.

START_REPORT_SESSION — enter report file collection mode; files the user sends next will be tagged for the report
  params: {}
  reply: tell user to send reference files and then give the report topic (in their language)
  note: use when user says "我要准备一份报告/报告模式/start report/report mode/先发参考文件/
        I'll upload files for the report/收集报告文件". Distinct from DRAFT_REPORT — this just
        opens a collection window; the actual draft happens when user sends the topic command.

DRAFT_REPORT — compile a deep multi-source report from market agents + uploaded reference files
  params: {"topic": "report title / subject",
           "markets": ["spot", "bess-map"],
           "outline": "user's notes, thoughts, key points — include verbatim"}
  reply: brief acknowledgment that report is being drafted (in user's language), mention ~1-2 min wait
  note: use when user says "帮我起草/生成/写一份深度报告/研究报告/分析报告/会议材料",
        "draft a report", "conference report/materials", "prepare a report on X".
        Extract topic from the message. Put ALL user notes/outline/instructions into outline param.
        Default markets: ["spot","bess-map"] for China reports; add "mengxi" for IM ops;
        add "internet" for web-research-backed reports; adjust to international codes for global scope.
        If the user uploaded reference files before sending this command, they are automatically included.

WRITE_DEV_REQUEST — record a development request as a structured .md file to OneDrive for laptop-side development with company Claude token
  params: {"message": "user's full request message verbatim"}
  reply: brief acknowledgment that the request is being processed
  note: use when user says "记录需求", "记录一个需求", "save dev request", "记录开发需求",
        "写一个需求文档", "帮我记录", "development request", "dev request", "需求文档".
        Pass the user's FULL original message in the message param — ThinkingAgent will structure it.

EXPORT_ANSWER — export the last assistant answer to a file and save to OneDrive
  params: {"title": "document title", "fmt": "docx|pdf|png", "folder": "/OneDrive path (default /Hermes Exports)"}
  reply: you will generate after seeing results
  note: use when user says "save as Word/PDF/PNG", "export", "导出", "保存为文件"

CLARIFY — you need more information to determine which market agent to use
  params: {"question": "clarifying question to ask the user"}
  reply: the clarifying question

REPLY — just reply, no board or drive action
  params: {}
  reply: response text (use knowledge base context if provided above)

Rules:
- For relative dates (tomorrow, Friday, next week) compute an absolute YYYY-MM-DD date.
- Current date/time is injected at the top of the system prompt — use it for all date calculations.
- For OneDrive_READ and ONEDRIVE_SEARCH, set reply to empty string — the caller will format results.
- When user wants to save a PDF, image, Excel, Word or any binary file to OneDrive to a specific folder, use SAVE_NEXT_FILE (not ONEDRIVE_UPLOAD). ONLY use SAVE_NEXT_FILE when user explicitly says they have a file/document to upload or save (e.g. "存这个文件", "save this to folder X", "上传到"). Never use SAVE_NEXT_FILE for factual questions about market data or storage capacity. If they mention a number (e.g. "5个文件"), set count accordingly.
- When user says "classify", "categorize", "归类", "分类", or "market fundamentals / 市场基础信息" for an upcoming file, use CLASSIFY_NEXT_FILE instead of SAVE_NEXT_FILE.
- When user says "add to knowledge base", "知识库", "让你学习", "分析这份文件", use INGEST_NEXT_FILE. ONLY use INGEST_NEXT_FILE when user explicitly says they want to ingest a file — never for factual data questions.
- When user sends capacity compensation / 容量补偿 screenshots or says to save capcomp data to knowledge base, use CAPCOMP_INGEST_NEXT_FILE (NOT INGEST_NEXT_FILE). This also extracts data into province_cap_comp so bess-map is updated.
- When user asks about email, inbox, unread messages, or says "邮件"/"收件箱"/"邮箱", use EMAIL_SUMMARY.
- When user says "from now on / automatically / whenever I send X" route files to a folder, use ADD_FILE_RULE.
- When ADD_FILE_RULE is used with "also ingest/add to knowledge base" or "also analyze/digest", set auto_kb=true and/or auto_digest=true accordingly.
- When ADD_FILE_RULE is used for 装机容量/储能装机/installed capacity Excel files, set auto_etl=true so data is automatically parsed into the province_installed_monthly DB table (used by bess-map BESS Demand tab).
- When user says "X 已完成", "已完成 X", "X done", "X finished", "X completed" — where X is a task name or partial description — use DONE action with X as the title.
- When user says "mark X as done/complete/finished" or "将X标为完成", use DONE.
- When user asks about market data, prices, revenues, BESS economics, or a specific market, use MARKET_AGENT.
- For MARKET_AGENT market keys: gb=GB/Great Britain (9 tools: system price, EPEX, ancillary, BESS leaderboard/revenue index/assets, Elexon ops, KB search), au=Australia NEM, ercot=Texas, caiso=California, pjm=PJM, ph=Philippines, po=Portugal (all via intl_market_common: spot price, ancillary, BESS leaderboard/revenue index/assets, KB search), bess-map=China BESS economics for all provinces (get_bess_economics, get_dispatch_detail, get_mengxi_capacity, get_irr_estimate), spot=China spot electricity prices + DA/RT spreads + fundamentals + BESS P&L + KB (7 tools), mengxi=Inner Mongolia BESS operations (P&L attribution waterfall, 15-min dispatch data, RT prices, strategy comparison, KB — use for trading ops questions about the 4 IM BESS assets), internet=web search/URL reading/GitHub/YouTube/RSS/Bilibili.
- Use MARKET_AGENT(mengxi) for operational/trading questions about the 4 Inner Mongolia BESS assets (景蓝乌尔图/悦杭独贵/景通四益堂储/裕昭沙子坝): P&L breakdown, dispatch execution, RT prices, strategy comparison. Use MARKET_AGENT(bess-map) for province-level BESS economics (capture rates, IRR, theoretical revenue) including 蒙西 as a market.
- When user asks to "search the web", "look up a URL", "read this link", "search GitHub", "find on YouTube/Bilibili/B站", "check RSS", or research any topic online, use MARKET_AGENT(internet). IMPORTANT: "搜索" (search) in Chinese is an explicit internet search request — use MARKET_AGENT(internet) whenever "搜索" appears in the message, especially combined with "最新" (latest) or "网上". This takes PRIORITY over all other routing rules (bess-map, spot, etc.).
- For GENERATE_CHART: use market=spot for ANY Chinese province spot price chart (现货价格/实时价格/日前价格/RT price/DA price/spot price — regardless of which province 陕西/山东/广东/蒙西/etc.). Use market=bess-map only for BESS economics charts (BESS revenues, capture rates, IRR, capacity). For international markets use the corresponding market code.
- If the message is exactly "蒙西储能日报", use REPLY with text "正在生成日报…" — the app handles this as a report trigger, do NOT use MARKET_AGENT.
- If the market is ambiguous and you cannot infer it from context, use CLARIFY.
- When user says "save as Word/PDF/PNG/file" or "export" about a previous answer, use EXPORT_ANSWER.
- When user says "提供源文件", "给我看原文", "新闻源文件", "原文链接", "找一下原文", "发给我原文", or "source file" in the context of a news article or KB document (e.g. mentioning a ★ rating, article title, or PDF name from a news digest), use MARKET_AGENT(spot) with a KB search question like "find article: <title>". Do NOT use ONEDRIVE_SEARCH or ONEDRIVE_LIST for news/KB content — these articles are stored in the knowledge base, not OneDrive.
- When user says "我要准备报告/报告模式/start report/先发参考文件/I'll upload files for the report", use START_REPORT_SESSION to open a file collection window before DRAFT_REPORT.
- When user says "帮我起草/生成/写一份深度报告/分析报告/研究报告/会议材料", "draft a report", "conference report/materials", "prepare a report on X", use DRAFT_REPORT. Extract the topic from the message. Put ALL user notes, outline details, and strategic thoughts into the outline param. Default markets ["spot","bess-map"] for China; add "mengxi" for Inner Mongolia ops; add "internet" for web-backed; adjust to intl codes for global scope. Uploaded reference files (pdf/ppt/txt) sent before this command are automatically included.
- When KNOWLEDGE BASE CONTEXT is provided above, use it to write an informed reply for REPLY actions.
- When user asks about "IRR", "NPV", "payback period", "project economics" for a specific province or market, use MARKET_AGENT(bess-map) — the bess-map agent has get_irr_estimate.
- When user says "meeting prep", "会议准备", "prepare me for a meeting about X", "briefing for X", use REPLY with KB context structured as: Background | Key Data Points | Talking Points | Questions to Prepare.
- When user says "structuring", "term sheet", "market entry", "project financing", "条款", use REPLY drawing from KB context with: Market Context | Key Economics | Risk Factors | Recommendation.
- When user says "Inner Mongolia", "内蒙古", "Mengxi", "蒙西" for operational data (P&L, dispatch), use MARKET_AGENT(bess-map) or REPLY with KB context if no specific data question.
- When user asks about "装机容量", "installed capacity", "total MW", "total GW", "总容量", "总装机", "储能装机", "独立储能" for ANY province (江苏, 山东, 广东, 湖北, 蒙西, etc.), use MARKET_AGENT(bess-map) — the bess-map agent has get_province_installed_capacity(province=<name>) which covers all provinces. For Mengxi-specific plant-level owner breakdown, it also has get_mengxi_capacity. EXCEPTION: if the message starts with "/capacity" or "/capacity-add" followed by actual province data (e.g. "/capacity 山东 9.7GW"), that is a WRITE command handled before the LLM — use REPLY and tell the user the data was accepted.
- When user asks what you can do in a certain area (e.g. "what can you do for X?"), use REPLY and describe the relevant capabilities from the CAPABILITY AREAS section above, with concrete examples.
- When user says "记录需求", "记录开发需求", "写需求文档", "save dev request", "development request", "dev request", use WRITE_DEV_REQUEST with the user's verbatim message in the message param.
- Always match the user's language in the reply field. If the user writes in Chinese (Simplified), reply in Chinese (Simplified). If in English, reply in English.
- Always respond with valid JSON only. No markdown fences, no extra text."""


class HermesAgent:
    def __init__(
        self,
        tasks: TasksClient,
        anthropic_api_key: str,
        onedrive: Optional[OneDriveClient] = None,
    ) -> None:
        self.tasks = tasks
        self.onedrive = onedrive
        self._api_key = anthropic_api_key
        self.client = _make_anthropic_client(anthropic_api_key)
        # Conversation memory (lazy-init; disabled gracefully if DB unavailable)
        self._memory: Optional[object] = None
        self._last_answer: str = ""  # track last substantive reply for EXPORT_ANSWER
        self._pending_chart_bytes: Optional[bytes] = None  # set by GENERATE_CHART

    # ── Model preference helpers ──────────────────────────────────────────────
    _MODEL_ALIASES: dict = {
        "gpt": "gpt", "gpt-4o": "gpt", "gpt-5.4": "gpt", "azure": "gpt", "openai": "gpt",
        "deepseek": "deepseek", "ds": "deepseek",
        "claude": "claude", "anthropic": "claude",
        "auto": "auto",
    }
    _MODEL_LABELS: dict = {
        "gpt":      "GPT-4o (Azure)",
        "deepseek": "DeepSeek",
        "claude":   "Claude Sonnet 4.6",
        "auto":     "Auto (GPT-4o → DeepSeek → Claude)",
    }

    def get_model_pref(self, chat_id: str) -> str:
        """Return stored model preference ('auto'|'gpt'|'deepseek'|'claude'). Default: gpt."""
        try:
            v = self.tasks.get_setting(f"llm_pref:{chat_id}")
            return v if v in self._MODEL_LABELS else "gpt"
        except Exception:
            return "gpt"

    def set_model_pref(self, chat_id: str, alias: str) -> str:
        """Persist model preference. Returns canonical name or raises ValueError."""
        canon = self._MODEL_ALIASES.get(alias.lower())
        if not canon:
            raise ValueError(f"Unknown model '{alias}'. Use: gpt, deepseek, claude, auto")
        self.tasks.set_setting(f"llm_pref:{chat_id}", canon)
        return canon

    def available_models(self) -> list:
        """Return configured model tiers (always includes claude)."""
        models = []
        if os.environ.get("AZURE_OPENAI_ENDPOINT") and os.environ.get("AZURE_OPENAI_API_KEY"):
            models.append("gpt")
        if os.environ.get("DEEPSEEK_API_KEY"):
            models.append("deepseek")
        models.append("claude")
        return models

    def _call_llm(self, system: str, user_text: str, preferred: str = "auto") -> str:
        """Call LLM respecting preferred model, falling back down chain on error.

        preferred: 'auto' | 'gpt' | 'deepseek' | 'claude'
        When preferred != 'auto', that tier is tried first; others follow as fallback.
        """
        az_endpoint = os.environ.get("AZURE_OPENAI_ENDPOINT", "")
        az_key      = os.environ.get("AZURE_OPENAI_API_KEY", "")
        ds_key      = os.environ.get("DEEPSEEK_API_KEY", "")

        def _try_gpt() -> Optional[str]:
            if not (az_endpoint and az_key):
                return None
            try:
                from openai import AzureOpenAI
                client = AzureOpenAI(
                    azure_endpoint=az_endpoint, api_key=az_key,
                    api_version=os.environ.get("AZURE_OPENAI_API_VERSION", "2025-01-01-preview"),
                )
                resp = client.chat.completions.create(
                    model=os.environ.get("AZURE_OPENAI_DEPLOYMENT", "gpt-4o"),
                    messages=[{"role": "system", "content": system},
                               {"role": "user",   "content": user_text}],
                    max_tokens=4096,
                    response_format={"type": "json_object"},
                )
                raw = resp.choices[0].message.content.strip()
                logger.info("LLM: GPT-5.4 (Azure) [%d chars]", len(raw))
                return raw
            except Exception as exc:
                logger.warning("GPT-4o failed: %s", exc, exc_info=True)
                return None

        def _try_deepseek() -> Optional[str]:
            if not ds_key:
                return None
            try:
                from openai import OpenAI as _OAI
                client = _OAI(api_key=ds_key, base_url="https://api.deepseek.com")
                model = os.environ.get("DEEPSEEK_MODEL", "deepseek-chat")
                resp = client.chat.completions.create(
                    model=model,
                    messages=[{"role": "system", "content": system},
                               {"role": "user",   "content": user_text}],
                    max_tokens=4096,
                    response_format={"type": "json_object"},
                )
                raw = resp.choices[0].message.content.strip()
                logger.info("LLM: DeepSeek (%s) [%d chars] raw=%s", model, len(raw), raw[:200])
                return raw
            except Exception as exc:
                logger.warning("DeepSeek failed: %s", exc, exc_info=True)
                return None

        def _try_claude() -> Optional[str]:
            try:
                resp = self.client.messages.create(
                    model="claude-sonnet-4-6",
                    max_tokens=8192,
                    system=system,
                    messages=[{"role": "user", "content": user_text}],
                )
                raw = resp.content[0].text.strip()
                logger.info("LLM: Claude Sonnet 4.6 [%d chars]", len(raw))
                return raw
            except Exception as exc:
                logger.warning("Claude Sonnet 4.6 failed: %s", exc, exc_info=True)
                return None

        # Build call order based on preference
        if preferred == "gpt":
            order = [_try_gpt, _try_deepseek, _try_claude]
        elif preferred == "deepseek":
            order = [_try_deepseek, _try_gpt, _try_claude]
        elif preferred == "claude":
            order = [_try_claude, _try_gpt, _try_deepseek]
        else:  # auto — default priority chain
            order = [_try_gpt, _try_deepseek, _try_claude]

        for fn in order:
            result = fn()
            if result is not None:
                return result
        raise RuntimeError("All LLM tiers failed — check API keys / Bedrock config")

    def _get_memory(self):
        if self._memory is None:
            try:
                from services.hermes.conversation_memory import HermesMemory
                pg_url = os.environ.get("PGURL") or os.environ.get("DATABASE_URL", "")
                self._memory = HermesMemory(pg_url=pg_url, api_key=self._api_key)
            except Exception:
                pass
        return self._memory

    def _retrieve_kb_context(self, query: str) -> str:
        """Lightweight DB-only knowledge base retrieval (no HyDE/rerank API calls)."""
        try:
            from services.knowledge_pool.advanced_retrieval import retrieve_for_agent
            ctx = retrieve_for_agent(
                query=query,
                api_key=self._api_key,
                app="strategist",
                use_hyde=False,
                use_rerank=False,
                top_k=5,
                include_policy_timeline=False,
            )
            if ctx and "No relevant knowledge" not in ctx:
                return ctx
        except Exception as exc:
            logger.debug("KB retrieval skipped: %s", exc)
        return ""

    # Keywords that unambiguously mean "search the internet"
    _INTERNET_PREFIXES = ("搜索", "帮我搜", "上网搜", "网上搜", "google", "search for", "web search")

    def process(self, msg: InboundMessage, chat_id: str = "") -> Action:
        import re
        from datetime import datetime, timezone, timedelta

        # Fast-path: explicit internet search keywords → bypass LLM routing
        _text_lower = msg.text.strip().lower()
        for _kw in self._INTERNET_PREFIXES:
            if _text_lower.startswith(_kw) or f" {_kw}" in _text_lower:
                logger.info("Internet fast-path triggered by keyword '%s'", _kw)
                return Action(
                    action="MARKET_AGENT",
                    params={"market": "internet", "question": msg.text},
                    reply="",
                )

        # Current Beijing time injected so the model always knows the date
        _bj = datetime.now(tz=timezone(timedelta(hours=8)))
        _weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        date_header = (
            f"Current date/time (Beijing UTC+8): {_bj.strftime('%Y-%m-%d %H:%M')} "
            f"({_weekdays[_bj.weekday()]})\n\n"
        )

        # Inject KB context when available
        kb_ctx = self._retrieve_kb_context(msg.text)
        if kb_ctx:
            system = (
                date_header
                + "--- KNOWLEDGE BASE CONTEXT (use this to answer market questions) ---\n"
                + kb_ctx
                + "\n--- END KNOWLEDGE BASE CONTEXT ---\n\n"
                + SYSTEM_PROMPT
            )
        else:
            system = date_header + SYSTEM_PROMPT

        # Inject relevant long-term insights from memory
        mem = self._get_memory()
        if mem:
            try:
                insights_ctx = mem.get_relevant_insights(msg.text)
                if insights_ctx:
                    system += f"\n\n--- MEMORY CONTEXT ---\n{insights_ctx}\n--- END MEMORY CONTEXT ---"
            except Exception:
                pass

        # Inject recent conversation history into system prompt (not as messages).
        # Using messages[] for history causes Claude to mimic plain-text assistant
        # turns and break JSON output. System-prompt injection avoids this.
        if chat_id and mem:
            try:
                history = mem.load_history(chat_id)
                if history:
                    hist_lines = [
                        f"{t['role'].capitalize()}: {t['content'][:300]}"
                        for t in history[-8:]
                    ]
                    system += (
                        "\n\n--- RECENT CONVERSATION HISTORY (shown for context only) ---\n"
                        + "\n".join(hist_lines)
                        + "\n--- END HISTORY ---"
                        "\n(Regardless of history format, always respond with valid JSON only.)"
                    )
            except Exception:
                pass

        # Save user turn
        if chat_id and mem:
            try:
                mem.save_turn(chat_id, "user", msg.text)
            except Exception:
                pass

        raw = self._call_llm(system, msg.text, preferred=self.get_model_pref(chat_id))

        def _try_parse(text: str) -> Optional[Action]:
            for sanitized in (text, re.sub(r'[\r\n]+', ' ', text)):
                try:
                    return Action.model_validate_json(sanitized)
                except Exception:
                    pass
            return None

        def _save_raw_to_history(json_text: str) -> None:
            """Save the raw JSON response as the assistant turn so history
            shows JSON format and Claude doesn't mimic plain-text next time."""
            if chat_id and mem:
                try:
                    mem.save_turn(chat_id, "assistant", json_text)
                except Exception:
                    pass

        # Try direct parse
        result = _try_parse(raw)
        if result:
            _save_raw_to_history(raw)
            return result
        # Extract JSON from markdown fences
        m = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', raw, re.DOTALL)
        if m:
            result = _try_parse(m.group(1))
            if result:
                _save_raw_to_history(m.group(1))
                return result
        # Last resort: find any JSON object in the response
        m = re.search(r'\{.*\}', raw, re.DOTALL)
        if m:
            result = _try_parse(m.group(0))
            if result:
                _save_raw_to_history(m.group(0))
                return result
        raise ValueError(f"No valid JSON in Claude response: {raw[:300]}")

    def execute(self, action: Action) -> str:
        """Execute the action and return an optional supplementary reply."""
        try:
            if action.action == "CREATE":
                self.tasks.create_card(**action.params)
            elif action.action == "LIST":
                cards = self.tasks.list_open_cards()
                if not cards:
                    return "No open cards."
                lines = [f"• {c['title']}" + (f" (due {c['due_date'].date()})" if c.get("due_date") else "") for c in cards]
                return "\n".join(lines)
            elif action.action == "DONE":
                self.tasks.complete_card(**action.params)
            elif action.action == "ONEDRIVE_LIST":
                if not self.onedrive:
                    return "OneDrive not configured."
                items = self.onedrive.list_items(action.params.get("folder_path", "/"))
                if not items:
                    return "Folder is empty."
                lines = []
                for it in items:
                    kind = "📁" if "folder" in it else "📄"
                    lines.append(f"{kind} {it['name']}")
                return "\n".join(lines)
            elif action.action == "ONEDRIVE_SEARCH":
                if not self.onedrive:
                    return "OneDrive not configured."
                results = self.onedrive.search(action.params.get("query", ""))
                if not results:
                    return "No files found."
                lines = []
                for r in results[:10]:
                    parent = r.get("parentReference", {}).get("path", "").replace("/drive/root:", "")
                    lines.append(f"📄 {r['name']}  (id: {r['id']}, path: {parent})")
                return "\n".join(lines)
            elif action.action == "ONEDRIVE_READ":
                if not self.onedrive:
                    return "OneDrive not configured."
                item_id = action.params.get("item_id", "")
                return self.onedrive.read_file_smart(item_id, max_chars=6000)
            elif action.action == "ADD_FILE_RULE":
                self.tasks.add_file_rule(
                    pattern=action.params.get("pattern", ""),
                    folder_template=action.params.get("folder_template", "Hermes Uploads"),
                    auto_kb=bool(action.params.get("auto_kb", False)),
                    auto_digest=bool(action.params.get("auto_digest", False)),
                    auto_etl=bool(action.params.get("auto_etl", False)),
                )
            elif action.action == "LIST_FILE_RULES":
                rules = self.tasks.get_file_rules()
                if not rules:
                    return "No file routing rules set."
                lines = [f"[{r['id']}] pattern='{r['pattern']}' → {r['folder_template']}" for r in rules]
                return "\n".join(lines)
            elif action.action == "DELETE_FILE_RULE":
                deleted = self.tasks.delete_file_rule(int(action.params.get("rule_id", 0)))
                return "Rule deleted." if deleted else "Rule not found."
            elif action.action == "ONEDRIVE_UPLOAD":
                if not self.onedrive:
                    return "OneDrive not configured."
                content = action.params.get("content", "").encode("utf-8")
                result = self.onedrive.upload_file(
                    folder_path=action.params.get("folder_path", "/"),
                    filename=action.params.get("filename", "hermes_upload.txt"),
                    content=content,
                )
                return f"Uploaded: {result.get('name')} ({result.get('size', '?')} bytes)"
            elif action.action == "INGEST_URL":
                return self._ingest_url(action.params.get("url", ""))
            elif action.action == "BAYESIAN_ANALYSIS":
                return self._run_bayesian(
                    action.params.get("question", ""),
                    chat_id=action.params.get("_chat_id", ""),
                )
            elif action.action == "MARKET_AGENT":
                return self._run_market_agent(
                    market=action.params.get("market", ""),
                    question=action.params.get("question", ""),
                )
            elif action.action == "SERVICE_CONTROL":
                return self._run_service_control(
                    market=action.params.get("market", "all"),
                    mode=action.params.get("mode", "status"),
                )
            elif action.action == "GENERATE_CHART":
                return self._generate_chart(
                    market=action.params.get("market", ""),
                    question=action.params.get("question", ""),
                    chart_type=action.params.get("chart_type", "line"),
                    title=action.params.get("title", ""),
                    y_label=action.params.get("y_label", ""),
                    x_label=action.params.get("x_label", "日期"),
                )
            elif action.action == "WRITE_DEV_REQUEST":
                msg = action.params.get("message", "")
                if not msg:
                    return "请说明需求内容。"
                from services.hermes.thinking_agent import ThinkingAgent
                thinker = ThinkingAgent(
                    anthropic_api_key=self._api_key,
                    pg_url=os.environ.get("PGURL") or os.environ.get("HERMES_DB_URL", ""),
                    feishu=None,
                    feishu_owner_open_id="",
                    onedrive=self.onedrive,
                )
                return thinker.write_dev_request_from_message(msg)
            elif action.action == "EXPORT_ANSWER":
                return self._export_answer(
                    title=action.params.get("title", "Hermes Answer"),
                    fmt=action.params.get("fmt", "pdf"),
                    folder=action.params.get("folder", "/Hermes Exports"),
                )
            elif action.action == "CLARIFY":
                return action.params.get("question", action.reply)
        except Exception as exc:
            logger.error("Action %s failed: %s", action.action, exc)
            return f"Error: {exc}"
        return ""

    def _ingest_url(self, url: str) -> str:
        if not url:
            return "No URL provided."
        try:
            from services.knowledge_pool.knowledge_docs import register_url
            doc_id, is_new, category = register_url(url=url, api_key=self._api_key)
            status = (
                f"✅ URL 已添加到知识库 (doc_id={doc_id}, category={category})"
                if is_new
                else f"ℹ️ URL 已存在于知识库中 (doc_id={doc_id})"
            )
        except Exception as exc:
            logger.error("INGEST_URL failed for %s: %s", url, exc)
            return f"添加失败：{exc}"

        # Summarise the ingested content so the user gets an immediate answer
        try:
            from services.knowledge_pool.knowledge_docs import get_conn
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT chunk_text FROM staging.spot_knowledge_chunks
                        WHERE doc_id = %s ORDER BY chunk_index LIMIT 30
                        """,
                        (doc_id,),
                    )
                    rows = cur.fetchall()
            if rows:
                full_text = "\n\n".join(r[0] for r in rows)[:12000]
                from shared.anthropic_client import make_client as _make_anthropic_client
                client = _make_anthropic_client(self._api_key)
                resp = client.messages.create(
                    model="claude-sonnet-4-6",  # haiku-4-5 requires use-case form on this Bedrock account
                    max_tokens=800,
                    system=(
                        "You are a research assistant. Summarise the article clearly and concisely. "
                        "Structure: 核心观点 | 主要内容 | 关键数据/结论. "
                        "Reply in the same language as the article (Chinese if Chinese)."
                    ),
                    messages=[{"role": "user", "content": f"请总结以下文章：\n\n{full_text}"}],
                )
                summary = resp.content[0].text.strip()
                return f"{status}\n\n{summary}"
        except Exception as exc:
            logger.warning("INGEST_URL summary failed for doc %s: %s", doc_id, exc)

        return status

    def _run_bayesian(self, question: str, chat_id: str = "") -> str:
        if not question:
            return "请告诉我你想分析的问题。"
        try:
            from services.hermes.bayesian_agent import BayesianAnalystAgent
            pg_url = os.environ.get("PGURL") or os.environ.get("HERMES_DB_URL", "")
            bay_model = "claude-sonnet-4-6"  # tool-use; global.anthropic.claude-sonnet-4-6 is the only confirmed-working model
            bay_agent = BayesianAnalystAgent(
                anthropic_api_key=self._api_key,
                pg_url=pg_url,
                model=bay_model,
            )
            return bay_agent.run(question)
        except Exception as exc:
            logger.error("BAYESIAN_ANALYSIS failed: %s", exc)
            return f"贝叶斯分析失败：{exc}"

    def _run_market_agent(self, market: str, question: str) -> str:
        if not market or not question:
            return "Market or question missing."
        try:
            from services.hermes.market_agent_bridge import run_market_query
            answer = run_market_query(market=market, question=question, api_key=self._api_key)
            self._last_answer = answer  # store for potential EXPORT_ANSWER
            return answer
        except Exception as exc:
            logger.error("MARKET_AGENT failed for %s: %s", market, exc)
            return f"Market agent error: {exc}"

    def _generate_chart(
        self,
        market: str,
        question: str,
        chart_type: str,
        title: str,
        y_label: str,
        x_label: str,
    ) -> str:
        """Fetch data via market agent, render chart PNG, store in _pending_chart_bytes."""
        self._pending_chart_bytes = None
        if not market or not question:
            return "缺少 market 或 question 参数。"
        try:
            from services.hermes.chart_utils import (
                generate_line_chart, generate_bar_chart, generate_spot_line_chart,
            )

            if market == "spot" and chart_type != "bar":
                # Spot market: query DB directly — avoids LLM response truncation on long date ranges
                pg_url = os.environ.get("PGURL") or os.environ.get("HERMES_DB_URL", "")
                chart_bytes = generate_spot_line_chart(
                    question=question, api_key=self._api_key, pg_url=pg_url,
                    title=title, y_label=y_label or "价格 (¥/MWh)",
                )
            else:
                # Other markets: ask market agent to return data as markdown table
                from services.hermes.market_agent_bridge import run_market_query
                data_question = (
                    question
                    + "\n\nIMPORTANT: Present ALL data as a clean markdown table. "
                    "Include column headers. Do not truncate or summarize rows."
                )
                data_text = run_market_query(
                    market=market, question=data_question, api_key=self._api_key
                )
                self._last_answer = data_text
                try:
                    if chart_type == "bar":
                        chart_bytes = generate_bar_chart(
                            data_text, title=title or question, y_label=y_label, x_label=x_label
                        )
                    else:
                        chart_bytes = generate_line_chart(
                            data_text, title=title or question, y_label=y_label, x_label=x_label
                        )
                except ValueError as table_err:
                    # Market agent returned prose instead of markdown table.
                    # If question looks like a spot price chart, fall back to direct DB query.
                    spot_keywords = ("价格", "price", "现货", "spot", "rt ", "da ", "实时", "日前")
                    if any(kw in question.lower() for kw in spot_keywords):
                        logger.warning(
                            "No markdown table from %s agent; falling back to spot DB: %s",
                            market, table_err,
                        )
                        pg_url = os.environ.get("PGURL") or os.environ.get("HERMES_DB_URL", "")
                        chart_bytes = generate_spot_line_chart(
                            question=question, api_key=self._api_key, pg_url=pg_url,
                            title=title, y_label=y_label or "价格 (¥/MWh)",
                        )
                    else:
                        raise

            self._pending_chart_bytes = chart_bytes
            return ""  # reply comes from action.reply; chart is sent separately by app.py
        except Exception as exc:
            logger.error("GENERATE_CHART failed: %s", exc, exc_info=True)
            return f"图表生成失败：{exc}"

    def _run_service_control(self, market: str, mode: str) -> str:
        try:
            import boto3
            from shared.service_control import get_all_status, set_service_mode, SERVICES
            _ecs = boto3.client("ecs", region_name="ap-southeast-1")
            _cluster = os.environ.get("ECS_CLUSTER", "bess-platform-cluster")

            if mode == "status":
                statuses = get_all_status(_ecs, _cluster)
                icons = {"web": "🟢", "scheduler": "🟡", "stopped": "🔴", "unknown": "❓"}
                return "\n".join(
                    f"{icons.get(s['mode'], '')} {s['label']}: {s['mode']}" for s in statuses
                )

            markets = list(SERVICES.keys()) if market == "all" else [market]
            lines = []
            for m in markets:
                if m not in SERVICES:
                    lines.append(f"Unknown market: {m}")
                    continue
                r = set_service_mode(m, mode, _ecs, _cluster)
                if mode == "web":
                    lines.append(f"✅ {r['label']} web starting — {r['web_url']} (ready ~90s)")
                else:
                    lines.append(f"✅ {r['label']} → {r['mode']}")
            return "\n".join(lines)
        except Exception as exc:
            logger.error("SERVICE_CONTROL failed: %s", exc)
            return f"Service control error: {exc}"

    def _export_answer(self, title: str, fmt: str, folder: str) -> str:
        if not self.onedrive:
            return "OneDrive not configured — cannot export file."
        text = self._last_answer
        if not text:
            return "No recent answer to export."
        try:
            from services.hermes.export_utils import export_answer
            item = export_answer(
                title=title,
                text=text,
                fmt=fmt,
                onedrive=self.onedrive,
                folder=folder,
            )
            web_url = item.get("webUrl", "")
            name = item.get("name", "file")
            if web_url:
                return f"Exported: [{name}]({web_url})"
            return f"Exported: {name}"
        except Exception as exc:
            logger.error("EXPORT_ANSWER failed: %s", exc)
            return f"Export failed: {exc}"

    def ingest_file_to_kb(self, filename: str, file_bytes: bytes, category: str = "") -> str:
        """Ingest a file into the knowledge base. Called from app.py on file receipt."""
        try:
            from services.knowledge_pool.knowledge_docs import register_and_ingest
            doc_id, is_new, detected_category = register_and_ingest(
                file_bytes=file_bytes,
                filename=filename,
                category_override=category or None,
                app="strategist",
                api_key=self._api_key,
                synthesize=True,
            )
            if is_new:
                return f"✅ 已添加到知识库：{filename}（分类：{detected_category}，doc_id={doc_id}）\n正在后台生成摘要和问答对…"
            else:
                return f"ℹ️ 该文件已在知识库中（doc_id={doc_id}）"
        except Exception as exc:
            logger.error("KB ingest failed for %s: %s", filename, exc)
            return f"知识库添加失败：{exc}"

    def generate_file_digest(self, filename: str, file_bytes: bytes) -> str:
        """Parse a file and generate a domain-expert analysis using Claude."""
        import io
        try:
            # Parse file content using onedrive_client's smart reader
            ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
            text = ""
            if ext in ("xlsx", "xlsm"):
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
                parts = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        cells = [str(c) if c is not None else "" for c in row]
                        if any(c.strip() for c in cells):
                            rows.append("\t".join(cells))
                        if len(rows) >= 150:
                            break
                    if rows:
                        parts.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows))
                text = "\n\n".join(parts)[:8000]
            elif ext == "pdf":
                import pdfplumber
                pages = []
                with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                    for page in pdf.pages[:10]:
                        t = page.extract_text()
                        if t:
                            pages.append(t)
                text = "\n\n".join(pages)[:8000]
            else:
                text = file_bytes.decode("utf-8", errors="replace")[:8000]

            if not text.strip():
                return f"⚠️ 无法读取文件内容：{filename}"

            resp = self.client.messages.create(
                model="claude-sonnet-4-6",  # haiku-4-5 requires use-case form on this Bedrock account
                max_tokens=1200,
                system=(
                    "You are a China electricity market expert and BESS investment analyst. "
                    "Analyze the provided file and write a concise domain expert digest in Chinese. "
                    "Structure your response as:\n"
                    "📋 **文件摘要**\n"
                    "🔍 **关键数据点**\n"
                    "💡 **市场洞察**\n"
                    "⚡ **对BESS投资的启示**\n"
                    "Keep it concise (300-400 Chinese characters)."
                ),
                messages=[{
                    "role": "user",
                    "content": f"文件名: {filename}\n\n文件内容:\n{text}",
                }],
            )
            digest = resp.content[0].text.strip()
            return f"🧠 专家分析 — {filename}\n\n{digest}"
        except Exception as exc:
            logger.error("generate_file_digest failed for %s: %s", filename, exc)
            return f"⚠️ 文件分析失败：{exc}"
