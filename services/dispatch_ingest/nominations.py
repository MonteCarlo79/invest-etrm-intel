"""Trader nomination ingestion (申报策略 folder tree → rm_nominations).

Layout survey (2026-08-28):
  01-苏右, 02-杭锦旗, 03-四子王, 05-乌海, 06-乌拉特:
      日期 | 时刻 | 预计划功率 | 正式申报
  04-谷山梁:
      日期 | 时刻 | 预策略D-2功率（MW） | 正式策略D-1功率（MW）
  07-巴盟:
      日期 | 时刻 | 预申报策略（MW） | 实际申报策略（MW）
  Single-column variant (seen in 01-苏右 April files):
      日期 | 时刻 | 申报功率（MW） | 爬坡校验 | ...
  Multi-sheet workbooks may carry a real nomination sheet (策略申报) plus a
  zeroed output template (输出模板) with the same header — parse_nomination_file
  merges by interval, letting the sheet with more actual values win.

Intervals are 5-min. Positive MW = discharge, negative = charge.
Header-driven column detection — no per-station hardcoding.
"""
from __future__ import annotations

import re
from typing import Any

import openpyxl
import pandas as pd

# 申报策略 station folder → rm_assets.name
NOMINATION_FOLDER_TO_ASSET = {
    "01-苏右": "景蓝乌尔图",
    "02-杭锦旗": "悦杭独贵",
    "03-四子王": "四子王旗",
    "04-谷山梁": "裕昭沙子坝",
    "05-乌海": "乌海康富",
    "06-乌拉特": "远景乌拉特",
    "07-巴盟": "景怡查干哈达",
}

# Header aliases → canonical field
_PLANNED_ALIASES = ("预计划功率", "预策略d-2功率", "预申报策略")
_NOMINATED_ALIASES = ("正式申报", "正式策略d-1功率", "实际申报策略", "实际申报功率", "申报功率", "功率")
_TIME_ALIASES = ("时刻", "时间")


def _norm(s: str) -> str:
    s = re.sub(r"[\s（(].*$", "", (s or "").strip().lower())
    return re.sub(r"^\d+[.、]", "", s)  # "2.申报功率（MW）" → "申报功率"


def _match_alias(cell: str, aliases: tuple[str, ...]) -> bool:
    n = _norm(cell)
    return any(n.startswith(a) for a in aliases)


def _find_columns(header_row: list[Any]) -> dict[str, int] | None:
    """Map header cells to field indices. Returns None if the row isn't a nomination header."""
    cols: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        text = str(cell or "").strip()
        if not text:
            continue
        n = _norm(text)
        if n == "日期" and "date" not in cols:
            cols["date"] = i
        elif n in _TIME_ALIASES and "time" not in cols:
            cols["time"] = i
        elif "planned" not in cols and _match_alias(text, _PLANNED_ALIASES):
            cols["planned"] = i
        elif "nominated" not in cols and _match_alias(text, _NOMINATED_ALIASES):
            cols["nominated"] = i
    if {"date", "time", "nominated"} <= cols.keys():
        return cols
    return None


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        m = re.search(r"-?\d[\d,]*\.?\d*", str(v))
        return float(m.group(0).replace(",", "")) if m else None


_SHEET_MONTH_RE = re.compile(r"(\d{1,2})月")
_MIN_PLAUSIBLE = pd.Timestamp("2024-01-01", tz="Asia/Shanghai")


def _repair_date(ts: "pd.Timestamp", sheet_title: str) -> "pd.Timestamp | None":
    """Repair known trader date typos; return None to drop the row.

    汇总 sheets: trader typed "06-01-26" meaning 2026-06-01; Excel stored
    YY-MM-DD → 2006-01-26, and following days advance the MONTH component
    (2006-02-26 = Jun 2 …). Verified on 【汇总】6月杭锦旗 workbook 2026-08-29.
    Other pre-2024 timestamps are junk tail cells (e.g. epoch-0) → drop.
    """
    if ts >= _MIN_PLAUSIBLE:
        return ts
    m = _SHEET_MONTH_RE.search(sheet_title or "")
    if ts.year == 2006 and m:
        month = int(m.group(1))
        try:
            return ts.replace(year=2026, month=month, day=ts.month)
        except ValueError:
            return None
    return None


def parse_nomination_sheet(ws) -> list[dict[str, Any]]:
    """Parse one nomination worksheet into interval dicts.

    Returns list of {interval_start, planned_mw, nominated_mw}; empty when the
    sheet has no nomination header.
    """
    rows = ws.iter_rows(values_only=True)
    cols = None
    header_row_idx = 0
    for idx, row in enumerate(rows):
        if idx > 5:
            break
        found = _find_columns(list(row))
        if found:
            cols = found
            header_row_idx = idx + 1
            break
    if cols is None:
        return []

    items = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        date_v = row[cols["date"]] if cols["date"] < len(row) else None
        time_v = row[cols["time"]] if cols["time"] < len(row) else None
        if date_v is None or time_v is None:
            continue
        try:
            ts = pd.Timestamp(f"{pd.to_datetime(date_v).date()} {time_v}", tz="Asia/Shanghai")
        except Exception:
            continue
        ts = _repair_date(ts, getattr(ws, "title", ""))
        if ts is None:
            continue
        items.append({
            "interval_start": ts,
            "planned_mw": _to_float(row[cols["planned"]]) if "planned" in cols else None,
            "nominated_mw": _to_float(row[cols["nominated"]]),
        })
    return items


def parse_nomination_file(file_path: str) -> list[dict[str, Any]]:
    """Parse nomination sheets in one workbook, merged by interval.

    Some workbooks carry both the real nomination sheet (策略申报) and a zeroed
    output template (输出模板) with an identical header. Sheets are applied in
    ascending order of non-null nomination count, so the sheet with the most
    actual data wins any overlapping interval.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    scored: list[tuple[tuple[int, int], list[dict[str, Any]]]] = []
    for ws in wb.worksheets:
        items = parse_nomination_sheet(ws)
        if items:
            # Template sheets are filled with literal 0s, so rank by non-zero
            # nominations; break ties toward the explicitly-named nomination sheet.
            nonzero = sum(1 for it in items if it["nominated_mw"] not in (None, 0))
            prefer = 1 if "策略申报" in ws.title else 0
            scored.append(((nonzero, prefer), items))
    wb.close()
    merged: dict[Any, dict[str, Any]] = {}
    for _, items in sorted(scored, key=lambda si: si[0]):
        for it in items:
            key = it["interval_start"]
            if key in merged and it["planned_mw"] is None and it["nominated_mw"] is None:
                continue
            merged[key] = it
    return list(merged.values())


def ingest_nominations(root: str, batch_id: str) -> dict[str, Any]:
    """Walk the 申报策略 tree and write nominations to rm_nominations.

    Returns a report dict: {assets: {asset_name: rows}, skipped: [...], errors: [...]}.
    """
    import os
    import uuid
    from shared.agents.db import get_conn

    report: dict[str, Any] = {"assets": {}, "skipped": [], "errors": []}
    with get_conn() as conn:
        with conn.cursor() as cur:
            for folder, asset_name in NOMINATION_FOLDER_TO_ASSET.items():
                cur.execute("SELECT id FROM marketdata.rm_assets WHERE name = %s", (asset_name,))
                row = cur.fetchone()
                if not row:
                    report["errors"].append(f"asset not found: {asset_name}")
                    continue
                asset_id = row[0]

                folder_path = os.path.join(root, folder)
                if not os.path.isdir(folder_path):
                    report["skipped"].append(f"missing folder: {folder}")
                    continue

                rows_written = 0
                pending: list[tuple] = []
                for dirpath, _, files in os.walk(folder_path):
                    for fname in sorted(files):
                        if not fname.endswith(".xlsx") or fname.startswith("~$"):
                            continue
                        fpath = os.path.join(dirpath, fname)
                        try:
                            items = parse_nomination_file(fpath)
                        except Exception as e:
                            report["errors"].append(f"{fname}: {e}")
                            continue
                        if not items:
                            report["skipped"].append(f"{fname}: no nomination sheets")
                            continue
                        rel = os.path.join(dirpath.replace(root, "").lstrip("/"), fname)
                        pending.extend(
                            (asset_id, it["interval_start"], it["planned_mw"],
                             it["nominated_mw"], rel, batch_id)
                            for it in items
                        )
                if pending:
                    from psycopg2.extras import execute_batch
                    # Batched upsert: row-by-row INSERTs over WAN cost hours at
                    # 5-min granularity; execute_batch cuts round trips ~2000x.
                    execute_batch(cur, """
                        INSERT INTO marketdata.rm_nominations
                            (asset_id, interval_start, planned_mw, nominated_mw,
                             source_file, upload_batch_id)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        ON CONFLICT (asset_id, interval_start) DO UPDATE SET
                            planned_mw = EXCLUDED.planned_mw,
                            nominated_mw = EXCLUDED.nominated_mw,
                            source_file = EXCLUDED.source_file,
                            upload_batch_id = EXCLUDED.upload_batch_id
                    """, pending, page_size=2000)
                    rows_written = len(pending)
                report["assets"][asset_name] = rows_written
                # Per-station commit: this network kills long transactions
                # (ERRORS.md bulk-load rule — a dead connection costs one chunk)
                conn.commit()
                print(f"[nominations] {asset_name}: {rows_written:,} rows committed", flush=True)
    return report
