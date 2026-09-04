"""Dispatch chain ingestion (调度计划表 folder tree → rm_dispatch_chain).

Per-day sheets ("5.01".."5.31") plus an aggregate sheet (电力交易调度计划):
  时间 | SOC（%）| 交易员申报计划(MW) | 日前出清(MW) | 实时调度出清(MW) | 实际执行功率(MW)

The time cell colour encodes the restriction window:
  green (FF00B050) or no fill → NULL (both charge+discharge allowed)
  orange fill                 → 'charge_only'
  red fill (FFFF0000)         → 'discharge_only'
(colour semantics confirmed by user 2026-08-28)
"""
from __future__ import annotations

import re
from typing import Any

import openpyxl

# 调度计划表 station folder → rm_assets.name
DISPATCH_FOLDER_TO_ASSET = {
    "1.巴盟-景怡查干哈达": "景怡查干哈达",
    "2.谷山梁-裕昭沙子坝": "裕昭沙子坝",
    "3.杭锦旗-悦杭独贵": "悦杭独贵",
    "4.四子王旗-景通四益堂": "四子王旗",
    "5.苏右-景蓝乌尔图": "景蓝乌尔图",
    "6.乌拉特": "远景乌拉特",
}

_GREEN = {"FF00B050", "00B050", "92D050"}
_RED = {"FFFF0000", "FF0000"}
_ORANGE_LIKE = {"FFFFA500", "FFA500", "FFFFC000", "FFC000", "FFFFCC00", "FFCC00", "FFF4B084", "F4B084"}


def restriction_from_fill(fill) -> str | None:
    """Map a cell fill to a restriction value (see module docstring for semantics)."""
    if not fill or fill.fill_type != "solid" or not fill.start_color:
        return None
    rgb = fill.start_color.rgb
    if not isinstance(rgb, str):
        return None  # theme colours return an RGB descriptor, not a hex string
    rgb = rgb.upper()
    if rgb in _RED:
        return "discharge_only"
    if rgb in _ORANGE_LIKE:
        return "charge_only"
    return None  # green, blue-grey, anything else


def _norm_header(s: str) -> str:
    return re.sub(r"[\s（(].*$", "", (s or "").strip().lower())


def _find_columns(header_row: list[Any]) -> dict[str, int] | None:
    cols: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        n = _norm_header(str(cell or ""))
        if n == "时间":
            cols["time"] = i
        elif n.startswith("soc") or n == "soc（%）":
            cols["soc"] = i
        elif n.startswith("交易员申报计划"):
            cols["nominated"] = i
        elif n.startswith("日前出清"):
            cols["da_cleared"] = i
        elif n.startswith("实时调度出清"):
            cols["rt_cleared"] = i
        elif n.startswith("实际执行功率"):
            cols["actual"] = i
    if {"time", "nominated", "da_cleared", "rt_cleared", "actual"} <= cols.keys():
        return cols
    return None


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        m = re.search(r"-?\d[\d,]*\.?\d*", str(v))
        return float(m.group(0).replace(",", "")) if m else None


def parse_dispatch_sheet(ws, drop_mismatch: bool = True) -> list[dict[str, Any]]:
    """Parse one per-day dispatch sheet into interval dicts (with restriction from time-cell fill).

    drop_mismatch=True (default, legacy): rows whose in-sheet date doesn't match
    the sheet name are dropped (template junk protection). False: keep all rows
    and let the caller apply workbook-level date logic (frozen-template recovery).
    """
    header = None
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx > 5:
            break
        found = _find_columns(list(row))
        if found:
            header = found
            break
    if header is None:
        return []

    import pandas as pd

    dm = _DAILY_SHEET_RE.match(ws.title.strip()) if hasattr(ws, "title") else None
    sheet_md = (int(dm.group(1)), int(dm.group(2))) if dm else None

    items = []
    for row in ws.iter_rows(min_row=2, max_col=max(header.values()) + 1):
        time_v = row[header["time"]].value
        if time_v is None:
            continue
        try:
            ts = pd.Timestamp(time_v)
            if ts.tz is None:
                ts = ts.tz_localize("Asia/Shanghai")
        except Exception:
            continue
        if drop_mismatch and sheet_md and (ts.month, ts.day) != sheet_md:
            continue
        items.append({
            "interval_start": ts,
            "soc_pct": _to_float(row[header["soc"]].value) if "soc" in header else None,
            "nominated_mw": _to_float(row[header["nominated"]].value),
            "da_cleared_mw": _to_float(row[header["da_cleared"]].value),
            "rt_cleared_mw": _to_float(row[header["rt_cleared"]].value),
            "actual_mw": _to_float(row[header["actual"]].value),
            "restriction": restriction_from_fill(row[header["time"]].fill),
        })
    return items


_DAILY_SHEET_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})$")

_DISPATCH_FIELDS = ("soc_pct", "nominated_mw", "da_cleared_mw", "rt_cleared_mw", "actual_mw")


def _restriction_from_rgb(rgb) -> str | None:
    if not rgb:
        return None
    r, g, b = rgb
    if r > 180 and g < 100 and b < 100:
        return "discharge_only"
    if r > 200 and 100 <= g <= 190 and b < 90:
        return "charge_only"
    return None


def parse_dispatch_xls(file_path: str) -> list[dict[str, Any]]:
    """Legacy .xls variant of the monthly dispatch workbook (xlrd, palette fills)."""
    import xlrd

    import pandas as pd

    book = xlrd.open_workbook(file_path, formatting_info=True)
    items: list[dict[str, Any]] = []
    for name in book.sheet_names():
        sh = book.sheet_by_name(name)
        header = None
        for r in range(min(6, sh.nrows)):
            cols = _find_columns([sh.cell_value(r, c) for c in range(sh.ncols)])
            if cols:
                header = (r, cols)
                break
        if not header:
            continue
        hr, cols = header
        for r in range(hr + 1, sh.nrows):
            tv = sh.cell_value(r, cols["time"])
            if tv in ("", None):
                continue
            try:
                if isinstance(tv, (int, float)):
                    ts = pd.Timestamp(xlrd.xldate.xldate_as_datetime(tv, book.datemode), tz="Asia/Shanghai")
                else:
                    ts = pd.Timestamp(tv)
                    ts = ts.tz_localize("Asia/Shanghai") if ts.tz is None else ts
            except Exception:
                continue
            rgb = book.colour_map.get(
                book.xf_list[sh.cell_xf_index(r, cols["time"])].background.pattern_colour_index
            )
            items.append({
                "interval_start": ts,
                "soc_pct": _to_float(sh.cell_value(r, cols["soc"])) if "soc" in cols else None,
                "nominated_mw": _to_float(sh.cell_value(r, cols["nominated"])),
                "da_cleared_mw": _to_float(sh.cell_value(r, cols["da_cleared"])),
                "rt_cleared_mw": _to_float(sh.cell_value(r, cols["rt_cleared"])),
                "actual_mw": _to_float(sh.cell_value(r, cols["actual"])),
                "restriction": _restriction_from_rgb(rgb),
            })
    return items


def parse_dispatch_file(file_path: str) -> list[dict[str, Any]]:
    """Parse dispatch sheets in one workbook, merged by interval.

    Workbooks hold per-day sheets ("5.01".."5.31") and may also carry an
    aggregate sheet (电力交易调度计划) covering the same intervals. The
    aggregate is applied first, per-day sheets on top, so day-level data (and
    its restriction colours) wins; all-empty rows never overwrite real data.

    Frozen-template recovery (巴盟 2026-08 workbooks): the trader fill-series'd
    one date (2026-07-04) into every daily sheet's time column, so no row's date
    matches its sheet name. When ≥2 daily sheets share the SAME wrong modal date
    (and their data is not byte-identical, i.e. not template copies), the date
    column is treated as frozen and each row's date is substituted with the
    sheet-name date (time-of-day preserved). A single rogue mismatched sheet is
    still dropped (the 07.15 junk-sheet protection).
    """
    from collections import Counter, defaultdict

    wb = openpyxl.load_workbook(file_path, data_only=True)
    daily: list[tuple[str, tuple[int, int], list[dict[str, Any]]]] = []
    other: list[list[dict[str, Any]]] = []
    for ws in wb.worksheets:
        items = parse_dispatch_sheet(ws, drop_mismatch=False)
        if not items:
            continue
        dm = _DAILY_SHEET_RE.match(ws.title.strip())
        if dm:
            daily.append((ws.title.strip(), (int(dm.group(1)), int(dm.group(2))), items))
        else:
            other.append(items)
    wb.close()

    # Modal date per daily sheet
    modal: dict[str, Any] = {}
    for title, md, items in daily:
        cnt = Counter(it["interval_start"].date() for it in items)
        modal[title] = cnt.most_common(1)[0][0] if cnt else None

    # Wrong-modal-date sheets grouped by their shared wrong date
    by_wrong_date: dict[Any, list[str]] = defaultdict(list)
    for title, md, items in daily:
        m = modal.get(title)
        if m and (m.month, m.day) != md:
            by_wrong_date[m].append(title)

    def _fingerprint(title: str) -> tuple:
        items = next(it for t, _, it in daily if t == title)
        return tuple((it["nominated_mw"], it["actual_mw"]) for it in items[:12])

    frozen_dates = set()
    for wrong_date, titles in by_wrong_date.items():
        if len(titles) < 2:
            continue
        fps = {_fingerprint(t) for t in titles}
        if len(fps) > 1:
            # ≥2 sheets share the wrong date but carry different data →
            # the date column is a frozen template, not copied junk sheets
            frozen_dates.add(wrong_date)

    scored: list[tuple[int, list[dict[str, Any]]]] = []
    for title, md, items in daily:
        m = modal.get(title)
        if m is None:
            continue
        if (m.month, m.day) == md:
            # Normal sheet: keep only rows whose in-sheet date matches the sheet
            # name (drops template junk rows, e.g. stray 2026-05-15 in a 07.15 sheet)
            scored.append((1, [
                it for it in items
                if (it["interval_start"].month, it["interval_start"].day) == md
            ]))
        elif m in frozen_dates:
            # Frozen template column: substitute sheet-name date, keep time-of-day
            scored.append((1, [
                {**it, "interval_start": it["interval_start"].replace(month=md[0], day=md[1])}
                for it in items
            ]))
        # else: single rogue mismatched sheet → dropped (junk protection)
    for items in other:
        scored.append((0, items))

    merged: dict[Any, dict[str, Any]] = {}
    for _, items in sorted(scored, key=lambda si: si[0]):
        for it in items:
            key = it["interval_start"]
            if key in merged and all(it[f] is None for f in _DISPATCH_FIELDS):
                continue
            merged[key] = it
    return list(merged.values())


def ingest_dispatch_chain(root: str, batch_id: str) -> dict[str, Any]:
    """Walk the 调度计划表 tree and write to rm_dispatch_chain."""
    import os
    from shared.agents.db import get_conn

    report: dict[str, Any] = {"assets": {}, "skipped": [], "errors": []}
    with get_conn() as conn:
        with conn.cursor() as cur:
            for folder, asset_name in DISPATCH_FOLDER_TO_ASSET.items():
                cur.execute("SELECT id FROM marketdata.rm_assets WHERE name = %s", (asset_name,))
                row = cur.fetchone()
                if not row:
                    report["errors"].append(f"asset not found: {asset_name}")
                    continue
                asset_id = row[0]

                folder_path = os.path.join(root, folder)
                if not os.path.isdir(folder_path):
                    report["skipped"].append(f"missing folder: {folder}")
                    continue

                rows_written = 0
                pending: list[tuple] = []
                for dirpath, _, files in os.walk(folder_path):
                    for fname in sorted(files):
                        if not fname.lower().endswith((".xlsx", ".xls")) or fname.startswith("~$"):
                            continue
                        fpath = os.path.join(dirpath, fname)
                        try:
                            if fname.lower().endswith(".xls"):
                                items = parse_dispatch_xls(fpath)
                            else:
                                items = parse_dispatch_file(fpath)
                        except Exception as e:
                            report["errors"].append(f"{fname}: {e}")
                            continue
                        if not items:
                            report["skipped"].append(f"{fname}: no dispatch sheets")
                            continue
                        rel = os.path.join(dirpath.replace(root, "").lstrip("/"), fname)
                        pending.extend(
                            (asset_id, it["interval_start"], it["soc_pct"],
                             it["nominated_mw"], it["da_cleared_mw"], it["rt_cleared_mw"],
                             it["actual_mw"], it["restriction"], rel, batch_id)
                            for it in items
                        )
                if pending:
                    from psycopg2.extras import execute_batch
                    # Batched upsert — see nominations.py for why (WAN round trips).
                    execute_batch(cur, """
                        INSERT INTO marketdata.rm_dispatch_chain
                            (asset_id, interval_start, soc_pct, nominated_mw,
                             da_cleared_mw, rt_cleared_mw, actual_mw, restriction,
                             source_file, upload_batch_id)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        ON CONFLICT (asset_id, interval_start) DO UPDATE SET
                            soc_pct = EXCLUDED.soc_pct,
                            nominated_mw = EXCLUDED.nominated_mw,
                            da_cleared_mw = EXCLUDED.da_cleared_mw,
                            rt_cleared_mw = EXCLUDED.rt_cleared_mw,
                            actual_mw = EXCLUDED.actual_mw,
                            restriction = EXCLUDED.restriction,
                            source_file = EXCLUDED.source_file,
                            upload_batch_id = EXCLUDED.upload_batch_id
                    """, pending, page_size=2000)
                    rows_written = len(pending)
                report["assets"][asset_name] = rows_written
                # Per-station commit: this network kills long transactions
                conn.commit()
                print(f"[dispatch_chain] {asset_name}: {rows_written:,} rows committed", flush=True)
    return report


# ---------------------------------------------------------------------------
# Reorganised tree ingest (data/raw/nomination/, 2026-08-31 layout)
# ---------------------------------------------------------------------------

# Reorganised per-asset folder → rm_assets.name
REORG_FOLDER_TO_ASSET = {
    "杭锦旗": "悦杭独贵",
    "四子王旗": "四子王旗",
    "乌拉特中旗": "远景乌拉特",
    "巴盟": "景怡查干哈达",
    "苏右": "景蓝乌尔图",
    "谷山梁": "裕昭沙子坝",
}

_SKIP_SUBDIRS = {"archived", "client-reports"}


def _month_token(fname: str) -> int | None:
    """Month number from a workbook filename.

    Handles '4月'/'5月'/【4月】 Chinese-month tokens and 巴盟 export-date style
    '05.31'/'08.29' (mm.dd in the name — the month part).
    """
    m = re.search(r"(\d{1,2})月", fname)
    if m:
        return int(m.group(1))
    m = re.search(r"(?<!\d)(\d{2})\.(\d{2})(?!\d)", fname)
    if m:
        return int(m.group(1))
    return None


def select_latest_file_per_month(folder_path: str) -> dict[int, str]:
    """{month: path} of the latest-mtime workbook per month in one asset folder.

    Each daily download is a full monthly workbook snapshot (later = more
    day-tabs filled), so the latest-mtime file for a month is the fullest.
    Top-level files only — archived/ and client-reports/ are excluded.
    """
    import os
    out: dict[int, tuple[float, str]] = {}
    if not os.path.isdir(folder_path):
        return {}
    for fname in os.listdir(folder_path):
        fpath = os.path.join(folder_path, fname)
        if not os.path.isfile(fpath) or fname.startswith("~$"):
            continue
        if not fname.lower().endswith((".xlsx", ".xls")):
            continue
        mo = _month_token(fname)
        if mo is None or not (1 <= mo <= 12):
            continue
        mt = os.path.getmtime(fpath)
        if mo not in out or mt > out[mo][0]:
            out[mo] = (mt, fpath)
    return {mo: path for mo, (mt, path) in sorted(out.items())}


def ingest_dispatch_tree_reorg(root: str, batch_id: str, dry_run: bool = False,
                               assets: list[str] | None = None) -> dict[str, Any]:
    """Ingest the reorganised data/raw/nomination/ tree into rm_dispatch_chain.

    Per asset folder, only the latest-mtime workbook per month is parsed
    (per user rule 2026-08-31: most recent file per month; archived and
    client-reports subfolders ignored; 魏桥 excluded from the asset map).
    Upsert semantics per (asset_id, interval_start); per-station commit.

    dry_run=True parses and reports intervals/days per asset-month, writes nothing.
    """
    import os
    import pandas as pd
    from psycopg2.extras import execute_batch
    from shared.agents.db import get_conn

    report: dict[str, Any] = {"assets": {}, "files": [], "skipped": [], "errors": []}
    with get_conn() as conn:
        with conn.cursor() as cur:
            for folder, asset_name in REORG_FOLDER_TO_ASSET.items():
                if assets and asset_name not in assets:
                    continue
                cur.execute("SELECT id FROM marketdata.rm_assets WHERE name = %s", (asset_name,))
                row = cur.fetchone()
                if not row:
                    report["errors"].append(f"asset not found: {asset_name}")
                    continue
                asset_id = row[0]

                folder_path = os.path.join(root, folder)
                chosen = select_latest_file_per_month(folder_path)
                if not chosen:
                    report["skipped"].append(f"{folder}: no workbooks found")
                    continue

                pending: list[tuple] = []
                for mo, fpath in chosen.items():
                    fname = os.path.basename(fpath)
                    try:
                        if fname.lower().endswith(".xls"):
                            items = parse_dispatch_xls(fpath)
                        else:
                            items = parse_dispatch_file(fpath)
                    except Exception as e:
                        report["errors"].append(f"{fname}: {type(e).__name__}: {e}")
                        continue
                    days = sorted({it["interval_start"].date().isoformat() for it in items})
                    rel = os.path.join(folder, fname)
                    report["files"].append({
                        "asset": asset_name, "month": mo, "file": rel,
                        "intervals": len(items), "days": len(days),
                        "first": days[0] if days else None, "last": days[-1] if days else None,
                    })
                    if not items:
                        continue
                    pending.extend(
                        (asset_id, it["interval_start"], it["soc_pct"],
                         it["nominated_mw"], it["da_cleared_mw"], it["rt_cleared_mw"],
                         it["actual_mw"], it["restriction"], rel, batch_id)
                        for it in items
                    )

                if dry_run:
                    report["assets"][asset_name] = f"dry-run: {len(pending):,} rows would upsert"
                    continue
                if pending:
                    execute_batch(cur, """
                        INSERT INTO marketdata.rm_dispatch_chain
                            (asset_id, interval_start, soc_pct, nominated_mw,
                             da_cleared_mw, rt_cleared_mw, actual_mw, restriction,
                             source_file, upload_batch_id)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        ON CONFLICT (asset_id, interval_start) DO UPDATE SET
                            soc_pct = EXCLUDED.soc_pct,
                            nominated_mw = EXCLUDED.nominated_mw,
                            da_cleared_mw = EXCLUDED.da_cleared_mw,
                            rt_cleared_mw = EXCLUDED.rt_cleared_mw,
                            actual_mw = EXCLUDED.actual_mw,
                            restriction = EXCLUDED.restriction,
                            source_file = EXCLUDED.source_file,
                            upload_batch_id = EXCLUDED.upload_batch_id
                    """, pending, page_size=2000)
                    conn.commit()
                report["assets"][asset_name] = len(pending)
                print(f"[reorg] {asset_name}: {len(pending):,} rows committed", flush=True)
    return report
