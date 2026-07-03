"""
Market Fundamentals loader.

Parses: data/market-fundamentals/2023-2025 全国各省电力市场基础信息汇总2026-03-30.xlsx

Key sections extracted per province sheet:
  - 需求13: Installed capacity by fuel type (万kW) for 2024 and 2025
  - 需求14: Generation by fuel type (亿kWh) for 2024 and 2025
  - 需求11: Peak load (MW) by season for 2024 and 2025
"""
from __future__ import annotations

import re
from functools import lru_cache
from pathlib import Path

_REPO = Path(__file__).resolve().parents[2]
_EXCEL_DIR = _REPO / "data" / "market-fundamentals"

FUEL_TYPES_CN = ["风电", "光伏", "火电", "水电", "核电", "储能"]

FUEL_EN = {
    "风电": "Wind",
    "光伏": "Solar",
    "火电": "Thermal",
    "水电": "Hydro",
    "核电": "Nuclear",
    "储能": "Storage",
}

FUEL_COLORS = {
    "Wind":    "#4C9BE8",
    "Solar":   "#F5C842",
    "Thermal": "#9E9E9E",
    "Hydro":   "#2196F3",
    "Nuclear": "#9C27B0",
    "Storage": "#4CAF50",
}

PROVINCE_EN: dict[str, str] = {
    "北京": "Beijing",    "天津": "Tianjin",   "冀北": "Hebei-N",   "冀南": "Hebei-S",
    "山西": "Shanxi",     "蒙西": "Mengxi",    "蒙东": "Mengdong",  "辽宁": "Liaoning",
    "吉林": "Jilin",      "黑龙江": "Heilongjiang", "上海": "Shanghai", "江苏": "Jiangsu",
    "浙江": "Zhejiang",   "安徽": "Anhui",     "福建": "Fujian",    "山东": "Shandong",
    "河南": "Henan",      "湖北": "Hubei",     "湖南": "Hunan",     "广东": "Guangdong",
    "广西": "Guangxi",    "海南": "Hainan",    "重庆": "Chongqing", "四川": "Sichuan",
    "贵州": "Guizhou",    "云南": "Yunnan",    "西藏": "Tibet",     "陕西": "Shaanxi",
    "甘肃": "Gansu",      "青海": "Qinghai",   "宁夏": "Ningxia",   "新疆": "Xinjiang",
    "江西": "Jiangxi",
}

_SEASON_MAP = {"度夏": "summer", "度冬": "winter", "其余月份": "other"}


def _latest_excel() -> Path | None:
    """Return the province-summary .xlsx from the data directory.

    Preference order:
      1. Any file whose name contains '全国' or '市场基础' (the national summary).
      2. Largest file by size (most data).
    """
    if not _EXCEL_DIR.exists():
        return None
    candidates = list(_EXCEL_DIR.glob("*.xlsx"))
    if not candidates:
        return None
    # Prefer files with province-summary keywords in the name
    preferred = [p for p in candidates if "全国" in p.name or "市场基础" in p.name]
    pool = preferred if preferred else candidates
    return max(pool, key=lambda p: p.stat().st_size)


def _sheet_province(name: str) -> str | None:
    """Extract Chinese province name from sheet name; return None for deprecated sheets."""
    if name.endswith("旧") or "暂未上线" in name:
        return None
    m = re.match(r"^[\d.]+(.+)$", name)
    return m.group(1).strip() if m else name.strip()


def _load_sheet(ws) -> dict[tuple[int, int], object]:
    """Read sheet into {(row, col): value} dict (1-indexed, skipping None)."""
    data: dict[tuple[int, int], object] = {}
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=60, min_col=1, max_col=40, values_only=True), 1):
        for ci, v in enumerate(row, 1):
            if v is not None:
                data[(ri, ci)] = v
    return data


def _find_anchors(data: dict[tuple[int, int], object], label: str) -> list[tuple[int, int]]:
    """Return sorted list of (row, col) where label appears as a substring."""
    found = [(ri, ci) for (ri, ci), v in data.items() if isinstance(v, str) and label in v]
    return sorted(found)


def _parse_fuel_rows(
    data: dict[tuple[int, int], object],
    anchor_row: int,
    anchor_col: int,
    stop_before_row: int = 9999,
) -> dict[str, dict]:
    """
    Parse 6 fuel-type rows anchored at (anchor_row, anchor_col).

    Column layout (consistent across all sheets):
      anchor_col + 1 = fuel type (Chinese)
      anchor_col + 2 = numeric value (capacity/generation)
      anchor_col + 3 = share (0–1 fraction)

    Fuel rows start 2–3 rows below the anchor and run for up to 10 rows.
    stop_before_row: exclusive upper bound (prevents capacity section from
                     overlapping into the generation section rows).
    """
    type_col  = anchor_col + 1
    val_col   = anchor_col + 2
    share_col = anchor_col + 3
    result: dict[str, dict] = {}
    for ri in range(anchor_row, min(anchor_row + 12, stop_before_row)):
        v = data.get((ri, type_col))
        if not (isinstance(v, str) and v in FUEL_TYPES_CN):
            continue
        raw_val   = data.get((ri, val_col))
        raw_share = data.get((ri, share_col))
        val   = float(raw_val)   if isinstance(raw_val, (int, float)) and raw_val not in ("-",) else None
        share = float(raw_share) if isinstance(raw_share, (int, float)) else None
        result[v] = {"value": val, "share": share}
    return result


def _parse_peak_load(
    data: dict[tuple[int, int], object],
    anchor_row: int,
    anchor_col: int,
    stop_before_row: int = 9999,
) -> dict[str, float]:
    """
    Parse peak-load rows anchored at (anchor_row, anchor_col).

    Column layout:
      anchor_col + 1 = season label (度夏 / 度冬 / 其余月份)
      anchor_col + 2 = MW value

    stop_before_row: do not read rows >= this value (prevents overlap with
                     the next 需求11 section in the same sheet).
    """
    label_col = anchor_col + 1
    val_col   = anchor_col + 2
    result: dict[str, float] = {}
    for ri in range(anchor_row, min(anchor_row + 8, stop_before_row)):
        label = data.get((ri, label_col))
        if not isinstance(label, str):
            continue
        for cn, en in _SEASON_MAP.items():
            if cn in label:
                val = data.get((ri, val_col))
                if isinstance(val, (int, float)):
                    result[en] = float(val)
                break
    return result


@lru_cache(maxsize=1)
def load_province_data() -> dict[str, dict]:
    """
    Load all province market fundamentals from the Excel file.

    Returns:
        {
            province_cn: {
                "province_en": str,
                "capacity":   {2024: {fuel_cn: {"value": float, "share": float}}, 2025: {...}},
                "generation": {2024: {fuel_cn: {"value": float, "share": float}}, 2025: {...}},
                "peak_load":  {2024: {"summer": float, "winter": float, "other": float}, 2025: {...}},
            }
        }
    """
    import openpyxl

    path = _latest_excel()
    if path is None:
        return {}

    import io as _io, shutil as _shutil, tempfile as _tempfile
    # On Windows with certain Unicode path encodings, io.open() may fail with
    # PermissionError even though shutil can read the file.  Copy to a temp
    # location with an ASCII name and load from there.
    try:
        _raw = path.read_bytes()
    except (PermissionError, OSError):
        _tmp = _tempfile.mktemp(suffix=".xlsx")
        _shutil.copy2(str(path), _tmp)
        with open(_tmp, "rb") as _fh:
            _raw = _fh.read()
        try:
            import os as _os; _os.unlink(_tmp)
        except OSError:
            pass
    wb = openpyxl.load_workbook(_io.BytesIO(_raw), read_only=True, data_only=True)
    result: dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        province_cn = _sheet_province(sheet_name)
        if province_cn is None:
            continue

        ws   = wb[sheet_name]
        data = _load_sheet(ws)

        # ── Installed capacity (需求13) ────────────────────────────────────────
        d13_anchors = sorted(_find_anchors(data, "需求13"), key=lambda x: x[1])[:2]
        # Stop capacity scan before the generation section to avoid row overlap
        d14_first_row = min((r for r, _ in _find_anchors(data, "需求14")), default=9999)
        capacity: dict[int, dict] = {}
        for i, (ar, ac) in enumerate(d13_anchors):
            year = 2024 + i
            capacity[year] = _parse_fuel_rows(data, ar, ac, stop_before_row=d14_first_row)

        # ── Generation (需求14) ────────────────────────────────────────────────
        d14_anchors = sorted(_find_anchors(data, "需求14"), key=lambda x: x[1])[:2]
        generation: dict[int, dict] = {}
        for i, (ar, ac) in enumerate(d14_anchors):
            year = 2024 + i
            generation[year] = _parse_fuel_rows(data, ar, ac)

        # ── Peak load (需求11) ─────────────────────────────────────────────────
        # Each 需求11 anchor has a year label in the adjacent cell.
        # Sort anchors by row so we can pass stop_before_row = next anchor's row.
        d11_anchors = sorted(_find_anchors(data, "需求11"), key=lambda x: x[0])
        peak_load: dict[int, dict] = {}
        for idx, (ar, ac) in enumerate(d11_anchors):
            year_label = data.get((ar, ac + 1), "")
            if not isinstance(year_label, str):
                continue
            year = 2025 if "2025" in year_label else (2024 if "2024" in year_label else None)
            if year is None:
                continue
            # Stop before the next 需求11 anchor to avoid row overlap
            next_anchor_row = d11_anchors[idx + 1][0] if idx + 1 < len(d11_anchors) else 9999
            pl = _parse_peak_load(data, ar, ac, stop_before_row=next_anchor_row)
            if pl:
                peak_load[year] = pl

        result[province_cn] = {
            "province_en": PROVINCE_EN.get(province_cn, province_cn),
            "capacity":    capacity,
            "generation":  generation,
            "peak_load":   peak_load,
        }

    return result


def load_latest_installed_monthly(dsn: str) -> dict[str, dict]:
    """
    Return the most recent monthly snapshot per province from
    province_installed_monthly (populated by scripts/scan_installed_capacity.py).

    Returns:
        {
            province_cn: {
                "year_month":  datetime.date,
                "wind_mw":     float | None,
                "solar_mw":    float | None,
                "thermal_mw":  float | None,
                "hydro_mw":    float | None,
                "nuclear_mw":  float | None,
                "bess_mw":     float | None,
                "total_mw":    float | None,
                "renew_mw":    float,          # wind + solar
                "source_file": str | None,
            }
        }
    """
    import psycopg2
    sql = """
        SELECT DISTINCT ON (province)
            province, year_month,
            wind_mw, solar_mw, thermal_mw, hydro_mw, nuclear_mw, bess_mw,
            total_mw, source_file
        FROM marketdata.province_installed_monthly
        ORDER BY province, year_month DESC
    """
    result: dict[str, dict] = {}
    try:
        conn = psycopg2.connect(dsn)
        try:
            with conn.cursor() as cur:
                cur.execute(sql)
                for row in cur.fetchall():
                    (prov, ym, wind, solar, thermal, hydro, nuclear, bess,
                     total, src) = row
                    wind   = float(wind)   if wind   is not None else None
                    solar  = float(solar)  if solar  is not None else None
                    thermal = float(thermal) if thermal is not None else None
                    hydro  = float(hydro)  if hydro  is not None else None
                    nuclear = float(nuclear) if nuclear is not None else None
                    bess   = float(bess)   if bess   is not None else None
                    total  = float(total)  if total  is not None else None
                    renew  = (wind or 0.0) + (solar or 0.0)
                    result[prov] = {
                        "year_month":  ym,
                        "wind_mw":     wind,
                        "solar_mw":    solar,
                        "thermal_mw":  thermal,
                        "hydro_mw":    hydro,
                        "nuclear_mw":  nuclear,
                        "bess_mw":     bess,
                        "total_mw":    total,
                        "renew_mw":    renew,
                        "source_file": src,
                    }
        finally:
            conn.close()
    except Exception:
        pass   # table may not exist yet; caller falls back to annual data
    return result


def get_fundamentals_summary(
    provinces: list[str] | None = None,
    year: int = 2025,
) -> dict:
    """
    Return a compact JSON-serialisable summary suitable for the agent tool.

    Args:
        provinces: List of province_en names (English) to filter.
                   If None, all provinces are returned.
        year:      2024 or 2025.

    Returns:
        {
          "year": int,
          "provinces": [
            {
              "province_cn": str,
              "province_en": str,
              "capacity_10kw": {"Wind": float, "Solar": float, ...},
              "capacity_share": {"Wind": float, ...},          # 0-1
              "generation_100gwh": {"Wind": float, ...},
              "generation_share": {"Wind": float, ...},
              "peak_summer_mw": float | None,
              "peak_winter_mw": float | None,
            },
            ...
          ]
        }
    """
    raw = load_province_data()
    rows = []
    for pcn, info in raw.items():
        pen = info["province_en"]
        if provinces and pen not in provinces:
            continue
        cap_raw = info["capacity"].get(year, {})
        gen_raw = info["generation"].get(year, {})
        pl      = info["peak_load"].get(year, {})

        cap_mw  = {FUEL_EN[k]: v["value"] for k, v in cap_raw.items() if v["value"] is not None}
        cap_sh  = {FUEL_EN[k]: v["share"] for k, v in cap_raw.items() if v["share"] is not None}
        gen_gwh = {FUEL_EN[k]: v["value"] for k, v in gen_raw.items() if v["value"] is not None}
        gen_sh  = {FUEL_EN[k]: v["share"] for k, v in gen_raw.items() if v["share"] is not None}

        rows.append({
            "province_cn":        pcn,
            "province_en":        pen,
            "capacity_10kw":      cap_mw,
            "capacity_share":     cap_sh,
            "generation_100gwh":  gen_gwh,
            "generation_share":   gen_sh,
            "peak_summer_mw":     pl.get("summer"),
            "peak_winter_mw":     pl.get("winter"),
        })

    rows.sort(key=lambda r: r["province_en"])
    return {"year": year, "provinces": rows}


# ── DB-backed loader ─────────────────────────────────────────────────────────

_FUEL_CN_FROM_COL = {
    "wind":    "风电",
    "solar":   "光伏",
    "thermal": "火电",
    "hydro":   "水电",
    "nuclear": "核电",
    "storage": "储能",
}


def load_province_data_from_db(dsn: str, schema: str = "marketdata") -> dict[str, dict]:
    """
    Load province fundamentals from the DB table (populated by
    scripts/ingest_province_fundamentals.py).

    Returns the same structure as load_province_data():
        {
            province_cn: {
                "province_en": str,
                "capacity":   {2024: {fuel_cn: {"value": float, "share": float}}, 2025: {...}},
                "generation": {2024: {fuel_cn: {"value": float, "share": float}}, 2025: {...}},
                "peak_load":  {2024: {"summer": float, "winter": float}, 2025: {...}},
            }
        }
    """
    import psycopg2
    sql = f"""
        SELECT province_cn, province_en, year,
               wind_cap_10kw, solar_cap_10kw, thermal_cap_10kw,
               hydro_cap_10kw, nuclear_cap_10kw, storage_cap_10kw,
               wind_gen_100gwh, solar_gen_100gwh, thermal_gen_100gwh,
               hydro_gen_100gwh, nuclear_gen_100gwh, storage_gen_100gwh,
               peak_summer_mw, peak_winter_mw, peak_other_mw
        FROM {schema}.province_fundamentals
        ORDER BY province_cn, year
    """
    result: dict[str, dict] = {}
    with psycopg2.connect(dsn) as conn:
        with conn.cursor() as cur:
            cur.execute(sql)
            for row in cur.fetchall():
                (pcn, pen, year,
                 w_cap, s_cap, th_cap, h_cap, n_cap, st_cap,
                 w_gen, s_gen, th_gen, h_gen, n_gen, st_gen,
                 p_sum, p_win, p_oth) = row

                if pcn not in result:
                    result[pcn] = {
                        "province_en": pen,
                        "capacity":   {},
                        "generation": {},
                        "peak_load":  {},
                    }

                cap_vals = [w_cap, s_cap, th_cap, h_cap, n_cap, st_cap]
                gen_vals = [w_gen, s_gen, th_gen, h_gen, n_gen, st_gen]
                fuel_keys = ["风电", "光伏", "火电", "水电", "核电", "储能"]

                result[pcn]["capacity"][year] = {
                    fuel: {"value": v, "share": None}
                    for fuel, v in zip(fuel_keys, cap_vals)
                    if v is not None
                }
                result[pcn]["generation"][year] = {
                    fuel: {"value": v, "share": None}
                    for fuel, v in zip(fuel_keys, gen_vals)
                    if v is not None
                }
                pl: dict[str, float] = {}
                if p_sum is not None:
                    pl["summer"] = p_sum
                if p_win is not None:
                    pl["winter"] = p_win
                if p_oth is not None:
                    pl["other"] = p_oth
                if pl:
                    result[pcn]["peak_load"][year] = pl

    return result
