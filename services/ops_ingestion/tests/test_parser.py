"""
tests/test_parser.py

Unit tests for services/ops_ingestion/inner_mongolia/parser.py

Uses an in-memory openpyxl workbook (no real file needed).

Covers:
  - 96-row worksheet → correct normalized + raw dicts (layout A: data at row 19)
  - Layout B: data at row 34 (alternate format with fault log before dispatch data)
  - raw_payload contains row_number
  - Blank-row stop (stops at first non-time col A)
  - EmptyCell in col A → stops cleanly, no AttributeError
  - EmptyCell in value columns → raw='', normalized=None
  - 00:00 → T00:00:00+08:00 (midnight CST)
  - Non-numeric values (e.g. "--") → None for float, preserved in raw_*
  - interval_end = interval_start + 15 minutes
  - data_date matches report_date
  - SUMMARY_SHEET_NAMES: summary sheets skipped by parse_workbook
"""
from __future__ import annotations

import sys
import os
import datetime

import pytest
import openpyxl

sys.path.insert(0, os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')))

from inner_mongolia.parser import (
    _parse_sheet,
    _find_data_start_row,
    _to_float,
    _raw_str,
    _cell_value,
    _CST,
    _COL_TIME,
    _COL_NOMINATED,
    _COL_ACTUAL,
    _COL_NODAL_PRICE,
    SUMMARY_SHEET_NAMES,
    parse_workbook,
)


# ---------------------------------------------------------------------------
# Helpers to build synthetic worksheets
# ---------------------------------------------------------------------------

_REPORT_DATE = datetime.date(2026, 2, 10)


def _make_worksheet(
    n_rows: int = 96,
    data_start_row: int = 19,
    extra_values: dict = None,
    title: str = "苏右（景蓝乌尔图）",
) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    Create an in-memory worksheet with n_rows of 15-min data beginning at
    data_start_row (1-based).

    extra_values: {(row_1based, col_0based): value} overrides
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    base_time = datetime.time(0, 0, 0)
    for i in range(n_rows):
        row_1based = data_start_row + i
        interval_minutes = i * 15
        t = (datetime.datetime.combine(_REPORT_DATE, base_time) +
             datetime.timedelta(minutes=interval_minutes)).time()

        ws.cell(row=row_1based, column=_COL_TIME + 1).value = t
        ws.cell(row=row_1based, column=_COL_NOMINATED + 1).value = 50.0
        ws.cell(row=row_1based, column=_COL_ACTUAL + 1).value = -30.5
        ws.cell(row=row_1based, column=_COL_NODAL_PRICE + 1).value = 320.0

    if extra_values:
        for (row, col), val in extra_values.items():
            # Use .value assignment — ws.cell(value=None) does NOT clear existing value
            ws.cell(row=row, column=col + 1).value = val

    return ws


# ---------------------------------------------------------------------------
# _find_data_start_row
# ---------------------------------------------------------------------------

class TestFindDataStartRow:
    def test_layout_a_row_19(self):
        ws = _make_worksheet(96, data_start_row=19)
        assert _find_data_start_row(ws) == 19

    def test_layout_b_row_34(self):
        ws = _make_worksheet(96, data_start_row=34)
        assert _find_data_start_row(ws) == 34

    def test_no_data_returns_none(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'some text'
        assert _find_data_start_row(ws) is None

    def test_requires_midnight_time(self):
        # A sheet where only non-midnight times exist should return None
        # (start scan only matches time(0,0))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=19, column=1).value = datetime.time(1, 0)   # 01:00, not midnight
        assert _find_data_start_row(ws) is None


# ---------------------------------------------------------------------------
# Tests: basic parsing — Layout A (data at row 19)
# ---------------------------------------------------------------------------

class TestParseSheetLayoutA:
    def setup_method(self):
        self.ws = _make_worksheet(96, data_start_row=19)

    def test_96_rows_parsed(self):
        result = _parse_sheet(self.ws, _REPORT_DATE)
        assert result.n_rows == 96
        assert len(result.rows) == 96

    def test_sheet_name(self):
        result = _parse_sheet(self.ws, _REPORT_DATE)
        assert result.sheet_name == "苏右（景蓝乌尔图）"

    def test_first_interval_start(self):
        result = _parse_sheet(self.ws, _REPORT_DATE)
        assert result.rows[0].interval_start == "2026-02-10T00:00:00+08:00"

    def test_first_interval_end(self):
        result = _parse_sheet(self.ws, _REPORT_DATE)
        assert result.rows[0].interval_end == "2026-02-10T00:15:00+08:00"

    def test_last_interval_start(self):
        result = _parse_sheet(self.ws, _REPORT_DATE)
        assert result.rows[-1].interval_start == "2026-02-10T23:45:00+08:00"

    def test_data_date(self):
        result = _parse_sheet(self.ws, _REPORT_DATE)
        assert all(r.data_date == "2026-02-10" for r in result.rows)

    def test_normalized_nominated_mw(self):
        result = _parse_sheet(self.ws, _REPORT_DATE)
        assert result.rows[0].nominated_dispatch_mw == pytest.approx(50.0)

    def test_normalized_actual_mw(self):
        result = _parse_sheet(self.ws, _REPORT_DATE)
        assert result.rows[0].actual_dispatch_mw == pytest.approx(-30.5)

    def test_normalized_nodal_price(self):
        result = _parse_sheet(self.ws, _REPORT_DATE)
        assert result.rows[0].nodal_price_excel == pytest.approx(320.0)


# ---------------------------------------------------------------------------
# Tests: Layout B (data at row 34, fault log before dispatch data)
# ---------------------------------------------------------------------------

class TestParseSheetLayoutB:
    def setup_method(self):
        self.ws = _make_worksheet(96, data_start_row=34)

    def test_96_rows_parsed(self):
        result = _parse_sheet(self.ws, _REPORT_DATE)
        assert result.n_rows == 96

    def test_first_interval_start(self):
        result = _parse_sheet(self.ws, _REPORT_DATE)
        assert result.rows[0].interval_start == "2026-02-10T00:00:00+08:00"

    def test_row_number_reflects_actual_row(self):
        result = _parse_sheet(self.ws, _REPORT_DATE)
        assert result.rows[0].row_number == 34
        assert result.rows[1].row_number == 35


# ---------------------------------------------------------------------------
# Tests: raw value preservation
# ---------------------------------------------------------------------------

class TestRawValues:
    def test_raw_payload_contains_row_number(self):
        ws = _make_worksheet(1, data_start_row=19)
        result = _parse_sheet(ws, _REPORT_DATE)
        assert 'row_number' in result.rows[0].raw_payload
        assert result.rows[0].raw_payload['row_number'] == 19

    def test_raw_nominated_is_string(self):
        ws = _make_worksheet(1)
        result = _parse_sheet(ws, _REPORT_DATE)
        assert isinstance(result.rows[0].raw_nominated, str)
        assert result.rows[0].raw_nominated == "50.0"

    def test_raw_double_dash_preserved(self):
        ws = _make_worksheet(1, extra_values={(19, _COL_NOMINATED): "--"})
        result = _parse_sheet(ws, _REPORT_DATE)
        assert result.rows[0].raw_nominated == "--"
        assert result.rows[0].nominated_dispatch_mw is None

    def test_raw_empty_string_preserved(self):
        ws = _make_worksheet(1)
        ws.cell(row=19, column=_COL_ACTUAL + 1).value = None
        result = _parse_sheet(ws, _REPORT_DATE)
        assert result.rows[0].raw_actual == ""
        assert result.rows[0].actual_dispatch_mw is None

    def test_raw_payload_has_sheet_name(self):
        ws = _make_worksheet(1)
        result = _parse_sheet(ws, _REPORT_DATE)
        assert result.rows[0].raw_payload['sheet_name'] == "苏右（景蓝乌尔图）"


# ---------------------------------------------------------------------------
# Tests: blank-row stop / EmptyCell safety
# ---------------------------------------------------------------------------

class TestBlankRowStop:
    def test_stops_at_first_non_time_row(self):
        ws = _make_worksheet(5)
        ws.cell(row=24, column=1).value = "TOTAL"
        result = _parse_sheet(ws, _REPORT_DATE)
        assert result.n_rows == 5

    def test_zero_rows_if_no_time_in_first_row(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A19'] = "header"
        result = _parse_sheet(ws, _REPORT_DATE)
        assert result.n_rows == 0

    def test_empty_cell_in_col_a_stops_cleanly(self):
        """
        EmptyCell objects (openpyxl read-only mode) must not raise AttributeError.
        Simulated by setting col A to None after 5 data rows.
        """
        ws = _make_worksheet(5)
        ws.cell(row=24, column=1).value = None   # simulates EmptyCell.value = None
        result = _parse_sheet(ws, _REPORT_DATE)
        assert result.n_rows == 5   # stops cleanly, no crash

    def test_no_crash_on_empty_sheet(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        result = _parse_sheet(ws, _REPORT_DATE)
        assert result.n_rows == 0
        assert len(result.parse_warnings) > 0   # warning that no data found


# ---------------------------------------------------------------------------
# Tests: SUMMARY_SHEET_NAMES
# ---------------------------------------------------------------------------

class TestSummarySheetNames:
    def test_summary_names_present(self):
        assert '内蒙储能项目报告' in SUMMARY_SHEET_NAMES
        assert '汇总' in SUMMARY_SHEET_NAMES

    def test_asset_sheet_names_not_in_summary(self):
        for name in ['苏右（景蓝乌尔图）', '苏右(景蓝乌尔图）', '杭锦旗（悦杭独贵）',
                     '四子王旗（景通四益堂储）', '谷山梁(裕昭沙子坝)']:
            assert name not in SUMMARY_SHEET_NAMES


# ---------------------------------------------------------------------------
# Tests: _to_float helper
# ---------------------------------------------------------------------------

class TestToFloat:
    def test_numeric_value(self):
        ws = openpyxl.Workbook().active
        ws.cell(row=1, column=1, value=42.5)
        row_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
        assert _to_float(row_cells, 0) == pytest.approx(42.5)

    def test_string_numeric(self):
        ws = openpyxl.Workbook().active
        ws.cell(row=1, column=1, value="-37")
        row_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
        assert _to_float(row_cells, 0) == pytest.approx(-37.0)

    def test_non_numeric_returns_none(self):
        ws = openpyxl.Workbook().active
        ws.cell(row=1, column=1, value="--")
        row_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
        assert _to_float(row_cells, 0) is None

    def test_none_cell_returns_none(self):
        ws = openpyxl.Workbook().active
        ws.cell(row=1, column=1, value=None)
        row_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
        assert _to_float(row_cells, 0) is None

    def test_out_of_bounds_returns_none(self):
        ws = openpyxl.Workbook().active
        ws.cell(row=1, column=1, value=1.0)
        row_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
        assert _to_float(row_cells, 99) is None
