"""
services/ops_ingestion/inner_mongolia/parser.py

Parse one Inner Mongolia BESS daily operations Excel workbook.

Workbook layout (observed across 93 production files)
------------------------------------------------------
Two workbook layouts exist in the wild:

  Layout A (67/93 files) — 6 sheets:
    Index 0: '内蒙储能项目报告'   ← summary, skipped
    Index 1: '汇总'               ← summary, skipped
    Index 2: '苏右(景蓝乌尔图）'  ← asset data
    Index 3: '杭锦旗(悦杭独贵)'   ← asset data
    Index 4: '四子王旗（景通四益堂储）' ← asset data
    Index 5: '谷山梁(裕昭沙子坝)' ← asset data

  Layout B (23/93 files) — 4 sheets (no summary sheets):
    Index 0: '谷山梁(裕昭沙子坝)'
    Index 1: '杭锦旗(悦杭独贵)'
    Index 2: '苏右(景蓝乌尔图）'
    Index 3: '四子王旗（景通四益堂储）'

  Occasional extras (1/93): extra asset sheets beyond the 4 core ones.
  Sheet2 files (2/93): likely corrupted — yield 0 data rows.

Summary sheets are identified by name in SUMMARY_SHEET_NAMES and skipped
before parsing.  All remaining sheets are parsed; the matcher decides which
are asset sheets.

Sheet naming: bracket characters are mixed — e.g. '苏右(景蓝乌尔图）' uses
ASCII '(' and full-width '）'. The matcher normalises these, so the parser
does not need to care.

Data layout per asset sheet
---------------------------
- Row 18 (1-based): header row — skipped
- Rows 19–114: 96 × 15-min data rows (stop at first non-time col A)
  Col A (index 0): datetime.time  — interval time
  Col B (index 1): nominated dispatch MW (申报曲线)
  Col D (index 3): actual dispatch MW (实际充放曲线)
  Col E (index 4): nodal electricity price CNY/MWh (节点电价)

EmptyCell handling
------------------
openpyxl in read-only mode returns EmptyCell objects for rows that are absent
from the worksheet XML.  EmptyCell.value is None but EmptyCell has NO .row
attribute.  The parser NEVER accesses cell.row; row numbers are tracked via
enumerate starting from DATA_START_ROW (1-based 19).

Data stop condition
-------------------
Parsing stops at the first row where col A is not a datetime.time value.
EmptyCell in col A has value=None → isinstance(None, datetime.time) is False
→ parser stops cleanly without crashing.
"""
from __future__ import annotations

import datetime
from dataclasses import dataclass, field
from typing import List, Optional

import openpyxl


# ---------------------------------------------------------------------------
# Result types
# ---------------------------------------------------------------------------

@dataclass
class ParsedRow:
    """One 15-min interval from one asset sheet."""
    sheet_name: str
    row_number: int              # 1-based row number in the workbook
    interval_start: str          # ISO 8601 with +08:00, e.g. "2026-02-10T00:00:00+08:00"
    interval_end: str            # interval_start + 15 minutes
    data_date: str               # ISO date, e.g. "2026-02-10"
    nominated_dispatch_mw: Optional[float]
    actual_dispatch_mw: Optional[float]
    nodal_price_excel: Optional[float]
    # Raw values
    raw_nominated: str
    raw_actual: str
    raw_nodal_price: str
    raw_payload: dict            # full row dict for JSONB storage


@dataclass
class SheetParseResult:
    """Parse output for one sheet in the workbook."""
    sheet_name: str
    rows: List[ParsedRow] = field(default_factory=list)
    n_rows: int = 0
    parse_warnings: List[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Column indices (0-based)
_COL_TIME = 0
_COL_NOMINATED = 1
_COL_ACTUAL = 3
_COL_NODAL_PRICE = 4

# Data start row scan range (1-based, inclusive).
# Layout A: data at row 19.  Layout B: data at row 34.
# The parser auto-detects by looking for the first row where col A is
# datetime.time(0,0) within this range.
_DATA_SCAN_MIN_ROW = 5
_DATA_SCAN_MAX_ROW = 50

# Sheet names that are summaries — skip these entirely, not passed to matcher
SUMMARY_SHEET_NAMES: frozenset = frozenset({
    '内蒙储能项目报告',
    '汇总',
})

# CST timezone offset
_CST = datetime.timezone(datetime.timedelta(hours=8))


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_workbook(path: str, report_date: datetime.date) -> List[SheetParseResult]:
    """
    Parse all non-summary sheets in the workbook.

    Summary sheets named in SUMMARY_SHEET_NAMES are silently skipped.
    All remaining sheets are parsed and returned; the matcher downstream
    decides which are asset sheets vs. unknown extras.

    Parameters
    ----------
    path : str
        Path to the .xlsx file.
    report_date : datetime.date
        The date to assign to each interval (from date_parser.parse_date).

    Returns
    -------
    List[SheetParseResult]
        One entry per non-summary sheet.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    results = []
    try:
        for ws in wb.worksheets:
            if ws.title in SUMMARY_SHEET_NAMES:
                continue   # silently skip known summary sheets
            result = _parse_sheet(ws, report_date)
            results.append(result)
    finally:
        wb.close()
    return results


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _find_data_start_row(ws) -> int | None:
    """
    Auto-detect the 1-based row where 15-min dispatch data begins.

    Looks for the first row in [_DATA_SCAN_MIN_ROW, _DATA_SCAN_MAX_ROW] where
    col A contains datetime.time(0, 0) (the midnight interval).

    Returns None if no such row is found (sheet has no dispatch data).
    """
    target = datetime.time(0, 0)
    for row_cells in ws.iter_rows(
        min_row=_DATA_SCAN_MIN_ROW,
        max_row=_DATA_SCAN_MAX_ROW,
        values_only=False,
    ):
        val = _cell_value(row_cells, _COL_TIME)
        if val == target:
            # Return the actual row number of col A if available;
            # fall back to inferring it from iter position.
            row_num = getattr(row_cells[0], 'row', None) if row_cells else None
            if row_num is not None:
                return row_num
            # Estimate from the scan (less precise but EmptyCell-safe)
            # We won't reach here for data rows — they're ReadOnlyCell
    return None


def _parse_sheet(ws, report_date: datetime.date) -> SheetParseResult:
    """
    Parse one worksheet into a SheetParseResult.

    Auto-detects the data start row (handles both row-19 and row-34 layouts).
    Row numbers are tracked via enumerate — never via cell.row on loop rows —
    so that EmptyCell objects (no .row attribute) are handled safely.
    """
    result = SheetParseResult(sheet_name=ws.title)

    data_start = _find_data_start_row(ws)
    if data_start is None:
        result.parse_warnings.append(
            f"No 15-min dispatch data found in sheet {ws.title!r} "
            f"(scanned rows {_DATA_SCAN_MIN_ROW}–{_DATA_SCAN_MAX_ROW})"
        )
        return result

    for row_idx, row_cells in enumerate(
        ws.iter_rows(min_row=data_start, values_only=False)
    ):
        # Track 1-based row number without relying on cell.row
        row_num = data_start + row_idx

        # Col A must be a datetime.time to be a data row.
        # EmptyCell.value is None → isinstance check returns False → stops cleanly.
        time_cell_value = _cell_value(row_cells, _COL_TIME)
        if not isinstance(time_cell_value, datetime.time):
            break

        # Construct timestamps with explicit CST offset
        dt_naive = datetime.datetime.combine(report_date, time_cell_value)
        dt_cst = dt_naive.replace(tzinfo=_CST)
        dt_end_cst = dt_cst + datetime.timedelta(minutes=15)

        interval_start = dt_cst.isoformat()
        interval_end = dt_end_cst.isoformat()

        # Extract raw cell values (as strings)
        raw_nominated = _raw_str(row_cells, _COL_NOMINATED)
        raw_actual = _raw_str(row_cells, _COL_ACTUAL)
        raw_nodal_price = _raw_str(row_cells, _COL_NODAL_PRICE)

        # Coerce to float (None on failure)
        nominated_mw = _to_float(row_cells, _COL_NOMINATED)
        actual_mw = _to_float(row_cells, _COL_ACTUAL)
        nodal_price = _to_float(row_cells, _COL_NODAL_PRICE)

        raw_payload = {
            "row_number": row_num,
            "sheet_name": ws.title,
            "time_str": str(time_cell_value),
            "col_b_raw": raw_nominated,
            "col_d_raw": raw_actual,
            "col_e_raw": raw_nodal_price,
        }

        parsed_row = ParsedRow(
            sheet_name=ws.title,
            row_number=row_num,
            interval_start=interval_start,
            interval_end=interval_end,
            data_date=report_date.isoformat(),
            nominated_dispatch_mw=nominated_mw,
            actual_dispatch_mw=actual_mw,
            nodal_price_excel=nodal_price,
            raw_nominated=raw_nominated,
            raw_actual=raw_actual,
            raw_nodal_price=raw_nodal_price,
            raw_payload=raw_payload,
        )
        result.rows.append(parsed_row)

    result.n_rows = len(result.rows)
    return result


def _cell_value(row_cells, col_idx: int):
    """
    Safely get a cell value by column index.

    Returns None for:
    - out-of-bounds index
    - EmptyCell (which has value=None but may lack other attributes)
    """
    if col_idx >= len(row_cells):
        return None
    return getattr(row_cells[col_idx], 'value', None)


def _raw_str(row_cells, col_idx: int) -> str:
    """Return str(cell.value) for the given column, or '' if out of bounds / None."""
    val = _cell_value(row_cells, col_idx)
    if val is None:
        return ""
    return str(val)


def _to_float(row_cells, col_idx: int) -> Optional[float]:
    """
    Coerce a cell value to float.  Returns None for:
      - out-of-bounds column
      - None cell value
      - non-numeric strings (e.g. "--", "N/A", empty string)
    """
    val = _cell_value(row_cells, col_idx)
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return None
