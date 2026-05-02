"""
Excel read/write and cross-check for the spot market summary workbook.

File: data/spot reports/2026/电力现货市场价格与运行日报数据汇总-apr26.xlsx

Observed structure (from inspection):
  Row 0: title (province name)
  Row 1: section headers  (日期 | 实时市场 | 日前市场)
  Row 2: column headers   (exact field names)
  Row 3+: data rows       (col 0 = datetime)

Column layout (col indices, typical sheet):
  0   日期           → date
  1   实时市场出清均价  → rt_avg
  3   实时市场最高价    → rt_max
  5   实时市场最低价    → rt_min
  7   日前市场出清均价  → da_avg (companion in RT section — same value as col 9)
  9   日前市场出清均价  → da_avg (actual DA section — use LAST occurrence)
  11  日前市场最高价    → da_max
  13  日前市场最低价    → da_min

Some sheets have an extra "运行情况" column at position 1, shifting all cols by 1.
The keyword-based column detection handles this automatically.

Units: Excel, PDF, and DB all use yuan/kWh (values like 0.2–0.6).
No unit conversion needed — compare directly.

Primary reader: pandas + python-calamine (fast, handles complex Excel, works
               even when the file is open in Excel via temp-file copy)
Primary writer: openpyxl (write-back with blank-cell-only policy)

Public API:
  read_province_sheet(wb_path, province_cn) -> {date: {da_avg, ...}}
  find_excel_row(wb_path, province_cn, date) -> dict | None
  update_excel_row(wb_path, province_cn, date, data)   (no overwrite of existing)
  cross_check(pdf_data, excel_data, db_data) -> list[str]
  get_excel_path(repo_root, year) -> Path
"""
from __future__ import annotations

import datetime as dt
import logging
import shutil
import tempfile
import threading
from pathlib import Path
from typing import Dict, List, Optional

_log = logging.getLogger(__name__)
_excel_write_lock = threading.Lock()


# ── Column header keywords ────────────────────────────────────────────────────
# Row 2 of each sheet contains the detailed column names in Chinese.

_RT_AVG_KW = ["实时市场出清均价", "实时均价"]
_RT_MAX_KW = ["实时市场最高价"]
_RT_MIN_KW = ["实时市场最低价"]
# DA avg appears TWICE in the header row (once as companion in RT section,
# once in the actual DA section). We want the last occurrence.
_DA_AVG_KW = ["日前市场出清均价", "日前均价"]
_DA_MAX_KW = ["日前市场最高价"]
_DA_MIN_KW = ["日前市场最低价"]
_DATE_KW   = ["日期"]

# Fields to find with LAST-match logic (da_avg duplicated in RT section)
_LAST_MATCH_FIELDS = {"da_avg"}

_FIELD_KEYWORDS = {
    "date":   _DATE_KW,
    "rt_avg": _RT_AVG_KW,
    "rt_max": _RT_MAX_KW,
    "rt_min": _RT_MIN_KW,
    "da_avg": _DA_AVG_KW,
    "da_max": _DA_MAX_KW,
    "da_min": _DA_MIN_KW,
}


def _find_col(row: list, keywords: list, last: bool = False) -> Optional[int]:
    """Return 0-based column index matching any keyword. If last=True, return last match."""
    result = None
    for i, cell in enumerate(row):
        cell_str = str(cell or "").strip()
        if any(kw in cell_str for kw in keywords):
            if not last:
                return i
            result = i  # keep scanning for last match
    return result


def _detect_columns(rows: list) -> Optional[tuple]:
    """
    Find the detailed-header row (row with 实时市场出清均价 etc.) and build a
    column map.  Returns (header_row_index, col_map) or None.

    The date column header ('日期') is in the row ABOVE the detailed headers,
    but the date values are always in column 0.  We use column 0 for dates
    regardless of where '日期' appears.
    """
    for row_idx, row in enumerate(rows[:6]):
        row_s = [str(c or "").strip() for c in row]
        # The detailed header row contains at least one of these
        if any(any(kw in cell for kw in _RT_AVG_KW) for cell in row_s):
            col_map: dict[str, int] = {"date": 0}
            for field, keywords in _FIELD_KEYWORDS.items():
                if field == "date":
                    continue
                last = field in _LAST_MATCH_FIELDS
                idx = _find_col(row_s, keywords, last=last)
                if idx is not None:
                    col_map[field] = idx
            return row_idx, col_map
    return None


# ── Type coercion ─────────────────────────────────────────────────────────────

def _coerce_date(val) -> Optional[dt.date]:
    if val is None:
        return None
    if isinstance(val, dt.datetime):
        return val.date()
    if isinstance(val, dt.date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _coerce_float(val) -> Optional[float]:
    if val is None:
        return None
    try:
        f = float(val)
        return None if f != f else f  # NaN → None
    except (TypeError, ValueError):
        return None


# ── Row parsing (shared between calamine / openpyxl paths) ───────────────────

def _parse_sheet_rows(rows: list) -> Dict[dt.date, dict]:
    """
    Convert list-of-lists into {date: {da_avg, da_max, da_min, rt_avg, rt_max, rt_min}}
    with prices in yuan/kWh (same units as PDF and DB).
    """
    result = _detect_columns(rows)
    if result is None:
        return {}
    header_row_idx, col_map = result

    out: Dict[dt.date, dict] = {}
    for row_idx, row in enumerate(rows):
        if row_idx <= header_row_idx:
            continue
        if not row:
            continue
        cell_date = _coerce_date(row[0] if row else None)
        if cell_date is None:
            continue
        prices: dict = {}
        for field in ("da_avg", "da_max", "da_min", "rt_avg", "rt_max", "rt_min"):
            col_idx = col_map.get(field)
            raw = row[col_idx] if (col_idx is not None and col_idx < len(row)) else None
            prices[field] = _coerce_float(raw)
        out[cell_date] = prices
    return out


# ── calamine reader (primary) ─────────────────────────────────────────────────

def _read_calamine(wb_path: Path, sheet_name: str) -> Dict[dt.date, dict]:
    """Read using pandas + python-calamine. Works on files open in Excel."""
    try:
        import pandas as pd  # type: ignore
    except ImportError:
        _log.warning("pandas not available; falling back to openpyxl reader")
        return _read_openpyxl(wb_path, sheet_name)

    # Copy to temp file so we can read even if the file is open in Excel
    tmp = None
    try:
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        shutil.copy2(wb_path, tmp.name)

        try:
            xl = pd.ExcelFile(tmp.name, engine="calamine")
        except Exception:
            # calamine not installed; try openpyxl engine
            try:
                xl = pd.ExcelFile(tmp.name, engine="openpyxl")
            except Exception as e:
                _log.error("pandas ExcelFile failed: %s", e)
                return {}

        sheets = xl.sheet_names
        # Match sheet name — exact first, then partial
        matched = None
        for s in sheets:
            if s == sheet_name:
                matched = s
                break
        if matched is None:
            for s in sheets:
                if sheet_name in s or s in sheet_name:
                    matched = s
                    break
        if matched is None:
            _log.warning("Sheet '%s' not found in %s (available: %s)",
                         sheet_name, wb_path.name, sheets)
            return {}

        df = xl.parse(matched, header=None)
        rows = df.values.tolist()
        return _parse_sheet_rows(rows)

    except Exception as e:
        _log.error("calamine read failed for %s / %s: %s", wb_path.name, sheet_name, e)
        return {}
    finally:
        if tmp:
            try:
                import os
                os.unlink(tmp.name)
            except Exception:
                pass


# ── openpyxl reader / writer ──────────────────────────────────────────────────

def _read_openpyxl(wb_path: Path, sheet_name: str) -> Dict[dt.date, dict]:
    """Fallback reader using openpyxl (may fail on files with complex styles)."""
    try:
        import openpyxl  # type: ignore
        wb = openpyxl.load_workbook(str(wb_path), read_only=True, data_only=True)
        matched = None
        for s in wb.sheetnames:
            if s == sheet_name or sheet_name in s or s in sheet_name:
                matched = s
                break
        if matched is None:
            _log.warning("Sheet '%s' not found in %s", sheet_name, wb_path.name)
            wb.close()
            return {}
        ws = wb[matched]
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        wb.close()
        return _parse_sheet_rows(rows)
    except Exception as e:
        _log.error("openpyxl read failed: %s", e)
        return {}


def _write_openpyxl(
    wb_path: Path,
    sheet_name: str,
    date: dt.date,
    updates_mwh: dict,
) -> None:
    """Write only blank cells for the matching date row (values in yuan/kWh)."""
    with _excel_write_lock:
        try:
            import openpyxl  # type: ignore
            wb = openpyxl.load_workbook(str(wb_path), data_only=False)
        except Exception as e:
            _log.error("openpyxl write: cannot open %s: %s", wb_path.name, e)
            return

        matched = None
        for s in wb.sheetnames:
            if s == sheet_name or sheet_name in s or s in sheet_name:
                matched = s
                break
        if matched is None:
            _log.warning("Write: sheet '%s' not found", sheet_name)
            return

        ws = wb[matched]
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        result = _detect_columns(rows)
        if result is None:
            _log.warning("Write: cannot detect column layout in sheet '%s'", sheet_name)
            return

        header_row_idx, col_map = result

        for row_idx, row in enumerate(rows):
            if row_idx <= header_row_idx:
                continue
            cell_date = _coerce_date(row[0] if row else None)
            if cell_date != date:
                continue
            excel_row = row_idx + 1  # 1-based for openpyxl
            for field in ("da_avg", "da_max", "da_min", "rt_avg", "rt_max", "rt_min"):
                col_idx = col_map.get(field)
                if col_idx is None:
                    continue
                new_val_mwh = updates_mwh.get(field)
                if new_val_mwh is None:
                    continue
                existing = row[col_idx] if col_idx < len(row) else None
                if existing is not None and str(existing).strip() not in ("", "None"):
                    continue  # never overwrite
                ws.cell(row=excel_row, column=col_idx + 1).value = new_val_mwh
            break

        # Write to a temp file first, then atomically replace the original.
        # This ensures wb_path is never left in a partial state if the process
        # is killed mid-save.
        # On Windows with OneDrive, os.replace can fail with WinError 5
        # (access denied) when the target file is momentarily locked. Retry
        # a few times with short sleeps before giving up.
        import os
        import time as _time
        tmp_path = wb_path.with_suffix(".write_tmp.xlsx")
        try:
            wb.save(str(tmp_path))
            last_exc = None
            for _attempt in range(5):
                try:
                    os.replace(str(tmp_path), str(wb_path))
                    last_exc = None
                    break
                except OSError as _e:
                    last_exc = _e
                    _time.sleep(0.5)
            if last_exc is not None:
                # Fallback: copy then delete (works even when os.replace is blocked)
                try:
                    shutil.copy2(str(tmp_path), str(wb_path))
                    last_exc = None
                except Exception as _e2:
                    last_exc = _e2
            if last_exc is not None:
                raise last_exc
            _log.info("Excel updated for %s %s", sheet_name, date)
        except Exception as e:
            _log.error("openpyxl save failed: %s", e)
            try:
                os.unlink(str(tmp_path))
            except Exception:
                pass


# ── Public API ────────────────────────────────────────────────────────────────

def read_province_sheet(wb_path: Path, province_cn: str) -> Dict[dt.date, dict]:
    """
    Read all rows for a province sheet.
    Returns {date: {da_avg, da_max, da_min, rt_avg, rt_max, rt_min}} in yuan/kWh.
    """
    return _read_calamine(wb_path, province_cn)


def find_excel_row(wb_path: Path, province_cn: str, date: dt.date) -> Optional[dict]:
    """Read a single date's price row from Excel. Returns dict in yuan/kWh, or None."""
    return read_province_sheet(wb_path, province_cn).get(date)


def update_excel_row(wb_path: Path, province_cn: str, date: dt.date, data_mwh: dict) -> None:
    """
    Write price data (in yuan/kWh) to the Excel row for (province_cn, date).
    Only fills blank cells — never overwrites existing manual data.
    """
    _write_openpyxl(wb_path, province_cn, date, data_mwh)


def cross_check(
    pdf_data: dict,
    excel_data: Optional[dict],
    db_data: Optional[dict],
    tolerance: float = 0.02,
) -> List[str]:
    """
    Compare prices across three sources (all in yuan/kWh).
    Returns list of discrepancy strings where relative diff > tolerance.
    """
    issues: List[str] = []
    for field in ("da_avg", "da_max", "da_min", "rt_avg", "rt_max", "rt_min"):
        pdf_val = _coerce_float(pdf_data.get(field))
        xl_val  = _coerce_float((excel_data or {}).get(field))
        db_val  = _coerce_float((db_data or {}).get(field))

        ref = abs(xl_val or pdf_val or 1.0) or 1.0

        if xl_val is not None and pdf_val is not None:
            if abs(xl_val - pdf_val) / ref > tolerance:
                issues.append(
                    f"{field}: PDF={pdf_val:.1f} vs Excel={xl_val:.1f} "
                    f"({abs(xl_val - pdf_val) / ref * 100:.1f}% diff)"
                )

        if xl_val is not None and db_val is not None:
            if abs(xl_val - db_val) / ref > tolerance:
                issues.append(
                    f"{field}: DB={db_val:.1f} vs Excel={xl_val:.1f} "
                    f"({abs(xl_val - db_val) / ref * 100:.1f}% diff)"
                )

    return issues


def get_excel_path(repo_root: Path, year: int = 2026) -> Path:
    """Return the path to the spot reports Excel summary file."""
    base = repo_root / "data" / "spot reports" / str(year)
    # Try the known filename first, then fall back to any matching xlsx
    known = base / "电力现货市场价格与运行日报数据汇总-apr26.xlsx"
    if known.exists():
        return known
    # Fallback: first xlsx containing 数据汇总 in the name
    candidates = sorted(base.glob("*数据汇总*.xlsx"))
    if candidates:
        _log.info("Excel path resolved via glob: %s", candidates[0].name)
        return candidates[0]
    return known  # return the expected path even if missing (caller will warn)
