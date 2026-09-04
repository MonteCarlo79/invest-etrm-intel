"""Tests for dispatch_ingest parsers (nominations + dispatch chain)."""
import datetime
import pytest

from services.dispatch_ingest.nominations import (
    _find_columns as _nom_cols, parse_nomination_sheet, NOMINATION_FOLDER_TO_ASSET,
)
from services.dispatch_ingest.dispatch_chain import (
    _find_columns as _dc_cols, restriction_from_fill, DISPATCH_FOLDER_TO_ASSET,
)


class _FakeFill:
    def __init__(self, rgb):
        self.fill_type = "solid"
        self.start_color = type("C", (), {"rgb": rgb})()


class _FakeCell:
    def __init__(self, value, fill=None):
        self.value = value
        self.fill = fill


class _FakeSheet:
    def __init__(self, header, data_rows):
        self._rows = [header] + data_rows

    def iter_rows(self, values_only=False, min_row=1, max_col=None):
        if values_only:
            for r in self._rows[min_row - 1:]:
                yield r
        else:
            for r in self._rows[min_row - 1:]:
                yield [_FakeCell(v) if not isinstance(v, _FakeCell) else v for v in r]


# --- nominations ---

def test_nom_columns_suyou_layout():
    cols = _nom_cols(["日期", "时刻", "预计划功率", "正式申报"])
    assert cols == {"date": 0, "time": 1, "planned": 2, "nominated": 3}


def test_nom_columns_gushanliang_layout():
    cols = _nom_cols(["日期", "时刻", "预策略D-2功率（MW）", "正式策略D-1功率（MW）"])
    assert cols == {"date": 0, "time": 1, "planned": 2, "nominated": 3}


def test_nom_columns_bameng_layout():
    cols = _nom_cols(["日期", "时刻", "预申报策略（MW）", "实际申报策略（MW）"])
    assert cols == {"date": 0, "time": 1, "planned": 2, "nominated": 3}


def test_nom_columns_rejects_other():
    assert _nom_cols(["时间", "SOC（%）", "交易员申报计划 (MW)"]) is None


def test_nom_columns_single_power_variant():
    # April 苏右 layout: one 申报功率 column, no planned/nominated split
    cols = _nom_cols(["日期", "时刻", "申报功率（MW）", "爬坡校验"])
    assert cols == {"date": 0, "time": 1, "nominated": 2}


def test_parse_nomination_file_template_sheet_loses(tmp_path):
    """A zeroed 输出模板 sheet with the same header must not overwrite 策略申报."""
    import openpyxl
    from services.dispatch_ingest.nominations import parse_nomination_file

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "策略申报"
    ws1.append(["日期", "时刻", "申报功率（MW）", "爬坡校验"])
    ws1.append(["2026-04-22", "00:05:00", 16, 16])
    ws1.append(["2026-04-22", "00:10:00", 32, 16])
    ws2 = wb.create_sheet("输出模板")
    ws2.append(["日期", "时刻", "申报功率（MW）", "D-H列输出时请删除"])
    ws2.append(["2026-04-22", "00:05:00", 0, ""])
    ws2.append(["2026-04-22", "00:10:00", 0, ""])
    path = tmp_path / "nom.xlsx"
    wb.save(path)

    items = parse_nomination_file(str(path))
    assert len(items) == 2
    assert [it["nominated_mw"] for it in items] == [16.0, 32.0]



def test_parse_nomination_sheet_suyou():
    ws = _FakeSheet(
        ["日期", "时刻", "预计划功率", "正式申报"],
        [["2026-07-01", "00:05:00", 16, 16],
         ["2026-07-01", "00:10:00", -32, -32]],
    )
    items = parse_nomination_sheet(ws)
    assert len(items) == 2
    assert items[0]["nominated_mw"] == 16.0
    assert items[1]["nominated_mw"] == -32.0
    assert items[0]["interval_start"].year == 2026
    assert items[0]["interval_start"].minute == 5


# --- dispatch chain ---

def test_dc_columns():
    cols = _dc_cols(["时间", "SOC（%）", "交易员申报计划 (MW)", "日前出清 (MW)", "实时调度出清 (MW)", "实际执行功率 (MW)"])
    assert cols == {"time": 0, "soc": 1, "nominated": 2, "da_cleared": 3, "rt_cleared": 4, "actual": 5}


def test_restriction_color_map():
    # user-confirmed semantics 2026-08-28: green/none=NULL, orange=charge_only, red=discharge_only
    assert restriction_from_fill(_FakeFill("FFFF0000")) == "discharge_only"
    assert restriction_from_fill(_FakeFill("FF00B050")) is None
    assert restriction_from_fill(_FakeFill("FFFFA500")) == "charge_only"
    assert restriction_from_fill(_FakeFill("FF629BB7")) is None  # blue-grey = no restriction


def test_folder_maps_cover_known_stations():
    assert NOMINATION_FOLDER_TO_ASSET["01-苏右"] == "景蓝乌尔图"
    assert NOMINATION_FOLDER_TO_ASSET["07-巴盟"] == "景怡查干哈达"
    assert DISPATCH_FOLDER_TO_ASSET["4.四子王旗-景通四益堂"] == "四子王旗"
    assert DISPATCH_FOLDER_TO_ASSET["2.谷山梁-裕昭沙子坝"] == "裕昭沙子坝"
